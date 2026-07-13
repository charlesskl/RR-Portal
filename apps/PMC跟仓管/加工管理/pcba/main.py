import os
import io
import re
from datetime import date, datetime
from typing import List, Optional
from urllib.parse import quote

from fastapi import FastAPI, HTTPException, Request, Depends, File, Form, UploadFile
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side
from openpyxl.utils.datetime import from_excel
from starlette.middleware.sessions import SessionMiddleware
from pydantic import BaseModel

from pcba import db
from pcba.auth import hash_password, verify_password
from pcba.summary import (
    compute_material_totals,
    compute_public_summary,
    compute_sticker_type_totals,
    compute_summary,
)
from pcba.db import DEPARTMENTS, LOCATIONS
from pcba.export import build_workbook
from pcba.shaoyang_reconcile import (
    build_shaoyang_cd_reconcile,
    build_shaoyang_issue_export_workbook,
)

app = FastAPI(title="唱片机管理系统")
BASE_PATH = os.environ.get("PCBA_BASE_PATH", "").rstrip("/")
app.add_middleware(
    SessionMiddleware,
    secret_key=os.environ.get("PCBA_SECRET", "pcba-local-secret-change-me"),
    session_cookie=os.environ.get("PCBA_COOKIE_NAME", "session"),
    path=os.environ.get("PCBA_COOKIE_PATH", "/"),
    https_only=False,
)

STATIC_DIR = os.path.join(os.path.dirname(__file__), "static")


def _render_html(filename: str):
    with open(os.path.join(STATIC_DIR, filename), encoding="utf-8") as f:
        html = f.read()
    inject = f'<script>window.APP_BASE="{BASE_PATH}";</script>'
    html = html.replace("<head>", "<head>\n" + inject, 1)
    if BASE_PATH:
        html = html.replace('"/static/', f'"{BASE_PATH}/static/')
        html = html.replace('href="/"', f'href="{BASE_PATH}/"')
    return HTMLResponse(html)


@app.middleware("http")
async def _no_cache_static(request: Request, call_next):
    response = await call_next(request)
    if request.url.path.startswith("/static") or request.url.path == "/":
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response


@app.on_event("startup")
def _startup():
    db.init_db()


@app.get("/api/health")
def health():
    return {"status": "ok"}


# ---------- 鉴权辅助 ----------
def current_user(request: Request):
    uid = request.session.get("uid")
    department = request.session.get("department")
    if not uid or not department:
        raise HTTPException(status_code=401, detail="未登录")
    conn = db.get_conn()
    try:
        row = conn.execute(
            "SELECT id, username, role, department FROM users WHERE id=?", (uid,)
        ).fetchone()
    finally:
        conn.close()
    if not row:
        raise HTTPException(status_code=401, detail="未登录")
    if department not in DEPARTMENTS:
        raise HTTPException(status_code=401, detail="未登录")
    user = dict(row)
    if user["role"] == "admin":
        user["department"] = department
        return user
    db_department = user.get("department")
    if db_department not in DEPARTMENTS:
        raise HTTPException(status_code=403, detail="账号未分配有效部门")
    user["department"] = db_department
    if department != db_department:
        request.session["department"] = db_department
    return user


def require_admin(user=Depends(current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    return user


# ---------- 登录 ----------
class LoginIn(BaseModel):
    username: str
    password: str
    department: str


class DepartmentSwitchIn(BaseModel):
    department: str


def _validate_department(department: Optional[str], required=True):
    if not department:
        if required:
            raise HTTPException(status_code=400, detail="请选择部门")
        return None
    if department not in DEPARTMENTS:
        raise HTTPException(status_code=400, detail="部门无效")
    return department


def _validate_date(value: Optional[str], field: str):
    if not value:
        return None
    try:
        date.fromisoformat(value)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"{field}格式必须为 YYYY-MM-DD")
    return value


def _date_filter(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    doc_no: Optional[str] = None,
):
    start = _validate_date(date_from, "date_from")
    end = _validate_date(date_to, "date_to")
    if start and end and start > end:
        raise HTTPException(status_code=400, detail="开始日期不能晚于结束日期")
    filters = {"date_from": start, "date_to": end}
    doc_no = _clean_optional(doc_no)
    if doc_no:
        filters["doc_no"] = doc_no
    return filters


def _append_date_filter(sql: str, params: list, filters: dict):
    if filters.get("date_from"):
        sql += " AND r.rec_date >= ?"
        params.append(filters["date_from"])
    if filters.get("date_to"):
        sql += " AND r.rec_date <= ?"
        params.append(filters["date_to"])
    if filters.get("doc_no"):
        sql += " AND COALESCE(r.doc_no, '') LIKE ?"
        params.append(f"%{filters['doc_no']}%")
    return sql, params


def _clean_optional(value: Optional[str]):
    return value.strip() if value and value.strip() else None


@app.get("/api/departments")
def list_departments():
    return DEPARTMENTS


@app.post("/api/login")
def login(body: LoginIn, request: Request):
    department = _validate_department(body.department)
    conn = db.get_conn()
    try:
        row = conn.execute(
            "SELECT id, username, role, password_hash, department "
            "FROM users WHERE username=?",
            (body.username,),
        ).fetchone()
    finally:
        conn.close()
    if not row or not verify_password(body.password, row["password_hash"]):
        raise HTTPException(status_code=401, detail="账号或密码错误")
    if row["role"] != "admin" and row["department"] != department:
        raise HTTPException(status_code=403, detail="账号不属于所选部门")
    request.session["uid"] = row["id"]
    request.session["department"] = department
    return {"username": row["username"], "role": row["role"], "department": department}


@app.post("/api/logout")
def logout(request: Request):
    request.session.clear()
    return {"ok": True}


@app.get("/api/me")
def me(user=Depends(current_user)):
    return user


@app.post("/api/me/department")
def switch_current_department(
    body: DepartmentSwitchIn,
    request: Request,
    user=Depends(require_admin),
):
    department = _validate_department(body.department)
    request.session["department"] = department
    result = dict(user)
    result["department"] = department
    return result


# ---------- 用户管理（仅 admin） ----------
class UserIn(BaseModel):
    username: str
    password: str
    role: str = "operator"
    department: Optional[str] = None


class UserUpdateIn(BaseModel):
    role: str = "operator"
    department: Optional[str] = None


@app.get("/api/users")
def list_users(_admin=Depends(require_admin)):
    conn = db.get_conn()
    try:
        rows = conn.execute(
            "SELECT id, username, role, department, created_at FROM users ORDER BY id"
        ).fetchall()
    finally:
        conn.close()
    return [dict(r) for r in rows]


@app.post("/api/users")
def create_user(body: UserIn, _admin=Depends(require_admin)):
    if body.role not in ("admin", "operator"):
        raise HTTPException(status_code=400, detail="角色无效")
    if len(body.password) < 6:
        raise HTTPException(status_code=400, detail="密码至少 6 位")
    department = _validate_department(
        body.department, required=(body.role == "operator")
    )
    conn = db.get_conn()
    try:
        exists = conn.execute(
            "SELECT 1 FROM users WHERE username=?", (body.username,)
        ).fetchone()
        if exists:
            raise HTTPException(status_code=400, detail="账号已存在")
        conn.execute(
            "INSERT INTO users(username, password_hash, role, department) VALUES (?,?,?,?)",
            (body.username, hash_password(body.password), body.role, department),
        )
        conn.commit()
    finally:
        conn.close()
    return {"ok": True}


class PasswordIn(BaseModel):
    password: str


@app.put("/api/me/password")
def change_my_password(body: PasswordIn, user=Depends(current_user)):
    if len(body.password) < 6:
        raise HTTPException(status_code=400, detail="密码至少 6 位")
    conn = db.get_conn()
    try:
        conn.execute(
            "UPDATE users SET password_hash=? WHERE id=?",
            (hash_password(body.password), user["id"]),
        )
        conn.commit()
    finally:
        conn.close()
    return {"ok": True}


@app.put("/api/users/{user_id}")
def update_user(user_id: int, body: UserUpdateIn, _admin=Depends(require_admin)):
    if body.role not in ("admin", "operator"):
        raise HTTPException(status_code=400, detail="角色无效")
    department = _validate_department(
        body.department, required=(body.role == "operator")
    )
    if body.role == "admin":
        department = None
    conn = db.get_conn()
    try:
        row = conn.execute("SELECT id FROM users WHERE id=?", (user_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="用户不存在")
        conn.execute(
            "UPDATE users SET role=?, department=? WHERE id=?",
            (body.role, department, user_id),
        )
        conn.commit()
    finally:
        conn.close()
    return {"ok": True}


@app.put("/api/users/{user_id}/password")
def reset_password(user_id: int, body: PasswordIn, _admin=Depends(require_admin)):
    if len(body.password) < 6:
        raise HTTPException(status_code=400, detail="密码至少 6 位")
    conn = db.get_conn()
    try:
        conn.execute(
            "UPDATE users SET password_hash=? WHERE id=?",
            (hash_password(body.password), user_id),
        )
        conn.commit()
    finally:
        conn.close()
    return {"ok": True}


# ---------- 静态前端 ----------
@app.get("/")
def index():
    return _render_html("login.html")


@app.get("/static/app.html")
def app_page():
    return _render_html("app.html")


@app.get("/static/public-summary.html")
def public_summary_page():
    return _render_html("public-summary.html")


VALID_TYPES = (
    "inbound_raw", "issue", "finished", "semi_finished",
    "semi_inbound", "semi_outbound",
)
SUPPLIER_DEPARTMENT = "兴信B来料仓"
ASSEMBLY_DEPARTMENT = "东莞车间"
SEMI_FINISHED_DEPARTMENT = "碟片半成品"
SEMI_FINISHED_FILENAME_KEYWORD = "半成品"
OUTSOURCE_DEPARTMENT = "东莞加工厂利鸿"
HONGYA_DEPARTMENT = "东莞加工厂鸿亚"
OUTSOURCE_DEPARTMENTS = (OUTSOURCE_DEPARTMENT, HONGYA_DEPARTMENT)
HEYUAN_DEPARTMENT = "河源华兴"
SHAOYANG_DEPARTMENT = "邵阳华登"
XINSHAO_DEPARTMENT = "新邵"
PO_CUSTOMER_DEPARTMENTS = (SHAOYANG_DEPARTMENT, XINSHAO_DEPARTMENT)
PROCESSING_BALANCE_DEPARTMENTS = (
    HEYUAN_DEPARTMENT,
    SHAOYANG_DEPARTMENT,
    XINSHAO_DEPARTMENT,
)
NFC_MATERIAL = "NFC贴纸"
PCBA_MATERIAL = "77794-PCBA板"
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
RECORD_IMPORT_HEADERS = [
    "类型", "物料名称", "贴纸类型", "加工点", "供应商",
    "日期", "单据编号", "数量", "备注", "PO", "客名",
]


def _is_outsource_department(department):
    return department in OUTSOURCE_DEPARTMENTS


def _xlsx_response(wb: Workbook, filename: str):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


def _cell_text(value):
    if value is None:
        return ""
    return str(value).replace("\r", "").replace("\n", "").strip()


def _load_upload_workbook(file: UploadFile):
    try:
        return load_workbook(file.file, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Excel 文件无法读取")


def _require_filename_contains(file: UploadFile, required_text: str):
    filename = _cell_text(getattr(file, "filename", ""))
    if required_text not in filename:
        raise HTTPException(
            status_code=400,
            detail=f"导入文件名必须包含{required_text}",
        )


def _worksheet_rows(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_cell_text(value) for value in rows[0]]
    result = []
    for offset, values in enumerate(rows[1:], start=2):
        if not any(_cell_text(value) for value in values):
            continue
        row = {}
        for i, header in enumerate(headers):
            if header:
                row[header] = values[i] if i < len(values) else None
        result.append((offset, row))
    return result


def _sheet_rows(file: UploadFile):
    return _worksheet_rows(_load_upload_workbook(file).active)


def _first_value(row: dict, *headers):
    for header in headers:
        value = row.get(header)
        text = _cell_text(value)
        if text:
            return text
    return ""


def _int_value(value, row_no: int, field: str):
    if value is None or _cell_text(value) == "":
        raise HTTPException(status_code=400, detail=f"第{row_no}行：{field}不能为空")
    try:
        return int(value)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail=f"第{row_no}行：{field}必须是数字")


def _date_value(value, row_no: int):
    if value is None or _cell_text(value) == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _cell_text(value)
    try:
        date.fromisoformat(text)
    except ValueError:
        raise HTTPException(status_code=400, detail=f"第{row_no}行：日期格式必须为 YYYY-MM-DD")
    return text


def _record_type_from_import(value, department):
    text = _cell_text(value)
    if text in VALID_TYPES:
        return text
    department_maps = {
        SUPPLIER_DEPARTMENT: {"入库": "inbound_raw", "来料入库": "inbound_raw", "出库": "issue", "领料": "issue"},
        SEMI_FINISHED_DEPARTMENT: {"入库": "semi_inbound", "半成品入库": "semi_inbound", "出库": "semi_outbound", "半成品出库": "semi_outbound"},
        ASSEMBLY_DEPARTMENT: {"领料": "issue", "成品入库": "finished", "半成品入库": "semi_finished"},
        OUTSOURCE_DEPARTMENT: {"领料": "issue", "成品入库": "finished", "半成品入库": "semi_finished"},
        HONGYA_DEPARTMENT: {"领料": "issue", "成品入库": "finished", "半成品入库": "semi_finished"},
        HEYUAN_DEPARTMENT: {"领料": "issue", "成品入库": "finished"},
        SHAOYANG_DEPARTMENT: {"领料": "issue", "成品入库": "finished"},
        XINSHAO_DEPARTMENT: {"领料": "issue", "成品入库": "finished"},
    }
    rec_type = department_maps.get(department, {}).get(text)
    if rec_type:
        return rec_type
    fallback = {
        "来料入库": "inbound_raw",
        "领料": "issue",
        "出库": "issue",
        "成品入库": "finished",
        "半成品入库": "semi_finished",
    }.get(text)
    if fallback:
        return fallback
    raise HTTPException(status_code=400, detail=f"类型无效：{text or '空'}")


def _location_id_from_name(conn, name, row_no: int):
    name = _cell_text(name)
    if not name:
        return None
    row = conn.execute("SELECT id FROM locations WHERE name=?", (name,)).fetchone()
    if not row:
        raise HTTPException(status_code=400, detail=f"第{row_no}行：加工点不存在")
    return row["id"]


def _simple_workbook(headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    return wb


def _read_name_import(file: UploadFile, *headers):
    names = []
    for row_no, row in _sheet_rows(file):
        name = _first_value(row, *headers)
        if not name:
            raise HTTPException(status_code=400, detail=f"第{row_no}行：名称不能为空")
        names.append(name)
    return names


def _insert_unique_names(conn, table, names):
    imported = 0
    skipped = 0
    for name in names:
        exists = conn.execute(f"SELECT 1 FROM {table} WHERE name=?", (name,)).fetchone()
        if exists:
            skipped += 1
            continue
        conn.execute(f"INSERT INTO {table}(name) VALUES (?)", (name,))
        imported += 1
    return {"imported": imported, "skipped": skipped}


def _legacy_excel_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date().isoformat()
        except Exception:
            return None
    text = _cell_text(value)
    if not text or text == "日期":
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def _legacy_cutoff_date(value, default_year=None):
    rec_date = _legacy_excel_date(value)
    if rec_date:
        return rec_date
    text = _cell_text(value)
    if not text:
        return None
    match = re.search(
        r"(?:(19\d{2}|20\d{2})\s*[年/-])?(\d{1,2})\s*(?:月|/|-)\s*(\d{1,2})\s*(?:日|号)?",
        text,
    )
    if not match:
        return None
    year = int(match.group(1) or default_year or date.today().year)
    month = int(match.group(2))
    day = int(match.group(3))
    try:
        return date(year, month, day).isoformat()
    except ValueError:
        return None


def _legacy_workbook_year(wb):
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                rec_date = _legacy_excel_date(cell.value)
                if rec_date:
                    year = date.fromisoformat(rec_date).year
                    if 2000 <= year <= 2100:
                        return year
    return date.today().year


def _find_legacy_item_header_row(ws):
    for row_no in range(1, ws.max_row + 1):
        if _cell_text(ws.cell(row_no, 1).value) == "物料名称":
            return row_no
    return None


def _legacy_monthly_column(ws, header_row, expected_header):
    if not header_row:
        return None
    for col_no in range(1, ws.max_column + 1):
        if _cell_text(ws.cell(header_row, col_no).value) == expected_header:
            return col_no
    return None


def _legacy_opening_stock_columns(ws, header_row):
    if not header_row:
        return []
    columns = []
    for col_no in range(1, ws.max_column + 1):
        text = _cell_text(ws.cell(header_row, col_no).value)
        if "盘点截数" in text or "入库截数" in text:
            columns.append(col_no)
    return columns


def _legacy_detail_columns(ws):
    columns = []
    for col_no in range(1, ws.max_column + 1):
        rec_date = _legacy_excel_date(ws.cell(1, col_no).value)
        doc_no = _cell_text(ws.cell(2, col_no).value)
        if rec_date and doc_no:
            columns.append((col_no, rec_date, doc_no))
    return columns


def _legacy_matrix_detail_columns(ws, default_year=None):
    columns = []
    for col_no in range(1, ws.max_column + 1):
        rec_date = _legacy_cutoff_date(ws.cell(1, col_no).value, default_year)
        doc_no = _cell_text(ws.cell(2, col_no).value)
        if rec_date and doc_no:
            columns.append((col_no, rec_date, doc_no))
    return columns


def _legacy_row_headers(ws, row_no=1):
    return {
        _cell_text(ws.cell(row_no, col_no).value)
        for col_no in range(1, ws.max_column + 1)
        if _cell_text(ws.cell(row_no, col_no).value)
    }


def _legacy_sheet_has_headers(ws, *headers):
    actual_headers = _legacy_row_headers(ws)
    return all(header in actual_headers for header in headers)


def _legacy_sheet_has_top_header(ws, header):
    for col_no in range(1, ws.max_column + 1):
        if _cell_text(ws.cell(1, col_no).value) == header:
            return True
    return False


def _is_legacy_semi_finished_workbook(wb):
    for ws in wb.worksheets:
        header_row = _find_legacy_item_header_row(ws)
        if not header_row:
            continue
        headers = {
            _cell_text(ws.cell(header_row, col_no).value)
            for col_no in range(1, ws.max_column + 1)
        }
        if "当月入仓总数" in headers or "当月出仓总数" in headers:
            return True
    return False


def _is_legacy_outsource_workbook(wb):
    sheet_names = set(wb.sheetnames)
    if "领料明细" not in sheet_names or "半成品入仓明细" not in sheet_names:
        return False
    return _legacy_sheet_has_headers(
        wb["领料明细"], "日期", "领料编号", "物料名称", "领料数"
    ) and _legacy_sheet_has_headers(
        wb["半成品入仓明细"], "日期", "送货单号", "品名/规格"
    )


def _is_legacy_outsource_nfc_workbook(wb):
    sheet_names = set(wb.sheetnames)
    if not {"总表", "领料明细", "入仓明细"}.issubset(sheet_names):
        return False
    return bool(
        _find_legacy_item_header_row(wb["领料明细"])
        and _find_legacy_item_header_row(wb["入仓明细"])
    )


def _is_legacy_assembly_workbook(wb):
    sheet_names = set(wb.sheetnames)
    if "领料明细" not in sheet_names or "半成品入仓明细" not in sheet_names:
        return False
    issue_ws = wb["领料明细"]
    semi_ws = wb["半成品入仓明细"]
    return bool(
        _find_legacy_item_header_row(issue_ws)
        and _find_legacy_item_header_row(semi_ws)
        and _legacy_sheet_has_top_header(issue_ws, "当月领料总数")
        and _legacy_sheet_has_top_header(semi_ws, "当月入仓总数")
    )


def _is_legacy_heyuan_workbook(wb):
    return (
        "成品入仓明细" in wb.sheetnames
        and any("领料明细" in name for name in wb.sheetnames)
    )


def _is_legacy_supplier_workbook(wb):
    sheet_names = set(wb.sheetnames)
    if "入仓明细" in sheet_names:
        return True
    return {"总表", "入库明细", "出库明细"}.issubset(sheet_names)


def _legacy_record_type(sheet_name):
    if "入库" in sheet_name:
        return "semi_inbound", "monthly_inbound", "当月入仓总数"
    if "领料" in sheet_name or "出库" in sheet_name:
        return "semi_outbound", "monthly_outbound", "当月出仓总数"
    return None, None, None


def _legacy_int(value, row_no, field, allow_blank=True):
    if value is None or _cell_text(value) == "":
        return None if allow_blank else _int_value(value, row_no, field)
    try:
        return int(value)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail=f"第{row_no}行：{field}必须是数字")


def _parse_legacy_semi_finished_workbook(conn, wb, department):
    if department != SEMI_FINISHED_DEPARTMENT:
        raise HTTPException(status_code=400, detail="半成品台账只能在碟片半成品部门导入")

    bodies = []
    monthly_totals = {}
    for ws in wb.worksheets:
        if ws.title == "总表":
            continue
        rec_type, monthly_key, monthly_header = _legacy_record_type(ws.title)
        if not rec_type:
            continue
        header_row = _find_legacy_item_header_row(ws)
        if not header_row:
            continue

        location_id = None
        if rec_type == "semi_outbound":
            location_id = _location_id_from_name(
                conn, _legacy_location_from_sheet(ws.title), 1
            )
        detail_columns = _legacy_detail_columns(ws)
        monthly_col = _legacy_monthly_column(ws, header_row, monthly_header)
        opening_stock_columns = _legacy_opening_stock_columns(ws, header_row)
        for row_no in range(header_row + 1, ws.max_row + 1):
            sticker_type = _cell_text(ws.cell(row_no, 1).value)
            if not sticker_type:
                continue
            if "#NFC" not in sticker_type:
                continue
            sticker_type = _normalize_sticker_type(conn, NFC_MATERIAL, sticker_type)
            key = (NFC_MATERIAL, sticker_type)

            if monthly_col:
                monthly_qty = _legacy_int(
                    ws.cell(row_no, monthly_col).value,
                    row_no,
                    monthly_header,
                )
                if monthly_qty is not None:
                    monthly = monthly_totals.setdefault(
                        key,
                        {
                            "material": NFC_MATERIAL,
                            "sticker_type": sticker_type,
                            "opening_stock": 0,
                            "monthly_inbound": 0,
                            "monthly_outbound": 0,
                        },
                    )
                    monthly[monthly_key] += monthly_qty

            opening_stock_candidate = 0
            has_opening_stock = False
            for opening_col in opening_stock_columns:
                opening_qty = _legacy_int(
                    ws.cell(row_no, opening_col).value,
                    row_no,
                    "上月库存数",
                )
                if opening_qty is not None:
                    has_opening_stock = True
                    opening_stock_candidate += opening_qty
            if has_opening_stock:
                monthly = monthly_totals.setdefault(
                    key,
                    {
                        "material": NFC_MATERIAL,
                        "sticker_type": sticker_type,
                        "opening_stock": 0,
                        "monthly_inbound": 0,
                        "monthly_outbound": 0,
                    },
                )
                monthly["opening_stock"] = max(
                    monthly["opening_stock"],
                    opening_stock_candidate,
                )

            for col_no, rec_date, doc_no in detail_columns:
                qty = _legacy_int(ws.cell(row_no, col_no).value, row_no, "数量")
                if qty is None or qty <= 0:
                    continue
                body = RecordIn(
                    rec_type=rec_type,
                    rec_date=rec_date,
                    doc_no=doc_no,
                    material=NFC_MATERIAL,
                    sticker_type=sticker_type,
                    location_id=location_id,
                    qty=qty,
                    remark=f"{ws.title}导入",
                )
                _validate_record(body, department)
                bodies.append(body)

    if not bodies and not monthly_totals:
        raise HTTPException(status_code=400, detail="半成品台账没有可导入的数据")
    return bodies, list(monthly_totals.values())


def _parse_legacy_assembly_workbook(conn, wb, department):
    if department != ASSEMBLY_DEPARTMENT:
        raise HTTPException(status_code=400, detail="东莞车间台账只能在东莞车间部门导入")

    bodies = []
    location_id = _location_id_from_name(conn, "东莞车间", 1)
    workbook_year = _legacy_workbook_year(wb)
    sheet_types = {
        "领料明细": "issue",
        "半成品入仓明细": "semi_finished",
    }

    for sheet_name, rec_type in sheet_types.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header_row = _find_legacy_item_header_row(ws)
        if not header_row:
            continue
        detail_columns = _legacy_matrix_detail_columns(ws, workbook_year)
        for row_no in range(header_row + 1, ws.max_row + 1):
            sticker_type = _cell_text(ws.cell(row_no, 1).value)
            if not sticker_type or "#NFC" not in sticker_type:
                continue
            sticker_type = _normalize_sticker_type(conn, NFC_MATERIAL, sticker_type)
            for col_no, rec_date, doc_no in detail_columns:
                qty = _legacy_int(ws.cell(row_no, col_no).value, row_no, "数量")
                if qty is None or qty <= 0:
                    continue
                body = RecordIn(
                    rec_type=rec_type,
                    location_id=location_id,
                    rec_date=rec_date,
                    doc_no=doc_no,
                    material=NFC_MATERIAL,
                    sticker_type=sticker_type,
                    qty=qty,
                    remark=f"{ws.title}导入",
                )
                _validate_record(body, department)
                bodies.append(body)

    if not bodies:
        raise HTTPException(status_code=400, detail="东莞车间台账没有可导入的数据")
    return bodies


def _legacy_header_map(ws):
    headers = {}
    for col_no in range(1, ws.max_column + 1):
        header = _cell_text(ws.cell(1, col_no).value)
        if header:
            headers[header] = col_no
    return headers


def _legacy_header_value(ws, headers, row_no, *names):
    for name in names:
        col_no = headers.get(name)
        if col_no:
            return ws.cell(row_no, col_no).value
    return None


def _normalize_pcba_material(value=None):
    text = _cell_text(value)
    if not text or "PCBA" in text or "PCB" in text or "主板" in text:
        return PCBA_MATERIAL
    return text


def _outsource_date_value(value):
    rec_date = _legacy_excel_date(value)
    return rec_date


def _legacy_doc_no(value, date_value, fallback):
    doc_no = _cell_text(value)
    if doc_no:
        return doc_no
    return _cell_text(date_value) or fallback


def _legacy_location_from_sheet(sheet_name):
    if "河源" in sheet_name:
        return HEYUAN_DEPARTMENT
    if "邵阳" in sheet_name:
        return "邵阳华登"
    if "新邵" in sheet_name:
        return XINSHAO_DEPARTMENT
    if "鸿亚" in sheet_name:
        return HONGYA_DEPARTMENT
    if "利鸿" in sheet_name or "加工厂" in sheet_name:
        return "东莞加工厂利鸿"
    if "东莞" in sheet_name or "车间" in sheet_name:
        return "东莞车间"
    return "东莞车间"


def _add_record_body(bodies, body, department, validate_positive=True):
    if body.qty > 0 or validate_positive:
        _validate_record(body, department)
    bodies.append(body)


def _parse_legacy_outsource_workbook(conn, wb, department):
    if not _is_outsource_department(department):
        raise HTTPException(status_code=400, detail="东莞加工厂台账只能在对应加工厂部门导入")

    bodies = []
    issue_ws = wb["领料明细"] if "领料明细" in wb.sheetnames else None
    if issue_ws:
        headers = _legacy_header_map(issue_ws)
        for row_no in range(2, issue_ws.max_row + 1):
            material = _cell_text(_legacy_header_value(issue_ws, headers, row_no, "物料名称"))
            if not material or "小计" in material:
                continue
            qty = _legacy_int(
                _legacy_header_value(issue_ws, headers, row_no, "领料数"),
                row_no,
                "领料数",
            )
            if qty is None or qty == 0:
                continue
            date_value = _legacy_header_value(issue_ws, headers, row_no, "日期")
            doc_no = _cell_text(_legacy_header_value(issue_ws, headers, row_no, "领料编号"))
            if not doc_no:
                doc_no = _cell_text(date_value)
            body = RecordIn(
                rec_type="issue",
                location_id=_location_id_from_name(conn, department, 1),
                rec_date=_outsource_date_value(date_value),
                doc_no=doc_no,
                material=_normalize_pcba_material(material),
                qty=qty,
                remark=_first_value(
                    {"备注": _legacy_header_value(issue_ws, headers, row_no, "备注")},
                    "备注",
                ) or "领料明细导入",
            )
            _validate_record(body, department)
            bodies.append(body)

    inbound_ws = wb["半成品入仓明细"] if "半成品入仓明细" in wb.sheetnames else None
    if inbound_ws:
        headers = _legacy_header_map(inbound_ws)
        for row_no in range(2, inbound_ws.max_row + 1):
            name = _cell_text(_legacy_header_value(inbound_ws, headers, row_no, "品名/规格"))
            if not name or "小计" in name:
                continue
            qty = _legacy_int(
                _legacy_header_value(inbound_ws, headers, row_no, "数量（pcs）", "数量"),
                row_no,
                "数量",
            )
            if qty is None or qty == 0:
                continue
            contract = _cell_text(_legacy_header_value(inbound_ws, headers, row_no, "合同号"))
            item_no = _cell_text(_legacy_header_value(inbound_ws, headers, row_no, "货号"))
            body = RecordIn(
                rec_type="semi_finished",
                rec_date=_outsource_date_value(
                    _legacy_header_value(inbound_ws, headers, row_no, "日期")
                ),
                doc_no=_cell_text(
                    _legacy_header_value(inbound_ws, headers, row_no, "送货单号")
                ),
                material=PCBA_MATERIAL,
                qty=qty,
                remark=" ".join(
                    part for part in [contract, item_no, name, "半成品入仓明细导入"] if part
                ),
            )
            _validate_record(body, department)
            bodies.append(body)

    if not bodies:
        raise HTTPException(status_code=400, detail=f"{department}台账没有可导入的数据")
    return bodies


def _legacy_outsource_nfc_detail_columns(ws, default_year=None):
    header_row = _find_legacy_item_header_row(ws)
    if not header_row:
        return []
    columns = []
    for col_no in range(1, ws.max_column + 1):
        rec_date = _legacy_cutoff_date(ws.cell(1, col_no).value, default_year)
        if not rec_date:
            continue
        doc_no = _cell_text(ws.cell(header_row, col_no).value) or rec_date
        columns.append((col_no, rec_date, doc_no))
    return columns


def _parse_legacy_outsource_nfc_workbook(conn, wb, department):
    if not _is_outsource_department(department):
        raise HTTPException(status_code=400, detail="加工厂贴纸台账只能在对应加工厂部门导入")

    bodies = []
    workbook_year = _legacy_workbook_year(wb)
    sheet_types = {
        "领料明细": "issue",
        "入仓明细": "finished",
    }
    issue_location_id = _location_id_from_name(conn, department, 1)
    for sheet_name, rec_type in sheet_types.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header_row = _find_legacy_item_header_row(ws)
        if not header_row:
            continue
        detail_columns = _legacy_outsource_nfc_detail_columns(ws, workbook_year)
        for row_no in range(header_row + 1, ws.max_row + 1):
            sticker_type = _cell_text(ws.cell(row_no, 1).value)
            if not sticker_type or "#NFC" not in sticker_type:
                continue
            sticker_type = _normalize_sticker_type(conn, NFC_MATERIAL, sticker_type)
            for col_no, rec_date, doc_no in detail_columns:
                qty = _legacy_int(ws.cell(row_no, col_no).value, row_no, "数量")
                if qty is None or qty <= 0:
                    continue
                body = RecordIn(
                    rec_type=rec_type,
                    location_id=issue_location_id if rec_type == "issue" else None,
                    rec_date=rec_date,
                    doc_no=doc_no,
                    material=NFC_MATERIAL,
                    sticker_type=sticker_type,
                    qty=qty,
                    remark=f"{ws.title}导入",
                )
                _validate_record(body, department)
                bodies.append(body)

    if not bodies:
        raise HTTPException(status_code=400, detail=f"{department}贴纸台账没有可导入的数据")
    return bodies


def _parse_legacy_heyuan_workbook(conn, wb, department):
    if department != HEYUAN_DEPARTMENT:
        raise HTTPException(status_code=400, detail="河源华兴台账只能在河源华兴部门导入")

    bodies = []
    location_id = _location_id_from_name(conn, HEYUAN_DEPARTMENT, 1)
    for ws in wb.worksheets:
        if "领料明细" not in ws.title:
            continue
        headers = _legacy_header_map(ws)
        for row_no in range(2, ws.max_row + 1):
            material = _cell_text(_legacy_header_value(ws, headers, row_no, "物料名称"))
            if not material or "小计" in material:
                continue
            material = _normalize_pcba_material(material)
            if material != PCBA_MATERIAL:
                continue
            qty = _legacy_int(
                _legacy_header_value(ws, headers, row_no, "领料数"),
                row_no,
                "领料数",
            )
            if qty is None or qty == 0:
                continue
            date_value = _legacy_header_value(ws, headers, row_no, "日期")
            doc_no = _cell_text(_legacy_header_value(ws, headers, row_no, "领料编号"))
            if not doc_no:
                doc_no = _cell_text(date_value)
            body = RecordIn(
                rec_type="issue",
                location_id=location_id,
                rec_date=_outsource_date_value(date_value),
                doc_no=doc_no,
                material=material,
                qty=qty,
                remark=_first_value(
                    {"备注": _legacy_header_value(ws, headers, row_no, "备注")},
                    "备注",
                ) or f"{ws.title}导入",
            )
            if qty > 0:
                _validate_record(body, department)
            bodies.append(body)

    if "成品入仓明细" in wb.sheetnames:
        ws = wb["成品入仓明细"]
        headers = _legacy_header_map(ws)
        for row_no in range(2, ws.max_row + 1):
            name = _cell_text(_legacy_header_value(ws, headers, row_no, "品名/规格"))
            if not name or "小计" in name:
                continue
            qty = _legacy_int(
                _legacy_header_value(ws, headers, row_no, "数量（pcs）", "数量"),
                row_no,
                "数量",
            )
            if qty is None or qty <= 0:
                continue
            contract = _cell_text(_legacy_header_value(ws, headers, row_no, "合同号"))
            item_no = _cell_text(_legacy_header_value(ws, headers, row_no, "货号"))
            body = RecordIn(
                rec_type="finished",
                location_id=location_id,
                rec_date=_outsource_date_value(
                    _legacy_header_value(ws, headers, row_no, "日期")
                ),
                doc_no=_cell_text(_legacy_header_value(ws, headers, row_no, "送货单号")),
                material=PCBA_MATERIAL,
                qty=qty,
                remark=" ".join(
                    part for part in [contract, item_no, name, "成品入仓明细导入"] if part
                ),
            )
            _validate_record(body, department)
            bodies.append(body)

    if not bodies:
        raise HTTPException(status_code=400, detail="河源华兴台账没有可导入的数据")
    return bodies


def _pcba_summary_key_from_label(label):
    text = _cell_text(label)
    if not text:
        return None
    compact = re.sub(r"\s+", "", text)
    if "入仓" in compact or "入库" in compact:
        return ("inbound_raw", None)
    if "邵阳" in compact:
        return ("issue", "邵阳华登")
    if "河源" in compact:
        return ("issue", "河源华兴")
    if "鸿亚" in compact:
        return ("issue", HONGYA_DEPARTMENT)
    if "利鸿" in compact or "加工厂" in compact:
        return ("issue", "东莞加工厂利鸿")
    if "东莞车间" in compact or "车间领料" in compact:
        return ("issue", "东莞车间")
    return None


def _pcba_summary_month_from_header(value):
    text = _cell_text(value)
    match = re.search(r"(\d{1,2})月", text)
    if not match:
        return None
    month = int(match.group(1))
    return month if month in SUMMARY_MONTHS else None


def _pcba_summary_month_totals(wb):
    if "总表" not in wb.sheetnames:
        return {}
    ws = wb["总表"]
    month_columns = []
    for col_no in range(1, ws.max_column + 1):
        month = _pcba_summary_month_from_header(ws.cell(1, col_no).value)
        if month:
            month_columns.append((month, col_no))
    if not month_columns:
        return {}

    totals = {}
    for row_no in range(2, ws.max_row + 1):
        key = _pcba_summary_key_from_label(
            ws.cell(row_no, 2).value or ws.cell(row_no, 1).value
        )
        if not key:
            continue
        month_values = {}
        for month, col_no in month_columns:
            month_values[month] = _legacy_int(
                ws.cell(row_no, col_no).value,
                row_no,
                f"{month}月",
            ) or 0
        totals[key] = month_values
    return totals


def _assign_pcba_summary_months(groups, totals):
    for key, bodies in groups.items():
        month_totals = totals.get(key)
        if not month_totals:
            continue
        months = [
            month for month in SUMMARY_MONTHS
            if int(month_totals.get(month) or 0) != 0
        ]
        if not months:
            continue
        month_index = 0
        remaining = int(month_totals[months[month_index]] or 0)
        for body in bodies:
            while month_index < len(months) and remaining == 0:
                month_index += 1
                if month_index < len(months):
                    remaining = int(month_totals[months[month_index]] or 0)
            if month_index >= len(months):
                break
            body.summary_month = months[month_index]
            remaining -= int(body.qty or 0)


def _parse_legacy_supplier_pcba_workbook(conn, wb, department):
    bodies = []
    groups = {}
    inbound_ws = wb["入仓明细"] if "入仓明细" in wb.sheetnames else None
    if inbound_ws:
        headers = _legacy_header_map(inbound_ws)
        for row_no in range(2, inbound_ws.max_row + 1):
            material = _cell_text(
                _legacy_header_value(inbound_ws, headers, row_no, "物料名称")
            )
            if not material or "小计" in material:
                continue
            qty = _legacy_int(
                _legacy_header_value(inbound_ws, headers, row_no, "入仓数", "入库数"),
                row_no,
                "入仓数",
            )
            if qty is None or qty == 0:
                continue
            date_value = _legacy_header_value(inbound_ws, headers, row_no, "日期")
            body = RecordIn(
                rec_type="inbound_raw",
                rec_date=_outsource_date_value(date_value),
                doc_no=_legacy_doc_no(
                    _legacy_header_value(inbound_ws, headers, row_no, "入仓单号", "入库单号"),
                    date_value,
                    f"{inbound_ws.title}-{row_no}",
                ),
                material=_normalize_pcba_material(material),
                qty=qty,
                remark=_first_value(
                    {"备注": _legacy_header_value(inbound_ws, headers, row_no, "备注")},
                    "备注",
                ) or f"{inbound_ws.title}导入",
            )
            _add_record_body(bodies, body, department, validate_positive=False)
            groups.setdefault(("inbound_raw", None), []).append(body)

    for ws in wb.worksheets:
        if "领料" not in ws.title:
            continue
        location_name = _legacy_location_from_sheet(ws.title)
        location_id = _location_id_from_name(conn, location_name, 1)
        headers = _legacy_header_map(ws)
        for row_no in range(2, ws.max_row + 1):
            material = _cell_text(_legacy_header_value(ws, headers, row_no, "物料名称"))
            if not material or "小计" in material:
                continue
            material = _normalize_pcba_material(material)
            if material != PCBA_MATERIAL:
                continue
            qty = _legacy_int(
                _legacy_header_value(ws, headers, row_no, "领料数"),
                row_no,
                "领料数",
            )
            if qty is None or qty == 0:
                continue
            date_value = _legacy_header_value(ws, headers, row_no, "日期")
            body = RecordIn(
                rec_type="issue",
                location_id=location_id,
                rec_date=_outsource_date_value(date_value),
                doc_no=_legacy_doc_no(
                    _legacy_header_value(ws, headers, row_no, "领料单号", "领料编号"),
                    date_value,
                    f"{ws.title}-{row_no}",
                ),
                material=material,
                qty=qty,
                remark=_first_value(
                    {"备注": _legacy_header_value(ws, headers, row_no, "备注")},
                    "备注",
                ) or f"{ws.title}导入",
            )
            _add_record_body(bodies, body, department, validate_positive=False)
            groups.setdefault(("issue", location_name), []).append(body)
    _assign_pcba_summary_months(groups, _pcba_summary_month_totals(wb))
    return bodies


def _nfc_total_row_values(ws, row_no):
    sticker_type = _cell_text(ws.cell(row_no, 1).value)
    if "#NFC" not in sticker_type:
        return None
    return {
        "sticker_type": sticker_type,
        "opening_inbound": _legacy_int(
            ws.cell(row_no, 3).value, row_no, "期初入仓"
        ) or 0,
        "dongguan_opening_outbound": _legacy_int(
            ws.cell(row_no, 13).value, row_no, "东莞期初出仓"
        ) or 0,
        "shaoyang_opening_outbound": _legacy_int(
            ws.cell(row_no, 14).value, row_no, "邵阳期初领料"
        ) or 0,
    }


def _parse_supplier_nfc_total_rows(conn, wb, department):
    bodies = []
    if "总表" not in wb.sheetnames:
        return bodies
    ws = wb["总表"]
    workbook_year = _legacy_workbook_year(wb)
    opening_inbound_date = _legacy_cutoff_date(ws.cell(2, 3).value, workbook_year)
    dongguan_opening_date = _legacy_cutoff_date(ws.cell(2, 13).value, workbook_year)
    shaoyang_opening_date = (
        _legacy_cutoff_date(ws.cell(2, 14).value, workbook_year)
        or dongguan_opening_date
        or opening_inbound_date
    )
    dongguan_id = _location_id_from_name(conn, "东莞车间", 1)
    shaoyang_id = _location_id_from_name(conn, "邵阳华登", 1)
    for row_no in range(3, ws.max_row + 1):
        values = _nfc_total_row_values(ws, row_no)
        if not values:
            continue
        sticker_type = _normalize_sticker_type(
            conn, NFC_MATERIAL, values["sticker_type"]
        )
        if values["opening_inbound"]:
            _add_record_body(
                bodies,
                RecordIn(
                    rec_type="inbound_raw",
                    material=NFC_MATERIAL,
                    sticker_type=sticker_type,
                    rec_date=opening_inbound_date,
                    doc_no=f"{sticker_type}-期初入仓",
                    qty=values["opening_inbound"],
                    remark="总表期初入仓导入",
                ),
                department,
            )
        if values["dongguan_opening_outbound"]:
            _add_record_body(
                bodies,
                RecordIn(
                    rec_type="issue",
                    location_id=dongguan_id,
                    material=NFC_MATERIAL,
                    sticker_type=sticker_type,
                    rec_date=dongguan_opening_date,
                    doc_no=f"{sticker_type}-东莞期初出仓",
                    qty=values["dongguan_opening_outbound"],
                    remark="总表东莞期初出仓导入",
                ),
                department,
            )
        if values["shaoyang_opening_outbound"]:
            _add_record_body(
                bodies,
                RecordIn(
                    rec_type="issue",
                    location_id=shaoyang_id,
                    material=NFC_MATERIAL,
                    sticker_type=sticker_type,
                    rec_date=shaoyang_opening_date,
                    doc_no=f"{sticker_type}-邵阳期初领料",
                    qty=values["shaoyang_opening_outbound"],
                    remark="总表邵阳期初领料导入",
                ),
                department,
            )
    return bodies


def _parse_supplier_nfc_detail_sheet(conn, ws, rec_type, department):
    bodies = []
    header_row = _find_legacy_item_header_row(ws)
    if not header_row:
        return bodies
    location_id = None
    if rec_type == "issue":
        location_id = _location_id_from_name(conn, "东莞车间", 1)
    for row_no in range(header_row + 1, ws.max_row + 1):
        sticker_type = _cell_text(ws.cell(row_no, 1).value)
        if "#NFC" not in sticker_type:
            continue
        sticker_type = _normalize_sticker_type(conn, NFC_MATERIAL, sticker_type)
        for col_no in range(3, ws.max_column + 1):
            qty = _legacy_int(ws.cell(row_no, col_no).value, row_no, "数量")
            if qty is None or qty == 0:
                continue
            date_value = ws.cell(1, col_no).value
            body = RecordIn(
                rec_type=rec_type,
                location_id=location_id,
                rec_date=_outsource_date_value(date_value),
                doc_no=_legacy_doc_no(
                    ws.cell(header_row, col_no).value,
                    date_value,
                    f"{ws.title}-{row_no}-{col_no}",
                ),
                material=NFC_MATERIAL,
                sticker_type=sticker_type,
                qty=qty,
                remark=f"{ws.title}导入",
            )
            _add_record_body(
                bodies, body, department, validate_positive=(qty > 0)
            )
    return bodies


def _parse_legacy_supplier_workbook(conn, wb, department):
    if department != SUPPLIER_DEPARTMENT:
        raise HTTPException(status_code=400, detail="来料仓台账只能在兴信B来料仓导入")

    bodies = []
    if "入仓明细" in wb.sheetnames:
        bodies.extend(_parse_legacy_supplier_pcba_workbook(conn, wb, department))
    if {"总表", "入库明细", "出库明细"}.issubset(set(wb.sheetnames)):
        bodies.extend(_parse_supplier_nfc_total_rows(conn, wb, department))
        bodies.extend(
            _parse_supplier_nfc_detail_sheet(
                conn, wb["入库明细"], "inbound_raw", department
            )
        )
        bodies.extend(
            _parse_supplier_nfc_detail_sheet(
                conn, wb["出库明细"], "issue", department
            )
        )
    if not bodies:
        raise HTTPException(status_code=400, detail="来料仓台账没有可导入的数据")
    return bodies


def _record_duplicate_id(conn, body, user):
    row = conn.execute(
        "SELECT id FROM records WHERE department=? AND rec_type=? "
        "AND COALESCE(rec_date, '')=? AND COALESCE(doc_no, '')=? "
        "AND material=? AND COALESCE(sticker_type, '')=? AND qty=? "
        "AND COALESCE(location_id, 0)=? AND COALESCE(remark, '')=?",
        (
            user["department"],
            body.rec_type,
            body.rec_date or "",
            body.doc_no or "",
            body.material,
            body.sticker_type or "",
            body.qty,
            body.location_id or 0,
            body.remark or "",
        ),
    ).fetchone()
    if row:
        return row["id"]
    if body.rec_date is None and body.doc_no:
        legacy_row = conn.execute(
            "SELECT id FROM records WHERE department=? AND rec_type=? "
            "AND COALESCE(rec_date, '')='' AND COALESCE(doc_no, '')='' "
            "AND material=? AND COALESCE(sticker_type, '')=? AND qty=? "
            "AND COALESCE(location_id, 0)=? AND COALESCE(remark, '')=?",
            (
                user["department"],
                body.rec_type,
                body.material,
                body.sticker_type or "",
                body.qty,
                body.location_id or 0,
                body.remark or "",
            ),
        ).fetchone()
        if legacy_row:
            return legacy_row["id"]
    return None


def _insert_record_body(conn, body, user, skip_duplicate=False):
    if skip_duplicate:
        duplicate_id = _record_duplicate_id(conn, body, user)
        if duplicate_id:
            if body.summary_month is not None:
                conn.execute(
                    "UPDATE records SET summary_month=? WHERE id=?",
                    (body.summary_month, duplicate_id),
                )
            return None
    supplier, po_no, customer_name = _record_extras(body, user["department"])
    sticker_type = _normalize_sticker_type(conn, body.material, body.sticker_type)
    cur = conn.execute(
        "INSERT INTO records(rec_type, location_id, rec_date, doc_no, "
        "material, sticker_type, qty, remark, supplier, po_no, customer_name, "
        "summary_month, department, created_by) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        (body.rec_type, body.location_id, body.rec_date, body.doc_no,
         body.material, sticker_type, body.qty, body.remark, supplier, po_no, customer_name,
         body.summary_month,
         user["department"], user["id"]),
    )
    return cur.lastrowid


def _upsert_semi_finished_monthly_totals(conn, department, rows):
    for row in rows:
        conn.execute(
            "INSERT INTO semi_finished_monthly_totals("
            "department, material, sticker_type, opening_stock, monthly_inbound, monthly_outbound"
            ") VALUES (?,?,?,?,?,?) "
            "ON CONFLICT(department, material, sticker_type) DO UPDATE SET "
            "opening_stock=excluded.opening_stock, "
            "monthly_inbound=excluded.monthly_inbound, "
            "monthly_outbound=excluded.monthly_outbound, "
            "updated_at=datetime('now')",
            (
                department,
                row["material"],
                row["sticker_type"],
                int(row["opening_stock"] or 0),
                int(row["monthly_inbound"] or 0),
                int(row["monthly_outbound"] or 0),
            ),
        )


@app.get("/api/locations")
def list_locations(_user=Depends(current_user)):
    conn = db.get_conn()
    try:
        rows = conn.execute(
            "SELECT id, name, sort FROM locations ORDER BY sort"
        ).fetchall()
    finally:
        conn.close()
    return [dict(r) for r in rows]


@app.get("/api/materials")
def list_materials(_user=Depends(current_user)):
    conn = db.get_conn()
    try:
        rows = conn.execute(
            "SELECT id, name FROM materials ORDER BY id"
        ).fetchall()
    finally:
        conn.close()
    return [dict(r) for r in rows]


@app.get("/api/materials/export")
def export_materials(_user=Depends(current_user)):
    conn = db.get_conn()
    try:
        rows = conn.execute("SELECT name FROM materials ORDER BY id").fetchall()
    finally:
        conn.close()
    wb = _simple_workbook(["物料名称"], [[row["name"]] for row in rows])
    return _xlsx_response(wb, "materials.xlsx")


@app.post("/api/materials/import")
def import_materials(file: UploadFile = File(...), _user=Depends(current_user)):
    names = _read_name_import(file, "物料名称", "物料", "名称")
    conn = db.get_conn()
    try:
        result = _insert_unique_names(conn, "materials", names)
        conn.commit()
    finally:
        conn.close()
    return result


class MaterialIn(BaseModel):
    name: str


class SupplierIn(BaseModel):
    name: str


class StickerTypeIn(BaseModel):
    name: str


@app.get("/api/suppliers")
def list_suppliers(_user=Depends(current_user)):
    conn = db.get_conn()
    try:
        rows = conn.execute(
            "SELECT id, name, created_at FROM suppliers ORDER BY id"
        ).fetchall()
    finally:
        conn.close()
    return [dict(r) for r in rows]


@app.get("/api/suppliers/export")
def export_suppliers(_user=Depends(current_user)):
    conn = db.get_conn()
    try:
        rows = conn.execute("SELECT name FROM suppliers ORDER BY id").fetchall()
    finally:
        conn.close()
    wb = _simple_workbook(["供应商名称"], [[row["name"]] for row in rows])
    return _xlsx_response(wb, "suppliers.xlsx")


@app.post("/api/suppliers/import")
def import_suppliers(file: UploadFile = File(...), _user=Depends(current_user)):
    names = _read_name_import(file, "供应商名称", "供应商", "名称")
    conn = db.get_conn()
    try:
        result = _insert_unique_names(conn, "suppliers", names)
        conn.commit()
    finally:
        conn.close()
    return result


@app.post("/api/suppliers")
def create_supplier(body: SupplierIn, _user=Depends(current_user)):
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="供应商名称不能为空")
    conn = db.get_conn()
    try:
        exists = conn.execute(
            "SELECT 1 FROM suppliers WHERE name=?", (name,)
        ).fetchone()
        if exists:
            raise HTTPException(status_code=400, detail="供应商名称已存在")
        cur = conn.execute("INSERT INTO suppliers(name) VALUES (?)", (name,))
        conn.commit()
        new_id = cur.lastrowid
    finally:
        conn.close()
    return {"id": new_id, "name": name}


@app.put("/api/suppliers/{supplier_id}")
def update_supplier(supplier_id: int, body: SupplierIn, _user=Depends(current_user)):
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="供应商名称不能为空")
    conn = db.get_conn()
    try:
        row = conn.execute(
            "SELECT id FROM suppliers WHERE id=?", (supplier_id,)
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="供应商不存在")
        exists = conn.execute(
            "SELECT 1 FROM suppliers WHERE name=? AND id<>?",
            (name, supplier_id),
        ).fetchone()
        if exists:
            raise HTTPException(status_code=400, detail="供应商名称已存在")
        conn.execute(
            "UPDATE suppliers SET name=? WHERE id=?",
            (name, supplier_id),
        )
        conn.commit()
    finally:
        conn.close()
    return {"id": supplier_id, "name": name}


@app.delete("/api/suppliers/{supplier_id}")
def delete_supplier(supplier_id: int, _user=Depends(current_user)):
    conn = db.get_conn()
    try:
        conn.execute("DELETE FROM suppliers WHERE id=?", (supplier_id,))
        conn.commit()
    finally:
        conn.close()
    return {"ok": True}


@app.get("/api/sticker-types")
def list_sticker_types(_user=Depends(current_user)):
    conn = db.get_conn()
    try:
        rows = conn.execute(
            "SELECT id, name, sort FROM sticker_types ORDER BY sort, id"
        ).fetchall()
    finally:
        conn.close()
    return [dict(r) for r in rows]


@app.get("/api/sticker-types/export")
def export_sticker_types(_user=Depends(current_user)):
    conn = db.get_conn()
    try:
        rows = conn.execute(
            "SELECT sort, name FROM sticker_types ORDER BY sort, id"
        ).fetchall()
    finally:
        conn.close()
    wb = _simple_workbook(["排序", "贴纸类型"], [[row["sort"], row["name"]] for row in rows])
    return _xlsx_response(wb, "sticker_types.xlsx")


@app.post("/api/sticker-types/import")
def import_sticker_types(file: UploadFile = File(...), _user=Depends(current_user)):
    incoming = []
    seen_names = set()
    seen_sorts = set()
    for index, (row_no, row) in enumerate(_sheet_rows(file), start=1):
        name = _first_value(row, "贴纸类型", "贴纸名称", "名称")
        if not name:
            raise HTTPException(status_code=400, detail=f"第{row_no}行：贴纸类型不能为空")
        if name in seen_names:
            raise HTTPException(status_code=400, detail=f"第{row_no}行：贴纸类型重复")
        seen_names.add(name)
        sort_text = _first_value(row, "排序")
        try:
            sort = int(sort_text) if sort_text else index
        except ValueError:
            raise HTTPException(status_code=400, detail=f"第{row_no}行：排序必须是数字")
        if sort in seen_sorts:
            raise HTTPException(status_code=400, detail=f"第{row_no}行：排序重复")
        seen_sorts.add(sort)
        incoming.append((sort, name))

    imported = 0
    skipped = 0
    conn = db.get_conn()
    try:
        for sort, name in incoming:
            by_sort = conn.execute(
                "SELECT id, name FROM sticker_types WHERE sort=?", (sort,)
            ).fetchone()
            by_name = conn.execute(
                "SELECT id, name FROM sticker_types WHERE name=?", (name,)
            ).fetchone()
            if by_sort:
                old_name = by_sort["name"]
                if old_name == name:
                    skipped += 1
                    continue
                if by_name and by_name["id"] != by_sort["id"]:
                    raise HTTPException(status_code=400, detail=f"贴纸类型已存在：{name}")
                conn.execute(
                    "UPDATE sticker_types SET name=? WHERE id=?",
                    (name, by_sort["id"]),
                )
                conn.execute(
                    "UPDATE records SET sticker_type=? WHERE sticker_type=?",
                    (name, old_name),
                )
                imported += 1
            elif by_name:
                conn.execute(
                    "UPDATE sticker_types SET sort=? WHERE id=?",
                    (sort, by_name["id"]),
                )
                imported += 1
            else:
                conn.execute(
                    "INSERT INTO sticker_types(name, sort) VALUES (?, ?)",
                    (name, sort),
                )
                imported += 1
        conn.commit()
    finally:
        conn.close()
    return {"imported": imported, "skipped": skipped}


@app.post("/api/sticker-types")
def create_sticker_type(body: StickerTypeIn, _user=Depends(current_user)):
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="贴纸类型不能为空")
    conn = db.get_conn()
    try:
        exists = conn.execute(
            "SELECT 1 FROM sticker_types WHERE name=?", (name,)
        ).fetchone()
        if exists:
            raise HTTPException(status_code=400, detail="贴纸类型已存在")
        next_sort = conn.execute(
            "SELECT COALESCE(MAX(sort), 0) + 1 AS next_sort FROM sticker_types"
        ).fetchone()["next_sort"]
        cur = conn.execute(
            "INSERT INTO sticker_types(name, sort) VALUES (?, ?)",
            (name, next_sort),
        )
        conn.commit()
        new_id = cur.lastrowid
    finally:
        conn.close()
    return {"id": new_id, "name": name, "sort": next_sort}


@app.put("/api/sticker-types/{sticker_type_id}")
def update_sticker_type(
    sticker_type_id: int,
    body: StickerTypeIn,
    _user=Depends(current_user),
):
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="贴纸类型不能为空")
    conn = db.get_conn()
    try:
        row = conn.execute(
            "SELECT id, sort FROM sticker_types WHERE id=?", (sticker_type_id,)
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="贴纸类型不存在")
        exists = conn.execute(
            "SELECT 1 FROM sticker_types WHERE name=? AND id<>?",
            (name, sticker_type_id),
        ).fetchone()
        if exists:
            raise HTTPException(status_code=400, detail="贴纸类型已存在")
        conn.execute(
            "UPDATE sticker_types SET name=? WHERE id=?",
            (name, sticker_type_id),
        )
        conn.commit()
        sort = row["sort"]
    finally:
        conn.close()
    return {"id": sticker_type_id, "name": name, "sort": sort}


@app.delete("/api/sticker-types/{sticker_type_id}")
def delete_sticker_type(sticker_type_id: int, _user=Depends(current_user)):
    conn = db.get_conn()
    try:
        conn.execute("DELETE FROM sticker_types WHERE id=?", (sticker_type_id,))
        conn.commit()
    finally:
        conn.close()
    return {"ok": True}


@app.post("/api/materials")
def create_material(body: MaterialIn, _user=Depends(current_user)):
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="物料名称不能为空")
    conn = db.get_conn()
    try:
        exists = conn.execute(
            "SELECT 1 FROM materials WHERE name=?", (name,)
        ).fetchone()
        if exists:
            raise HTTPException(status_code=400, detail="物料名称已存在")
        cur = conn.execute("INSERT INTO materials(name) VALUES (?)", (name,))
        conn.commit()
        new_id = cur.lastrowid
    finally:
        conn.close()
    return {"id": new_id, "name": name}


@app.put("/api/materials/{material_id}")
def update_material(material_id: int, body: MaterialIn, _user=Depends(current_user)):
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="物料名称不能为空")
    conn = db.get_conn()
    try:
        row = conn.execute(
            "SELECT id FROM materials WHERE id=?", (material_id,)
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="物料不存在")
        exists = conn.execute(
            "SELECT 1 FROM materials WHERE name=? AND id<>?",
            (name, material_id),
        ).fetchone()
        if exists:
            raise HTTPException(status_code=400, detail="物料名称已存在")
        conn.execute(
            "UPDATE materials SET name=? WHERE id=?",
            (name, material_id),
        )
        conn.commit()
    finally:
        conn.close()
    return {"id": material_id, "name": name}


@app.delete("/api/materials/{material_id}")
def delete_material(material_id: int, _user=Depends(current_user)):
    conn = db.get_conn()
    try:
        conn.execute("DELETE FROM materials WHERE id=?", (material_id,))
        conn.commit()
    finally:
        conn.close()
    return {"ok": True}


class RecordIn(BaseModel):
    rec_type: str
    location_id: Optional[int] = None
    rec_date: Optional[str] = None
    doc_no: Optional[str] = None
    material: str = PCBA_MATERIAL
    sticker_type: Optional[str] = None
    qty: int
    remark: Optional[str] = None
    supplier: Optional[str] = None
    po_no: Optional[str] = None
    customer_name: Optional[str] = None
    summary_month: Optional[int] = None


def _validate_record(body: RecordIn, department: Optional[str] = None):
    if body.rec_type not in VALID_TYPES:
        raise HTTPException(status_code=400, detail="类型无效")
    if _is_outsource_department(department) and body.rec_type not in ("issue", "finished", "semi_finished"):
        raise HTTPException(status_code=400, detail="东莞加工厂只能录入领料/成品/半成品入库")
    if department == HEYUAN_DEPARTMENT and body.rec_type not in ("issue", "finished"):
        raise HTTPException(status_code=400, detail="河源华兴只能录入领料/成品入库")
    if department == SHAOYANG_DEPARTMENT and body.rec_type not in ("issue", "finished"):
        raise HTTPException(status_code=400, detail="邵阳华登只能录入领料/成品入库")
    if department == XINSHAO_DEPARTMENT and body.rec_type not in ("issue", "finished"):
        raise HTTPException(status_code=400, detail="新邵只能录入领料/成品入库")
    if department == SUPPLIER_DEPARTMENT and body.rec_type == "finished":
        raise HTTPException(status_code=400, detail="兴信B来料仓只能录入入库/出库")
    if body.rec_type == "semi_finished" and department not in (ASSEMBLY_DEPARTMENT, *OUTSOURCE_DEPARTMENTS):
        raise HTTPException(status_code=400, detail="半成品入库仅限东莞车间/东莞加工厂部门")
    if body.rec_type in ("semi_inbound", "semi_outbound") and department != SEMI_FINISHED_DEPARTMENT:
        raise HTTPException(status_code=400, detail="半成品仓出入库仅限碟片半成品部门")
    if body.qty is None or body.qty < 0:
        raise HTTPException(status_code=400, detail="数量必须为非负整数")
    if body.rec_type in ("issue", "semi_outbound") and not body.location_id:
        raise HTTPException(status_code=400, detail="出库必须选择目标部门")
    if (
        body.rec_type in ("finished", "semi_finished")
        and not _is_outsource_department(department)
        and not body.location_id
    ):
        raise HTTPException(status_code=400, detail="领料/入库必须选择加工点")
    if body.rec_type in ("inbound_raw", "semi_inbound") or (
        _is_outsource_department(department)
        and body.rec_type in ("finished", "semi_finished")
    ):
        body.location_id = None


class StickerRecordItemIn(BaseModel):
    sticker_type: str
    qty: int


class RecordBatchIn(BaseModel):
    rec_type: str
    location_id: Optional[int] = None
    rec_date: Optional[str] = None
    doc_no: Optional[str] = None
    material: str = NFC_MATERIAL
    remark: Optional[str] = None
    supplier: Optional[str] = None
    po_no: Optional[str] = None
    customer_name: Optional[str] = None
    items: List[StickerRecordItemIn]


class RecordBulkDeleteIn(BaseModel):
    ids: List[int]


class RecordClearIn(BaseModel):
    department: str
    material: str


def _record_extras(body, department):
    supplier = (body.supplier or "").strip() or None
    if department != SUPPLIER_DEPARTMENT:
        supplier = None
    po_no = (body.po_no or "").strip() or None
    customer_name = (body.customer_name or "").strip() or None
    if department not in PO_CUSTOMER_DEPARTMENTS or body.rec_type != "finished":
        po_no = None
        customer_name = None
    return supplier, po_no, customer_name


def _normalize_sticker_type(conn, material, sticker_type):
    if material != NFC_MATERIAL:
        return None
    sticker_type = _clean_optional(sticker_type)
    if not sticker_type:
        return None
    row = conn.execute(
        "SELECT name FROM sticker_types WHERE name=?", (sticker_type,)
    ).fetchone()
    if not row:
        raise HTTPException(status_code=400, detail="贴纸类型无效")
    return row["name"]


def _location_name_from_id(conn, location_id):
    if not location_id:
        return None
    row = conn.execute(
        "SELECT name FROM locations WHERE id=?", (location_id,)
    ).fetchone()
    return row["name"] if row else None


def _auto_record_remark(source_department, target_department, original_remark):
    prefix = f"自动联动：{source_department}->{target_department}"
    original_remark = _clean_optional(original_remark)
    return f"{prefix}；{original_remark}" if original_remark else prefix


def _auto_flow_record_body(
    source_body,
    source_department,
    target_department,
    rec_type,
    location_id=None,
):
    body = RecordIn(
        rec_type=rec_type,
        location_id=location_id,
        rec_date=source_body.rec_date,
        doc_no=source_body.doc_no,
        material=source_body.material,
        sticker_type=source_body.sticker_type,
        qty=source_body.qty,
        remark=_auto_record_remark(
            source_department, target_department, source_body.remark
        ),
        summary_month=source_body.summary_month,
    )
    return body


def _auto_receive_record_type(target_department):
    if target_department == SUPPLIER_DEPARTMENT:
        return "inbound_raw"
    if target_department == SEMI_FINISHED_DEPARTMENT:
        return "semi_inbound"
    return "issue"


def _auto_receive_location_id(conn, target_department, rec_type):
    if rec_type != "issue":
        return None
    return _location_id_from_name(conn, target_department, 1)


def _append_department_transfer_target(
    conn,
    targets,
    source_body,
    source_department,
    source_location,
):
    if source_body.rec_type not in ("issue", "semi_outbound"):
        return
    if source_location not in DEPARTMENTS or source_location == source_department:
        return
    rec_type = _auto_receive_record_type(source_location)
    targets.append((
        source_location,
        _auto_flow_record_body(
            source_body,
            source_department,
            source_location,
            rec_type,
            _auto_receive_location_id(conn, source_location, rec_type),
        ),
        "department_transfer",
    ))


def _auto_flow_targets(conn, source_body, source_department):
    if source_body.qty is None or source_body.qty <= 0:
        return []

    if source_body.material == NFC_MATERIAL and source_body.sticker_type:
        _normalize_sticker_type(conn, source_body.material, source_body.sticker_type)
    source_location = _location_name_from_id(conn, source_body.location_id)
    targets = []

    _append_department_transfer_target(
        conn, targets, source_body, source_department, source_location
    )

    if source_department == ASSEMBLY_DEPARTMENT and source_body.rec_type == "semi_finished":
        targets.append((
            SEMI_FINISHED_DEPARTMENT,
            _auto_flow_record_body(
                source_body,
                source_department,
                SEMI_FINISHED_DEPARTMENT,
                "semi_inbound",
            ),
            "assembly_to_semi_finished",
        ))

    if (
        source_department == HONGYA_DEPARTMENT
        and source_body.rec_type in ("finished", "semi_finished")
    ):
        targets.append((
            SEMI_FINISHED_DEPARTMENT,
            _auto_flow_record_body(
                source_body,
                source_department,
                SEMI_FINISHED_DEPARTMENT,
                "semi_inbound",
            ),
            "hongya_to_semi_finished",
        ))

    return targets


def _insert_auto_record(conn, target_department, body, source_record_id, source_flow, created_by):
    _validate_record(body, target_department)
    supplier, po_no, customer_name = _record_extras(body, target_department)
    sticker_type = _normalize_sticker_type(conn, body.material, body.sticker_type)
    conn.execute(
        "INSERT INTO records(rec_type, location_id, rec_date, doc_no, "
        "material, sticker_type, qty, remark, supplier, po_no, customer_name, "
        "summary_month, department, created_by, source_record_id, source_flow) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        (
            body.rec_type,
            body.location_id,
            body.rec_date,
            body.doc_no,
            body.material,
            sticker_type,
            body.qty,
            body.remark,
            supplier,
            po_no,
            customer_name,
            body.summary_month,
            target_department,
            created_by,
            source_record_id,
            source_flow,
        ),
    )


def _sync_auto_linked_records(conn, source_record_id, source_body, user):
    conn.execute("DELETE FROM records WHERE source_record_id=?", (source_record_id,))
    for target_department, target_body, source_flow in _auto_flow_targets(
        conn, source_body, user["department"]
    ):
        _insert_auto_record(
            conn,
            target_department,
            target_body,
            source_record_id,
            source_flow,
            user["id"],
        )


def _reject_auto_record_direct_edit(row):
    if row["source_record_id"]:
        raise HTTPException(
            status_code=400,
            detail="自动生成记录不能直接修改或删除，请修改原始记录",
        )


@app.get("/api/records")
def list_records(
    user=Depends(current_user),
    rec_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    doc_no: Optional[str] = None,
):
    filters = _date_filter(date_from, date_to, doc_no)
    conn = db.get_conn()
    try:
        sql = (
            "SELECT r.*, l.name AS location_name, u.username AS created_by_name "
            "FROM records r "
            "LEFT JOIN locations l ON r.location_id = l.id "
            "LEFT JOIN users u ON r.created_by = u.id "
            "WHERE r.department = ?"
        )
        params = [user["department"]]
        if rec_type:
            sql += " AND r.rec_type = ?"
            params.append(rec_type)
        sql += " AND COALESCE(r.remark, '') NOT LIKE ?"
        params.append("总表%导入")
        sql, params = _append_date_filter(sql, params, filters)
        sql += (
            " ORDER BY r.rec_date IS NULL, r.rec_date DESC, "
            "CASE WHEN r.sticker_type GLOB '[0-9]*#*' THEN 0 ELSE 1 END, "
            "CASE WHEN r.sticker_type GLOB '[0-9]*#*' THEN CAST(r.sticker_type AS INTEGER) END, "
            "r.id DESC"
        )
        rows = conn.execute(sql, params).fetchall()
    finally:
        conn.close()
    return [dict(r) for r in rows]


@app.post("/api/records")
def create_record(body: RecordIn, user=Depends(current_user)):
    _validate_record(body, user["department"])
    conn = db.get_conn()
    try:
        supplier, po_no, customer_name = _record_extras(body, user["department"])
        sticker_type = _normalize_sticker_type(conn, body.material, body.sticker_type)
        cur = conn.execute(
            "INSERT INTO records(rec_type, location_id, rec_date, doc_no, "
            "material, sticker_type, qty, remark, supplier, po_no, customer_name, "
            "summary_month, department, created_by) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (body.rec_type, body.location_id, body.rec_date, body.doc_no,
             body.material, sticker_type, body.qty, body.remark, supplier, po_no, customer_name,
             body.summary_month,
            user["department"], user["id"]),
        )
        new_id = cur.lastrowid
        _sync_auto_linked_records(conn, new_id, body, user)
        conn.commit()
    finally:
        conn.close()
    return {"id": new_id}


@app.post("/api/records/batch")
def create_records_batch(body: RecordBatchIn, user=Depends(current_user)):
    if body.material != NFC_MATERIAL:
        raise HTTPException(status_code=400, detail="批量录入仅支持 NFC贴纸")
    if not body.items:
        raise HTTPException(status_code=400, detail="请选择至少一种贴纸")
    common = RecordIn(
        rec_type=body.rec_type,
        location_id=body.location_id,
        rec_date=body.rec_date,
        doc_no=body.doc_no,
        material=body.material,
        qty=sum(int(item.qty or 0) for item in body.items),
        remark=body.remark,
        supplier=body.supplier,
        po_no=body.po_no,
        customer_name=body.customer_name,
    )
    _validate_record(common, user["department"])
    conn = db.get_conn()
    try:
        supplier, po_no, customer_name = _record_extras(common, user["department"])
        seen = set()
        valid_items = []
        for item in body.items:
            if item.qty is None or item.qty <= 0:
                raise HTTPException(status_code=400, detail="贴纸数量必须大于 0")
            sticker_type = _normalize_sticker_type(conn, NFC_MATERIAL, item.sticker_type)
            if not sticker_type:
                raise HTTPException(status_code=400, detail="贴纸类型不能为空")
            if sticker_type in seen:
                raise HTTPException(status_code=400, detail="贴纸类型不能重复")
            seen.add(sticker_type)
            valid_items.append((sticker_type, item.qty))
        ids = []
        for sticker_type, qty in valid_items:
            item_body = RecordIn(
                rec_type=common.rec_type,
                location_id=common.location_id,
                rec_date=common.rec_date,
                doc_no=common.doc_no,
                material=NFC_MATERIAL,
                sticker_type=sticker_type,
                qty=qty,
                remark=common.remark,
                supplier=common.supplier,
                po_no=common.po_no,
                customer_name=common.customer_name,
                summary_month=common.summary_month,
            )
            cur = conn.execute(
                "INSERT INTO records(rec_type, location_id, rec_date, doc_no, "
                "material, sticker_type, qty, remark, supplier, po_no, customer_name, "
                "summary_month, department, created_by) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (item_body.rec_type, item_body.location_id, item_body.rec_date, item_body.doc_no,
                 NFC_MATERIAL, sticker_type, qty, item_body.remark, supplier, po_no, customer_name,
                 item_body.summary_month,
                 user["department"], user["id"]),
            )
            record_id = cur.lastrowid
            ids.append(record_id)
            _sync_auto_linked_records(conn, record_id, item_body, user)
        conn.commit()
    finally:
        conn.close()
    return {"ids": ids}


@app.get("/api/records/import-template")
def record_import_template(_user=Depends(current_user)):
    wb = Workbook()
    ws = wb.active
    ws.title = "流水导入模板"
    ws.append(RECORD_IMPORT_HEADERS)
    ws.append([
        "入库", "NFC贴纸", "1#NFC贴纸", "", "供应商A",
        "2026-07-08", "示例单号001", 100, "示例行，可删除", "", "",
    ])
    return _xlsx_response(wb, "records_import_template.xlsx")


def _record_export_type_label(rec_type, department):
    if rec_type == "issue":
        return "出库" if department == SUPPLIER_DEPARTMENT else "领料"
    labels = {
        "inbound_raw": "入库",
        "finished": "成品入库",
        "semi_finished": "半成品入库",
        "semi_inbound": "入库",
        "semi_outbound": "出库",
    }
    return labels.get(rec_type, rec_type)


def _records_export_workbook(records, department):
    wb = Workbook()
    ws = wb.active
    ws.title = "流水导出"
    ws.append(RECORD_IMPORT_HEADERS)
    for row in records:
        ws.append([
            _record_export_type_label(row["rec_type"], department),
            row["material"],
            row["sticker_type"],
            row["location_name"],
            row["supplier"],
            _record_export_date(row),
            row["doc_no"],
            row["qty"],
            row["remark"],
            row["po_no"],
            row["customer_name"],
        ])
    return wb


EXPORT_MONTHS = tuple(range(7, 13))
SUMMARY_MONTHS = (6, *EXPORT_MONTHS)
SUMMARY_MONTH_LABELS = {
    6: "6月月结",
    **{month: f"{month}月" for month in EXPORT_MONTHS},
}


def _export_date(value):
    if not value:
        return None
    try:
        return date.fromisoformat(str(value))
    except ValueError:
        return value


def _legacy_opening_record_date(row):
    if row.get("rec_date"):
        return None
    text = f"{row.get('doc_no') or ''} {row.get('remark') or ''}"
    if "期初" in text:
        return date(date.today().year, 6, 27).isoformat()
    return None


def _record_export_date(row):
    return row.get("rec_date") or _legacy_opening_record_date(row)


def _export_month(value):
    if not value:
        return 6
    try:
        return date.fromisoformat(str(value)).month
    except ValueError:
        return 6


def _sum_qty(records):
    return sum(int(row.get("qty") or 0) for row in records)


def _month_sum(records, month):
    return _sum_qty([row for row in records if _record_month(row) == month])


def _record_month(row):
    summary_month = row.get("summary_month")
    if summary_month:
        try:
            month = int(summary_month)
            if month in SUMMARY_MONTHS:
                return month
        except (TypeError, ValueError):
            pass
    return _export_month(_record_export_date(row))


def _monthly_flow_values(records):
    values = []
    for month in SUMMARY_MONTHS:
        issue = _sum_qty([
            row for row in records
            if row["rec_type"] == "issue" and _record_month(row) == month
        ])
        finished = _sum_qty([
            row for row in records
            if row["rec_type"] in ("finished", "semi_finished")
            and _record_month(row) == month
        ])
        values.append({
            "issue": issue,
            "finished": finished,
            "balance": issue - finished,
        })
    return values


def _monthly_raw_values(records):
    values = []
    for month in SUMMARY_MONTHS:
        inbound = _sum_qty([
            row for row in records
            if row["rec_type"] == "inbound_raw" and _record_month(row) == month
        ])
        outbound = _sum_qty([
            row for row in records
            if row["rec_type"] == "issue" and _record_month(row) == month
        ])
        values.append({
            "inbound": inbound,
            "outbound": outbound,
            "balance": inbound - outbound,
        })
    return values


def _sum_month_values(values, key):
    return sum(int(row.get(key) or 0) for row in values)


def _summary_material(row):
    material = (row.get("material") or PCBA_MATERIAL).strip()
    return NFC_MATERIAL if material == NFC_MATERIAL else material


def _summary_materials(records):
    preferred = [NFC_MATERIAL, PCBA_MATERIAL]
    names = sorted({_summary_material(row) for row in records})
    return sorted(
        names,
        key=lambda name: (
            preferred.index(name) if name in preferred else len(preferred),
            name,
        ),
    )


def _monthly_location_summary(records, location_names):
    locations = []
    for location in location_names:
        location_records = [
            row for row in records if row.get("location") == location
        ]
        for material in _summary_materials(location_records):
            material_records = [
                row for row in location_records
                if _summary_material(row) == material
            ]
            values = _monthly_flow_values(material_records)
            locations.append({
                "location": location,
                "material": material,
                "issue": _sum_month_values(values, "issue"),
                "finished": _sum_month_values(values, "finished"),
                "balance": _sum_month_values(values, "balance"),
                "values": values,
            })
    subtotal_values = []
    for index, _month in enumerate(SUMMARY_MONTHS):
        issue = sum(row["values"][index]["issue"] for row in locations)
        finished = sum(row["values"][index]["finished"] for row in locations)
        subtotal_values.append({
            "issue": issue,
            "finished": finished,
            "balance": issue - finished,
        })
    raw_values = _monthly_raw_values(records)
    return {
        "months": [
            {"month": month, "label": SUMMARY_MONTH_LABELS[month]}
            for month in SUMMARY_MONTHS
        ],
        "locations": locations,
        "subtotal": {
            "issue": _sum_month_values(subtotal_values, "issue"),
            "finished": _sum_month_values(subtotal_values, "finished"),
            "balance": _sum_month_values(subtotal_values, "balance"),
            "values": subtotal_values,
        },
        "raw": {
            "inbound": _sum_month_values(raw_values, "inbound"),
            "outbound": _sum_month_values(raw_values, "outbound"),
            "balance": _sum_month_values(raw_values, "balance"),
            "values": raw_values,
        },
    }


def _format_nfc_sticker_name(name):
    name = name or ""
    return name.replace("NFC贴纸", "NFC\n贴纸")


def _apply_legacy_sheet_style(ws, max_row=None, max_col=None):
    max_row = max_row or ws.max_row
    max_col = max_col or ws.max_column
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.alignment = alignment
            cell.border = border
    for col_no in range(1, max_col + 1):
        letter = ws.cell(1, col_no).column_letter
        ws.column_dimensions[letter].width = 13
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16


SUPPLIER_PCBA_SUMMARY_ROWS = (
    ("邵阳华登", "邵阳华登领料总数", "PCB主板", "邵阳华登领料"),
    ("河源华兴", "河源华兴领料总数", "PCB主板", "河源华兴领料"),
    ("东莞加工厂利鸿", "加工厂利鸿领料总数", "PCB主板", "加工厂利鸿领料"),
    ("东莞加工厂鸿亚", "加工厂鸿亚领料总数", "PCB主板", "加工厂鸿亚领料"),
    ("东莞车间", "东莞车间领料", None, "东莞车间领料"),
)


def _supplier_pcba_month_sums(records):
    return [_month_sum(records, month) for month in (6, *EXPORT_MONTHS)]


def _supplier_pcba_activity_row(material_label, label, records):
    month_sums = _supplier_pcba_month_sums(records)
    month_values = [value or None for value in month_sums]
    return [material_label, label, sum(month_sums), *month_values, None]


def _supplier_pcba_detail_sheet(wb, title, records, doc_header, qty_header):
    ws = wb.create_sheet(title[:31])
    ws.append(["日期", doc_header, "物料名称", qty_header, "备注"])
    for row in records:
        ws.append([
            _record_export_date(row),
            row.get("doc_no"),
            PCBA_MATERIAL,
            row.get("qty"),
            row.get("remark"),
        ])
    ws.append([None, None, "小计：", _sum_qty(records), None])
    _apply_legacy_sheet_style(ws, ws.max_row, 5)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["E"].width = 20


def _pcba_inbound_detail_sheet(wb, title, records):
    ws = wb.create_sheet(title[:31])
    ws.append(["日期", "送货单号", "合同号", "货号", "品名/规格", "数量（pcs）", "备注"])
    for row in records:
        ws.append([
            _record_export_date(row),
            row.get("doc_no"),
            None,
            "77794",
            row.get("material") or PCBA_MATERIAL,
            row.get("qty"),
            row.get("remark"),
        ])
    ws.append([None, None, None, None, "小计：", _sum_qty(records), None])
    _apply_legacy_sheet_style(ws, ws.max_row, 7)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["G"].width = 22


def _outsource_pcba_export_workbook(records):
    wb = Workbook()
    ws = wb.active
    ws.title = "总表"
    issue = [row for row in records if row["rec_type"] == "issue"]
    finished = [row for row in records if row["rec_type"] == "finished"]
    semi_finished = [row for row in records if row["rec_type"] == "semi_finished"]
    headers = ["物料名称", None, "累计出入数", "截6月月结"]
    headers += [f"{month}月" for month in EXPORT_MONTHS]
    headers.append("备注")
    ws.append(headers)
    ws.append(["PCBA主板", "领料总数", _sum_qty(issue), *_supplier_pcba_month_sums(issue), None])
    ws.append([
        None, "半成品入仓总数", _sum_qty(semi_finished),
        *_supplier_pcba_month_sums(semi_finished), None,
    ])
    if finished:
        ws.append([
            None, "成品入仓总数", _sum_qty(finished),
            *_supplier_pcba_month_sums(finished), None,
        ])
    _apply_legacy_sheet_style(ws, ws.max_row, len(headers))
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18

    _supplier_pcba_detail_sheet(wb, "领料明细", issue, "领料编号", "领料数")
    _pcba_inbound_detail_sheet(wb, "半成品入仓明细", semi_finished)
    if finished:
        _pcba_inbound_detail_sheet(wb, "成品入仓明细", finished)
    return wb


def _heyuan_pcba_export_workbook(records):
    wb = Workbook()
    ws = wb.active
    ws.title = "总表"
    issue = [row for row in records if row["rec_type"] == "issue"]
    finished = [row for row in records if row["rec_type"] == "finished"]
    headers = ["物料名称", "领料总数", "截6月月结"]
    headers += [f"{month}月" for month in EXPORT_MONTHS]
    headers.append("备注")
    ws.append(headers)
    ws.append(["PCBA主板", _sum_qty(issue), *_supplier_pcba_month_sums(issue), None])
    if finished:
        ws.append(["成品入仓总数", _sum_qty(finished), *_supplier_pcba_month_sums(finished), None])
    _apply_legacy_sheet_style(ws, ws.max_row, len(headers))
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14

    _supplier_pcba_detail_sheet(wb, "36#CD领料明细", [], "领料编号", "领料数")
    _supplier_pcba_detail_sheet(wb, "PCB板领料明细", issue, "领料编号", "领料数")
    _pcba_inbound_detail_sheet(wb, "成品入仓明细", finished)
    return wb


def _supplier_pcba_export_workbook(records):
    wb = Workbook()
    ws = wb.active
    ws.title = "总表"
    inbound = [row for row in records if row["rec_type"] == "inbound_raw"]
    issues = [row for row in records if row["rec_type"] == "issue"]
    headers = ["物料名称", None, "累计出入总数", "6月月结"]
    headers += [f"{month}月" for month in EXPORT_MONTHS]
    headers.append("备注")
    ws.append(headers)
    ws.append(_supplier_pcba_activity_row("PCB主板", "入仓总数", inbound))
    issue_month_sums = [0] * 7
    for row_no, (location, label, material_label, _sheet_title) in enumerate(
        SUPPLIER_PCBA_SUMMARY_ROWS, start=3
    ):
        location_rows = [row for row in issues if row.get("location_name") == location]
        ws.append(_supplier_pcba_activity_row(material_label, label, location_rows))
        issue_month_sums = [
            current + value
            for current, value in zip(issue_month_sums, _supplier_pcba_month_sums(location_rows))
        ]
    ws.append([None] * len(headers))
    balance_rows = [
        inbound_value - issue_value
        for inbound_value, issue_value in zip(_supplier_pcba_month_sums(inbound), issue_month_sums)
    ]
    ws.append([None, "应存数", sum(balance_rows), *balance_rows, None])
    _apply_legacy_sheet_style(ws, 8, len(headers))
    ws["C8"].fill = PatternFill(fill_type="solid", fgColor="FFFFFF00")
    ws.column_dimensions["A"].width = 14.625
    ws.column_dimensions["B"].width = 17.25
    ws.column_dimensions["C"].width = 15.5
    ws.column_dimensions["D"].width = 14.75
    ws.column_dimensions["K"].width = 9

    _supplier_pcba_detail_sheet(wb, "入仓明细", inbound, "入仓单号", "入仓数")
    for location, _label, _material_label, sheet_title in SUPPLIER_PCBA_SUMMARY_ROWS:
        location_rows = [row for row in issues if row.get("location_name") == location]
        _supplier_pcba_detail_sheet(
            wb, sheet_title, location_rows, "领料单号", "领料数"
        )
    wb.create_sheet("Sheet2")
    return wb


def _supplier_nfc_sticker_names(records, sticker_types):
    names = list(sticker_types)
    for row in records:
        name = row.get("sticker_type")
        if name and name not in names:
            names.append(name)
    return names


def _supplier_nfc_export_doc_no(row):
    text = f"{row.get('doc_no') or ''} {row.get('remark') or ''}"
    if "东莞期初出仓" in text:
        return "东莞期初出仓"
    if "邵阳期初领料" in text:
        return "邵阳期初领料"
    if "期初入仓" in text:
        return "期初入仓"
    if "期初出仓" in text:
        return "期初出仓"
    return row.get("doc_no")


def _supplier_nfc_detail_groups(records):
    groups = {}
    for row in records:
        key = (_record_export_date(row), _supplier_nfc_export_doc_no(row))
        if key not in groups:
            groups[key] = {
                "rec_date": key[0],
                "doc_no": key[1],
                "records": [],
            }
        groups[key]["records"].append(row)
    return list(groups.values())


def _supplier_nfc_detail_sheet(wb, title, records, sticker_names, total_header):
    ws = wb.create_sheet(title)
    records = [
        row for row in records
        if _export_month(_record_export_date(row)) in (6, *EXPORT_MONTHS)
    ]
    groups = _supplier_nfc_detail_groups(records)
    ws.cell(1, 2).value = total_header
    ws.cell(2, 1).value = "物料名称"
    for offset, group in enumerate(groups, start=3):
        ws.cell(1, offset).value = group["rec_date"]
        ws.cell(2, offset).value = group["doc_no"]
    for row_no, sticker_type in enumerate(sticker_names, start=3):
        ws.cell(row_no, 1).value = _format_nfc_sticker_name(sticker_type)
        sticker_records = [
            row for row in records if row.get("sticker_type") == sticker_type
        ]
        ws.cell(row_no, 2).value = _sum_qty(sticker_records) or 0
        for col_no, group in enumerate(groups, start=3):
            group_sticker_records = [
                row for row in group["records"]
                if row.get("sticker_type") == sticker_type
            ]
            qty = _sum_qty(group_sticker_records)
            if qty:
                ws.cell(row_no, col_no).value = qty
    total_row = len(sticker_names) + 3
    ws.cell(total_row, 1).value = "小计："
    ws.cell(total_row, 2).value = _sum_qty(records) or 0
    for col_no, group in enumerate(groups, start=3):
        ws.cell(total_row, col_no).value = _sum_qty(group["records"]) or None
    _apply_legacy_sheet_style(ws, total_row, max(2, len(groups) + 2))
    return ws


def _supplier_nfc_export_workbook(records, sticker_types):
    wb = Workbook()
    ws = wb.active
    ws.title = "总表"
    inbound = [row for row in records if row["rec_type"] == "inbound_raw"]
    outbound = [row for row in records if row["rec_type"] == "issue"]
    sticker_names = _supplier_nfc_sticker_names(records, sticker_types)
    ws.cell(1, 2).value = "累计入仓总数"
    ws.cell(2, 1).value = "物料名称"
    ws.cell(2, 3).value = "截止6月27号"
    for i, month in enumerate(EXPORT_MONTHS, start=4):
        ws.cell(1, i).value = f"{month}月入仓\n总数"
    ws.cell(1, 11).value = "应存数"
    ws.cell(1, 12).value = "累计出仓总数"
    ws.cell(1, 13).value = "东莞"
    ws.cell(2, 13).value = "截止6月27号"
    ws.cell(1, 14).value = "邵阳领料"
    for i, month in enumerate(EXPORT_MONTHS, start=15):
        ws.cell(1, i).value = f"{month}月出仓\n总数"

    for row_no, sticker_type in enumerate(sticker_names, start=3):
        sticker_inbound = [row for row in inbound if row.get("sticker_type") == sticker_type]
        sticker_outbound = [row for row in outbound if row.get("sticker_type") == sticker_type]
        ws.cell(row_no, 1).value = _format_nfc_sticker_name(sticker_type)
        ws.cell(row_no, 2).value = _sum_qty(sticker_inbound) or 0
        ws.cell(row_no, 3).value = _month_sum(sticker_inbound, 6) or 0
        for col_no, month in enumerate(EXPORT_MONTHS, start=4):
            ws.cell(row_no, col_no).value = _month_sum(sticker_inbound, month) or None
        ws.cell(row_no, 11).value = _sum_qty(sticker_inbound) - _sum_qty(sticker_outbound)
        ws.cell(row_no, 12).value = _sum_qty(sticker_outbound) or 0
        dongguan_opening = [
            row for row in sticker_outbound
            if row.get("location_name") == "东莞车间" and _export_month(row.get("rec_date")) == 6
        ]
        shaoyang_opening = [
            row for row in sticker_outbound
            if row.get("location_name") == "邵阳华登" and _export_month(row.get("rec_date")) == 6
        ]
        ws.cell(row_no, 13).value = _sum_qty(dongguan_opening) or 0
        ws.cell(row_no, 14).value = _sum_qty(shaoyang_opening) or 0
        for col_no, month in enumerate(EXPORT_MONTHS, start=15):
            ws.cell(row_no, col_no).value = _month_sum(sticker_outbound, month) or None
    _apply_legacy_sheet_style(ws, len(sticker_names) + 2, 20)
    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 11
    ws.column_dimensions["C"].width = 12

    _supplier_nfc_detail_sheet(wb, "入库明细", inbound, sticker_names, "当月入仓总数")
    _supplier_nfc_detail_sheet(wb, "出库明细", outbound, sticker_names, "当月出仓总数")
    return wb


def _monthly_totals_map(monthly_totals):
    return {
        row["sticker_type"]: {
            "opening_stock": int(row.get("opening_stock") or 0),
            "monthly_inbound": int(row.get("monthly_inbound") or 0),
            "monthly_outbound": int(row.get("monthly_outbound") or 0),
        }
        for row in monthly_totals
    }


def _semi_finished_sticker_names(records, sticker_types, monthly_totals):
    names = list(sticker_types)
    for row in monthly_totals:
        name = row.get("sticker_type")
        if name and name not in names:
            names.append(name)
    for row in records:
        name = row.get("sticker_type")
        if name and name not in names:
            names.append(name)
    return names


def _semi_month_total(total_row, records, key, rec_type):
    value = int((total_row or {}).get(key) or 0)
    if value:
        return value
    return _sum_qty([row for row in records if row["rec_type"] == rec_type])


def _semi_finished_detail_sheet(
    wb, title, records, sticker_names, totals_by_sticker, is_inbound
):
    ws = wb.create_sheet(title)
    rec_type = "semi_inbound" if is_inbound else "semi_outbound"
    total_key = "monthly_inbound" if is_inbound else "monthly_outbound"
    total_header = "当月入仓总数" if is_inbound else "当月出仓总数"
    sheet_records = [row for row in records if row["rec_type"] == rec_type]
    start_col = 5 if is_inbound else 4
    header_row = 3 if is_inbound else 5
    data_row_start = header_row + 1

    if is_inbound:
        ws.cell(1, 4).value = "日期"
        ws.cell(2, 4).value = "入库单号"
        ws.cell(3, 1).value = "物料名称"
        ws.cell(3, 2).value = total_header
        ws.cell(3, 3).value = "6/24\n东莞车间入库截数"
        ws.cell(3, 4).value = "6/24\n鸿亚入库截数"
    else:
        ws.cell(5, 1).value = "物料名称"
        ws.cell(5, 2).value = total_header
        ws.cell(5, 3).value = "6/24盘点截数"

    for offset, row in enumerate(sheet_records, start=start_col):
        ws.cell(1, offset).value = _record_export_date(row)
        ws.cell(2, offset).value = row.get("doc_no")

    for row_no, sticker_type in enumerate(sticker_names, start=data_row_start):
        sticker_records = [
            row for row in sheet_records if row.get("sticker_type") == sticker_type
        ]
        total_row = totals_by_sticker.get(sticker_type, {})
        ws.cell(row_no, 1).value = _format_nfc_sticker_name(sticker_type)
        ws.cell(row_no, 2).value = _semi_month_total(
            total_row, sticker_records, total_key, rec_type
        ) or 0
        ws.cell(row_no, 3).value = int(total_row.get("opening_stock") or 0) or None
        for col_no, record in enumerate(sheet_records, start=start_col):
            if record.get("sticker_type") == sticker_type:
                ws.cell(row_no, col_no).value = record.get("qty")

    total_row_no = data_row_start + len(sticker_names)
    ws.cell(total_row_no, 1).value = "小计："
    ws.cell(total_row_no, 2).value = sum(
        ws.cell(row_no, 2).value or 0
        for row_no in range(data_row_start, total_row_no)
    )
    ws.cell(total_row_no, 3).value = sum(
        ws.cell(row_no, 3).value or 0
        for row_no in range(data_row_start, total_row_no)
    ) or None
    for col_no, record in enumerate(sheet_records, start=start_col):
        ws.cell(total_row_no, col_no).value = record.get("qty")

    _apply_legacy_sheet_style(
        ws, total_row_no, max(start_col - 1, len(sheet_records) + start_col - 1)
    )
    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 13
    return ws


def _semi_finished_export_workbook(records, sticker_types, monthly_totals):
    wb = Workbook()
    ws = wb.active
    ws.title = "总表"
    sticker_names = _semi_finished_sticker_names(records, sticker_types, monthly_totals)
    totals_by_sticker = _monthly_totals_map(monthly_totals)
    inbound = [row for row in records if row["rec_type"] == "semi_inbound"]
    outbound = [row for row in records if row["rec_type"] == "semi_outbound"]

    ws.cell(1, 2).value = "累计入仓总数"
    ws.cell(2, 1).value = "物料名称"
    ws.cell(2, 3).value = "截止6月27号"
    for col_no, month in enumerate(EXPORT_MONTHS, start=4):
        ws.cell(1, col_no).value = f"{month}月入仓\n总数"
    ws.cell(1, 11).value = "应存数"
    ws.cell(1, 12).value = "累计出仓总数"
    ws.cell(1, 13).value = "东莞"
    ws.cell(2, 13).value = "截止6月27号"
    ws.cell(1, 14).value = "邵阳生产"
    for col_no, month in enumerate(EXPORT_MONTHS, start=15):
        ws.cell(1, col_no).value = f"{month}月出仓\n总数"

    for row_no, sticker_type in enumerate(sticker_names, start=3):
        total_row = totals_by_sticker.get(sticker_type, {})
        sticker_inbound = [row for row in inbound if row.get("sticker_type") == sticker_type]
        sticker_outbound = [row for row in outbound if row.get("sticker_type") == sticker_type]
        inbound_total = _semi_month_total(
            total_row, sticker_inbound, "monthly_inbound", "semi_inbound"
        )
        outbound_total = _semi_month_total(
            total_row, sticker_outbound, "monthly_outbound", "semi_outbound"
        )
        ws.cell(row_no, 1).value = _format_nfc_sticker_name(sticker_type)
        ws.cell(row_no, 2).value = inbound_total or 0
        ws.cell(row_no, 3).value = int(total_row.get("opening_stock") or 0) or None
        ws.cell(row_no, 4).value = inbound_total or None
        ws.cell(row_no, 11).value = inbound_total - outbound_total
        ws.cell(row_no, 12).value = outbound_total or 0
        ws.cell(row_no, 14).value = int(total_row.get("opening_stock") or 0) or None
        ws.cell(row_no, 15).value = outbound_total or None

    _apply_legacy_sheet_style(ws, len(sticker_names) + 2, 20)
    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 11
    ws.column_dimensions["C"].width = 12

    _semi_finished_detail_sheet(
        wb, "入库明细", records, sticker_names, totals_by_sticker, True
    )
    _semi_finished_detail_sheet(
        wb, "邵阳领料", records, sticker_names, totals_by_sticker, False
    )
    wb.create_sheet("河源华兴36#CD领料")
    wb.create_sheet("车间36#CD领料")
    return wb


@app.get("/api/records/export")
def export_records(
    user=Depends(current_user),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    doc_no: Optional[str] = None,
    material: Optional[str] = None,
):
    filters = _date_filter(date_from, date_to, doc_no)
    material = _clean_optional(material)
    conn = db.get_conn()
    try:
        sql = (
            "SELECT r.id, r.rec_type, l.name AS location_name, r.rec_date, r.doc_no, "
            "r.material, r.sticker_type, r.supplier, r.po_no, r.customer_name, "
            "r.qty, r.remark, r.summary_month "
            "FROM records r "
            "LEFT JOIN locations l ON r.location_id = l.id "
            "WHERE r.department=?"
        )
        params = [user["department"]]
        sql, params = _append_date_filter(sql, params, filters)
        if material:
            sql += " AND r.material=?"
            params.append(material)
        sql += " ORDER BY r.id"
        rows = conn.execute(sql, params).fetchall()
        sticker_types = [
            row["name"] for row in conn.execute(
                "SELECT name FROM sticker_types ORDER BY sort, id"
            ).fetchall()
        ]
        totals_sql = (
            "SELECT material, sticker_type, opening_stock, "
            "monthly_inbound, monthly_outbound "
            "FROM semi_finished_monthly_totals WHERE department=?"
        )
        totals_params = [user["department"]]
        if material:
            totals_sql += " AND material=?"
            totals_params.append(material)
        monthly_totals = conn.execute(totals_sql, totals_params).fetchall()
    finally:
        conn.close()
    records = [dict(r) for r in rows]
    monthly_totals = [dict(r) for r in monthly_totals]
    if user["department"] == SEMI_FINISHED_DEPARTMENT and material in (None, NFC_MATERIAL):
        return _xlsx_response(
            _semi_finished_export_workbook(records, sticker_types, monthly_totals),
            "塑胶仓77772#CD半成品出入明细.xlsx",
        )
    if user["department"] == SUPPLIER_DEPARTMENT and material == PCBA_MATERIAL:
        return _xlsx_response(
            _supplier_pcba_export_workbook(records),
            "来料仓77794PCB主板出入明细.xlsx",
        )
    if user["department"] == SUPPLIER_DEPARTMENT and material == NFC_MATERIAL:
        return _xlsx_response(
            _supplier_nfc_export_workbook(records, sticker_types),
            "来料仓77772#NFC出入明细.xlsx",
        )
    if _is_outsource_department(user["department"]) and material == PCBA_MATERIAL:
        return _xlsx_response(
            _outsource_pcba_export_workbook(records),
            f"{user['department']}77794PCB主板出入明细.xlsx",
        )
    if user["department"] == HEYUAN_DEPARTMENT and material == PCBA_MATERIAL:
        return _xlsx_response(
            _heyuan_pcba_export_workbook(records),
            "河源华兴77794PCB主板出入明细.xlsx",
        )
    return _xlsx_response(
        _records_export_workbook(records, user["department"]),
        "records_export.xlsx",
    )


def _record_body_from_import_row(conn, row_no, row, department):
    rec_type = _record_type_from_import(_first_value(row, "类型"), department)
    material = _first_value(row, "物料名称", "物料") or PCBA_MATERIAL
    sticker_type = _first_value(row, "贴纸类型", "贴纸名称")
    if material == NFC_MATERIAL and not sticker_type:
        raise HTTPException(status_code=400, detail=f"第{row_no}行：NFC贴纸必须填写贴纸类型")
    body = RecordIn(
        rec_type=rec_type,
        location_id=_location_id_from_name(
            conn,
            _first_value(row, "加工点", "地点"),
            row_no,
        ),
        rec_date=_date_value(row.get("日期"), row_no),
        doc_no=_first_value(row, "单据编号", "单号"),
        material=material,
        sticker_type=sticker_type,
        qty=_int_value(row.get("数量"), row_no, "数量"),
        remark=_first_value(row, "备注"),
        supplier=_first_value(row, "供应商"),
        po_no=_first_value(row, "PO"),
        customer_name=_first_value(row, "客名", "客户名"),
    )
    _validate_record(body, department)
    return body


@app.post("/api/records/import")
def import_records(file: UploadFile = File(...), user=Depends(current_user)):
    wb = _load_upload_workbook(file)
    conn = db.get_conn()
    try:
        legacy_semi_finished_import = _is_legacy_semi_finished_workbook(wb)
        legacy_assembly_import = _is_legacy_assembly_workbook(wb)
        legacy_outsource_import = _is_legacy_outsource_workbook(wb)
        legacy_outsource_nfc_import = (
            _is_outsource_department(user["department"])
            and _is_legacy_outsource_nfc_workbook(wb)
        )
        legacy_heyuan_import = _is_legacy_heyuan_workbook(wb)
        legacy_supplier_import = _is_legacy_supplier_workbook(wb)
        legacy_import = (
            legacy_semi_finished_import
            or legacy_assembly_import
            or legacy_outsource_import
            or legacy_outsource_nfc_import
            or legacy_heyuan_import
            or legacy_supplier_import
        )
        if legacy_semi_finished_import:
            _require_filename_contains(file, SEMI_FINISHED_FILENAME_KEYWORD)
            bodies, monthly_totals = _parse_legacy_semi_finished_workbook(
                conn, wb, user["department"]
            )
        elif legacy_assembly_import:
            _require_filename_contains(file, ASSEMBLY_DEPARTMENT)
            bodies = _parse_legacy_assembly_workbook(conn, wb, user["department"])
            monthly_totals = []
        elif legacy_outsource_import:
            _require_filename_contains(file, user["department"])
            bodies = _parse_legacy_outsource_workbook(conn, wb, user["department"])
            monthly_totals = []
        elif legacy_outsource_nfc_import:
            _require_filename_contains(file, user["department"])
            bodies = _parse_legacy_outsource_nfc_workbook(
                conn, wb, user["department"]
            )
            monthly_totals = []
        elif legacy_heyuan_import:
            bodies = _parse_legacy_heyuan_workbook(conn, wb, user["department"])
            monthly_totals = []
        elif legacy_supplier_import:
            bodies = _parse_legacy_supplier_workbook(conn, wb, user["department"])
            monthly_totals = []
        else:
            rows = _worksheet_rows(wb.active)
            if not rows:
                raise HTTPException(status_code=400, detail="Excel 没有可导入的数据")
            bodies = [
                _record_body_from_import_row(conn, row_no, row, user["department"])
                for row_no, row in rows
            ]
            monthly_totals = []
        ids = []
        skipped = 0
        for body in bodies:
            record_id = _insert_record_body(
                conn,
                body,
                user,
                skip_duplicate=legacy_import,
            )
            if record_id is None:
                skipped += 1
            else:
                ids.append(record_id)
                _sync_auto_linked_records(conn, record_id, body, user)
        if monthly_totals:
            _upsert_semi_finished_monthly_totals(
                conn, user["department"], monthly_totals
            )
        conn.commit()
    finally:
        conn.close()
    return {
        "created": len(ids),
        "ids": ids,
        "monthly_totals": len(monthly_totals),
        "skipped": skipped,
    }


@app.post("/api/shaoyang-cd/reconcile")
def shaoyang_cd_reconcile(
    month: int = Form(7),
    issue_file: UploadFile = File(...),
    finished_file: UploadFile = File(...),
    _user=Depends(current_user),
):
    try:
        return build_shaoyang_cd_reconcile(
            issue_file.file.read(),
            issue_file.filename or "",
            finished_file.file.read(),
            finished_file.filename or "",
            month,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@app.post("/api/shaoyang-cd/export-issue")
def shaoyang_cd_export_issue(
    month: int = Form(7),
    issue_file: UploadFile = File(...),
    finished_file: UploadFile = File(...),
    _user=Depends(current_user),
):
    try:
        wb = build_shaoyang_issue_export_workbook(
            issue_file.file.read(),
            issue_file.filename or "",
            finished_file.file.read(),
            finished_file.filename or "",
            month,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    return _xlsx_response(wb, "邵阳77772#CD领料明细-已填成品入仓.xlsx")


def _get_record_or_404(conn, record_id, department):
    row = conn.execute(
        "SELECT * FROM records WHERE id=? AND department=?",
        (record_id, department),
    ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="记录不存在")
    return row


def _check_owner(row, user):
    if user["role"] != "admin" and row["created_by"] != user["id"]:
        raise HTTPException(status_code=403, detail="只能修改自己录入的记录")


def _normalized_record_ids(ids):
    result = []
    seen = set()
    for value in ids or []:
        try:
            record_id = int(value)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="记录 ID 无效")
        if record_id <= 0:
            raise HTTPException(status_code=400, detail="记录 ID 无效")
        if record_id not in seen:
            seen.add(record_id)
            result.append(record_id)
    if not result:
        raise HTTPException(status_code=400, detail="请选择要删除的记录")
    return result


def _delete_record_ids_with_links(conn, record_ids):
    if not record_ids:
        return 0
    deleted = 0
    chunk_size = 500
    for start in range(0, len(record_ids), chunk_size):
        chunk = record_ids[start:start + chunk_size]
        placeholders = ",".join("?" for _ in chunk)
        linked_count = conn.execute(
            f"SELECT COUNT(*) AS c FROM records WHERE source_record_id IN ({placeholders})",
            chunk,
        ).fetchone()["c"]
        conn.execute(
            f"DELETE FROM records WHERE source_record_id IN ({placeholders})",
            chunk,
        )
        deleted += int(linked_count or 0)
    for start in range(0, len(record_ids), chunk_size):
        chunk = record_ids[start:start + chunk_size]
        placeholders = ",".join("?" for _ in chunk)
        cur = conn.execute(
            f"DELETE FROM records WHERE id IN ({placeholders})",
            chunk,
        )
        deleted += int(cur.rowcount or 0)
    return deleted


@app.put("/api/records/{record_id}")
def update_record(record_id: int, body: RecordIn, user=Depends(current_user)):
    conn = db.get_conn()
    try:
        row = _get_record_or_404(conn, record_id, user["department"])
        _reject_auto_record_direct_edit(row)
        _check_owner(row, user)
        _validate_record(body, user["department"])
        supplier, po_no, customer_name = _record_extras(body, user["department"])
        sticker_type = _normalize_sticker_type(conn, body.material, body.sticker_type)
        conn.execute(
            "UPDATE records SET rec_type=?, location_id=?, rec_date=?, doc_no=?, "
            "material=?, sticker_type=?, qty=?, remark=?, supplier=?, po_no=?, "
            "customer_name=?, summary_month=? WHERE id=?",
            (body.rec_type, body.location_id, body.rec_date, body.doc_no,
             body.material, sticker_type, body.qty, body.remark, supplier, po_no, customer_name,
             body.summary_month,
             record_id),
        )
        _sync_auto_linked_records(conn, record_id, body, user)
        conn.commit()
    finally:
        conn.close()
    return {"ok": True}


@app.delete("/api/records/{record_id}")
def delete_record(record_id: int, user=Depends(current_user)):
    conn = db.get_conn()
    try:
        row = _get_record_or_404(conn, record_id, user["department"])
        _reject_auto_record_direct_edit(row)
        _check_owner(row, user)
        deleted = _delete_record_ids_with_links(conn, [record_id])
        conn.commit()
    finally:
        conn.close()
    return {"ok": True, "deleted": deleted}


@app.post("/api/records/bulk-delete")
def bulk_delete_records(body: RecordBulkDeleteIn, user=Depends(current_user)):
    record_ids = _normalized_record_ids(body.ids)
    placeholders = ",".join("?" for _ in record_ids)
    conn = db.get_conn()
    try:
        rows = conn.execute(
            f"SELECT * FROM records WHERE department=? AND id IN ({placeholders})",
            [user["department"], *record_ids],
        ).fetchall()
        rows_by_id = {row["id"]: row for row in rows}
        if len(rows_by_id) != len(record_ids):
            raise HTTPException(status_code=404, detail="部分记录不存在")
        for record_id in record_ids:
            row = rows_by_id[record_id]
            _reject_auto_record_direct_edit(row)
            _check_owner(row, user)
        deleted = _delete_record_ids_with_links(conn, record_ids)
        conn.commit()
    finally:
        conn.close()
    return {"ok": True, "deleted": deleted}


@app.post("/api/records/clear")
def clear_records(body: RecordClearIn, _admin=Depends(require_admin)):
    department = _validate_department(body.department)
    material = (body.material or "").strip()
    if not material:
        raise HTTPException(status_code=400, detail="请选择物料")
    conn = db.get_conn()
    try:
        rows = conn.execute(
            "SELECT id FROM records WHERE department=? AND material=?",
            (department, material),
        ).fetchall()
        record_ids = [row["id"] for row in rows]
        deleted = _delete_record_ids_with_links(conn, record_ids)
        conn.commit()
    finally:
        conn.close()
    return {"ok": True, "matched": len(record_ids), "deleted": deleted}


def _all_records_for_summary(conn, department, filters=None):
    sql = (
        "SELECT r.rec_type AS rec_type, l.name AS location, r.qty AS qty, "
        "r.material AS material, r.sticker_type AS sticker_type, "
        "r.rec_date AS rec_date, r.doc_no AS doc_no, r.remark AS remark, "
        "r.summary_month AS summary_month, r.department AS department "
        "FROM records r LEFT JOIN locations l ON r.location_id = l.id "
        "WHERE r.department=?"
    )
    params = [department]
    sql, params = _append_date_filter(sql, params, filters or {})
    rows = conn.execute(sql, params).fetchall()
    return [dict(r) for r in rows]


def _public_records_for_summary(conn, filters=None, material=None, department=None):
    sql = (
        "SELECT r.rec_type AS rec_type, r.qty AS qty, "
        "r.material AS material, r.sticker_type AS sticker_type, r.department AS department "
        "FROM records r WHERE 1=1"
    )
    params = []
    sql, params = _append_date_filter(sql, params, filters or {})
    if material:
        sql += " AND r.material = ?"
        params.append(material)
    if department:
        sql += " AND r.department = ?"
        params.append(department)
    sql += " ORDER BY r.department, r.material, r.id"
    rows = conn.execute(sql, params).fetchall()
    return [dict(r) for r in rows]


def _public_semi_finished_monthly_totals(conn, material=None, department=None):
    if department and department != SEMI_FINISHED_DEPARTMENT:
        return []
    sql = (
        "SELECT department, material, sticker_type, opening_stock, "
        "monthly_inbound, monthly_outbound, "
        "(opening_stock + monthly_inbound - monthly_outbound) AS monthly_balance "
        "FROM semi_finished_monthly_totals WHERE department=?"
    )
    params = [SEMI_FINISHED_DEPARTMENT]
    if material:
        sql += " AND material=?"
        params.append(material)
    sql += " ORDER BY sticker_type"
    return [dict(r) for r in conn.execute(sql, params).fetchall()]


def _with_common_summary_fields(summary, records, filters, reverse_departments=()):
    reverse_departments = set(reverse_departments or ())
    result = dict(summary)
    result["materials"] = compute_material_totals(records, reverse_departments)
    result["sticker_types"] = compute_sticker_type_totals(
        records, reverse_departments
    )
    result["monthly_locations"] = _monthly_location_summary(records, LOCATIONS)
    result["filters"] = filters
    return result


def _compute_semi_finished_warehouse_summary(records):
    inbound = sum(int(r["qty"] or 0) for r in records if r["rec_type"] == "semi_inbound")
    outbound = sum(int(r["qty"] or 0) for r in records if r["rec_type"] == "semi_outbound")
    return {
        "locations": [
            {"location": name, "issue": 0, "finished": 0, "balance": 0}
            for name in LOCATIONS
        ],
        "subtotal": {"issue": 0, "finished": 0, "balance": 0},
        "raw": {"inbound": inbound, "outbound": outbound, "balance": inbound - outbound},
    }


def _compute_outsource_inbound_summary(records):
    issue = sum(int(r["qty"] or 0) for r in records if r["rec_type"] == "issue")
    finished = sum(int(r["qty"] or 0) for r in records if r["rec_type"] == "finished")
    semi_finished = sum(int(r["qty"] or 0) for r in records if r["rec_type"] == "semi_finished")
    inbound = finished + semi_finished
    balance = issue - inbound if issue else inbound
    return {
        "locations": [
            {"location": name, "issue": 0, "finished": 0, "balance": 0}
            for name in LOCATIONS
        ],
        "subtotal": {"issue": 0, "finished": 0, "balance": 0},
        "raw": {
            "issue": issue,
            "finished_inbound": finished,
            "semi_finished_inbound": semi_finished,
            "inbound": inbound,
            "outbound": issue,
            "balance": balance,
        },
    }


def _compute_processing_department_summary(records):
    summary = compute_summary(records, LOCATIONS)
    summary["raw"] = {
        "inbound": summary["subtotal"]["finished"],
        "outbound": summary["subtotal"]["issue"],
        "balance": summary["subtotal"]["balance"],
    }
    return summary


@app.get("/api/public-summary")
def public_summary(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    material: Optional[str] = None,
    department: Optional[str] = None,
):
    filters = _date_filter(date_from, date_to)
    material = _clean_optional(material)
    department = _validate_department(_clean_optional(department), required=False)
    public_filters = {
        **filters,
        "material": material,
        "department": department,
    }
    conn = db.get_conn()
    try:
        records = _public_records_for_summary(
            conn,
            filters,
            material=material,
            department=department,
        )
        semi_finished_monthly_totals = _public_semi_finished_monthly_totals(
            conn,
            material=material,
            department=department,
        )
    finally:
        conn.close()
    result = compute_public_summary(
        records,
        DEPARTMENTS,
        public_filters,
        reverse_departments={
            *OUTSOURCE_DEPARTMENTS,
            *PROCESSING_BALANCE_DEPARTMENTS,
        },
    )
    result["semi_finished_monthly_totals"] = semi_finished_monthly_totals
    return result


@app.get("/api/summary")
def summary(
    user=Depends(current_user),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    doc_no: Optional[str] = None,
):
    filters = _date_filter(date_from, date_to, doc_no)
    conn = db.get_conn()
    try:
        records = _all_records_for_summary(conn, user["department"], filters)
    finally:
        conn.close()
    if user["department"] == SEMI_FINISHED_DEPARTMENT:
        summary_data = _compute_semi_finished_warehouse_summary(records)
        return _with_common_summary_fields(summary_data, records, filters)
    if _is_outsource_department(user["department"]):
        summary_data = _compute_outsource_inbound_summary(records)
        return _with_common_summary_fields(
            summary_data, records, filters, {user["department"]}
        )
    if user["department"] in PROCESSING_BALANCE_DEPARTMENTS:
        summary_data = _compute_processing_department_summary(records)
        return _with_common_summary_fields(
            summary_data, records, filters, {user["department"]}
        )
    return _with_common_summary_fields(compute_summary(records, LOCATIONS), records, filters)


@app.get("/api/export")
def export(
    user=Depends(current_user),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    doc_no: Optional[str] = None,
):
    filters = _date_filter(date_from, date_to, doc_no)
    conn = db.get_conn()
    try:
        summary_records = _all_records_for_summary(conn, user["department"], filters)
        sql = (
            "SELECT r.rec_type, l.name AS location_name, r.rec_date, r.doc_no, "
            "r.material, r.sticker_type, r.supplier, r.po_no, r.customer_name, "
            "r.qty, r.remark, r.summary_month FROM records r "
            "LEFT JOIN locations l ON r.location_id = l.id "
            "WHERE r.department=?"
        )
        params = [user["department"]]
        sql, params = _append_date_filter(sql, params, filters)
        sql += " ORDER BY r.id"
        detail_rows = conn.execute(sql, params).fetchall()
    finally:
        conn.close()
    if user["department"] == SEMI_FINISHED_DEPARTMENT:
        summary = _compute_semi_finished_warehouse_summary(summary_records)
    elif _is_outsource_department(user["department"]):
        summary = _compute_outsource_inbound_summary(summary_records)
    elif user["department"] in PROCESSING_BALANCE_DEPARTMENTS:
        summary = _compute_processing_department_summary(summary_records)
    else:
        summary = compute_summary(summary_records, LOCATIONS)
    detail_records = [dict(r) for r in detail_rows]
    buf = build_workbook(
        summary,
        detail_records,
        LOCATIONS,
        include_supplier=(user["department"] == SUPPLIER_DEPARTMENT),
        warehouse_mode=(user["department"] == SEMI_FINISHED_DEPARTMENT),
        outsource_mode=_is_outsource_department(user["department"]),
        outsource_label=user["department"],
        shaoyang_mode=(user["department"] in PO_CUSTOMER_DEPARTMENTS),
    )
    headers = {"Content-Disposition": "attachment; filename=record_player_summary.xlsx"}
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")
