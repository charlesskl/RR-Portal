import io
import re

from openpyxl import load_workbook

try:
    import xlrd
except ImportError:  # pragma: no cover - only reached when .xls support is absent.
    xlrd = None


def _cell_text(value):
    if value is None:
        return ""
    return str(value).replace("\r", "").replace("\n", "").strip()


def _compact_text(value):
    return re.sub(r"\s+", "", _cell_text(value))


def _number(value):
    if value is None or _cell_text(value) == "":
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(round(value))
    text = _cell_text(value).replace(",", "")
    try:
        return int(round(float(text)))
    except ValueError:
        return 0


def _sticker_no(value):
    match = re.search(r"(\d+)\s*#", _cell_text(value))
    return int(match.group(1)) if match else None


def _vinyl_no(value):
    match = re.search(r"VINYL-S1-(\d+)", _cell_text(value), re.IGNORECASE)
    return int(match.group(1)) if match else None


def _read_sheets(content, filename):
    suffix = (filename or "").lower().rsplit(".", 1)[-1]
    if suffix == "xls":
        if xlrd is None:
            raise ValueError("当前环境缺少 xlrd，无法读取 .xls 文件")
        book = xlrd.open_workbook(file_contents=content)
        sheets = []
        for sheet in book.sheets():
            rows = [
                [sheet.cell_value(row_no, col_no) for col_no in range(sheet.ncols)]
                for row_no in range(sheet.nrows)
            ]
            sheets.append((sheet.name, rows))
        return sheets

    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    return [
        (ws.title, [list(row) for row in ws.iter_rows(values_only=True)])
        for ws in wb.worksheets
    ]


def _value_at(row, col_index):
    return row[col_index] if col_index is not None and col_index < len(row) else None


def _month_inbound_col(rows, month):
    wanted = f"{month}月"
    for row in rows[:8]:
        for index, value in enumerate(row):
            text = _compact_text(value)
            if wanted in text and "入仓" in text and "成品" in text:
                return index
    for row in rows[:8]:
        for index, value in enumerate(row):
            text = _compact_text(value)
            if wanted in text and "入仓" in text:
                return index
    return None


def _issue_month_inbound(content, filename, month):
    for _sheet_name, rows in _read_sheets(content, filename):
        month_col = _month_inbound_col(rows, month)
        if month_col is None:
            continue
        items = {}
        for row in rows:
            no = _sticker_no(row[0] if row else None)
            if no is None:
                continue
            items[no] = {
                "sticker_name": _compact_text(row[0]),
                "issue_month_inbound": _number(_value_at(row, month_col)),
            }
        if items:
            return items
    raise ValueError(f"领料表总表里找不到 {month} 月成品入仓总数")


def _header_index(headers, *candidates):
    compact_headers = [_compact_text(header).lower() for header in headers]
    for candidate in candidates:
        text = candidate.lower()
        for index, header in enumerate(compact_headers):
            if text in header:
                return index
    return None


def _finished_totals(content, filename):
    for _sheet_name, rows in _read_sheets(content, filename):
        for header_no, row in enumerate(rows[:12]):
            item_col = _header_index(row, "itemno")
            subtotal_col = _header_index(row, "小计")
            if item_col is None or subtotal_col is None:
                continue
            name_col = _header_index(row, "minisname", "name")
            result = {}
            for values in rows[header_no + 1:]:
                item_no = _cell_text(_value_at(values, item_col))
                no = _vinyl_no(item_no)
                if no is None:
                    continue
                result[no] = {
                    "item_no": item_no,
                    "minis_name": _cell_text(_value_at(values, name_col)),
                    "finished_total": _number(_value_at(values, subtotal_col)),
                }
            if result:
                return result
    raise ValueError("成品入仓表总表里找不到 Item No. 和小计")


def _validate_month(month):
    month = int(month)
    if month < 1 or month > 12:
        raise ValueError("月份必须在 1 到 12 之间")
    return month


def _load_editable_xlsx(content, filename):
    suffix = (filename or "").lower().rsplit(".", 1)[-1]
    if suffix != "xlsx":
        raise ValueError("导出领料表只支持 .xlsx 领料文件")
    try:
        return load_workbook(io.BytesIO(content))
    except Exception as exc:
        raise ValueError("领料表无法读取") from exc


def _find_month_inbound_worksheet(wb, month):
    for ws in wb.worksheets:
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        month_col = _month_inbound_col(rows, month)
        if month_col is not None:
            return ws, month_col + 1
    raise ValueError(f"领料表总表里找不到 {month} 月成品入仓总数")


def _is_subtotal_row(ws, row_no):
    for col_no in range(1, min(ws.max_column, 6) + 1):
        if "小计" in _compact_text(ws.cell(row_no, col_no).value):
            return True
    return False


def build_shaoyang_issue_export_workbook(
    issue_content, issue_filename, finished_content, finished_filename, month
):
    month = _validate_month(month)
    wb = _load_editable_xlsx(issue_content, issue_filename)
    ws, month_col = _find_month_inbound_worksheet(wb, month)
    finished_rows = _finished_totals(finished_content, finished_filename)

    written_total = 0
    for row_no in range(1, ws.max_row + 1):
        no = _sticker_no(ws.cell(row_no, 1).value)
        if no is None:
            continue
        qty = int((finished_rows.get(no) or {}).get("finished_total") or 0)
        ws.cell(row_no, month_col).value = qty
        written_total += qty

    for row_no in range(1, ws.max_row + 1):
        if _is_subtotal_row(ws, row_no):
            ws.cell(row_no, month_col).value = written_total
            break

    return wb


def build_shaoyang_cd_reconcile(issue_content, issue_filename, finished_content, finished_filename, month):
    month = _validate_month(month)

    issue_rows = _issue_month_inbound(issue_content, issue_filename, month)
    finished_rows = _finished_totals(finished_content, finished_filename)
    rows = []
    for no in sorted(set(issue_rows) | set(finished_rows)):
        issue = issue_rows.get(no, {})
        finished = finished_rows.get(no, {})
        issue_qty = int(issue.get("issue_month_inbound") or 0)
        finished_qty = int(finished.get("finished_total") or 0)
        rows.append({
            "sticker_no": no,
            "sticker_name": issue.get("sticker_name") or f"{no}#NFC贴纸",
            "item_no": finished.get("item_no") or f"VINYL-S1-{no:03d}",
            "minis_name": finished.get("minis_name") or "",
            "issue_month_inbound": issue_qty,
            "finished_total": finished_qty,
            "difference": finished_qty - issue_qty,
        })

    totals = {
        "issue_month_inbound": sum(row["issue_month_inbound"] for row in rows),
        "finished_total": sum(row["finished_total"] for row in rows),
    }
    totals["difference"] = totals["finished_total"] - totals["issue_month_inbound"]
    return {"month": month, "rows": rows, "totals": totals}
