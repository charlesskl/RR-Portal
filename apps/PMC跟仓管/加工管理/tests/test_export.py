import io
import openpyxl


DEFAULT_DEPARTMENT = "兴信B来料仓"


def admin_login(client, department=DEFAULT_DEPARTMENT):
    client.post(
        "/api/login",
        json={"username": "admin", "password": "admin123", "department": department},
    )


def loc_id(client, name):
    locs = client.get("/api/locations").json()
    return next(l["id"] for l in locs if l["name"] == name)


def summary_row(rows, scope, item):
    return next(row for row in rows if row and row[0] == scope and row[1] == item)


def test_export_returns_xlsx_with_summary(client):
    admin_login(client, "东莞车间")
    sy = loc_id(client, "邵阳华登")
    client.post("/api/records", json={"rec_type": "inbound_raw", "qty": 1000000})
    client.post("/api/records", json={
        "rec_type": "issue", "location_id": sy, "qty": 1690242})
    client.post("/api/records", json={
        "rec_type": "finished", "location_id": sy, "qty": 928113})

    r = client.get("/api/export")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]

    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    assert "总表" in wb.sheetnames
    ws = wb["总表"]
    # 找到邵阳华登行，校验应存数 = 762129
    found = False
    for row in ws.iter_rows(values_only=True):
        if row and row[0] == "邵阳华登":
            assert row[3] == 762129
            found = True
    assert found


def test_export_is_scoped_to_current_department(client):
    admin_login(client, "东莞车间")
    sy = loc_id(client, "邵阳华登")
    client.post("/api/records", json={"rec_type": "inbound_raw", "qty": 1000000})
    client.post("/api/records", json={
        "rec_type": "issue", "location_id": sy, "qty": 1690242})

    client.post("/api/logout")
    admin_login(client, "兴信B来料仓")
    client.post("/api/records", json={"rec_type": "inbound_raw", "qty": 5})

    r = client.get("/api/export")
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    rows = list(wb["总表"].iter_rows(values_only=True))
    inbound_row = summary_row(rows, "来料仓", "入仓总数")
    outbound_row = summary_row(rows, "来料仓", "出库总数")
    assert inbound_row[2] == 5
    assert inbound_row[3] == 5
    assert outbound_row[2] == 0


def test_xingxin_export_summary_is_split_by_month(client):
    admin_login(client, "兴信B来料仓")
    dg = loc_id(client, "东莞车间")
    client.post("/api/records", json={
        "rec_type": "inbound_raw", "rec_date": "2026-06-27", "qty": 40})
    client.post("/api/records", json={
        "rec_type": "inbound_raw", "rec_date": "2026-07-05", "qty": 60})
    client.post("/api/records", json={
        "rec_type": "issue", "location_id": dg, "rec_date": "2026-06-27", "qty": 10})
    client.post("/api/records", json={
        "rec_type": "issue", "location_id": dg, "rec_date": "2026-07-06", "qty": 20})

    r = client.get("/api/export")
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    ws = wb["总表"]
    rows = list(ws.iter_rows(values_only=True))
    keyed = {(row[0], row[1]): row for row in rows if row and row[1]}

    assert [ws.cell(row=2, column=col).value for col in range(1, 12)] == [
        "范围", "项目", "累计总数", "6月月结", "7月", "8月",
        "9月", "10月", "11月", "12月", "备注",
    ]
    assert keyed[("东莞车间", "领料数")][2:5] == (30, 10, 20)
    assert keyed[("东莞车间", "成品完成数")][2:5] == (0, 0, 0)
    assert keyed[("东莞车间", "应存数")][2:5] == (30, 10, 20)
    assert keyed[("小计", "领料数")][2:5] == (30, 10, 20)
    assert keyed[("来料仓", "入仓总数")][2:5] == (100, 40, 60)
    assert keyed[("来料仓", "出库总数")][2:5] == (30, 10, 20)
    assert keyed[("货仓", "应存")][2:5] == (70, 30, 40)


def test_export_can_filter_by_date_range(client):
    admin_login(client, "兴信B来料仓")
    client.post("/api/records", json={
        "rec_type": "inbound_raw", "rec_date": "2026-07-01", "qty": 10})
    client.post("/api/records", json={
        "rec_type": "inbound_raw", "rec_date": "2026-07-05", "qty": 20})

    r = client.get("/api/export?date_from=2026-07-02&date_to=2026-07-31")
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    rows = list(wb["总表"].iter_rows(values_only=True))
    raw_row = summary_row(rows, "来料仓", "入仓总数")

    assert raw_row[2] == 20
    assert raw_row[4] == 20


def test_export_can_filter_by_doc_no(client):
    admin_login(client, "兴信B来料仓")
    client.post("/api/records", json={
        "rec_type": "inbound_raw", "doc_no": "KEEP-001", "qty": 12})
    client.post("/api/records", json={
        "rec_type": "inbound_raw", "doc_no": "DROP-001", "qty": 88})

    r = client.get("/api/export?doc_no=KEEP")
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    rows = list(wb["总表"].iter_rows(values_only=True))
    raw_row = summary_row(rows, "来料仓", "入仓总数")

    assert raw_row[2] == 12
    assert raw_row[3] == 12


def test_xingxin_export_detail_includes_supplier_column(client):
    admin_login(client, "兴信B来料仓")
    client.post("/api/suppliers", json={"name": "供应商A"})
    client.post("/api/records", json={
        "rec_type": "inbound_raw", "material": "NFC贴纸",
        "qty": 100, "supplier": "供应商A"})

    r = client.get("/api/export")
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    ws = wb["登信入仓"]
    headers = [cell.value for cell in ws[1]]
    assert headers == ["日期", "单据编号", "物料名称", "供应商", "入仓数", "备注"]
    assert ws.cell(row=2, column=4).value == "供应商A"


def test_export_includes_sticker_type_column_when_present(client):
    admin_login(client, "兴信B来料仓")
    client.post("/api/records/batch", json={
        "rec_type": "inbound_raw",
        "material": "NFC贴纸",
        "items": [{"sticker_type": "1#NFC贴纸", "qty": 12}],
    })

    r = client.get("/api/export")
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    ws = wb["登信入仓"]
    headers = [cell.value for cell in ws[1]]
    assert headers == ["日期", "单据编号", "物料名称", "贴纸类型", "供应商", "入仓数", "备注"]
    assert ws.cell(row=2, column=4).value == "1#NFC贴纸"
    assert ws.cell(row=2, column=6).value == 12


def test_export_fills_legacy_opening_date_when_record_date_is_blank(client):
    admin_login(client, "兴信B来料仓")
    dg = loc_id(client, "东莞车间")
    client.post("/api/records", json={
        "rec_type": "issue",
        "location_id": dg,
        "material": "NFC贴纸",
        "sticker_type": "1#NFC贴纸",
        "doc_no": "1#NFC贴纸-东莞期初出仓",
        "qty": 60,
        "remark": "总表东莞期初出仓导入",
    })

    r = client.get("/api/export")
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    ws = wb["东莞车间领料"]

    assert ws.cell(row=2, column=1).value == "2026-06-27"


def test_non_xingxin_export_detail_omits_supplier_column(client):
    admin_login(client, "东莞车间")
    dg = loc_id(client, "东莞加工厂利鸿")
    client.post("/api/records", json={
        "rec_type": "issue", "location_id": dg, "material": "PCBA板",
        "qty": 100, "supplier": "供应商A"})

    r = client.get("/api/export")
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    ws = wb["东莞加工厂利鸿领料"]
    headers = [cell.value for cell in ws[1]]
    assert headers == ["日期", "单据编号", "物料名称", "领料数", "备注"]


def test_export_includes_semi_finished_detail_sheet(client):
    admin_login(client, "东莞车间")
    dg = loc_id(client, "东莞加工厂利鸿")
    client.post("/api/records", json={
        "rec_type": "semi_finished", "location_id": dg,
        "material": "PCBA板", "qty": 20})

    r = client.get("/api/export")
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    assert "东莞加工厂利鸿半成品入库" in wb.sheetnames
    ws = wb["东莞加工厂利鸿半成品入库"]
    assert [cell.value for cell in ws[1]] == ["日期", "单据编号", "物料名称", "入仓数", "备注"]
    assert ws.cell(row=2, column=4).value == 20


def test_semi_finished_department_export_has_inbound_and_outbound_sheets(client):
    admin_login(client, "碟片半成品")
    client.post("/api/records", json={
        "rec_type": "semi_inbound", "material": "PCBA板", "qty": 80})
    client.post("/api/records", json={
        "rec_type": "semi_outbound", "material": "PCBA板", "qty": 30})

    r = client.get("/api/export")
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    assert "半成品入库" in wb.sheetnames
    assert "半成品出库" in wb.sheetnames
    inbound = wb["半成品入库"]
    outbound = wb["半成品出库"]
    assert inbound.cell(row=2, column=4).value == 80
    assert outbound.cell(row=2, column=4).value == 30


def test_outsource_export_has_finished_and_semi_finished_sheets(client):
    admin_login(client, "东莞加工厂利鸿")
    client.post("/api/records", json={
        "rec_type": "finished", "material": "PCBA板", "qty": 70})
    client.post("/api/records", json={
        "rec_type": "semi_finished", "material": "NFC贴纸", "qty": 30})

    r = client.get("/api/export")
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    rows = list(wb["总表"].iter_rows(values_only=True))
    header_index = next(i for i, row in enumerate(rows) if row and row[0] == "成品入库总数")
    total_row = rows[header_index + 1]
    assert total_row == (70, 30, 100)
    assert "东莞加工厂利鸿成品入库" in wb.sheetnames
    assert "东莞加工厂利鸿半成品入库" in wb.sheetnames
    assert wb["东莞加工厂利鸿成品入库"].cell(row=2, column=4).value == 70
    assert wb["东莞加工厂利鸿半成品入库"].cell(row=2, column=4).value == 30


def test_shaoyang_finished_export_includes_po_and_customer_columns(client):
    admin_login(client, "邵阳华登")
    sy = loc_id(client, "邵阳华登")
    client.post("/api/records", json={
        "rec_type": "finished", "location_id": sy, "material": "PCBA板",
        "qty": 45, "po_no": "PO-001", "customer_name": "客户A"})

    r = client.get("/api/export")
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    ws = wb["邵阳华登成品入仓"]
    assert [cell.value for cell in ws[1]] == [
        "日期", "单据编号", "物料名称", "PO", "客名", "入仓数", "备注"]
    assert ws.cell(row=2, column=4).value == "PO-001"
    assert ws.cell(row=2, column=5).value == "客户A"
    assert ws.cell(row=2, column=6).value == 45


def test_xinshao_finished_export_includes_po_and_customer_columns(client):
    admin_login(client, "新邵")
    lid = loc_id(client, "新邵")
    client.post("/api/records", json={
        "rec_type": "finished", "location_id": lid, "material": "PCBA板",
        "qty": 45, "po_no": "PO-X01", "customer_name": "客户X"})

    r = client.get("/api/export")
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    ws = wb["新邵成品入仓"]
    assert [cell.value for cell in ws[1]] == [
        "日期", "单据编号", "物料名称", "PO", "客名", "入仓数", "备注"]
    assert ws.cell(row=2, column=4).value == "PO-X01"
    assert ws.cell(row=2, column=5).value == "客户X"
    assert ws.cell(row=2, column=6).value == 45

