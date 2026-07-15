import io

import openpyxl


DEFAULT_DEPARTMENT = "兴信B来料仓"


def login(client, username="admin", password="admin123", department=DEFAULT_DEPARTMENT):
    return client.post(
        "/api/login",
        json={"username": username, "password": password, "department": department},
    )


def loc_id(client, name):
    locations = client.get("/api/locations").json()
    return next(item["id"] for item in locations if item["name"] == name)


def workbook_bytes(headers, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def upload_xlsx(client, path, headers, rows):
    content = workbook_bytes(headers, rows)
    return client.post(
        path,
        files={
            "file": (
                "import.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )


def legacy_semi_finished_workbook_bytes():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "入库明细"
    ws["D1"] = "日期"
    ws["E1"] = "2026-07-01"
    ws["F1"] = "2026-07-02"
    ws["D2"] = "入库单号"
    ws["E2"] = "RK-001"
    ws["F2"] = "RK-002"
    ws["A3"] = "物料名称"
    ws["B3"] = "当月入仓总数"
    ws["C3"] = "6/24东莞车间入库截数"
    ws["D3"] = "6/24鸿亚入库截数"
    ws["A4"] = "1#NFC\n贴纸"
    ws["B4"] = 100
    ws["C4"] = 10
    ws["D4"] = 15
    ws["E4"] = 30
    ws["F4"] = 60
    ws["A5"] = "小计："
    ws["B5"] = 100
    ws["E5"] = 30

    out_ws = wb.create_sheet("邵阳领料")
    out_ws["D1"] = "2026-07-03"
    out_ws["D2"] = "LL-001"
    out_ws["A5"] = "物料名称"
    out_ws["B5"] = "当月出仓总数"
    out_ws["C5"] = "6/24盘点截数"
    out_ws["A6"] = "1#NFC\n贴纸"
    out_ws["B6"] = 40
    out_ws["C6"] = 5
    out_ws["D6"] = 35
    out_ws["A7"] = "小计："
    out_ws["B7"] = 40
    out_ws["D7"] = 35

    total_ws = wb.create_sheet("总表")
    total_ws["A1"] = "总表不用导入"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def routed_semi_finished_workbook_bytes(shaoyang_qty=10):
    wb = openpyxl.Workbook()
    inbound = wb.active
    inbound.title = "入库明细"
    inbound["D1"] = "日期"
    inbound["E1"] = "2026-07-01"
    inbound["D2"] = "入库单号"
    inbound["E2"] = "RK-ROUTE-001"
    inbound["A3"] = "物料名称"
    inbound["B3"] = "当月入仓总数"
    inbound["A4"] = "36#NFC\n贴纸"
    inbound["B4"] = 100
    inbound["E4"] = 100
    inbound["A5"] = "37#NFC\n贴纸"
    inbound["B5"] = 50
    inbound["E5"] = 50

    shaoyang = wb.create_sheet("邵阳领料")
    shaoyang["D1"] = "2026-07-02"
    shaoyang["D2"] = "LL-SY-001"
    shaoyang["A5"] = "物料名称"
    shaoyang["B5"] = "当月出仓总数"
    shaoyang["A6"] = "36#NFC\n贴纸"
    shaoyang["B6"] = shaoyang_qty
    shaoyang["D6"] = shaoyang_qty

    heyuan = wb.create_sheet("河源华兴36#CD领料")
    heyuan["C1"] = "2026-07-02"
    heyuan["C2"] = "LL-HY-001"
    heyuan["A3"] = "供应商"
    heyuan["B3"] = "上级名称"
    heyuan["C3"] = "河源华兴"
    heyuan["A4"] = "数量"
    heyuan["B4"] = 20
    heyuan["C4"] = 20

    workshop = wb.create_sheet("车间36#CD领料")
    workshop["C1"] = "2026-07-02"
    workshop["C2"] = "LL-DG-001"
    workshop["A3"] = "供应商"
    workshop["B3"] = "上级名称"
    workshop["C3"] = "HD"
    workshop["A4"] = "25#NFC贴纸"
    workshop["A5"] = "36#NFC贴纸"
    workshop["B5"] = 30
    workshop["C5"] = 30
    workshop["A6"] = "44#NFC贴纸"

    total = wb.create_sheet("总表", 0)
    total["A1"] = "总表不用导入"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def legacy_assembly_nfc_workbook_bytes():
    wb = openpyxl.Workbook()
    total_ws = wb.active
    total_ws.title = "总表"
    total_ws["A2"] = "物料名称"
    total_ws["B1"] = "累计入仓总数"
    total_ws["C1"] = "东莞"
    total_ws["C2"] = "截止6月27号"
    total_ws["D1"] = "7月领料\n总数"
    total_ws["E1"] = "2026-07-01"
    total_ws["A3"] = "1#NFC\n贴纸"
    total_ws["B3"] = 100
    total_ws["C3"] = 40
    total_ws["D3"] = 60

    issue_ws = wb.create_sheet("领料明细")
    issue_ws["B1"] = "当月领料总数"
    issue_ws["C1"] = "6月29日"
    issue_ws["A2"] = "物料名称"
    issue_ws["C2"] = "DG-LL-001"
    issue_ws["A3"] = "1#NFC\n贴纸"
    issue_ws["B3"] = 60
    issue_ws["C3"] = 60

    semi_ws = wb.create_sheet("半成品入仓明细")
    semi_ws["B1"] = "当月入仓总数"
    semi_ws["C1"] = "6月30日"
    semi_ws["A2"] = "物料名称"
    semi_ws["C2"] = "DG-RK-001"
    semi_ws["A3"] = "1#NFC\n贴纸"
    semi_ws["B3"] = 25
    semi_ws["C3"] = 25

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def legacy_assembly_pcba_workbook_bytes():
    wb = openpyxl.Workbook()
    total_ws = wb.active
    total_ws.title = "总表"
    total_ws.append(["物料名称", "领料总数", "截6月月结", "7月"])
    total_ws.append(["PCBA主板", 20800, None, 20800])

    issue_ws = wb.create_sheet("PCB主板领料明细")
    issue_ws.append(["日期", "领料编号", "物料名称", "领料数", "备注"])
    issue_ws.append(["2026-07-06", 2202830, "PCBA主板", 20800, None])

    finished_ws = wb.create_sheet("成品入仓明细")
    finished_ws.append(
        ["日期", "送货单号", "合同号", "货号", "品名/规格", "数量（pcs）", "备注"]
    )
    finished_ws.append(
        ["2026-07-07", 2510160, 4500204119, "MSLD182-77794", None, 11328, None]
    )
    finished_ws.append(
        ["2026-07-09", 2510080, 4500204119, "MSLD182-77794", None, 17088, None]
    )
    finished_ws.append(
        ["2026-07-09", 2510079, 4500204118, "MSLD182-77794", None, 17280, None]
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def legacy_outsource_workbook_bytes():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "领料明细"
    ws.append(["日期", "领料编号", "物料名称", "领料数", "备注"])
    ws.append(["截止到6月17号", None, "77794-PCBA板", 90, None])
    ws.append(["2026-07-01", 2518831, "77794-PCBA板", 10, "7月领料"])
    ws.append([None, None, "7月小计：", 10, None])

    inbound_ws = wb.create_sheet("半成品入仓明细")
    inbound_ws.append(["日期", "送货单号", "合同号", "货号", "品名/规格", "数量（pcs）", "备注"])
    inbound_ws.append(["2026-07-03", "BS2026070301", "LH202607", 77794, "光身唱机", 8, "返仓备注"])
    inbound_ws.append([None, None, None, None, "7月小计：", 8, None])

    total_ws = wb.create_sheet("总表")
    total_ws.append(["物料名称", None, "累计出入数", "截6月月结", "7月"])
    total_ws.append(["PCBA主板", "领料总数", 100, 90, 10])
    total_ws.append([None, "半成品入仓总数", 8, 0, 8])
    total_ws.append([None, "应存数", 92, 90, 2])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def legacy_heyuan_workbook_bytes():
    wb = openpyxl.Workbook()
    total_ws = wb.active
    total_ws.title = "总表"
    total_ws.append(["物料名称", "领料总数", "截6月月结", "7月"])
    total_ws.append(["PCBA主板", 125, 100, 25])
    total_ws.append(["36#唱片CD", 75, 50, 25])

    ws = wb.create_sheet("PCB板领料明细")
    ws.append(["日期", "领料编号", "物料名称", "领料数", "备注"])
    ws.append(["2026-06-20", 2519360, "77794-PCBA板", 100, None])
    ws.append([None, None, "6月小计：", 100, None])
    ws.append(["2026-06-30", 2518791, "77794-PCBA板", 30, None])
    ws.append(["2026-07-02", "退不良品", "77794-PCBA板", -5, None])
    ws.append([None, None, "7月小计：", 25, None])

    cd_ws = wb.create_sheet("36#CD领料明细")
    cd_ws.append(["日期", "领料编号", "物料名称", "领料数", "备注"])
    cd_ws.append(["6月", "月结", "36#唱片CD", 50, None])
    cd_ws.append([None, None, "6月小计：", 50, None])
    cd_ws.append(["2026-07-08", 2611533, "36#唱片CD", 25, None])
    cd_ws.append([None, None, "7月小计：", 25, None])

    finished_ws = wb.create_sheet("成品入仓明细")
    finished_ws.append(["日期", "送货单号", "合同号", "货号", "品名/规格", "数量（pcs）", "备注"])
    finished_ws.append(["2026-07-03", "HY2026070301", "HYHT", 77794, "成品唱机", 20, None])
    finished_ws.append([None, None, None, None, "7月小计：", 20, None])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def legacy_supplier_pcba_workbook_bytes():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "入仓明细"
    ws.append(["日期", "入仓单号", "物料名称", "入仓数", "备注"])
    ws.append(["截止到6月17号", None, "77794-PCBA板", 100, None])
    ws.append(["2026-07-01", 2534693, "77794-PCBA板", 40, "转单：A"])
    ws.append(["2026-07-01", 2534693, "77794-PCBA板", 40, "转单：B"])
    ws.append(["2026-07-02", 2651574, "77794-PCBA板", -5, "退货"])
    ws.append([None, None, "7月小计：", 75, None])

    issue_ws = wb.create_sheet("河源华兴领料")
    issue_ws.append(["日期", "领料单号", "物料名称", "领料数", "备注"])
    issue_ws.append(["2026-06-30", 2518000, "77794-PCBA板", 18, None])
    issue_ws.append(["2026-07-02", 2518791, "77794-PCBA板", 30, None])
    issue_ws.append(["2026-07-03", "退不良品", "77794-PCBA板", -5, None])
    issue_ws.append([None, None, "7月小计：", 43, None])

    total_ws = wb.create_sheet("总表")
    total_ws.append(["物料名称", None, "累计出入总数", "6月月结", "7月"])
    total_ws.append(["PCB主板", "入仓总数", 175, 100, 75])
    total_ws.append(["PCB主板", "河源华兴领料总数", 43, None, 43])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def legacy_supplier_nfc_workbook_bytes():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "总表"
    ws.cell(1, 2).value = "累计入仓总数"
    ws.cell(2, 1).value = "物料名称"
    ws.cell(2, 3).value = "截止6月27号"
    ws.cell(1, 4).value = "7月入仓\n总数"
    ws.cell(1, 11).value = "应存数"
    ws.cell(1, 12).value = "累计出仓总数"
    ws.cell(1, 13).value = "东莞"
    ws.cell(2, 13).value = "截止6月27号"
    ws.cell(1, 14).value = "邵阳领料"
    ws.cell(1, 15).value = "7月出仓\n总数"
    ws.cell(3, 1).value = "1#NFC\n贴纸"
    ws.cell(3, 2).value = 180
    ws.cell(3, 3).value = 100
    ws.cell(3, 4).value = 80
    ws.cell(3, 11).value = 75
    ws.cell(3, 12).value = 105
    ws.cell(3, 13).value = 60
    ws.cell(3, 14).value = 15
    ws.cell(3, 15).value = 30

    inbound_ws = wb.create_sheet("入库明细")
    inbound_ws.cell(1, 2).value = "当月入仓总数"
    inbound_ws.cell(1, 3).value = "2026-07-01"
    inbound_ws.cell(1, 4).value = "2026-07-02"
    inbound_ws.cell(2, 1).value = "物料名称"
    inbound_ws.cell(2, 3).value = "RK-1"
    inbound_ws.cell(2, 4).value = "RK-2"
    inbound_ws.cell(3, 1).value = "1#NFC\n贴纸"
    inbound_ws.cell(3, 2).value = 80
    inbound_ws.cell(3, 3).value = 50
    inbound_ws.cell(3, 4).value = 30
    inbound_ws.cell(4, 1).value = "小计："
    inbound_ws.cell(4, 2).value = 80

    outbound_ws = wb.create_sheet("出库明细")
    outbound_ws.cell(1, 2).value = "当月出仓总数"
    outbound_ws.cell(1, 3).value = "2026-07-03"
    outbound_ws.cell(2, 1).value = "物料名称"
    outbound_ws.cell(2, 3).value = "CK-1"
    outbound_ws.cell(3, 1).value = "1#NFC\n贴纸"
    outbound_ws.cell(3, 2).value = 30
    outbound_ws.cell(3, 3).value = 30

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def legacy_hongya_nfc_workbook_bytes():
    wb = openpyxl.Workbook()
    total_ws = wb.active
    total_ws.title = "总表"
    total_ws.cell(1, 2).value = "累计领料总数"
    total_ws.cell(1, 4).value = "7月领料\n总数"
    total_ws.cell(1, 11).value = "应存数"
    total_ws.cell(1, 12).value = "累计入仓总数"
    total_ws.cell(1, 14).value = "7月入仓\n总数"
    total_ws.cell(2, 1).value = "物料名称"
    total_ws.cell(3, 1).value = "1#NFC\n贴纸"
    total_ws.cell(3, 2).value = 30
    total_ws.cell(3, 4).value = 30
    total_ws.cell(3, 11).value = 10
    total_ws.cell(3, 12).value = 20
    total_ws.cell(3, 14).value = 20

    issue_ws = wb.create_sheet("领料明细")
    issue_ws.cell(1, 2).value = "当月领料总数"
    issue_ws.cell(1, 3).value = "2026-07-01"
    issue_ws.cell(1, 4).value = "2026-07-01"
    issue_ws.cell(2, 1).value = "物料名称"
    issue_ws.cell(3, 1).value = "1#NFC\n贴纸"
    issue_ws.cell(3, 2).value = 30
    issue_ws.cell(3, 3).value = 10
    issue_ws.cell(3, 4).value = 20

    inbound_ws = wb.create_sheet("入仓明细")
    inbound_ws.cell(1, 2).value = "当月入仓总数"
    inbound_ws.cell(1, 3).value = "2026-07-02"
    inbound_ws.cell(1, 4).value = "2026-07-02"
    inbound_ws.cell(2, 1).value = "物料名称"
    inbound_ws.cell(3, 1).value = "1#NFC\n贴纸"
    inbound_ws.cell(3, 2).value = 20
    inbound_ws.cell(3, 3).value = 5
    inbound_ws.cell(3, 4).value = 15

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def upload_bytes(client, path, content, filename="legacy.xlsx", data=None):
    return client.post(
        path,
        data=data or {},
        files={
            "file": (
                filename,
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )


def test_record_import_template_exports_expected_headers(client):
    login(client)

    r = client.get("/api/records/import-template")
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    headers = [cell.value for cell in wb.active[1]]

    assert r.status_code == 200
    assert headers == [
        "类型", "物料名称", "贴纸类型", "加工点", "供应商",
        "日期", "单据编号", "数量", "备注", "PO", "客名",
    ]


def test_non_supplier_record_export_matches_import_template_headers(client):
    login(client, "东莞车间", "123456", "东莞车间")
    dongguan = loc_id(client, "东莞加工厂利鸿")
    client.post("/api/records", json={
        "rec_type": "issue",
        "location_id": dongguan,
        "material": "77794-PCBA板",
        "rec_date": "2026-07-08",
        "doc_no": "EXP-001",
        "qty": 18,
        "remark": "导出测试",
    })

    r = client.get("/api/records/export")
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    first_row = [cell.value for cell in ws[2]]

    assert r.status_code == 200
    assert headers == [
        "类型", "物料名称", "贴纸类型", "加工点", "供应商",
        "日期", "单据编号", "数量", "备注", "PO", "客名",
    ]
    assert first_row == [
        "领料", "77794-PCBA板", None, "东莞加工厂利鸿", None,
        "2026-07-08", "EXP-001", 18, "导出测试", None, None,
    ]


def test_xingxin_pcba_export_uses_legacy_warehouse_workbook(client):
    login(client, "兴信B来料仓", "123456", DEFAULT_DEPARTMENT)
    heyuan = loc_id(client, "河源华兴")
    client.post("/api/records", json={
        "rec_type": "inbound_raw",
        "material": "77794-PCBA板",
        "rec_date": "2026-07-01",
        "doc_no": "RK-PCBA-1",
        "qty": 40,
        "remark": "入仓测试",
    })
    client.post("/api/records", json={
        "rec_type": "issue",
        "location_id": heyuan,
        "material": "77794-PCBA板",
        "rec_date": "2026-07-02",
        "doc_no": "LL-PCBA-1",
        "qty": 25,
        "remark": "领料测试",
    })

    r = client.get("/api/records/export?material=77794-PCBA板")
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)

    assert r.status_code == 200
    assert wb.sheetnames[:8] == [
        "总表", "入仓明细", "邵阳华登领料", "河源华兴领料",
        "加工厂利鸿领料", "加工厂鸿亚领料", "东莞车间领料", "Sheet2",
    ]
    assert [cell.value for cell in wb["入仓明细"][1]] == [
        "日期", "入仓单号", "物料名称", "入仓数", "备注"]
    assert [cell.value for cell in wb["河源华兴领料"][1]] == [
        "日期", "领料单号", "物料名称", "领料数", "备注"]
    assert wb["入仓明细"].cell(2, 2).value == "RK-PCBA-1"
    assert wb["入仓明细"].cell(2, 4).value == 40
    assert wb["河源华兴领料"].cell(2, 2).value == "LL-PCBA-1"
    assert wb["河源华兴领料"].cell(2, 4).value == 25
    total = wb["总表"]
    assert [total.cell(1, col).value for col in range(1, 12)] == [
        "物料名称", None, "累计出入总数", "6月月结", "7月", "8月",
        "9月", "10月", "11月", "12月", "备注",
    ]
    assert [total.cell(row, 2).value for row in range(2, 10)] == [
        "入仓总数", "邵阳华登领料总数", "河源华兴领料总数",
        "加工厂利鸿领料总数", "加工厂鸿亚领料总数", "东莞车间领料",
        None, "应存数",
    ]
    assert total.cell(2, 1).value == "PCB主板"
    assert total.cell(2, 3).value == 40
    assert total.cell(2, 5).value == 40
    assert total.cell(4, 3).value == 25
    assert total.cell(4, 5).value == 25
    assert total.cell(9, 3).value == 15
    assert total.cell(9, 4).value == 0
    assert total.cell(9, 5).value == 15


def test_xingxin_nfc_export_uses_legacy_matrix_workbook(client):
    login(client, "兴信B来料仓", "123456", DEFAULT_DEPARTMENT)
    dongguan = loc_id(client, "东莞车间")
    client.post("/api/records/batch", json={
        "rec_type": "inbound_raw",
        "material": "NFC贴纸",
        "rec_date": "2026-07-01",
        "doc_no": "RK-NFC-1",
        "items": [{"sticker_type": "1#NFC贴纸", "qty": 50}],
    })
    client.post("/api/records/batch", json={
        "rec_type": "issue",
        "location_id": dongguan,
        "material": "NFC贴纸",
        "rec_date": "2026-07-02",
        "doc_no": "CK-NFC-1",
        "items": [{"sticker_type": "1#NFC贴纸", "qty": 20}],
    })

    r = client.get("/api/records/export?material=NFC贴纸")
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)

    assert r.status_code == 200
    assert wb.sheetnames[:3] == ["总表", "入库明细", "出库明细"]
    assert wb["总表"].cell(1, 2).value == "累计入仓总数"
    assert wb["总表"].cell(2, 1).value == "物料名称"
    assert wb["总表"].cell(3, 1).value == "1#NFC\n贴纸"
    assert wb["总表"].cell(3, 2).value == 50
    assert wb["总表"].cell(3, 11).value == 30
    assert wb["总表"].cell(3, 12).value == 20
    assert wb["入库明细"].cell(1, 2).value == "当月入仓总数"
    assert wb["入库明细"].cell(2, 1).value == "物料名称"
    assert wb["入库明细"].cell(2, 3).value == "RK-NFC-1"
    assert wb["入库明细"].cell(3, 2).value == 50
    assert wb["入库明细"].cell(3, 3).value == 50
    assert wb["出库明细"].cell(1, 2).value == "当月出仓总数"
    assert wb["出库明细"].cell(2, 3).value == "CK-NFC-1"
    assert wb["出库明细"].cell(3, 2).value == 20
    assert wb["出库明细"].cell(3, 3).value == 20


def test_xingxin_nfc_record_export_fills_legacy_opening_date(client):
    login(client, "兴信B来料仓", "123456", DEFAULT_DEPARTMENT)
    dongguan = loc_id(client, "东莞车间")
    client.post("/api/records", json={
        "rec_type": "issue",
        "location_id": dongguan,
        "material": "NFC贴纸",
        "sticker_type": "1#NFC贴纸",
        "doc_no": "1#NFC贴纸-东莞期初出仓",
        "qty": 60,
        "remark": "总表东莞期初出仓导入",
    })

    r = client.get("/api/records/export?material=NFC贴纸")
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)

    assert r.status_code == 200
    assert wb["出库明细"].cell(1, 3).value == "2026-06-27"


def test_xingxin_nfc_export_groups_opening_stock_by_document_column(client):
    login(client, "兴信B来料仓", "123456", DEFAULT_DEPARTMENT)
    client.post("/api/records", json={
        "rec_type": "inbound_raw",
        "material": "NFC贴纸",
        "sticker_type": "1#NFC贴纸",
        "rec_date": "2026-06-27",
        "doc_no": "1#NFC贴纸-期初入仓",
        "qty": 846669,
        "remark": "总表期初入仓导入",
    })
    client.post("/api/records", json={
        "rec_type": "inbound_raw",
        "material": "NFC贴纸",
        "sticker_type": "2#NFC贴纸",
        "rec_date": "2026-06-27",
        "doc_no": "2#NFC贴纸-期初入仓",
        "qty": 865000,
        "remark": "总表期初入仓导入",
    })

    r = client.get("/api/records/export?material=NFC贴纸")
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    ws = wb["入库明细"]

    assert r.status_code == 200
    assert ws.cell(1, 3).value == "2026-06-27"
    assert ws.cell(2, 3).value == "期初入仓"
    assert ws.cell(3, 3).value == 846669
    assert ws.cell(4, 3).value == 865000
    assert ws.cell(1, 4).value is None
    assert ws.cell(2, 4).value is None


def test_xingxin_nfc_import_does_not_treat_quantity_as_workbook_year(client):
    login(client, "兴信B来料仓", "123456", DEFAULT_DEPARTMENT)
    wb = openpyxl.load_workbook(io.BytesIO(legacy_supplier_nfc_workbook_bytes()))
    wb["总表"].cell(3, 2).value = 49855
    wb["总表"].cell(3, 3).value = 49855
    content = io.BytesIO()
    wb.save(content)

    imported = upload_bytes(
        client,
        "/api/records/import",
        content.getvalue(),
        "来料仓77772#NFC出入明细.xlsx",
    )
    assert imported.status_code == 200

    exported = client.get("/api/records/export?material=NFC贴纸")
    exported_wb = openpyxl.load_workbook(
        io.BytesIO(exported.content), data_only=True
    )
    assert exported_wb["入库明细"].cell(1, 3).value == "2026-06-27"


def test_xingxin_nfc_export_sorts_columns_by_date_and_document(client):
    login(client, "兴信B来料仓", "123456", DEFAULT_DEPARTMENT)
    documents = [
        ("2026-07-11", "RK-10"),
        ("2026-06-29", "RK-2"),
        ("2026-07-04", "RK-1"),
        ("2026-07-11", "RK-2"),
    ]
    for index, (rec_date, doc_no) in enumerate(documents, start=1):
        response = client.post("/api/records", json={
            "rec_type": "inbound_raw",
            "material": "NFC贴纸",
            "sticker_type": "1#NFC贴纸",
            "rec_date": rec_date,
            "doc_no": doc_no,
            "qty": index,
        })
        assert response.status_code == 200

    exported = client.get("/api/records/export?material=NFC贴纸")
    wb = openpyxl.load_workbook(io.BytesIO(exported.content), data_only=True)
    ws = wb["入库明细"]

    assert [ws.cell(1, col).value for col in range(3, 7)] == [
        "2026-06-29", "2026-07-04", "2026-07-11", "2026-07-11",
    ]
    assert [ws.cell(2, col).value for col in range(5, 7)] == [
        "RK-2", "RK-10",
    ]


def test_xingxin_nfc_export_repairs_legacy_future_opening_date(client):
    login(client, "兴信B来料仓", "123456", DEFAULT_DEPARTMENT)
    response = client.post("/api/records", json={
        "rec_type": "inbound_raw",
        "material": "NFC贴纸",
        "sticker_type": "1#NFC贴纸",
        "rec_date": "2036-06-27",
        "doc_no": "1#NFC贴纸-期初入仓",
        "qty": 100,
        "remark": "总表期初入仓导入",
    })
    assert response.status_code == 200

    exported = client.get("/api/records/export?material=NFC贴纸")
    wb = openpyxl.load_workbook(io.BytesIO(exported.content), data_only=True)
    assert wb["入库明细"].cell(1, 3).value == "2026-06-27"


def test_semi_finished_export_uses_legacy_matrix_workbook(client):
    login(client, "碟片半成品", "123456", "碟片半成品")
    upload = upload_bytes(
        client,
        "/api/records/import",
        legacy_semi_finished_workbook_bytes(),
        "塑胶仓77772#CD半成品出入明细.xlsx",
    )
    assert upload.status_code == 200

    r = client.get("/api/records/export?material=NFC贴纸")
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)

    assert r.status_code == 200
    assert wb.sheetnames[:5] == [
        "总表", "入库明细", "邵阳领料", "河源华兴36#CD领料", "车间36#CD领料"]
    assert wb["总表"].cell(1, 2).value == "累计入仓总数"
    assert wb["总表"].cell(1, 11).value == "应存数"
    assert wb["总表"].cell(1, 14).value == "湖南"
    assert wb["总表"].cell(3, 1).value == "1#NFC\n贴纸"
    assert wb["总表"].cell(3, 2).value == 100
    assert wb["总表"].cell(3, 11).value == 60
    assert wb["总表"].cell(3, 12).value == 40

    inbound = wb["入库明细"]
    assert inbound.cell(1, 4).value == "日期"
    assert inbound.cell(2, 4).value == "入库单号"
    assert inbound.cell(3, 1).value == "物料名称"
    assert inbound.cell(3, 2).value == "当月入仓总数"
    assert inbound.cell(3, 3).value == "6/24\n东莞车间入库截数"
    assert inbound.cell(4, 1).value == "1#NFC\n贴纸"
    assert inbound.cell(4, 2).value == 100
    assert inbound.cell(3, 4).value == "6/24\n鸿亚入库截数"
    assert inbound.cell(4, 3).value == 10
    assert inbound.cell(4, 4).value == 15
    assert inbound.cell(1, 5).value == "2026-07-01"
    assert inbound.cell(2, 5).value == "RK-001"
    assert inbound.cell(4, 5).value == 30

    outbound = wb["邵阳领料"]
    assert outbound.cell(5, 1).value == "物料名称"
    assert outbound.cell(5, 2).value == "当月出仓总数"
    assert outbound.cell(6, 1).value == "1#NFC\n贴纸"
    assert outbound.cell(6, 2).value == 40
    assert outbound.cell(6, 3).value == 25
    assert outbound.cell(1, 4).value == "2026-07-03"
    assert outbound.cell(2, 4).value == "LL-001"
    assert outbound.cell(6, 4).value == 35


def test_semi_finished_import_routes_outbound_sheets_to_departments(client):
    login(client, "碟片半成品", "123456", "碟片半成品")

    imported = upload_bytes(
        client,
        "/api/records/import",
        routed_semi_finished_workbook_bytes(),
        "塑胶仓77772#CD半成品出入明细.xlsx",
    )
    records = client.get("/api/records").json()
    outbound = {
        row["doc_no"]: row["location_name"]
        for row in records
        if row["rec_type"] == "semi_outbound"
    }

    assert imported.status_code == 200
    assert outbound == {
        "LL-SY-001": "邵阳华登",
        "LL-HY-001": "河源华兴",
        "LL-DG-001": "东莞车间",
    }


def test_record_import_check_groups_duplicate_documents_without_writing(client):
    login(client, "碟片半成品", "123456", "碟片半成品")
    content = routed_semi_finished_workbook_bytes()
    first = upload_bytes(
        client,
        "/api/records/import",
        content,
        "塑胶仓77772#CD半成品出入明细.xlsx",
    )
    before = client.get("/api/records").json()

    checked = upload_bytes(
        client,
        "/api/records/import-check",
        content,
        "塑胶仓77772#CD半成品出入明细.xlsx",
    )
    after = client.get("/api/records").json()

    assert first.status_code == 200
    assert checked.status_code == 200
    result = checked.json()
    assert result["documents"] == 4
    assert result["duplicates"] == 4
    assert {row["doc_no"] for row in result["duplicate_documents"]} == {
        "RK-ROUTE-001", "LL-SY-001", "LL-HY-001", "LL-DG-001",
    }
    inbound_duplicate = next(
        row for row in result["duplicate_documents"]
        if row["doc_no"] == "RK-ROUTE-001"
    )
    assert inbound_duplicate["file_rows"] == 2
    assert len(after) == len(before)


def test_record_import_skip_mode_skips_whole_duplicate_documents(client):
    login(client, "碟片半成品", "123456", "碟片半成品")
    content = routed_semi_finished_workbook_bytes()
    first = upload_bytes(
        client,
        "/api/records/import",
        content,
        "塑胶仓77772#CD半成品出入明细.xlsx",
    )
    second = upload_bytes(
        client,
        "/api/records/import",
        content,
        "塑胶仓77772#CD半成品出入明细.xlsx",
        data={"mode": "skip"},
    )

    assert first.status_code == 200
    assert first.json()["created"] == 5
    assert second.status_code == 200
    assert second.json()["created"] == 0
    assert second.json()["skipped_documents"] == 4
    assert len(client.get("/api/records").json()) == 5


def test_record_import_replace_mode_replaces_document_and_auto_records(client):
    login(client, "碟片半成品", "123456", "碟片半成品")
    filename = "塑胶仓77772#CD半成品出入明细.xlsx"
    first = upload_bytes(
        client,
        "/api/records/import",
        routed_semi_finished_workbook_bytes(shaoyang_qty=10),
        filename,
    )
    replaced = upload_bytes(
        client,
        "/api/records/import",
        routed_semi_finished_workbook_bytes(shaoyang_qty=15),
        filename,
        data={"mode": "replace"},
    )
    source_rows = client.get("/api/records").json()

    assert first.status_code == 200
    assert replaced.status_code == 200
    assert replaced.json()["replaced_documents"] == 4
    shaoyang_source = [
        row for row in source_rows if row["doc_no"] == "LL-SY-001"
    ]
    assert len(shaoyang_source) == 1
    assert shaoyang_source[0]["qty"] == 15

    login(client, "邵阳华登", "123456", "邵阳华登")
    target_rows = [
        row for row in client.get("/api/records").json()
        if row["doc_no"] == "LL-SY-001"
    ]
    assert len(target_rows) == 1
    assert target_rows[0]["qty"] == 15
    assert target_rows[0]["source_record_id"] is not None


def test_record_import_replace_allows_same_department_records_owned_by_another_user(client):
    login(client)
    created = client.post("/api/records", json={
        "rec_type": "inbound_raw",
        "material": "77794-PCBA板",
        "rec_date": "2026-07-01",
        "doc_no": "ADMIN-OWNED-001",
        "qty": 10,
    })
    assert created.status_code == 200

    client.post("/api/logout")
    login(client, "兴信B来料仓", "123456", DEFAULT_DEPARTMENT)
    replacing = workbook_bytes(
        ["类型", "物料名称", "日期", "单据编号", "数量"],
        [["入库", "77794-PCBA板", "2026-07-01", "ADMIN-OWNED-001", 20]],
    )
    replaced = upload_bytes(
        client,
        "/api/records/import",
        replacing,
        data={"mode": "replace"},
    )

    assert replaced.status_code == 200
    rows = client.get("/api/records?doc_no=ADMIN-OWNED-001").json()
    assert len(rows) == 1
    assert rows[0]["qty"] == 20


def test_record_import_same_document_number_on_different_dates_is_duplicate(client):
    login(client, "兴信B来料仓", "123456", DEFAULT_DEPARTMENT)
    headers = ["类型", "物料名称", "日期", "单据编号", "数量"]
    first = upload_xlsx(
        client,
        "/api/records/import",
        headers,
        [["入库", "77794-PCBA板", "2026-07-01", "PERIODIC-001", 10]],
    )
    second_content = workbook_bytes(
        headers,
        [["入库", "77794-PCBA板", "2026-08-01", "PERIODIC-001", 20]],
    )
    second = upload_bytes(
        client,
        "/api/records/import",
        second_content,
        data={"mode": "skip"},
    )

    assert first.status_code == 200
    assert second.status_code == 200
    assert second.json()["created"] == 0
    assert second.json()["skipped_documents"] == 1
    rows = client.get("/api/records?doc_no=PERIODIC-001").json()
    assert [(row["rec_date"], row["qty"]) for row in rows] == [
        ("2026-07-01", 10),
    ]


def test_semi_finished_export_splits_outbound_by_target_department(client):
    login(client, "碟片半成品", "123456", "碟片半成品")
    locations = {
        name: loc_id(client, name)
        for name in ("邵阳华登", "河源华兴", "东莞车间")
    }
    for location_name, doc_no, qty in (
        ("邵阳华登", "LL-SY-001", 10),
        ("河源华兴", "LL-HY-001", 20),
        ("东莞车间", "LL-DG-001", 30),
    ):
        response = client.post("/api/records", json={
            "rec_type": "semi_outbound",
            "location_id": locations[location_name],
            "material": "NFC贴纸",
            "sticker_type": "36#NFC贴纸",
            "rec_date": "2026-07-02",
            "doc_no": doc_no,
            "qty": qty,
        })
        assert response.status_code == 200

    exported = client.get("/api/records/export?material=NFC贴纸")
    wb = openpyxl.load_workbook(io.BytesIO(exported.content), data_only=True)

    def sticker_qty(sheet_name):
        ws = wb[sheet_name]
        row_no = next(
            row for row in range(1, ws.max_row + 1)
            if ws.cell(row, 1).value == "36#NFC\n贴纸"
        )
        return ws.cell(row_no, 4).value

    assert wb["邵阳领料"].cell(2, 4).value == "LL-SY-001"
    assert sticker_qty("邵阳领料") == 10
    assert wb["河源华兴36#CD领料"].cell(2, 4).value == "LL-HY-001"
    assert sticker_qty("河源华兴36#CD领料") == 20
    assert wb["车间36#CD领料"].cell(2, 4).value == "LL-DG-001"
    assert sticker_qty("车间36#CD领料") == 30


def test_outsource_record_export_uses_legacy_pcba_workbook(client):
    login(client, "东莞加工厂利鸿", "123456", "东莞加工厂利鸿")
    semi = loc_id(client, "碟片半成品")
    client.post("/api/records", json={
        "rec_type": "issue",
        "location_id": semi,
        "material": "77794-PCBA板",
        "rec_date": "2026-07-01",
        "doc_no": "LL-LH-001",
        "qty": 100,
    })
    client.post("/api/records", json={
        "rec_type": "semi_finished",
        "material": "77794-PCBA板",
        "rec_date": "2026-07-02",
        "doc_no": "BS-LH-001",
        "qty": 35,
        "remark": "LH202607 77794 光身唱机",
    })

    r = client.get("/api/records/export?material=77794-PCBA板")
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)

    assert r.status_code == 200
    assert wb.sheetnames[:3] == ["总表", "领料明细", "半成品入仓明细"]
    assert [cell.value for cell in wb["领料明细"][1][:5]] == [
        "日期", "领料编号", "物料名称", "领料数", "备注"]
    assert [cell.value for cell in wb["半成品入仓明细"][1][:7]] == [
        "日期", "送货单号", "合同号", "货号", "品名/规格", "数量（pcs）", "备注"]
    assert wb["领料明细"].cell(2, 2).value == "LL-LH-001"
    assert wb["领料明细"].cell(2, 4).value == 100
    assert wb["半成品入仓明细"].cell(2, 2).value == "BS-LH-001"
    assert wb["半成品入仓明细"].cell(2, 6).value == 35
    assert wb["总表"].cell(2, 2).value == "领料总数"
    assert wb["总表"].cell(2, 3).value == 100
    assert wb["总表"].cell(3, 2).value == "半成品入仓总数"
    assert wb["总表"].cell(3, 3).value == 35


def test_heyuan_record_export_uses_legacy_pcba_workbook(client):
    login(client, "河源华兴", "123456", "河源华兴")
    heyuan = loc_id(client, "河源华兴")
    client.post("/api/records", json={
        "rec_type": "issue",
        "location_id": heyuan,
        "material": "77794-PCBA板",
        "rec_date": "2026-06-20",
        "doc_no": "LL-HY-JUNE",
        "qty": 40,
    })
    client.post("/api/records", json={
        "rec_type": "issue",
        "location_id": heyuan,
        "material": "77794-PCBA板",
        "rec_date": "2026-07-03",
        "doc_no": "LL-HY-001",
        "qty": 88,
    })
    client.post("/api/records", json={
        "rec_type": "issue",
        "location_id": heyuan,
        "material": "NFC贴纸",
        "sticker_type": "36#NFC贴纸",
        "rec_date": "2026-06-27",
        "doc_no": "月结",
        "qty": 50,
    })
    client.post("/api/records", json={
        "rec_type": "issue",
        "location_id": heyuan,
        "material": "NFC贴纸",
        "sticker_type": "36#NFC贴纸",
        "rec_date": "2026-07-08",
        "doc_no": "2611533",
        "qty": 25,
    })
    client.post("/api/records", json={
        "rec_type": "finished",
        "location_id": heyuan,
        "material": "77794-PCBA板",
        "rec_date": "2026-07-04",
        "doc_no": "BS-HY-001",
        "qty": 66,
        "remark": "HY202607 77794 光身唱机",
    })

    r = client.get("/api/records/export?material=77794-PCBA板")
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)

    assert r.status_code == 200
    assert wb.sheetnames[:4] == ["总表", "36#CD领料明细", "PCB板领料明细", "成品入仓明细"]
    assert [cell.value for cell in wb["PCB板领料明细"][1][:5]] == [
        "日期", "领料编号", "物料名称", "领料数", "备注"]
    assert [cell.value for cell in wb["成品入仓明细"][1][:7]] == [
        "日期", "送货单号", "合同号", "货号", "品名/规格", "数量（pcs）", "备注"]
    assert wb["PCB板领料明细"].cell(2, 2).value == "LL-HY-JUNE"
    assert wb["PCB板领料明细"].cell(2, 4).value == 40
    assert wb["PCB板领料明细"].cell(5, 3).value == "6月小计："
    assert wb["PCB板领料明细"].cell(5, 4).value == 40
    assert wb["PCB板领料明细"].cell(6, 2).value == "LL-HY-001"
    assert wb["PCB板领料明细"].cell(6, 4).value == 88
    assert wb["PCB板领料明细"].cell(22, 3).value == "7月小计："
    assert wb["PCB板领料明细"].cell(22, 4).value == 88
    assert wb["36#CD领料明细"].cell(2, 2).value == "月结"
    assert wb["36#CD领料明细"].cell(2, 4).value == 50
    assert wb["36#CD领料明细"].cell(3, 3).value == "6月小计："
    assert wb["36#CD领料明细"].cell(4, 2).value == "2611533"
    assert wb["36#CD领料明细"].cell(4, 4).value == 25
    assert wb["36#CD领料明细"].cell(18, 3).value == "7月小计："
    assert wb["36#CD领料明细"].cell(18, 4).value == 25
    assert wb["成品入仓明细"].cell(2, 2).value == "BS-HY-001"
    assert wb["成品入仓明细"].cell(2, 6).value == 66
    assert wb["总表"].cell(2, 1).value == "PCBA主板"
    assert wb["总表"].cell(2, 2).value == 128
    assert wb["总表"].cell(2, 3).value == 40
    assert wb["总表"].cell(2, 4).value == 88
    assert wb["总表"].cell(3, 1).value == "36#唱片CD"
    assert wb["总表"].cell(3, 2).value == 75
    assert wb["总表"].cell(3, 3).value == 50
    assert wb["总表"].cell(3, 4).value == 25
    assert wb["总表"].cell(5, 2).value == "成品总数"
    assert wb["总表"].cell(6, 2).value == 66
    assert wb["总表"].cell(9, 2).value == "理论结存数"
    assert wb["总表"].cell(10, 2).value == 62
    assert wb["总表"].cell(11, 2).value == 75


def test_heyuan_export_naturally_sorts_same_day_document_numbers(client):
    login(client, "河源华兴", "123456", "河源华兴")
    heyuan = loc_id(client, "河源华兴")
    for doc_no in ("LL-10", "LL-2"):
        response = client.post("/api/records", json={
            "rec_type": "issue",
            "location_id": heyuan,
            "material": "77794-PCBA板",
            "rec_date": "2026-07-08",
            "doc_no": doc_no,
            "qty": 10,
        })
        assert response.status_code == 200
    for doc_no in ("BS-10", "BS-2"):
        response = client.post("/api/records", json={
            "rec_type": "finished",
            "location_id": heyuan,
            "material": "77794-PCBA板",
            "rec_date": "2026-07-09",
            "doc_no": doc_no,
            "qty": 5,
        })
        assert response.status_code == 200

    exported = client.get("/api/records/export?material=77794-PCBA板")
    wb = openpyxl.load_workbook(io.BytesIO(exported.content), data_only=True)
    issue_docs = [
        row[1]
        for row in wb["PCB板领料明细"].iter_rows(min_row=2, values_only=True)
        if isinstance(row[1], str) and row[1].startswith("LL-")
    ]
    finished_docs = [
        row[1]
        for row in wb["成品入仓明细"].iter_rows(min_row=2, values_only=True)
        if isinstance(row[1], str) and row[1].startswith("BS-")
    ]

    assert issue_docs == ["LL-2", "LL-10"]
    assert finished_docs == ["BS-2", "BS-10"]


def test_heyuan_empty_finished_export_keeps_header_only(client):
    login(client, "河源华兴", "123456", "河源华兴")

    r = client.get("/api/records/export?material=77794-PCBA板")
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)

    assert r.status_code == 200
    assert wb["成品入仓明细"].max_row == 1


def test_operator_can_import_record_rows_from_xlsx(client):
    login(client, "兴信B来料仓", "123456", DEFAULT_DEPARTMENT)

    r = upload_xlsx(
        client,
        "/api/records/import",
        ["类型", "物料名称", "贴纸类型", "加工点", "供应商", "日期", "单据编号", "数量", "备注"],
        [["入库", "NFC贴纸", "1#NFC\n贴纸", None, "供应商A", "2026-07-08", "IMP-001", 18, "导入测试"]],
    )
    records = client.get("/api/records?doc_no=IMP-001").json()

    assert r.status_code == 200
    assert r.json()["created"] == 1
    assert records[0]["material"] == "NFC贴纸"
    assert records[0]["sticker_type"] == "1#NFC贴纸"
    assert records[0]["qty"] == 18


def test_record_import_rejects_nfc_without_sticker_type(client):
    login(client, "兴信B来料仓", "123456", DEFAULT_DEPARTMENT)

    r = upload_xlsx(
        client,
        "/api/records/import",
        ["类型", "物料名称", "贴纸类型", "数量"],
        [["入库", "NFC贴纸", "", 18]],
    )

    assert r.status_code == 400
    assert "贴纸类型" in r.json()["detail"]


def test_semi_finished_legacy_workbook_imports_detail_rows_and_monthly_totals(client):
    login(client, "碟片半成品", "123456", "碟片半成品")

    r = upload_bytes(
        client,
        "/api/records/import",
        legacy_semi_finished_workbook_bytes(),
        "塑胶仓77772#CD半成品出入明细.xlsx",
    )
    records = client.get("/api/records").json()

    assert r.status_code == 200
    assert r.json()["created"] == 3
    assert r.json()["monthly_totals"] == 1
    rows = {(row["rec_type"], row["doc_no"]): row for row in records}
    assert rows[("semi_inbound", "RK-001")]["qty"] == 30
    assert rows[("semi_inbound", "RK-002")]["qty"] == 60
    assert rows[("semi_outbound", "LL-001")]["qty"] == 35
    assert rows[("semi_inbound", "RK-001")]["sticker_type"] == "1#NFC贴纸"

    public = client.get("/api/public-summary").json()
    semi_rows = public["semi_finished_monthly_totals"]
    assert semi_rows == [{
        "department": "碟片半成品",
        "material": "NFC贴纸",
        "sticker_type": "1#NFC贴纸",
        "opening_stock": 25,
        "monthly_inbound": 100,
        "monthly_outbound": 40,
        "monthly_balance": 85,
    }]

    second = upload_bytes(
        client,
        "/api/records/import",
        legacy_semi_finished_workbook_bytes(),
        "塑胶仓77772#CD半成品出入明细.xlsx",
    )
    assert second.status_code == 200
    assert second.json()["created"] == 0
    assert second.json()["skipped"] == 3
    assert len(client.get("/api/records").json()) == 3


def test_semi_finished_export_keeps_stored_total_on_matching_outbound_sheet(client):
    login(client, "碟片半成品", "123456", "碟片半成品")
    imported = upload_bytes(
        client,
        "/api/records/import",
        legacy_semi_finished_workbook_bytes(),
        "塑胶仓77772#CD半成品出入明细.xlsx",
    )
    exported = client.get("/api/records/export?material=NFC贴纸")
    wb = openpyxl.load_workbook(io.BytesIO(exported.content), data_only=True)

    assert imported.status_code == 200
    assert exported.status_code == 200
    assert wb["邵阳领料"].cell(6, 2).value == 40
    assert wb["河源华兴36#CD领料"].cell(6, 2).value == 0
    assert wb["车间36#CD领料"].cell(6, 2).value == 0


def test_semi_finished_legacy_workbook_rejects_filename_without_department(client):
    login(client, "碟片半成品", "123456", "碟片半成品")

    r = upload_bytes(
        client,
        "/api/records/import",
        legacy_semi_finished_workbook_bytes(),
        "legacy.xlsx",
    )

    assert r.status_code == 400
    assert "半成品" in r.json()["detail"]


def test_assembly_nfc_legacy_workbook_rejects_filename_without_department(client):
    login(client, "东莞车间", "123456", "东莞车间")

    r = upload_bytes(
        client,
        "/api/records/import",
        legacy_assembly_nfc_workbook_bytes(),
        "legacy.xlsx",
    )

    assert r.status_code == 400
    assert "东莞车间" in r.json()["detail"]


def test_assembly_nfc_legacy_workbook_imports_matrix_rows(client):
    login(client, "东莞车间", "123456", "东莞车间")

    r = upload_bytes(
        client,
        "/api/records/import",
        legacy_assembly_nfc_workbook_bytes(),
        "东莞车间77772#NFC贴纸出入明细.xlsx",
    )
    records = client.get("/api/records").json()

    assert r.status_code == 200
    assert r.json()["created"] == 3
    assert r.json()["monthly_totals"] == 0
    rows = {(row["rec_type"], row["doc_no"]): row for row in records}
    assert rows[("issue", "DG-LL-001")]["qty"] == 60
    assert rows[("issue", "DG-LL-001")]["rec_date"] == "2026-06-29"
    assert rows[("issue", "DG-LL-001")]["summary_month"] == 7
    assert rows[("issue", "DG-LL-001")]["material"] == "NFC贴纸"
    assert rows[("issue", "DG-LL-001")]["sticker_type"] == "1#NFC贴纸"
    assert rows[("issue", "DG-LL-001")]["location_name"] == "东莞车间"
    assert rows[("semi_finished", "DG-RK-001")]["qty"] == 25
    assert rows[("semi_finished", "DG-RK-001")]["rec_date"] == "2026-06-30"
    assert rows[("semi_finished", "DG-RK-001")]["summary_month"] == 7
    assert rows[("semi_finished", "DG-RK-001")]["location_name"] == "东莞车间"


def test_assembly_nfc_export_matches_import_workbook_format(client):
    login(client, "东莞车间", "123456", "东莞车间")
    imported = upload_bytes(
        client,
        "/api/records/import",
        legacy_assembly_nfc_workbook_bytes(),
        "东莞车间77772#NFC贴纸出入明细.xlsx",
    )

    exported = client.get("/api/records/export?material=NFC贴纸")
    wb = openpyxl.load_workbook(io.BytesIO(exported.content), data_only=True)

    assert imported.status_code == 200
    assert exported.status_code == 200
    assert wb.sheetnames == ["总表", "领料明细", "半成品入仓明细"]
    assert wb["总表"].cell(1, 2).value == "累计入仓总数"
    assert wb["总表"].cell(1, 4).value == "7月领料\n总数"
    assert wb["总表"].cell(1, 11).value == "应存数"
    assert wb["总表"].cell(1, 12).value == "累计出仓总数"
    assert wb["总表"].cell(1, 15).value == "7月入仓\n总数"
    assert wb["总表"].cell(3, 1).value == "1#NFC\n贴纸"
    assert wb["总表"].cell(3, 2).value == 100
    assert wb["总表"].cell(3, 3).value == 40
    assert wb["总表"].cell(3, 4).value == 60
    assert wb["总表"].cell(3, 11).value == 75
    assert wb["总表"].cell(3, 12).value == 25
    assert wb["总表"].cell(3, 15).value == 25

    issue = wb["领料明细"]
    assert issue.cell(1, 2).value == "当月领料总数"
    assert issue.cell(1, 3).value == "2026-06-29"
    assert issue.cell(2, 3).value == "DG-LL-001"
    assert issue.cell(3, 1).value == "1#NFC\n贴纸"
    assert issue.cell(3, 2).value == 60
    assert issue.cell(3, 3).value == 60

    inbound = wb["半成品入仓明细"]
    assert inbound.cell(1, 2).value == "当月入仓总数"
    assert inbound.cell(1, 3).value == "2026-06-30"
    assert inbound.cell(2, 3).value == "DG-RK-001"
    assert inbound.cell(3, 1).value == "1#NFC\n贴纸"
    assert inbound.cell(3, 2).value == 25
    assert inbound.cell(3, 3).value == 25

    reimported = upload_bytes(
        client,
        "/api/records/import",
        exported.content,
        "东莞车间77772#NFC贴纸出入明细.xlsx",
    )
    assert reimported.status_code == 200
    assert reimported.json()["created"] == 0
    assert reimported.json()["skipped"] == 3


def test_assembly_nfc_export_sorts_document_columns_by_date(client):
    login(client, "东莞车间", "123456", "东莞车间")
    assembly = loc_id(client, "东莞车间")
    dated_documents = [
        ("2026-07-10", "10"),
        ("2026-07-03", "10"),
        ("2026-06-29", "29"),
        ("2026-07-03", "2"),
        ("2026-07-01", "1"),
    ]
    for rec_type, prefix in (("issue", "LL"), ("semi_finished", "RK")):
        for index, (rec_date, document_number) in enumerate(
            dated_documents, start=1
        ):
            response = client.post("/api/records", json={
                "rec_type": rec_type,
                "location_id": assembly,
                "material": "NFC贴纸",
                "sticker_type": "1#NFC贴纸",
                "rec_date": rec_date,
                "doc_no": f"{prefix}-{document_number}",
                "qty": index,
                "summary_month": 7,
            })
            assert response.status_code == 200

    exported = client.get("/api/records/export?material=NFC贴纸")
    wb = openpyxl.load_workbook(io.BytesIO(exported.content), data_only=True)

    assert [wb["领料明细"].cell(1, col).value for col in range(3, 8)] == [
        "2026-06-29", "2026-07-01", "2026-07-03", "2026-07-03", "2026-07-10",
    ]
    assert [wb["领料明细"].cell(2, col).value for col in range(5, 7)] == [
        "LL-2", "LL-10",
    ]
    assert [
        wb["半成品入仓明细"].cell(1, col).value for col in range(3, 8)
    ] == [
        "2026-06-29", "2026-07-01", "2026-07-03", "2026-07-03", "2026-07-10",
    ]
    assert [
        wb["半成品入仓明细"].cell(2, col).value for col in range(5, 7)
    ] == ["RK-2", "RK-10"]


def test_assembly_pcba_legacy_workbook_imports_issue_and_finished_rows(client):
    login(client, "东莞车间", "123456", "东莞车间")

    r = upload_bytes(
        client,
        "/api/records/import",
        legacy_assembly_pcba_workbook_bytes(),
        "东莞车间77794#PCB主板出入明细.xlsx",
    )
    records = client.get("/api/records").json()

    assert r.status_code == 200
    assert r.json()["created"] == 4
    by_type_doc = {(row["rec_type"], row["doc_no"]): row for row in records}
    assert by_type_doc[("issue", "2202830")]["qty"] == 20800
    assert by_type_doc[("issue", "2202830")]["material"] == "77794-PCBA板"
    assert by_type_doc[("issue", "2202830")]["location_name"] == "东莞车间"
    assert by_type_doc[("finished", "2510160")]["qty"] == 11328
    assert by_type_doc[("finished", "2510160")]["item_no"] == "MSLD182-77794"
    assert by_type_doc[("finished", "2510160")]["location_name"] == "东莞车间"


def test_assembly_pcba_export_matches_import_workbook_format(client):
    login(client, "东莞车间", "123456", "东莞车间")
    imported = upload_bytes(
        client,
        "/api/records/import",
        legacy_assembly_pcba_workbook_bytes(),
        "东莞车间77794#PCB主板出入明细.xlsx",
    )

    exported = client.get("/api/records/export?material=77794-PCBA板")
    wb = openpyxl.load_workbook(io.BytesIO(exported.content), data_only=True)

    assert imported.status_code == 200
    assert exported.status_code == 200
    assert wb.sheetnames == [
        "总表", "36#CD领料明细", "PCB主板领料明细", "成品入仓明细",
    ]
    assert [cell.value for cell in wb["PCB主板领料明细"][1][:5]] == [
        "日期", "领料编号", "物料名称", "领料数", "备注",
    ]
    assert [cell.value for cell in wb["成品入仓明细"][1][:7]] == [
        "日期", "送货单号", "合同号", "货号", "品名/规格", "数量（pcs）", "备注",
    ]
    assert wb["PCB主板领料明细"].cell(2, 2).value == "2202830"
    assert wb["PCB主板领料明细"].cell(2, 3).value == "PCBA主板"
    assert wb["PCB主板领料明细"].cell(2, 4).value == 20800
    assert wb["成品入仓明细"].cell(2, 2).value == "2510160"
    assert wb["成品入仓明细"].cell(2, 3).value == "4500204119"
    assert wb["成品入仓明细"].cell(2, 4).value == "MSLD182-77794"
    assert wb["成品入仓明细"].cell(2, 5).value is None
    assert wb["成品入仓明细"].cell(2, 6).value == 11328
    assert wb["总表"].cell(1, 2).value == "领料总数"
    assert wb["总表"].cell(5, 2).value == "成品总数"
    assert wb["总表"].cell(9, 2).value == "理论结存数"

    reimported = upload_bytes(
        client,
        "/api/records/import",
        exported.content,
        "东莞车间77794#PCB主板出入明细.xlsx",
    )
    assert reimported.status_code == 200
    assert reimported.json()["created"] == 0
    assert reimported.json()["skipped"] == 4


def test_outsource_legacy_workbook_rejects_filename_without_department(client):
    login(client, "东莞加工厂利鸿", "123456", "东莞加工厂利鸿")

    r = upload_bytes(
        client,
        "/api/records/import",
        legacy_outsource_workbook_bytes(),
        "legacy.xlsx",
    )

    assert r.status_code == 400
    assert "东莞加工厂利鸿" in r.json()["detail"]


def test_outsource_legacy_workbook_imports_issue_and_semi_finished_rows(client):
    login(client, "东莞加工厂利鸿", "123456", "东莞加工厂利鸿")

    r = upload_bytes(
        client,
        "/api/records/import",
        legacy_outsource_workbook_bytes(),
        "东莞加工厂利鸿77794PCB主板出入明细.xlsx",
    )
    records = client.get("/api/records").json()

    assert r.status_code == 200
    assert r.json()["created"] == 3
    by_type_doc = {(row["rec_type"], row["doc_no"]): row for row in records}
    assert by_type_doc[("issue", "截止到6月17号")]["qty"] == 90
    assert by_type_doc[("issue", "2518831")]["qty"] == 10
    assert by_type_doc[("semi_finished", "BS2026070301")]["qty"] == 8
    assert by_type_doc[("semi_finished", "BS2026070301")]["contract_no"] == "LH202607"
    assert by_type_doc[("semi_finished", "BS2026070301")]["item_no"] == "77794"
    assert by_type_doc[("semi_finished", "BS2026070301")]["product_name"] == "光身唱机"
    assert by_type_doc[("semi_finished", "BS2026070301")]["remark"] == "返仓备注"
    assert by_type_doc[("issue", "2518831")]["material"] == "77794-PCBA板"

    summary = client.get("/api/summary").json()
    assert summary["raw"]["issue"] == 100
    assert summary["raw"]["semi_finished_inbound"] == 8
    assert summary["raw"]["balance"] == 92

    second = upload_bytes(
        client,
        "/api/records/import",
        legacy_outsource_workbook_bytes(),
        "东莞加工厂利鸿77794PCB主板出入明细.xlsx",
    )
    assert second.status_code == 200
    assert second.json()["created"] == 0
    assert second.json()["skipped"] == 3
    assert len(client.get("/api/records").json()) == 3


def test_lihong_legacy_workbook_round_trip_preserves_fields_and_skips_duplicates(client):
    login(client, "东莞加工厂利鸿", "123456", "东莞加工厂利鸿")
    filename = "东莞加工厂利鸿77794PCB主板出入明细.xlsx"
    first = upload_bytes(
        client,
        "/api/records/import",
        legacy_outsource_workbook_bytes(),
        filename,
    )

    exported = client.get("/api/records/export?material=77794-PCBA板")
    wb = openpyxl.load_workbook(io.BytesIO(exported.content), data_only=True)
    inbound = wb["半成品入仓明细"]

    assert first.status_code == 200
    assert first.json()["created"] == 3
    assert exported.status_code == 200
    assert wb.sheetnames == ["总表", "领料明细", "半成品入仓明细"]
    assert inbound.cell(2, 2).value == "BS2026070301"
    assert inbound.cell(2, 3).value == "LH202607"
    assert str(inbound.cell(2, 4).value) == "77794"
    assert inbound.cell(2, 5).value == "光身唱机"
    assert inbound.cell(2, 7).value == "返仓备注"

    second = upload_bytes(client, "/api/records/import", exported.content, filename)

    assert second.status_code == 200
    assert second.json()["created"] == 0
    assert second.json()["skipped"] == 3
    assert len(client.get("/api/records").json()) == 3


def test_hongya_nfc_legacy_workbook_rejects_filename_without_department(client):
    login(client, "东莞加工厂鸿亚", "123456", "东莞加工厂鸿亚")

    r = upload_bytes(
        client,
        "/api/records/import",
        legacy_hongya_nfc_workbook_bytes(),
        "legacy.xlsx",
    )

    assert r.status_code == 400
    assert "东莞加工厂鸿亚" in r.json()["detail"]


def test_hongya_nfc_legacy_workbook_imports_issue_and_finished_rows(client):
    login(client, "东莞加工厂鸿亚", "123456", "东莞加工厂鸿亚")

    r = upload_bytes(
        client,
        "/api/records/import",
        legacy_hongya_nfc_workbook_bytes(),
        "东莞加工厂鸿亚77772#NFC贴纸出入明细.xlsx",
    )
    records = client.get("/api/records").json()

    assert r.status_code == 200
    assert r.json()["created"] == 4
    rows = {(row["rec_type"], row["doc_no"]): row for row in records}
    assert rows[("issue", "2026-07-01#01")]["qty"] == 10
    assert rows[("issue", "2026-07-01#02")]["qty"] == 20
    assert rows[("finished", "2026-07-02#01")]["qty"] == 5
    assert rows[("finished", "2026-07-02#02")]["qty"] == 15
    assert rows[("issue", "2026-07-01#01")]["material"] == "NFC贴纸"
    assert rows[("issue", "2026-07-01#01")]["sticker_type"] == "1#NFC贴纸"
    assert rows[("issue", "2026-07-01#01")]["location_name"] == "东莞加工厂鸿亚"
    assert rows[("finished", "2026-07-02#01")]["location_name"] is None

    summary = client.get("/api/summary").json()
    assert summary["raw"]["issue"] == 30
    assert summary["raw"]["finished_inbound"] == 20
    assert summary["raw"]["balance"] == 10

    second = upload_bytes(
        client,
        "/api/records/import",
        legacy_hongya_nfc_workbook_bytes(),
        "东莞加工厂鸿亚77772#NFC贴纸出入明细.xlsx",
    )
    assert second.status_code == 200
    assert second.json()["created"] == 0
    assert second.json()["skipped"] == 4
    assert len(client.get("/api/records").json()) == 4


def test_hongya_nfc_export_matches_legacy_workbook_layout(client):
    login(client, "东莞加工厂鸿亚", "123456", "东莞加工厂鸿亚")
    imported = upload_bytes(
        client,
        "/api/records/import",
        legacy_hongya_nfc_workbook_bytes(),
        "东莞加工厂鸿亚77772#NFC贴纸出入明细.xlsx",
    )

    exported = client.get("/api/records/export?material=NFC贴纸")
    wb = openpyxl.load_workbook(io.BytesIO(exported.content), data_only=True)

    assert imported.status_code == 200
    assert exported.status_code == 200
    assert wb.sheetnames == ["总表", "领料明细", "入仓明细"]

    total = wb["总表"]
    assert total.cell(1, 2).value == "累计领料总数"
    assert total.cell(1, 11).value == "应存数"
    assert total.cell(1, 12).value == "累计入仓总数"
    assert total.cell(3, 1).value == "1#NFC\n贴纸"
    assert total.cell(3, 2).value == 30
    assert total.cell(3, 11).value == 10
    assert total.cell(3, 12).value == 20
    assert total.cell(48, 1).value is None
    assert total.cell(49, 1).value == "小计："

    issue = wb["领料明细"]
    assert issue.cell(1, 2).value == "当月入仓总数"
    assert [issue.cell(1, col).value for col in (3, 4)] == [
        "2026-07-01", "2026-07-01",
    ]
    assert issue.cell(2, 3).value is None
    assert issue.cell(2, 4).value is None
    assert [issue.cell(3, col).value for col in (3, 4)] == [10, 20]
    assert issue.cell(49, 1).value == "小计："

    inbound = wb["入仓明细"]
    assert [inbound.cell(1, col).value for col in (3, 4)] == [
        "2026-07-02", "2026-07-02",
    ]
    assert [inbound.cell(3, col).value for col in (3, 4)] == [5, 15]


def test_heyuan_legacy_workbook_imports_issue_corrections_and_finished_rows(client):
    login(client, "河源华兴", "123456", "河源华兴")

    r = upload_bytes(client, "/api/records/import", legacy_heyuan_workbook_bytes())
    records = client.get("/api/records").json()

    assert r.status_code == 200
    assert r.json()["created"] == 6
    by_type_doc = {(row["rec_type"], row["doc_no"]): row for row in records}
    assert by_type_doc[("issue", "2519360")]["qty"] == 100
    assert by_type_doc[("issue", "2518791")]["qty"] == 30
    assert by_type_doc[("issue", "2518791")]["rec_date"] == "2026-06-30"
    assert by_type_doc[("issue", "2518791")]["summary_month"] == 7
    assert by_type_doc[("issue", "退不良品")]["qty"] == -5
    assert by_type_doc[("finished", "HY2026070301")]["qty"] == 20
    assert by_type_doc[("finished", "HY2026070301")]["contract_no"] == "HYHT"
    assert by_type_doc[("finished", "HY2026070301")]["item_no"] == "77794"
    assert by_type_doc[("finished", "HY2026070301")]["product_name"] == "成品唱机"
    assert by_type_doc[("issue", "月结")]["material"] == "NFC贴纸"
    assert by_type_doc[("issue", "月结")]["sticker_type"] == "36#NFC贴纸"
    assert by_type_doc[("issue", "月结")]["qty"] == 50
    assert by_type_doc[("issue", "月结")]["rec_date"] == "2026-06-27"
    assert by_type_doc[("issue", "2611533")]["qty"] == 25
    assert by_type_doc[("issue", "2518791")]["location_name"] == "河源华兴"

    summary = client.get("/api/summary").json()
    by_location = {row["location"]: row for row in summary["locations"]}
    assert by_location["河源华兴"]["issue"] == 200
    assert by_location["河源华兴"]["finished"] == 20
    assert by_location["河源华兴"]["balance"] == 180
    monthly_rows = {
        (row["location"], row["material"]): row
        for row in summary["monthly_locations"]["locations"]
    }
    pcba_months = monthly_rows[("河源华兴", "77794-PCBA板")]["values"]
    assert pcba_months[0]["issue"] == 100
    assert pcba_months[1]["issue"] == 25

    second = upload_bytes(client, "/api/records/import", legacy_heyuan_workbook_bytes())
    assert second.status_code == 200
    assert second.json()["created"] == 0
    assert second.json()["skipped"] == 6
    assert len(client.get("/api/records").json()) == 6


def test_supplier_pcba_legacy_workbook_imports_inbound_and_issue_rows(client):
    login(client, "兴信B来料仓", "123456", DEFAULT_DEPARTMENT)

    r = upload_bytes(client, "/api/records/import", legacy_supplier_pcba_workbook_bytes())
    records = client.get("/api/records").json()

    assert r.status_code == 200
    assert r.json()["created"] == 7
    assert [row["qty"] for row in records if row["doc_no"] == "2534693"] == [40, 40]
    by_type_doc = {(row["rec_type"], row["doc_no"]): row for row in records}
    assert by_type_doc[("inbound_raw", "截止到6月17号")]["qty"] == 100
    assert by_type_doc[("inbound_raw", "2651574")]["qty"] == -5
    assert by_type_doc[("issue", "2518791")]["location_name"] == "河源华兴"
    assert by_type_doc[("issue", "退不良品")]["qty"] == -5

    summary = client.get("/api/summary").json()
    assert summary["raw"] == {"inbound": 175, "outbound": 43, "balance": 132}
    assert summary["materials"] == [
        {"material": "77794-PCBA板", "inbound": 175, "outbound": 43, "balance": 132}
    ]
    monthly_rows = {
        (row["location"], row["material"]): row
        for row in summary["monthly_locations"]["locations"]
    }
    assert monthly_rows[("河源华兴", "77794-PCBA板")]["values"][0]["issue"] == 0
    assert monthly_rows[("河源华兴", "77794-PCBA板")]["values"][1]["issue"] == 43

    second = upload_bytes(client, "/api/records/import", legacy_supplier_pcba_workbook_bytes())
    assert second.status_code == 200
    assert second.json()["created"] == 0
    assert second.json()["skipped"] == 7
    assert len(client.get("/api/records").json()) == 7


def test_supplier_nfc_legacy_workbook_keeps_summary_rows_out_of_record_list(client):
    login(client, "兴信B来料仓", "123456", DEFAULT_DEPARTMENT)

    r = upload_bytes(client, "/api/records/import", legacy_supplier_nfc_workbook_bytes())
    records = client.get("/api/records").json()

    assert r.status_code == 200
    assert r.json()["created"] == 6
    rows = {(row["rec_type"], row["doc_no"]): row for row in records}
    assert sorted(row["doc_no"] for row in records) == ["CK-1", "RK-1", "RK-2"]
    assert rows[("inbound_raw", "RK-1")]["qty"] == 50
    assert rows[("inbound_raw", "RK-2")]["qty"] == 30
    assert rows[("issue", "CK-1")]["qty"] == 30
    assert rows[("issue", "CK-1")]["location_name"] == "东莞车间"
    assert all(row["sticker_type"] == "1#NFC贴纸" for row in records)

    summary = client.get("/api/summary").json()
    sticker_rows = {row["sticker_type"]: row for row in summary["sticker_types"]}
    assert sticker_rows["1#NFC贴纸"] == {
        "sticker_type": "1#NFC贴纸",
        "inbound": 180,
        "outbound": 105,
        "balance": 75,
    }

    second = upload_bytes(client, "/api/records/import", legacy_supplier_nfc_workbook_bytes())
    assert second.status_code == 200
    assert second.json()["created"] == 0
    assert second.json()["skipped"] == 6
    assert len(client.get("/api/records").json()) == 3


def test_operator_can_import_and_export_materials(client):
    login(client, "兴信B来料仓", "123456", DEFAULT_DEPARTMENT)

    import_r = upload_xlsx(
        client,
        "/api/materials/import",
        ["物料名称"],
        [["测试导入物料"]],
    )
    export_r = client.get("/api/materials/export")
    wb = openpyxl.load_workbook(io.BytesIO(export_r.content), data_only=True)
    names = [row[0] for row in wb.active.iter_rows(min_row=2, values_only=True)]

    assert import_r.status_code == 200
    assert import_r.json()["imported"] == 1
    assert export_r.status_code == 200
    assert "测试导入物料" in names


def test_operator_can_import_and_export_suppliers(client):
    login(client, "兴信B来料仓", "123456", DEFAULT_DEPARTMENT)

    import_r = upload_xlsx(
        client,
        "/api/suppliers/import",
        ["供应商名称"],
        [["测试导入供应商"]],
    )
    export_r = client.get("/api/suppliers/export")
    wb = openpyxl.load_workbook(io.BytesIO(export_r.content), data_only=True)
    names = [row[0] for row in wb.active.iter_rows(min_row=2, values_only=True)]

    assert import_r.status_code == 200
    assert import_r.json()["imported"] == 1
    assert export_r.status_code == 200
    assert "测试导入供应商" in names


def test_operator_can_import_and_export_sticker_types_by_sort(client):
    login(client, "兴信B来料仓", "123456", DEFAULT_DEPARTMENT)

    import_r = upload_xlsx(
        client,
        "/api/sticker-types/import",
        ["排序", "贴纸类型"],
        [[1, "1#NFC贴纸-修改"]],
    )
    export_r = client.get("/api/sticker-types/export")
    wb = openpyxl.load_workbook(io.BytesIO(export_r.content), data_only=True)
    first_row = [cell.value for cell in wb.active[2]]

    assert import_r.status_code == 200
    assert import_r.json()["imported"] == 1
    assert export_r.status_code == 200
    assert first_row == [1, "1#NFC贴纸-修改"]
