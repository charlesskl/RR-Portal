import io

import openpyxl


def admin_login(client, department="邵阳"):
    client.post(
        "/api/login",
        json={"username": "admin", "password": "admin123", "department": department},
    )


def _workbook_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _shaoyang_issue_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "总表"
    ws.cell(1, 13).value = "累计入仓数"
    ws.cell(1, 15).value = "7月成品\n入仓总数"
    ws.cell(2, 1).value = "物料名称"
    ws.cell(3, 1).value = "1#NFC\n贴纸"
    ws.cell(3, 13).value = 12
    ws.cell(3, 15).value = 12
    ws.cell(4, 1).value = "2#NFC\n贴纸"
    ws.cell(4, 13).value = 5
    ws.cell(4, 15).value = 5
    ws.cell(5, 1).value = "小计："
    ws.cell(5, 15).value = 17
    return _workbook_bytes(wb)


def _shaoyang_finished_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "总表"
    ws.append(["Sku No.", "Item No.", "Minis Name", "Catalogue", "小计"])
    ws.append([77772, "VINYL-S1-001", "Song A", "C1", 12])
    ws.append([77772, "VINYL-S1-002", "Song B", "C1", 7])
    return _workbook_bytes(wb)


def test_shaoyang_reconcile_maps_sticker_number_to_vinyl_item(client):
    admin_login(client)

    r = client.post(
        "/api/shaoyang-cd/reconcile",
        data={"month": "7"},
        files={
            "issue_file": (
                "issue.xlsx",
                _shaoyang_issue_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            "finished_file": (
                "finished.xlsx",
                _shaoyang_finished_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )

    assert r.status_code == 200
    data = r.json()
    rows = {row["sticker_no"]: row for row in data["rows"]}
    assert data["month"] == 7
    assert data["totals"] == {
        "issue_month_inbound": 17,
        "finished_total": 19,
        "difference": 2,
    }
    assert rows[1]["sticker_name"] == "1#NFC贴纸"
    assert rows[1]["item_no"] == "VINYL-S1-001"
    assert rows[1]["minis_name"] == "Song A"
    assert rows[1]["issue_month_inbound"] == 12
    assert rows[1]["finished_total"] == 12
    assert rows[1]["difference"] == 0
    assert rows[2]["sticker_name"] == "2#NFC贴纸"
    assert rows[2]["item_no"] == "VINYL-S1-002"
    assert rows[2]["issue_month_inbound"] == 5
    assert rows[2]["finished_total"] == 7
    assert rows[2]["difference"] == 2


def test_shaoyang_export_issue_workbook_fills_month_inbound_from_finished_subtotal(client):
    admin_login(client)

    r = client.post(
        "/api/shaoyang-cd/export-issue",
        data={"month": "7"},
        files={
            "issue_file": (
                "issue.xlsx",
                _shaoyang_issue_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            "finished_file": (
                "finished.xlsx",
                _shaoyang_finished_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )

    assert r.status_code == 200
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    ws = wb["总表"]
    assert ws.cell(3, 15).value == 12
    assert ws.cell(4, 15).value == 7
    assert ws.cell(5, 15).value == 19
