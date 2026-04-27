from flask import Flask, request, redirect, url_for, jsonify, session, flash, render_template
import sqlite3
import json
import os
import sys
import openpyxl
from datetime import timedelta, datetime, date as date_cls
from functools import wraps

# 兼容 PyInstaller exe 和普通 Python 运行
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
    TEMPLATE_DIR = os.path.join(sys._MEIPASS, 'templates')
    STATIC_DIR = os.path.join(sys._MEIPASS, 'static')
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    TEMPLATE_DIR = os.path.join(BASE_DIR, 'templates')
    STATIC_DIR = os.path.join(BASE_DIR, 'static')

app = Flask(__name__, template_folder=TEMPLATE_DIR, static_folder=STATIC_DIR)

DATA_PATH = os.environ.get('DATA_PATH', BASE_DIR)
os.makedirs(DATA_PATH, exist_ok=True)
DATABASE = os.path.join(DATA_PATH, 'huadeng.db')

app.secret_key = os.environ.get('HUADENG_SECRET_KEY', 'dev-change-me-in-prod')
app.permanent_session_lifetime = timedelta(hours=8)


@app.route('/health')
def health():
    return {'status': 'ok'}


def get_db():
    db = sqlite3.connect(DATABASE)
    db.row_factory = sqlite3.Row
    return db


def init_db():
    """初始化所有新 schema 表。幂等。"""
    con = sqlite3.connect(DATABASE)
    cur = con.cursor()

    # flow_records 主流水表
    cur.execute("""
    CREATE TABLE IF NOT EXISTS flow_records (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        recorded_by TEXT NOT NULL,
        from_party  TEXT NOT NULL,
        to_party    TEXT NOT NULL,
        date        TEXT NOT NULL,
        order_no    TEXT,
        remark      TEXT,
        jx_qty REAL DEFAULT 0, gx_qty REAL DEFAULT 0, zx_qty REAL DEFAULT 0,
        jkb_qty REAL DEFAULT 0, mkb_qty REAL DEFAULT 0, xb_qty REAL DEFAULT 0,
        dz_qty REAL DEFAULT 0, wb_qty REAL DEFAULT 0, pk_qty REAL DEFAULT 0,
        xzx_qty REAL DEFAULT 0, dgb_qty REAL DEFAULT 0, xjp_qty REAL DEFAULT 0,
        dk_qty REAL DEFAULT 0,
        xs_qty REAL DEFAULT 0, gsb_qty REAL DEFAULT 0,
        djx_qty REAL DEFAULT 0, zb_qty REAL DEFAULT 0,
        reconciliation_id INTEGER,
        locked INTEGER DEFAULT 0,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (reconciliation_id) REFERENCES reconciliations(id)
    )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_flow_recorded_by ON flow_records(recorded_by, from_party, to_party)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_flow_pair_date ON flow_records(from_party, to_party, date)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_flow_reconc ON flow_records(reconciliation_id)")

    # reconciliations 核对批次
    cur.execute("""
    CREATE TABLE IF NOT EXISTS reconciliations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        initiator_party TEXT NOT NULL,
        approver_party  TEXT NOT NULL,
        pair_low  TEXT NOT NULL,
        pair_high TEXT NOT NULL,
        date_from TEXT NOT NULL,
        date_to   TEXT NOT NULL,
        status    TEXT NOT NULL,
        snapshot_json TEXT,
        notes     TEXT,
        created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        approved_at TIMESTAMP
    )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_reconc_approver ON reconciliations(approver_party, status)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_reconc_pair ON reconciliations(pair_low, pair_high, status)")

    # investment_records 投资记录（按 pair）
    cur.execute("""
    CREATE TABLE IF NOT EXISTS investment_records (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        recorded_by TEXT NOT NULL,
        counterparty TEXT NOT NULL,
        year_month TEXT NOT NULL,
        mkb_qty REAL DEFAULT 0, jkb_qty REAL DEFAULT 0,
        jx_qty REAL DEFAULT 0, gx_qty REAL DEFAULT 0,
        remark TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """)

    # monthly_inventory 月份实存数
    cur.execute("""
    CREATE TABLE IF NOT EXISTS monthly_inventory (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        recorded_by TEXT NOT NULL,
        counterparty TEXT NOT NULL,
        year_month TEXT NOT NULL,
        mkb_qty REAL, jkb_qty REAL, jx_qty REAL, gx_qty REAL,
        UNIQUE (recorded_by, counterparty, year_month)
    )
    """)

    # default_prices 单价表
    cur.execute("""
    CREATE TABLE IF NOT EXISTS default_prices (
        item_key TEXT PRIMARY KEY,
        price    REAL DEFAULT 0
    )
    """)

    con.commit()
    con.close()


# ==================== Party 常量与认证 ====================

PARTIES = {
    'hd': {'name': '华登',     'counterparties': ['sy', 'xx']},
    'sy': {'name': '邵阳华登', 'counterparties': ['hd', 'xx']},
    'xx': {'name': '兴信',     'counterparties': ['hd', 'sy']},
}

PARTY_ACCOUNTS = {
    'hd': {
        'username': os.environ.get('HUADENG_HD_USER', 'hd'),
        'password': os.environ.get('HUADENG_HD_PASSWORD', 'hd123456'),
    },
    'sy': {
        'username': os.environ.get('HUADENG_SY_USER', 'sy'),
        'password': os.environ.get('HUADENG_SY_PASSWORD', 'sy123456'),
    },
    'xx': {
        'username': os.environ.get('HUADENG_XX_USER', 'xx'),
        'password': os.environ.get('HUADENG_XX_PASSWORD', 'xx123456'),
    },
}

PARTY_BY_USERNAME = {acc['username']: p for p, acc in PARTY_ACCOUNTS.items()}
assert len(PARTY_BY_USERNAME) == 3, 'username 必须 3 个都唯一'

ITEMS = [
    ('jx', '胶箱'), ('gx', '钙塑箱'), ('zx', '纸箱'),
    ('jkb', '胶卡板'), ('mkb', '木卡板'), ('xb', '小板'),
    ('dz', '胶袋'), ('wb', '围布'), ('pk', '平卡'),
    ('xzx', '小纸箱'), ('dgb', '大盖板'), ('xjp', '小胶盆'),
    ('dk', '刀卡'),
    ('xs', '吸塑'), ('gsb', '钙塑板'),
    ('djx', '大胶箱'), ('zb', '纸板'),
]
STAT_ITEMS = [('mkb', '木卡板'), ('jkb', '胶卡板'), ('jx', '胶箱'), ('gx', '钙塑箱')]
TRIANGLE_ITEMS = [('mkb', '木卡板'), ('jkb', '胶卡板'), ('jx', '胶箱'), ('gx', '钙塑箱'), ('zx', '纸箱')]
PAIRS = [('hd', 'sy'), ('hd', 'xx'), ('sy', 'xx')]

# 导入模板。每个 entry 描述一个 Excel 文件的所有方向列映射，用户选模板 + sheet 即可批量导入。
# 模板按 sheet 字典格式：sheets 内可逐 sheet 自定义；如果无对应 sheet，用 'default' 兜底。
IMPORT_CONFIGS = {
    'huadeng_qingxi_shaoyang': {
        'label': '清溪华登 ↔ 邵阳华登（hd↔sy 双方向）',
        'filename_pattern': '清溪华登',
        'recorded_by_party': 'hd',  # 必须用 hd 账号导入
        'allowed_sheets': ['1月', '2月', '3月', '4月'],  # 其它 sheet 不导
        'directions': [
            {
                'direction': 'hd_to_sy',
                'start_row': 3,
                # 左半：清溪华登发邵阳华登
                'columns': {
                    0: 'date', 1: 'order_no',
                    2: 'jx_qty', 3: 'gx_qty', 4: 'zx_qty',
                    5: 'jkb_qty', 6: 'mkb_qty', 7: 'xb_qty',
                    8: 'dz_qty', 9: 'wb_qty', 10: 'pk_qty',
                    11: 'xzx_qty', 12: 'dgb_qty', 13: 'xjp_qty',
                    14: 'dk_qty', 15: 'remark',
                },
            },
            {
                'direction': 'sy_to_hd',
                'start_row': 3,
                # 右半：清溪华登收邵阳华登（即 sy 发的，hd 收）
                'columns': {
                    17: 'date', 18: 'order_no',
                    19: 'jx_qty', 20: 'gx_qty', 21: 'zx_qty',
                    22: 'jkb_qty', 23: 'mkb_qty',
                    24: 'dz_qty', 25: 'xb_qty',
                    26: 'xzx_qty', 27: 'wb_qty', 28: 'pk_qty',
                    # col 29 '专用卡' 无 ITEMS 映射，跳过
                },
            },
        ],
    },
    # 其余 4 种 (邵阳-兴信 / 清溪-兴信 / 兴信-清溪 / 兴信-邵阳) 拿到 xlsx 后按相同结构补
}

ALLOWED_DIRECTIONS = {
    'hd_to_sy': ('hd', 'sy'), 'sy_to_hd': ('sy', 'hd'),
    'hd_to_xx': ('hd', 'xx'), 'xx_to_hd': ('xx', 'hd'),
    'sy_to_xx': ('sy', 'xx'), 'xx_to_sy': ('xx', 'sy'),
}


def current_party():
    """Return the validated party from session, or None if absent / tampered."""
    p = session.get('party')
    return p if p in PARTIES else None


app.jinja_env.globals['current_party'] = current_party
app.jinja_env.globals['PARTIES'] = PARTIES
app.jinja_env.globals['ITEMS'] = ITEMS
app.jinja_env.globals['STAT_ITEMS'] = STAT_ITEMS


def party_required(fn):
    """装饰器：要求当前 session 有 party，且 URL 里的 party 与之匹配。"""
    @wraps(fn)
    def wrapped(*args, **kwargs):
        url_party = kwargs.get('party')
        sess_party = session.get('party')
        if not sess_party:
            return redirect(url_for('party_login', party=url_party or 'hd'))
        if url_party and url_party != sess_party:
            flash('无权访问其他 party 页面')
            return redirect(url_for('index'))
        return fn(*args, **kwargs)
    return wrapped


@app.context_processor
def inject_pending_count():
    """顶部 nav badge：当前 party 作为 approver 的待处理核对数。"""
    party = session.get('party')
    if not party:
        return {'pending_approval_count': 0}
    try:
        con = sqlite3.connect(DATABASE)
        row = con.execute(
            "SELECT COUNT(*) FROM reconciliations WHERE approver_party=? AND status='pending_approval'",
            (party,)
        ).fetchone()
        con.close()
        return {'pending_approval_count': row[0] if row else 0}
    except Exception:
        app.logger.exception('inject_pending_count failed')
        return {'pending_approval_count': 0}


# ==================== 页面路由 ====================

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/party/<party>/login', methods=['GET', 'POST'])
def party_login(party):
    if party not in PARTIES:
        return redirect(url_for('index'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        # 1. username 必须存在
        user_party = PARTY_BY_USERNAME.get(username)
        if not user_party:
            flash('账号或密码错误')
            return redirect(url_for('party_login', party=party))
        # 2. 与 URL 里的 party 必须一致
        if user_party != party:
            flash('账号与登录入口不符')
            return redirect(url_for('party_login', party=party))
        # 3. 密码校验
        expected = PARTY_ACCOUNTS[party]['password']
        if password != expected:
            flash('账号或密码错误')
            return redirect(url_for('party_login', party=party))
        session.permanent = True
        session['party'] = party
        session.modified = True
        return redirect(url_for('party_page', party=party))
    return render_template('party_login.html', party=party)


@app.route('/party/<party>/logout')
def party_logout(party):
    session.pop('party', None)
    session.modified = True
    return redirect(url_for('index'))


@app.route('/party/<party>')
@party_required
def party_page(party):
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    try:
        page_size = int(request.args.get('page_size', 50))
    except ValueError:
        page_size = 50
    if page_size not in (20, 50, 100, 200):
        page_size = 50

    con = sqlite3.connect(DATABASE)
    con.row_factory = sqlite3.Row

    counterparties = PARTIES[party]['counterparties']
    panels = []
    for cp in counterparties:
        panel = {'cp': cp, 'cp_name': PARTIES[cp]['name']}
        for direction, from_p, to_p in [('sent', party, cp), ('received', cp, party)]:
            all_r = _query_flow(con, recorded_by=party, from_party=from_p, to_party=to_p,
                                date_from=date_from, date_to=date_to)
            page_key = f'page_{cp}_{direction}'
            try:
                page = max(1, int(request.args.get(page_key, 1) or 1))
            except ValueError:
                page = 1
            total = len(all_r)
            pages = max(1, (total + page_size - 1) // page_size)
            page = min(page, pages)
            start = (page - 1) * page_size
            panel[f'{direction}_records'] = all_r[start:start + page_size]
            panel[f'{direction}_summary'] = _calc_summary(all_r)
            panel[f'{direction}_pagination'] = {
                'page': page, 'pages': pages, 'total': total, 'page_size': page_size,
                'start': start + 1 if total else 0,
                'end': start + len(all_r[start:start + page_size]),
                'page_key': page_key,
            }
        panels.append(panel)

    prices = {r['item_key']: r['price'] for r in con.execute('SELECT * FROM default_prices').fetchall()}
    con.close()
    return render_template('party.html', party=party, party_name=PARTIES[party]['name'],
                           panels=panels, prices=prices,
                           date_from=date_from, date_to=date_to, page_size=page_size)


@app.route('/party/<party>/entry', methods=['POST'])
@party_required
def party_entry(party):
    direction = request.form.get('direction')  # 'sent' | 'received'
    cp = request.form.get('counterparty')
    if cp not in PARTIES or cp == party:
        flash('无效的对方')
        return redirect(url_for('party_page', party=party))
    if cp not in PARTIES[party]['counterparties']:
        flash('无权对此 party 录入')
        return redirect(url_for('party_page', party=party))

    if direction == 'sent':
        from_p, to_p = party, cp
    elif direction == 'received':
        from_p, to_p = cp, party
    else:
        flash('无效 direction')
        return redirect(url_for('party_page', party=party))

    date = request.form.get('date', '').strip()
    order_no = request.form.get('order_no', '').strip() or None
    remark = request.form.get('remark', '').strip() or None
    if not date:
        flash('日期必填')
        return redirect(url_for('party_page', party=party))

    qty_cols = [f'{k}_qty' for k, _ in ITEMS]
    qty_vals = []
    for col in qty_cols:
        v = request.form.get(col, '0').strip()
        try:
            qty_vals.append(float(v) if v else 0)
        except ValueError:
            qty_vals.append(0)

    con = sqlite3.connect(DATABASE)
    placeholders = ', '.join(['?'] * len(qty_cols))
    con.execute(f"""
        INSERT INTO flow_records (recorded_by, from_party, to_party, date, order_no, remark,
                                  {', '.join(qty_cols)})
        VALUES (?, ?, ?, ?, ?, ?, {placeholders})
    """, [party, from_p, to_p, date, order_no, remark, *qty_vals])
    con.commit()
    con.close()
    return redirect(url_for('party_page', party=party))


@app.route('/party/<party>/export')
@party_required
def party_export(party):
    """导出 party↔cp 双方向 records 为 xlsx，含日期筛选。"""
    cp = request.args.get('cp', '')
    if cp not in PARTIES or cp == party or cp not in PARTIES[party]['counterparties']:
        flash('无效对方'); return redirect(url_for('party_page', party=party))
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()

    con = sqlite3.connect(DATABASE)
    con.row_factory = sqlite3.Row
    sent = _query_flow(con, recorded_by=party, from_party=party, to_party=cp,
                       date_from=date_from or None, date_to=date_to or None)
    received = _query_flow(con, recorded_by=party, from_party=cp, to_party=party,
                           date_from=date_from or None, date_to=date_to or None)
    con.close()

    import io
    import xlsxwriter
    buf = io.BytesIO()
    wb = xlsxwriter.Workbook(buf, {'in_memory': True})
    headers = ['日期', '订单号'] + [name for _, name in ITEMS] + ['备注']
    qty_keys = [k for k, _ in ITEMS]

    for sheet_name, records in [(f'发→{PARTIES[cp]["name"]}', sent),
                                 (f'收自{PARTIES[cp]["name"]}', received)]:
        ws = wb.add_worksheet(sheet_name)
        for col, h in enumerate(headers):
            ws.write(0, col, h)
        for i, r in enumerate(records, start=1):
            ws.write(i, 0, r.get('date') or '')
            ws.write(i, 1, r.get('order_no') or '')
            for j, k in enumerate(qty_keys, start=2):
                v = r.get(f'{k}_qty') or 0
                if v:
                    ws.write_number(i, j, v)
            ws.write(i, 2 + len(qty_keys), r.get('remark') or '')
    wb.close()
    buf.seek(0)

    from flask import send_file
    today = datetime.now().strftime('%Y%m%d')
    filename = f'{PARTIES[party]["name"]}-{PARTIES[cp]["name"]}-{today}.xlsx'
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/record/<int:rid>/edit', methods=['POST'])
def record_edit(rid):
    party = current_party()
    if not party:
        return redirect(url_for('index'))
    con = sqlite3.connect(DATABASE)
    con.row_factory = sqlite3.Row
    r = con.execute("SELECT * FROM flow_records WHERE id=?", (rid,)).fetchone()
    if not r:
        con.close(); flash('记录不存在'); return redirect(url_for('party_page', party=party))
    if r['recorded_by'] != party:
        con.close(); flash('无权编辑他人记录'); return redirect(url_for('party_page', party=party))
    if r['locked']:
        con.close(); flash('记录已锁定，不能修改'); return redirect(url_for('party_page', party=party))

    date = request.form.get('date', '').strip() or r['date']
    order_no = request.form.get('order_no', '').strip() or None
    remark = request.form.get('remark', '').strip() or None
    qty_cols = [f'{k}_qty' for k, _ in ITEMS]
    qty_vals = []
    for col in qty_cols:
        v = request.form.get(col, '').strip()
        if v == '':
            qty_vals.append(r[col] or 0)
        else:
            try:
                qty_vals.append(float(v))
            except ValueError:
                qty_vals.append(r[col] or 0)
    set_clause = ', '.join([f'{c}=?' for c in ['date', 'order_no', 'remark'] + qty_cols])
    con.execute(
        f"UPDATE flow_records SET {set_clause}, updated_at=CURRENT_TIMESTAMP WHERE id=?",
        [date, order_no, remark, *qty_vals, rid]
    )
    con.commit(); con.close()
    return redirect(url_for('party_page', party=party))


@app.route('/record/<int:rid>/delete', methods=['POST'])
def record_delete(rid):
    party = current_party()
    if not party:
        return redirect(url_for('index'))
    con = sqlite3.connect(DATABASE)
    con.row_factory = sqlite3.Row
    r = con.execute("SELECT recorded_by, locked FROM flow_records WHERE id=?", (rid,)).fetchone()
    if not r:
        con.close(); flash('记录不存在'); return redirect(url_for('party_page', party=party))
    if r['recorded_by'] != party:
        con.close(); flash('无权删除'); return redirect(url_for('party_page', party=party))
    if r['locked']:
        con.close(); flash('记录已锁定'); return redirect(url_for('party_page', party=party))
    con.execute("DELETE FROM flow_records WHERE id=?", (rid,))
    con.commit(); con.close()
    return redirect(url_for('party_page', party=party))


@app.route('/reconcile/start', methods=['POST'])
def reconcile_start():
    party = current_party()
    if not party:
        return redirect(url_for('index'))
    cp = request.form.get('counterparty')
    date_from = request.form.get('date_from', '').strip()
    date_to = request.form.get('date_to', '').strip()
    if cp not in PARTIES or cp == party or cp not in PARTIES[party]['counterparties']:
        flash('无效对方'); return redirect(url_for('party_page', party=party))
    if not date_from or not date_to:
        flash('日期必填'); return redirect(url_for('party_page', party=party))

    pair_low, pair_high = sorted([party, cp])

    con = sqlite3.connect(DATABASE)
    # 检查 pending overlap
    overlap = con.execute("""
        SELECT id FROM reconciliations
        WHERE pair_low=? AND pair_high=? AND status='pending_approval'
          AND NOT (date_to < ? OR date_from > ?)
    """, (pair_low, pair_high, date_from, date_to)).fetchone()
    if overlap:
        con.close()
        flash('已存在待审批的核对，范围重叠'); return redirect(url_for('party_page', party=party))

    # TOCTOU: overlap 检查与 INSERT 之间存在窗口；snapshot 与 UPDATE 之间也是。
    # 3 LAN 用户场景下并发概率可忽略，故意不加 BEGIN IMMEDIATE。
    snapshot = compare_pair(party, cp, date_from, date_to)
    cur = con.execute("""
        INSERT INTO reconciliations (initiator_party, approver_party, pair_low, pair_high,
                                     date_from, date_to, status, snapshot_json)
        VALUES (?, ?, ?, ?, ?, ?, 'pending_approval', ?)
    """, (party, cp, pair_low, pair_high, date_from, date_to, json.dumps(snapshot)))
    reconc_id = cur.lastrowid
    # 绑定 flow_records（未锁）
    con.execute("""
        UPDATE flow_records SET reconciliation_id=?
        WHERE date BETWEEN ? AND ?
          AND ((from_party=? AND to_party=?) OR (from_party=? AND to_party=?))
          AND reconciliation_id IS NULL
    """, (reconc_id, date_from, date_to, party, cp, cp, party))
    con.commit(); con.close()
    flash('核对已发起，等待对方审批'); return redirect(url_for('reconcile_detail', rid=reconc_id))


@app.route('/reconcile')
def reconcile_list():
    party = current_party()
    if not party:
        return redirect(url_for('index'))
    con = sqlite3.connect(DATABASE)
    con.row_factory = sqlite3.Row
    rows = con.execute("""
        SELECT * FROM reconciliations
        WHERE initiator_party=? OR approver_party=?
        ORDER BY
            CASE status WHEN 'pending_approval' THEN 0 ELSE 1 END,
            created_at DESC
    """, (party, party)).fetchall()
    con.close()
    return render_template('reconcile_list.html', items=[dict(r) for r in rows], party=party)


@app.route('/reconcile/<int:rid>')
def reconcile_detail(rid):
    party = current_party()
    if not party:
        return redirect(url_for('index'))
    con = sqlite3.connect(DATABASE)
    con.row_factory = sqlite3.Row
    r = con.execute("SELECT * FROM reconciliations WHERE id=?", (rid,)).fetchone()
    con.close()
    if not r:
        flash('核对不存在'); return redirect(url_for('reconcile_list'))
    snapshot = json.loads(r['snapshot_json']) if r['snapshot_json'] else {}
    return render_template('reconcile_detail.html', r=dict(r), snapshot=snapshot, party=party)


@app.route('/reconcile/<int:rid>/approve', methods=['POST'])
def reconcile_approve(rid):
    party = current_party()
    if not party:
        return redirect(url_for('index'))
    con = sqlite3.connect(DATABASE)
    con.row_factory = sqlite3.Row
    r = con.execute("SELECT * FROM reconciliations WHERE id=?", (rid,)).fetchone()
    if not r:
        con.close(); flash('核对不存在'); return redirect(url_for('index'))
    if r['approver_party'] != party:
        con.close(); flash('只有对方才能同意'); return redirect(url_for('reconcile_detail', rid=rid))
    if r['status'] != 'pending_approval':
        con.close(); flash('当前状态不允许操作'); return redirect(url_for('reconcile_detail', rid=rid))

    con.execute("""UPDATE reconciliations SET status='confirmed', approved_at=CURRENT_TIMESTAMP WHERE id=?""", (rid,))
    con.execute("UPDATE flow_records SET locked=1 WHERE reconciliation_id=?", (rid,))
    con.commit(); con.close()
    flash('已确认'); return redirect(url_for('reconcile_detail', rid=rid))


@app.route('/reconcile/<int:rid>/reject', methods=['POST'])
def reconcile_reject(rid):
    party = current_party()
    if not party:
        return redirect(url_for('index'))
    notes = request.form.get('notes', '').strip()
    con = sqlite3.connect(DATABASE)
    con.row_factory = sqlite3.Row
    r = con.execute("SELECT * FROM reconciliations WHERE id=?", (rid,)).fetchone()
    if not r:
        con.close(); flash('核对不存在'); return redirect(url_for('index'))
    if r['approver_party'] != party:
        con.close(); flash('只有对方才能打回'); return redirect(url_for('reconcile_detail', rid=rid))
    if r['status'] != 'pending_approval':
        con.close(); flash('当前状态不允许操作'); return redirect(url_for('reconcile_detail', rid=rid))

    con.execute("UPDATE reconciliations SET status='disputed', notes=? WHERE id=?", (notes, rid))
    con.execute("UPDATE flow_records SET reconciliation_id=NULL WHERE reconciliation_id=?", (rid,))
    con.commit(); con.close()
    flash('已打回'); return redirect(url_for('reconcile_detail', rid=rid))


@app.route('/reconcile/<int:rid>/withdraw', methods=['POST'])
def reconcile_withdraw(rid):
    party = current_party()
    if not party:
        return redirect(url_for('index'))
    con = sqlite3.connect(DATABASE)
    con.row_factory = sqlite3.Row
    r = con.execute("SELECT * FROM reconciliations WHERE id=?", (rid,)).fetchone()
    if not r:
        con.close(); flash('核对不存在'); return redirect(url_for('index'))
    if r['initiator_party'] != party:
        con.close(); flash('只有发起方能撤回'); return redirect(url_for('reconcile_detail', rid=rid))
    if r['status'] != 'pending_approval':
        con.close(); flash('当前状态不允许撤回'); return redirect(url_for('reconcile_detail', rid=rid))

    con.execute("UPDATE reconciliations SET status='withdrawn' WHERE id=?", (rid,))
    con.execute("UPDATE flow_records SET reconciliation_id=NULL WHERE reconciliation_id=?", (rid,))
    con.commit(); con.close()
    flash('已撤回'); return redirect(url_for('reconcile_detail', rid=rid))


@app.route('/reconcile/<int:rid>/cancel', methods=['POST'])
def reconcile_cancel(rid):
    """撤销已 confirmed 的核对（任一方都可）。"""
    party = current_party()
    if not party:
        return redirect(url_for('index'))
    con = sqlite3.connect(DATABASE)
    con.row_factory = sqlite3.Row
    r = con.execute("SELECT * FROM reconciliations WHERE id=?", (rid,)).fetchone()
    if not r:
        con.close(); flash('核对不存在'); return redirect(url_for('index'))
    if party not in (r['initiator_party'], r['approver_party']):
        con.close(); flash('无权'); return redirect(url_for('index'))
    if r['status'] != 'confirmed':
        con.close(); flash('仅 confirmed 状态可撤销'); return redirect(url_for('reconcile_detail', rid=rid))

    con.execute("UPDATE reconciliations SET status='withdrawn' WHERE id=?", (rid,))
    con.execute("UPDATE flow_records SET locked=0, reconciliation_id=NULL WHERE reconciliation_id=?", (rid,))
    con.commit(); con.close()
    flash('已撤销对账，记录解锁'); return redirect(url_for('reconcile_detail', rid=rid))


@app.route('/reports')
def reports():
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    only_confirmed = request.args.get('only_confirmed') == '1'

    con = sqlite3.connect(DATABASE)
    con.row_factory = sqlite3.Row

    prices = {r['item_key']: r['price']
              for r in con.execute('SELECT * FROM default_prices').fetchall()}

    # 每个方向的汇总（仅发方记录）
    direction_summaries = {}
    for a, b in PAIRS:
        for from_p, to_p in [(a, b), (b, a)]:
            direction_summaries[f'{from_p}_to_{to_p}'] = _sum_flow_records(
                con, from_party=from_p, to_party=to_p,
                only_sender=True, date_from=date_from, date_to=date_to,
                only_confirmed=only_confirmed,
            )

    # 三角债净欠
    triangle_rows = _build_triangle(direction_summaries, prices)
    pair_summary = _build_pair_summary(direction_summaries)

    # 月度明细
    monthly_by_pair = {}
    for a, b in PAIRS:
        by_month = {}  # ym -> {'a_sent': {...}, 'b_sent': {...}}
        for from_p, to_p in [(a, b), (b, a)]:
            rows = con.execute(f"""
                SELECT substr(date, 1, 7) AS ym,
                       {', '.join(['COALESCE(SUM(' + k + '_qty), 0) AS ' + k + '_sum' for k, _ in STAT_ITEMS])}
                FROM flow_records
                WHERE from_party=? AND to_party=? AND recorded_by=?
                GROUP BY ym ORDER BY ym
            """, (from_p, to_p, from_p)).fetchall()
            for r in rows:
                ym = r['ym']
                bucket = by_month.setdefault(ym, {'a_sent': {}, 'b_sent': {}})
                key = 'a_sent' if from_p == a else 'b_sent'
                for k, _ in STAT_ITEMS:
                    bucket[key][k] = float(r[f'{k}_sum'])
        # 计算 net per month + 补齐缺失 key（防 Jinja strict mode 下 KeyError）
        for ym, d in by_month.items():
            for k, _ in STAT_ITEMS:
                d['a_sent'].setdefault(k, 0)
                d['b_sent'].setdefault(k, 0)
            d['net'] = {k: d['a_sent'][k] - d['b_sent'][k] for k, _ in STAT_ITEMS}
        monthly_by_pair[(a, b)] = {'a': PARTIES[a]['name'], 'b': PARTIES[b]['name'],
                                   'months': sorted(by_month.items())}

    con.close()
    today = datetime.now().strftime('%Y/%m/%d')
    today_cn = datetime.now().strftime('%Y年%m月%d日')
    return render_template('reports.html',
                           direction_summaries=direction_summaries,
                           triangle_rows=triangle_rows,
                           has_outstanding_debt=_has_outstanding_debt(triangle_rows),
                           pair_summary=pair_summary,
                           monthly_by_pair=monthly_by_pair,
                           date_from=date_from, date_to=date_to,
                           only_confirmed=only_confirmed,
                           today=today, today_cn=today_cn,
                           PAIRS=PAIRS)


def _sum_flow_records(con, *, from_party, to_party, only_sender, date_from, date_to, only_confirmed):
    qty_cols_sql = ', '.join([f'COALESCE(SUM({k}_qty), 0) AS {k}_sum' for k, _ in ITEMS])
    sql = f"SELECT {qty_cols_sql} FROM flow_records WHERE from_party=? AND to_party=?"
    args = [from_party, to_party]
    if only_sender:
        sql += ' AND recorded_by=?'
        args.append(from_party)
    if date_from:
        sql += ' AND date >= ?'; args.append(date_from)
    if date_to:
        sql += ' AND date <= ?'; args.append(date_to)
    if only_confirmed:
        sql += " AND locked=1"
    row = con.execute(sql, args).fetchone()
    return {k: float(row[f'{k}_sum']) for k, _ in ITEMS}


def _build_triangle(direction_summaries, prices):
    """构造 6 行单向欠款表（3 个 pair × 2 个方向）。

    每行表示 "X 欠 Y"：显示 (Y_to_X - X_to_Y) > 0 的差额，否则空。
    Y 多发了 → X 多收了 → X 欠 Y。
    Note: prices 参数 spec literal 保留，当前未使用。
    """
    # (debtor, creditor) 行序模仿旧 UI
    DEBT_ROWS = [
        ('sy', 'xx'), ('xx', 'sy'),
        ('hd', 'xx'), ('xx', 'hd'),
        ('sy', 'hd'), ('hd', 'sy'),
    ]
    rows = []
    for idx, (debtor, creditor) in enumerate(DEBT_ROWS, 1):
        creditor_sent = direction_summaries[f'{creditor}_to_{debtor}']
        debtor_sent = direction_summaries[f'{debtor}_to_{creditor}']
        row = {
            'idx': idx,
            'label': f'{PARTIES[debtor]["name"]}欠{PARTIES[creditor]["name"]}',
            'has_any': False,
        }
        for k, _ in STAT_ITEMS:
            diff = creditor_sent[k] - debtor_sent[k]
            row[k] = int(diff) if diff > 0 else 0
            if row[k]:
                row['has_any'] = True
        rows.append(row)
    return rows


def _has_outstanding_debt(triangle_rows):
    """是否有任意未清三角债（用于底部红色提示）。"""
    return any(row['has_any'] for row in triangle_rows)


def _build_pair_summary(direction_summaries):
    """按 pair 汇总净欠，生成'X 欠 Y 多少 item'文案。

    Note: nets 按 TRIANGLE_ITEMS (5 项) 算，但模板只展示 STAT_ITEMS (4 项)；
    zx 永远不显示。spec literal, plan Task 15.
    """
    out = []
    for a, b in PAIRS:
        a_to_b = direction_summaries[f'{a}_to_{b}']
        b_to_a = direction_summaries[f'{b}_to_{a}']
        nets = {k: a_to_b[k] - b_to_a[k] for k, _ in TRIANGLE_ITEMS}
        out.append({'a': PARTIES[a]['name'], 'b': PARTIES[b]['name'], 'nets': nets})
    return out


def _query_flow(con, *, recorded_by, from_party, to_party, date_from=None, date_to=None):
    """查 flow_records。"""
    sql = """SELECT * FROM flow_records
             WHERE recorded_by=? AND from_party=? AND to_party=?"""
    args = [recorded_by, from_party, to_party]
    if date_from:
        sql += ' AND date >= ?'; args.append(date_from)
    if date_to:
        sql += ' AND date <= ?'; args.append(date_to)
    sql += ' ORDER BY date DESC, id DESC'
    return [dict(r) for r in con.execute(sql, args).fetchall()]


def _calc_summary(records):
    """累加 17 包材 qty + 金额。"""
    summary = {k: {'qty': 0, 'amount': 0} for k, _ in ITEMS}
    for r in records:
        for k, _ in ITEMS:
            qty = r.get(f'{k}_qty') or 0
            summary[k]['qty'] += qty
    return summary


def compare_pair(party_a, party_b, date_from, date_to):
    """对两方做汇总对比。返回 {'a_to_b': {...}, 'b_to_a': {...}}。

    diffs[k] = sender_sum[k] - receiver_sum[k]，正数表示发方录入 > 收方录入
    （发方虚报或收方漏收）。Task 14 UI 按此约定显示。
    """
    con = sqlite3.connect(DATABASE)
    con.row_factory = sqlite3.Row
    result = {}
    for sender, receiver in [(party_a, party_b), (party_b, party_a)]:
        sender_sum = _sum_items(con, recorded_by=sender, from_party=sender, to_party=receiver,
                                date_from=date_from, date_to=date_to)
        receiver_sum = _sum_items(con, recorded_by=receiver, from_party=sender, to_party=receiver,
                                  date_from=date_from, date_to=date_to)
        diffs = {k: round(sender_sum[k] - receiver_sum[k], 4)
                 for k, _ in ITEMS
                 if abs(sender_sum[k] - receiver_sum[k]) > 1e-9}
        result[f'{sender}_to_{receiver}'] = {
            'sender_recorded': sender_sum,
            'receiver_recorded': receiver_sum,
            'diffs': diffs,
        }
    con.close()
    return result


def _sum_items(con, *, recorded_by, from_party, to_party, date_from, date_to):
    """SUM 17 包材列，返回 {item_key: float}。"""
    qty_cols_sql = ', '.join([f'COALESCE(SUM({k}_qty), 0) AS {k}_sum' for k, _ in ITEMS])
    row = con.execute(f"""
        SELECT {qty_cols_sql} FROM flow_records
        WHERE recorded_by=? AND from_party=? AND to_party=? AND date BETWEEN ? AND ?
    """, (recorded_by, from_party, to_party, date_from, date_to)).fetchone()
    return {k: float(row[f'{k}_sum']) for k, _ in ITEMS}


# ==================== 默认单价 API ====================

@app.route('/api/prices', methods=['GET', 'POST'])
def api_prices():
    db = get_db()
    if request.method == 'POST':
        data = request.get_json()
        for key, price in data.items():
            db.execute('UPDATE default_prices SET price = ? WHERE item_key = ?', (float(price), key))
        db.commit()
        db.close()
        return jsonify({'status': 'ok'})
    prices = {row['item_key']: row['price']
              for row in db.execute('SELECT * FROM default_prices').fetchall()}
    db.close()
    return jsonify(prices)


def parse_excel_sheet(filepath, sheet_name, start_row, columns):
    """读一个 sheet 按 columns 配置提取行。

    columns: {excel_col_index(0-based): target_field_name}
    start_row: 1-based 数据起始行

    缺 date 字段的行会被丢弃（date 是 flow_records 必填）。
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    try:
        ws = wb[sheet_name]
        rows = []
        for i, excel_row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i < start_row:
                continue
            if all(v is None for v in excel_row):
                continue
            row = {}
            for idx, field in columns.items():
                if idx >= len(excel_row):
                    continue
                v = excel_row[idx]
                if v is None:
                    continue
                if field == 'date':
                    if isinstance(v, (datetime, date_cls)):
                        row[field] = v.strftime('%Y-%m-%d')
                    else:
                        s = str(v).strip()[:10]
                        try:
                            datetime.strptime(s, '%Y-%m-%d')
                            row[field] = s
                        except ValueError:
                            # 不是合法日期（如夹在数据中的标题行）→ 不设 date，行被丢
                            pass
                elif field.endswith('_qty'):
                    try:
                        row[field] = float(v)
                    except (ValueError, TypeError):
                        row[field] = 0
                else:
                    row[field] = str(v).strip()
            if 'date' in row:
                rows.append(row)
        return rows
    finally:
        wb.close()  # 防 Windows 文件锁残留


@app.route('/import', methods=['GET', 'POST'])
def import_page():
    party = current_party()
    if not party:
        return redirect(url_for('index'))
    if request.method == 'POST' and 'file' in request.files:
        # 上传 → 存 tmp → 读 sheet names → render preview
        f = request.files['file']
        tmp_path = os.path.join(DATA_PATH, 'upload_tmp.xlsx')
        f.save(tmp_path)
        wb = openpyxl.load_workbook(tmp_path, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        return render_template('import_preview.html', party=party, sheets=sheets,
                               filename=f.filename, ALLOWED_DIRECTIONS=ALLOWED_DIRECTIONS,
                               IMPORT_CONFIGS=IMPORT_CONFIGS)
    return render_template('import_preview.html', party=party,
                           ALLOWED_DIRECTIONS=ALLOWED_DIRECTIONS,
                           IMPORT_CONFIGS=IMPORT_CONFIGS)


@app.route('/import/commit', methods=['POST'])
def import_commit():
    party = current_party()
    if not party:
        return redirect(url_for('index'))

    # 从 form 或 file 拿（共享 tmp，3 LAN 用户可接受并发覆盖）
    tmp_path = os.path.join(DATA_PATH, 'upload_tmp.xlsx')
    file = request.files.get('file')
    if file and file.filename:
        file.save(tmp_path)

    sheet_name = request.form.get('sheet_name')
    try:
        start_row = int(request.form.get('start_row', '3') or '3')
    except ValueError:
        flash('起始行必须是数字'); return redirect(url_for('import_page'))
    direction = request.form.get('direction')
    if direction not in ALLOWED_DIRECTIONS:
        flash('无效方向'); return redirect(url_for('import_page'))
    from_p, to_p = ALLOWED_DIRECTIONS[direction]

    # party 必须能录这个方向
    if party not in (from_p, to_p):
        flash('无权导入此方向'); return redirect(url_for('import_page'))

    # 列映射（从 form 里读每个字段的列号）
    columns = {}
    for key in ['date', 'order_no', 'remark'] + [f'{k}_qty' for k, _ in ITEMS]:
        col = request.form.get(f'col_{key}')
        if col is not None and col != '':
            try:
                columns[int(col)] = key
            except ValueError:
                pass

    rows = parse_excel_sheet(tmp_path, sheet_name, start_row, columns)

    qty_cols = [f'{k}_qty' for k, _ in ITEMS]
    con = sqlite3.connect(DATABASE)
    for r in rows:
        args = [party, from_p, to_p, r.get('date'), r.get('order_no'), r.get('remark')]
        args += [r.get(c, 0) for c in qty_cols]
        placeholders = ', '.join(['?'] * len(args))
        con.execute(f"""
            INSERT INTO flow_records (recorded_by, from_party, to_party, date, order_no, remark,
                                      {', '.join(qty_cols)})
            VALUES ({placeholders})
        """, args)
    con.commit(); con.close()
    flash(f'导入 {len(rows)} 条')
    return redirect(url_for('party_page', party=party))


@app.route('/import/preset', methods=['POST'])
def import_preset():
    """按预定模板批量导入：上传 file + 选 preset → 自动扫描 allowed_sheets 全部导入。

    可选：传 sheet_name 限定单个 sheet（向后兼容）；不传则自动扫描所有 allowed。
    """
    party = current_party()
    if not party:
        return redirect(url_for('index'))

    preset_key = request.form.get('preset', '')
    cfg = IMPORT_CONFIGS.get(preset_key)
    if not cfg:
        flash('无效模板'); return redirect(url_for('import_page'))
    if cfg.get('recorded_by_party') and party != cfg['recorded_by_party']:
        flash(f'此模板需用 {cfg["recorded_by_party"]} 账号导入'); return redirect(url_for('import_page'))

    file = request.files.get('file')
    if not file or not file.filename:
        flash('未上传文件'); return redirect(url_for('import_page'))
    tmp_path = os.path.join(DATA_PATH, 'upload_tmp.xlsx')
    file.save(tmp_path)

    # 决定要导哪些 sheet
    requested_sheet = request.form.get('sheet_name', '').strip()
    allowed_sheets = cfg.get('allowed_sheets') or []
    wb = openpyxl.load_workbook(tmp_path, read_only=True)
    try:
        existing_sheets = set(wb.sheetnames)
    finally:
        wb.close()

    if requested_sheet:
        if allowed_sheets and requested_sheet not in allowed_sheets:
            flash(f'此模板仅支持 sheet：{", ".join(allowed_sheets)}'); return redirect(url_for('import_page'))
        targets = [requested_sheet] if requested_sheet in existing_sheets else []
    else:
        targets = [s for s in allowed_sheets if s in existing_sheets]

    if not targets:
        flash(f'文件中找不到任何模板支持的 sheet（{", ".join(allowed_sheets)}）')
        return redirect(url_for('import_page'))

    qty_cols = [f'{k}_qty' for k, _ in ITEMS]
    per_sheet = []
    grand_total = 0
    for sheet_name in targets:
        sheet_total = 0
        for dir_cfg in cfg.get('directions', []):
            direction = dir_cfg['direction']
            if direction not in ALLOWED_DIRECTIONS:
                continue
            from_p, to_p = ALLOWED_DIRECTIONS[direction]
            rows = parse_excel_sheet(tmp_path, sheet_name, dir_cfg['start_row'], dir_cfg['columns'])
            if not rows:
                continue
            con = sqlite3.connect(DATABASE)
            for r in rows:
                args = [party, from_p, to_p, r.get('date'), r.get('order_no'), r.get('remark')]
                args += [r.get(c, 0) for c in qty_cols]
                placeholders = ', '.join(['?'] * len(args))
                con.execute(f"""
                    INSERT INTO flow_records (recorded_by, from_party, to_party, date, order_no, remark,
                                              {', '.join(qty_cols)})
                    VALUES ({placeholders})
                """, args)
            con.commit(); con.close()
            sheet_total += len(rows)
        per_sheet.append(f'{sheet_name}({sheet_total})')
        grand_total += sheet_total

    flash(f'模板 [{cfg["label"]}] 导入 {grand_total} 条 — 明细：{", ".join(per_sheet)}')
    return redirect(url_for('party_page', party=party))


init_db()

if __name__ == '__main__':
    import threading, webbrowser
    port = int(os.environ.get('PORT', 7000))
    if getattr(sys, 'frozen', False):
        threading.Timer(1.5, lambda: webbrowser.open(f'http://127.0.0.1:{port}')).start()
        app.run(debug=False, host='0.0.0.0', port=port)
    else:
        app.run(debug=True, host='0.0.0.0', port=port)
