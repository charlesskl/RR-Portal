import sqlite3
import openpyxl
import app as app_module


def _login(client, p='hd'):
    with client.session_transaction() as s:
        s['party'] = p


def test_import_commits_rows(client, tmp_path):
    """上传 xlsx → 选 sheet + 方向 → 提交 → 入库。"""
    _login(client, 'hd')
    path = tmp_path / 'test.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = 'sheet1'
    ws['A2'] = '日期'; ws['B2'] = '订单号'; ws['C2'] = '胶箱'
    ws['A3'] = '2026-01-01'; ws['B3'] = 'ORD1'; ws['C3'] = 5
    wb.save(path)

    # 提交（内部流程：upload + commit 一步完成给测试用）
    with open(path, 'rb') as f:
        rv = client.post('/import/commit',
                         data={
                             'sheet_name': 'sheet1',
                             'start_row': '3',
                             'direction': 'hd_to_sy',
                             'col_date': '0', 'col_order_no': '1', 'col_jx_qty': '2',
                             'file': (f, 'test.xlsx'),
                         },
                         content_type='multipart/form-data')
    assert rv.status_code == 302
    con = sqlite3.connect(app_module.DATABASE)
    row = con.execute(
        "SELECT recorded_by, from_party, to_party, date, order_no, jx_qty FROM flow_records"
    ).fetchone()
    assert row == ('hd', 'hd', 'sy', '2026-01-01', 'ORD1', 5.0)


def test_import_rejects_garbage_start_row(client):
    """非数字 start_row 不应 500，而是 flash + redirect。"""
    _login(client, 'hd')
    rv = client.post('/import/commit',
                     data={'sheet_name': 'sheet1', 'start_row': 'abc', 'direction': 'hd_to_sy'},
                     follow_redirects=False)
    assert rv.status_code == 302
    con = sqlite3.connect(app_module.DATABASE)
    assert con.execute("SELECT COUNT(*) FROM flow_records").fetchone()[0] == 0


def test_import_rejects_invalid_direction(client):
    """direction 不在白名单 → flash + redirect 不入库。"""
    _login(client, 'hd')
    rv = client.post('/import/commit',
                     data={'sheet_name': 'sheet1', 'start_row': '3', 'direction': 'bogus'},
                     follow_redirects=False)
    assert rv.status_code == 302
    con = sqlite3.connect(app_module.DATABASE)
    assert con.execute("SELECT COUNT(*) FROM flow_records").fetchone()[0] == 0


def test_import_rejects_party_not_in_pair(client):
    """xx 想导入 hd_to_sy 方向（自己不在 pair 内）→ 拒。"""
    _login(client, 'xx')
    rv = client.post('/import/commit',
                     data={'sheet_name': 'sheet1', 'start_row': '3', 'direction': 'hd_to_sy'},
                     follow_redirects=False)
    assert rv.status_code == 302
    con = sqlite3.connect(app_module.DATABASE)
    assert con.execute("SELECT COUNT(*) FROM flow_records").fetchone()[0] == 0


def test_import_preset_dual_direction(client, tmp_path):
    """huadeng_qingxi_shaoyang 模板：同一 sheet 左右两块各 1 条 → 入库 2 条，方向正确。"""
    _login(client, 'hd')
    path = tmp_path / 'qx_sy.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = '1月'
    # row 1 (0-based) headers; row 2 起数据。spec start_row=3 → 第 3 行起
    # 左半 (0-15): hd→sy 一条 jx_qty=11
    # 右半 (17-29): sy→hd 一条 jx_qty=22
    ws.cell(3, 1, '2026-01-04')   # col 0, A 列
    ws.cell(3, 2, 'L1')           # col 1
    ws.cell(3, 3, 11)             # col 2 jx
    ws.cell(3, 18, '2026-01-05')  # col 17 right date
    ws.cell(3, 19, 'R1')          # col 18 right order_no
    ws.cell(3, 20, 22)            # col 19 right jx
    wb.save(path)

    with open(path, 'rb') as f:
        rv = client.post('/import/preset',
                         data={'preset': 'huadeng_qingxi_shaoyang',
                               'sheet_name': '1月',
                               'file': (f, 'qx_sy.xlsx')},
                         content_type='multipart/form-data',
                         follow_redirects=False)
    assert rv.status_code == 302
    con = sqlite3.connect(app_module.DATABASE)
    rows = con.execute(
        "SELECT recorded_by, from_party, to_party, date, order_no, jx_qty FROM flow_records ORDER BY id"
    ).fetchall()
    assert len(rows) == 2
    assert rows[0] == ('hd', 'hd', 'sy', '2026-01-04', 'L1', 11.0)
    assert rows[1] == ('hd', 'sy', 'hd', '2026-01-05', 'R1', 22.0)


def test_import_preset_rejects_wrong_party(client, tmp_path):
    """模板要求 hd 账号，sy 登录则拒。"""
    _login(client, 'sy')
    path = tmp_path / 'q.xlsx'
    wb = openpyxl.Workbook(); wb.active.title = '1月'; wb.save(path)
    with open(path, 'rb') as f:
        rv = client.post('/import/preset',
                         data={'preset': 'huadeng_qingxi_shaoyang',
                               'sheet_name': '1月',
                               'file': (f, 'q.xlsx')},
                         content_type='multipart/form-data',
                         follow_redirects=False)
    assert rv.status_code == 302
    con = sqlite3.connect(app_module.DATABASE)
    assert con.execute("SELECT COUNT(*) FROM flow_records").fetchone()[0] == 0


def test_import_preset_rejects_disallowed_sheet(client, tmp_path):
    """模板限定 1-4 月，'总表' 不应通过。"""
    _login(client, 'hd')
    path = tmp_path / 'q.xlsx'
    wb = openpyxl.Workbook(); wb.active.title = '总表'; wb.save(path)
    with open(path, 'rb') as f:
        rv = client.post('/import/preset',
                         data={'preset': 'huadeng_qingxi_shaoyang',
                               'sheet_name': '总表',
                               'file': (f, 'q.xlsx')},
                         content_type='multipart/form-data',
                         follow_redirects=False)
    assert rv.status_code == 302
    con = sqlite3.connect(app_module.DATABASE)
    assert con.execute("SELECT COUNT(*) FROM flow_records").fetchone()[0] == 0
