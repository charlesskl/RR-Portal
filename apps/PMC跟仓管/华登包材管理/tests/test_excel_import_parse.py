import openpyxl
import app as app_module


def test_parse_sheet_returns_rows(tmp_path):
    """构造一个迷你 xlsx 文件 → 跑 parse_sheet → 验证。"""
    path = tmp_path / 'test.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'test_sheet'
    ws['A1'] = '测试'
    ws['A2'] = '日期'; ws['B2'] = '订单号'; ws['C2'] = '胶箱'
    ws['A3'] = '2026-01-01'; ws['B3'] = 'ORD1'; ws['C3'] = 5
    ws['A4'] = '2026-01-02'; ws['B4'] = 'ORD2'; ws['C4'] = 10
    wb.save(path)

    rows = app_module.parse_excel_sheet(
        str(path), 'test_sheet',
        start_row=3,
        columns={0: 'date', 1: 'order_no', 2: 'jx_qty'}
    )
    assert len(rows) == 2
    assert rows[0]['date'] == '2026-01-01'
    assert rows[0]['jx_qty'] == 5


def test_parse_skips_embedded_header_rows(tmp_path):
    """同 sheet 中数据行后又夹了标题/列头行 → 不应被当数据导入。"""
    path = tmp_path / 'multi_block.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 's'
    ws['A1'] = '清溪华登发邵阳华登'   # 标题
    ws['A2'] = '日期'; ws['B2'] = '订单号'; ws['C2'] = '胶箱'  # 列头
    ws['A3'] = '2026-01-01'; ws['B3'] = 'O1'; ws['C3'] = 5     # 数据 1
    ws['A4'] = '2026-01-02'; ws['B4'] = 'O2'; ws['C4'] = 10    # 数据 2
    ws['A5'] = '清溪华登发邵阳华登'   # ← 嵌入标题行（应被跳）
    ws['A6'] = '日期'; ws['B6'] = '订单号'; ws['C6'] = '胶箱'  # ← 嵌入列头（应被跳）
    ws['A7'] = '2026-02-01'; ws['B7'] = 'O3'; ws['C7'] = 7     # 数据 3
    wb.save(path)

    rows = app_module.parse_excel_sheet(
        str(path), 's', start_row=3,
        columns={0: 'date', 1: 'order_no', 2: 'jx_qty'}
    )
    dates = [r['date'] for r in rows]
    assert dates == ['2026-01-01', '2026-01-02', '2026-02-01']
    # 嵌入的中文标题行不在结果中
    assert all('清溪华登发' not in r.get('date', '') for r in rows)
