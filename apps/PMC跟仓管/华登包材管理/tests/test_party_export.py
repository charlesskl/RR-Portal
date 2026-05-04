import sqlite3
import io
import openpyxl
import app as app_module


def _login(client, p='hd'):
    with client.session_transaction() as s:
        s['party'] = p


def _insert(recorded_by='hd', from_p='hd', to_p='sy', date='2026-05-01', jx_qty=0):
    con = sqlite3.connect(app_module.DATABASE)
    con.execute(
        "INSERT INTO flow_records (recorded_by, from_party, to_party, date, jx_qty) VALUES (?, ?, ?, ?, ?)",
        (recorded_by, from_p, to_p, date, jx_qty)
    )
    con.commit(); con.close()


def test_export_returns_xlsx(client):
    """hd 导出对 sy 的双向流水 → 200 + xlsx mime + 含两个 sheet。"""
    _login(client, 'hd')
    _insert(recorded_by='hd', from_p='hd', to_p='sy', date='2026-05-01', jx_qty=10)
    _insert(recorded_by='hd', from_p='sy', to_p='hd', date='2026-05-02', jx_qty=5)
    rv = client.get('/party/hd/export?cp=sy')
    assert rv.status_code == 200
    assert 'spreadsheetml' in rv.mimetype
    wb = openpyxl.load_workbook(io.BytesIO(rv.data))
    assert any('发' in name for name in wb.sheetnames)
    assert any('收自' in name for name in wb.sheetnames)


def test_export_rejects_invalid_cp(client):
    """cp 不在白名单 → flash + redirect 不返回 xlsx。"""
    _login(client, 'hd')
    rv = client.get('/party/hd/export?cp=hd', follow_redirects=False)  # 自己
    assert rv.status_code == 302
    rv = client.get('/party/hd/export?cp=zz', follow_redirects=False)  # 不存在
    assert rv.status_code == 302


def test_export_requires_login(client):
    """未登录 → @party_required 重定向到 login。"""
    rv = client.get('/party/hd/export?cp=sy', follow_redirects=False)
    assert rv.status_code == 302
    assert '/login' in rv.location
