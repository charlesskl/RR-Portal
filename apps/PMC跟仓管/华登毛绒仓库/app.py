"""
华登塑胶 · 库存管理系统(毛绒 + 戏服)
Flask 主程序

【骨架版】已实现:
- 登录 / 登出 / 当前用户查询
- 入库 API(查询 / 新增 / 删除)— 支持 category 筛选,作为示例
- 库存汇总 API(实时计算)

【待实现】给 Claude Code:
- 出库 API(仿照入库写,字段多 po 和 picker)
- 布标 API
- 用户管理 API
- 导出 CSV API
"""
import os
import re
import json
import shutil
import sqlite3
import csv
import io
import hashlib
from datetime import datetime
from urllib.parse import quote
from flask import Flask, render_template, request, jsonify, session, redirect, url_for, Response
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

import database as db


SCHEDULE_IMG_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                'static', 'schedule_images')
from auth import (
    hash_password, check_password,
    login_required, role_required,
    current_user
)


app = Flask(__name__)

# Session 密钥从环境变量注入；本地开发可不设，云端 compose 强制传入
app.secret_key = os.environ.get('HUADENG_MAORONG_SECRET_KEY') or 'dev-only-CHANGE-IN-PROD'

# Session 7 天有效
app.permanent_session_lifetime = 60 * 60 * 24 * 7

# 模块加载时确保表结构存在(CREATE IF NOT EXISTS,幂等)
db.init_database()


@app.route('/health')
def health():
    return {'status': 'ok'}, 200


# ==================== 工具函数 ====================

def validate_category(category):
    """验证品类参数,返回标准化后的值或 None"""
    if category in ('plush', 'costume'):
        return category
    return None


def validate_record_payload(data, allow_optional=()):
    """
    校验出入库 payload,返回 (data, None) 或 (None, (error_msg, status))
    新增 / 编辑 共用,出库可通过 allow_optional=('po','picker') 传入可空字段
    """
    category = data.get('category', 'plush')
    if category not in ('plush', 'costume'):
        return None, ({'error': '品类只能是 plush(毛绒) 或 costume(戏服)'}, 400)
    data['category'] = category
    # 戏服自动 flag = ''
    if category == 'costume':
        data['flag'] = ''
    required = ['date', 'billNo', 'sku', 'qty']
    if category == 'plush':
        # 毛绒:style(normal/rare)和 flag 都必填
        required.extend(['style', 'flag'])
    for field in required:
        if not data.get(field):
            return None, ({'error': f'缺少必填字段: {field}'}, 400)
    if category == 'plush' and data['style'] not in ('normal', 'rare'):
        return None, ({'error': '毛绒款式必须是 normal 或 rare'}, 400)
    # 戏服 style 可选,空时归一为空字符串(避免下游 NPE)
    if category == 'costume' and not data.get('style'):
        data['style'] = ''
    try:
        qty = int(data['qty'])
        if qty <= 0:
            return None, ({'error': '数量必须大于 0'}, 400)
        data['qty'] = qty
    except (ValueError, TypeError):
        return None, ({'error': '数量必须是数字'}, 400)
    if 'flag' not in data:
        data['flag'] = ''
    return data, None


# ==================== 页面路由 ====================

@app.route('/')
def index():
    """首页:已登录跳主程序,未登录跳登录页"""
    if 'user_id' not in session:
        return redirect(url_for('login_page'))
    return render_template('app.html', user=current_user())


@app.route('/login')
def login_page():
    """登录页面"""
    if 'user_id' in session:
        return redirect(url_for('index'))
    return render_template('login.html')


# ==================== 认证 API ====================

@app.route('/api/login', methods=['POST'])
def api_login():
    """登录"""
    data = request.get_json() or {}
    username = (data.get('username') or '').strip()
    password = data.get('password') or ''

    if not username or not password:
        return jsonify({'error': '请输入用户名和密码'}), 400

    user = db.get_user_by_username(username)
    if not user or not check_password(password, user['password_hash']):
        return jsonify({'error': '用户名或密码错误'}), 401

    # 写入 session
    session.permanent = True
    session['user_id'] = user['id']
    session['username'] = user['username']
    session['role'] = user['role']
    session['display_name'] = user['display_name'] or user['username']

    return jsonify({
        'success': True,
        'user': {
            'username': user['username'],
            'role': user['role'],
            'display_name': user['display_name']
        }
    })


@app.route('/api/logout', methods=['POST'])
def api_logout():
    """登出"""
    session.clear()
    return jsonify({'success': True})


@app.route('/api/me', methods=['GET'])
@login_required
def api_me():
    """获取当前登录用户"""
    return jsonify(current_user())


# ==================== 入库 API(示例完整实现) ====================

@app.route('/api/in', methods=['GET'])
@login_required
def api_in_list():
    """
    查询入库记录(所有角色都能看)
    Query 参数:
        category: 可选,'plush' 或 'costume',筛选品类
    """
    category = validate_category(request.args.get('category'))
    records = db.query_all_in_records(category=category)
    return jsonify(records)


@app.route('/api/in', methods=['POST'])
@role_required('admin', 'operator')
def api_in_create():
    """新增入库记录(主管 / 仓管员)"""
    data, err = validate_record_payload(request.get_json() or {})
    if err:
        body, status = err
        return jsonify(body), status
    record_id = db.insert_in_record(data, session.get('username'))
    return jsonify({'success': True, 'id': record_id})


@app.route('/api/in/<int:record_id>', methods=['PUT'])
@role_required('admin', 'operator')
def api_in_update(record_id):
    """编辑入库记录(主管 / 仓管员)"""
    data, err = validate_record_payload(request.get_json() or {})
    if err:
        body, status = err
        return jsonify(body), status
    affected = db.update_in_record(record_id, data)
    if affected == 0:
        return jsonify({'error': '记录不存在'}), 404
    return jsonify({'success': True})


@app.route('/api/in/<int:record_id>', methods=['DELETE'])
@role_required('admin')
def api_in_delete(record_id):
    """删除入库记录(仅主管)"""
    db.delete_in_record(record_id)
    return jsonify({'success': True})


@app.route('/api/in/batch-delete', methods=['POST'])
@role_required('admin')
def api_in_batch_delete():
    """批量删除入库记录(仅主管)。body: {ids: [int, ...]}"""
    ids = (request.get_json() or {}).get('ids') or []
    ids = [int(i) for i in ids]
    if not ids:
        return jsonify({'error': '没有要删除的记录'}), 400
    count = db.delete_in_records_batch(ids)
    return jsonify({'success': True, 'count': count})


# ==================== 出库 API ====================

@app.route('/api/out', methods=['GET'])
@login_required
def api_out_list():
    """
    查询出库记录(所有角色都能看)
    Query 参数:
        category: 可选,'plush' 或 'costume',筛选品类
    """
    category = validate_category(request.args.get('category'))
    records = db.query_all_out_records(category=category)
    return jsonify(records)


@app.route('/api/out', methods=['POST'])
@role_required('admin', 'operator')
def api_out_create():
    """新增出库记录(主管 / 仓管员)。出库多 po(可空)/ picker(可空)两个字段"""
    data, err = validate_record_payload(request.get_json() or {})
    if err:
        body, status = err
        return jsonify(body), status
    record_id = db.insert_out_record(data, session.get('username'))
    return jsonify({'success': True, 'id': record_id})


@app.route('/api/out/batch', methods=['POST'])
@role_required('admin', 'operator')
def api_out_batch_create():
    """批量出库(排期对比页一键出库用)。单事务,任一条失败全部回滚。"""
    payload = request.get_json() or {}
    records = payload.get('records') or []
    if not records:
        return jsonify({'error': '没有要出库的记录'}), 400
    cleaned = []
    for i, r in enumerate(records):
        d, err = validate_record_payload(r)
        if err:
            body, status = err
            return jsonify({'error': f'第 {i+1} 条: {body["error"]}'}), status
        cleaned.append(d)
    ids = db.insert_out_records_batch(cleaned, session.get('username'))
    return jsonify({'success': True, 'ids': ids, 'count': len(ids)})


@app.route('/api/out/<int:record_id>', methods=['PUT'])
@role_required('admin', 'operator')
def api_out_update(record_id):
    """编辑出库记录(主管 / 仓管员)"""
    data, err = validate_record_payload(request.get_json() or {})
    if err:
        body, status = err
        return jsonify(body), status
    affected = db.update_out_record(record_id, data)
    if affected == 0:
        return jsonify({'error': '记录不存在'}), 404
    return jsonify({'success': True})


@app.route('/api/out/<int:record_id>', methods=['DELETE'])
@role_required('admin')
def api_out_delete(record_id):
    """删除出库记录(仅主管)"""
    db.delete_out_record(record_id)
    return jsonify({'success': True})


@app.route('/api/out/batch-delete', methods=['POST'])
@role_required('admin')
def api_out_batch_delete():
    """批量删除出库记录(仅主管)。body: {ids: [int, ...]}"""
    ids = (request.get_json() or {}).get('ids') or []
    ids = [int(i) for i in ids]
    if not ids:
        return jsonify({'error': '没有要删除的记录'}), 400
    count = db.delete_out_records_batch(ids)
    return jsonify({'success': True, 'count': count})


# ==================== 库存查询 API ====================

@app.route('/api/stock', methods=['GET'])
@login_required
def api_stock_summary():
    """
    查询库存汇总(实时计算)
    Query 参数:
        category: 可选,筛选品类

    返回:list[dict],每条字段:
        category, sku, name, style, flag, in_total, out_total, stock
    """
    category = validate_category(request.args.get('category'))
    summary = db.get_stock_summary(category=category)
    return jsonify(summary)


# ==================== 布标 API ====================

@app.route('/api/flags', methods=['GET'])
@login_required
def api_flags_list():
    """查询所有布标(任何登录用户)"""
    return jsonify(db.query_all_flags())


@app.route('/api/flags', methods=['POST'])
@role_required('admin')
def api_flags_create():
    """新增布标(仅 admin),body: {"name": "荷兰"}"""
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': '布标名称不能为空'}), 400

    try:
        db.add_flag(name)
    except sqlite3.IntegrityError:
        return jsonify({'error': f'布标 "{name}" 已存在'}), 400

    return jsonify({'success': True})


@app.route('/api/flags/<path:name>', methods=['DELETE'])
@role_required('admin')
def api_flags_delete(name):
    """删除布标(仅 admin),URL 中的中文名 Flask 自动解码"""
    affected = db.delete_flag(name)
    if affected == 0:
        return jsonify({'error': f'布标 "{name}" 不存在'}), 404
    return jsonify({'success': True})


# ==================== 用户管理 API ====================

VALID_ROLES = ('admin', 'operator', 'viewer')


@app.route('/api/users', methods=['GET'])
@role_required('admin')
def api_users_list():
    """列出所有用户(仅 admin)"""
    return jsonify(db.query_all_users())


@app.route('/api/users', methods=['POST'])
@role_required('admin')
def api_users_create():
    """
    新增用户(仅 admin)
    Body: username(4-20), password(>=4), role, display_name(可空)
    """
    data = request.get_json() or {}
    username = (data.get('username') or '').strip()
    password = data.get('password') or ''
    role = data.get('role') or ''
    display_name = (data.get('display_name') or '').strip()

    if not (4 <= len(username) <= 20):
        return jsonify({'error': '用户名长度必须为 4-20 个字符'}), 400
    if len(password) < 4:
        return jsonify({'error': '密码至少 4 位'}), 400
    if role not in VALID_ROLES:
        return jsonify({'error': f'角色必须是 {VALID_ROLES} 之一'}), 400

    try:
        new_id = db.insert_user(username, hash_password(password), role, display_name)
    except sqlite3.IntegrityError:
        return jsonify({'error': f'用户名 "{username}" 已存在'}), 400

    return jsonify({'success': True, 'id': new_id})


@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@role_required('admin')
def api_users_delete(user_id):
    """删除用户(仅 admin,禁止删自己)"""
    if user_id == session.get('user_id'):
        return jsonify({'error': '不能删除当前登录的自己'}), 400

    affected = db.delete_user(user_id)
    if affected == 0:
        return jsonify({'error': '用户不存在'}), 404
    return jsonify({'success': True})


@app.route('/api/users/<int:user_id>/password', methods=['PUT'])
@login_required
def api_users_change_password(user_id):
    """
    改密码:
      - admin 可改任何人,old_password 可省略
      - 非 admin 只能改自己,必须验证 old_password
    Body: { "old_password": "...", "new_password": "..." }
    """
    data = request.get_json() or {}
    new_password = data.get('new_password') or ''
    old_password = data.get('old_password') or ''

    if len(new_password) < 4:
        return jsonify({'error': '新密码至少 4 位'}), 400

    current = current_user()
    target = db.get_user_by_id(user_id)
    if not target:
        return jsonify({'error': '用户不存在'}), 404

    # 非 admin:只能改自己 + 必须验证旧密码
    if current['role'] != 'admin':
        if user_id != current['id']:
            return jsonify({'error': '只能修改自己的密码'}), 403
        if not check_password(old_password, target['password_hash']):
            return jsonify({'error': '旧密码错误'}), 400

    db.update_user_password(user_id, hash_password(new_password))
    return jsonify({'success': True})


# ==================== 导出 CSV API ====================

def _category_label(c):
    return '毛绒' if c == 'plush' else '戏服' if c == 'costume' else (c or '')


def _style_label(s):
    """毛绒款式中文化,戏服直接返回原文本"""
    if s == 'normal': return '普通款'
    if s == 'rare': return '稀有款'
    return s or ''


def _csv_response(rows, header, filename):
    """
    生成 CSV 响应
    - 文件首字节写 BOM,Excel 打开中文不乱码
    - filename URL 编码,支持中文文件名
    """
    output = io.StringIO()
    output.write('﻿')
    writer = csv.writer(output)
    writer.writerow(header)
    for row in rows:
        writer.writerow(row)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{quote(filename)}"}
    )


def _filter_by_date(records, date_from, date_to):
    if date_from:
        records = [r for r in records if r['date'] >= date_from]
    if date_to:
        records = [r for r in records if r['date'] <= date_to]
    return records


@app.route('/api/export/in', methods=['GET'])
@login_required
def api_export_in():
    """导出入库流水。Query: category / from / to"""
    category = validate_category(request.args.get('category'))
    records = _filter_by_date(
        db.query_all_in_records(category=category),
        request.args.get('from'), request.args.get('to')
    )
    rows = [[
        _category_label(r['category']), r['date'], r['bill_no'], r['sku'],
        r.get('name', ''), _style_label(r['style']), r['flag'], r['qty']
    ] for r in records]
    header = ['品类', '日期', '单号', '货号', '物料名称', '款式', '布标', '数量']
    filename = f'入库流水_{datetime.now().strftime("%Y%m%d")}.csv'
    return _csv_response(rows, header, filename)


@app.route('/api/export/out', methods=['GET'])
@login_required
def api_export_out():
    """导出出库流水。Query: category / from / to"""
    category = validate_category(request.args.get('category'))
    records = _filter_by_date(
        db.query_all_out_records(category=category),
        request.args.get('from'), request.args.get('to')
    )
    rows = [[
        _category_label(r['category']), r['date'], r['bill_no'],
        r.get('po', '') or '', r.get('picker', '') or '',
        r['sku'], r.get('name', ''), _style_label(r['style']), r['flag'], r['qty']
    ] for r in records]
    header = ['品类', '日期', '单号', 'PO号', '领货人', '货号', '物料名称', '款式', '布标', '数量']
    filename = f'出库流水_{datetime.now().strftime("%Y%m%d")}.csv'
    return _csv_response(rows, header, filename)


@app.route('/api/export/stock', methods=['GET'])
@login_required
def api_export_stock():
    """导出当前库存。Query: category"""
    category = validate_category(request.args.get('category'))
    summary = db.get_stock_summary(category=category)
    rows = [[
        _category_label(s['category']), s['sku'], s.get('name', ''),
        _style_label(s['style']), s['flag'],
        s['in_total'], s['out_total'], s['stock']
    ] for s in summary]
    header = ['品类', '货号', '物料名称', '款式', '布标', '入库累计', '出库累计', '当前库存']
    filename = f'库存总览_{datetime.now().strftime("%Y%m%d")}.csv'
    return _csv_response(rows, header, filename)


# ==================== 排期表导入与查询 ====================

def _is_blue_font(font):
    """字体颜色是否属于"已完货蓝":R<80 && G<80 && B>150"""
    if not font or not font.color or font.color.type != 'rgb':
        return False
    rgb = (font.color.rgb or '').upper()
    if len(rgb) != 8:
        return False
    try:
        r = int(rgb[2:4], 16)
        g = int(rgb[4:6], 16)
        b = int(rgb[6:8], 16)
    except ValueError:
        return False
    return b > 150 and r < 80 and g < 80


def _find_schedule_header(ws, max_scan=12):
    """
    扫前 max_scan 行,跨行合并识别各种表头列(因为 AU/AV "每款普通/稀有" 表头
    在 R3,而 PO号 等核心列在 R4——必须合扫才能拿到完整字段)
    返回 (primary_row, cols) —— primary_row 是 PO号 所在的行(数据起始行 = primary_row + 1)
    """
    upper = min(ws.max_column + 1, 80)
    cols = {}
    primary_row = None
    for r in range(1, max_scan + 1):
        for c in range(1, upper):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v).strip()
            if s == 'PO号' and 'po' not in cols:
                cols['po'] = c
                primary_row = r
            elif 'ITEM#' in s and 'item' not in cols:
                cols['item'] = c
            elif s == 'SKU' and 'csku' not in cols:
                cols['csku'] = c
            elif 'PO数量' in s and 'qty' not in cols:
                cols['qty'] = c
            elif s == '第三方客户名称' and 'customer' not in cols:
                cols['customer'] = c
            elif s == '走货国家' and 'country' not in cols:
                cols['country'] = c
            elif s == '中文名' and 'name' not in cols:
                cols['name'] = c
            elif s == '计划出货期' and 'ship' not in cols:
                cols['ship'] = c
            elif s == '布标' and 'flag_type' not in cols:
                cols['flag_type'] = c
            elif s == 'MA布标' and 'ma_qty' not in cols:
                cols['ma_qty'] = c
            elif s == '客版布标' and 'kb_qty' not in cols:
                cols['kb_qty'] = c
            elif '每款普通' in s and 'ratio_normal' not in cols:
                cols['ratio_normal'] = c
            elif ('每款公仔稀有' in s or '每款稀有' in s) and 'ratio_rare' not in cols:
                cols['ratio_rare'] = c
    if primary_row and {'po', 'item', 'qty'} <= cols.keys():
        return primary_row, cols
    return None, None


def _parse_ratio(text, letter_qty):
    """
    根据排期字符串解析"普通款 / 稀有款"实际数量
    返回 (qty, mode):
      - mode='absolute': 字符串是纯数字,直接取 qty 不再 round
      - mode='fraction': 字符串含 N/M 分数,按 letter_qty * N/M round
      - mode='subtract_rare': 字符串带"减去稀有",由调用方算
      - mode=None: 读不出
    """
    if text is None:
        return None, None
    s = str(text).strip()
    if not s:
        return None, None
    # 1. 分数 N/M
    m = re.search(r'(\d+)\s*/\s*(\d+)', s)
    if m:
        num, den = int(m.group(1)), int(m.group(2))
        if den > 0:
            return int(round(letter_qty * num / den)), 'fraction'
    # 2. "减去稀有"
    if '减去稀有' in s:
        return None, 'subtract_rare'
    # 3. 纯数字
    try:
        return int(round(float(s))), 'absolute'
    except (ValueError, TypeError):
        return None, None


def _collect_letter_columns(ws, header_row):
    """
    收集字母款式列:R{header_row} 是单个大写字母的列
    返回 [(column_index, letter), ...]
    """
    out = []
    upper = min(ws.max_column + 1, 100)
    for c in range(20, upper):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        s = str(v).strip()
        if len(s) == 1 and s.isalpha() and s.isupper():
            out.append((c, s))
    return out


def _clean_material_from_column(s):
    """从列头如 '5"锁扣粉色无毛猫 Wrinkle McStinkles' 提取物料名 '粉色无毛猫'"""
    s = re.sub(r'^\d+["″“”]', '', s).strip()
    s = re.sub(r'^锁扣', '', s).strip()
    m = re.match(r'^([一-鿿]+)', s)
    return m.group(1) if m else (s.split()[0] if s else '')


def _collect_no_letter_style_columns(ws, name_row=3, marker_row=4, start_col=20):
    """
    无字母款式列(15789 模式):R{name_row} 列头是中文+英文(如 '5"锁扣粉色无毛猫 ...'),
    R{marker_row} 可能标 '稀有款'。每列对应一个独立款式。
    返回 [(col_index, material_name, style), ...]  style: 'normal'/'rare'
    """
    out = []
    upper = min(ws.max_column + 1, 100)
    for c in range(start_col, upper):
        v = ws.cell(row=name_row, column=c).value
        if v is None:
            continue
        s = str(v).strip()
        # 必须 <数字>"... 开头 + 含中文(15789 这种"5"锁扣XX猫"格式)
        if not re.match(r'^\d+["″“”]', s):
            continue
        if not re.search(r'[一-鿿]', s):
            continue
        name = _clean_material_from_column(s)
        if not name:
            continue
        marker = ws.cell(row=marker_row, column=c).value
        marker_s = str(marker).strip() if marker else ''
        style = 'rare' if '稀有' in marker_s else 'normal'
        out.append((c, name, style))
    return out


def _extract_letter_images(wb_fmt, sheet_letter_cols):
    """
    从 xlsx 中提取每个字母列上方的产品图,根据 anchor 的列号匹配字母
    sheet_letter_cols: {sheet_name: [(col_index_1based, letter), ...]}
    返回 {(sheet_name, letter): bytes}
    """
    out = {}
    for sheet_name, letter_cols in sheet_letter_cols.items():
        col_to_letter = {c: l for c, l in letter_cols}
        ws = wb_fmt[sheet_name]
        for img in getattr(ws, '_images', []) or []:
            anc = getattr(img, 'anchor', None)
            anchor_from = getattr(anc, '_from', None) if anc else None
            if anchor_from is None:
                continue
            col_1based = anchor_from.col + 1
            if col_1based in col_to_letter:
                try:
                    data = img._data() if callable(img._data) else img._data
                    out[(sheet_name, col_to_letter[col_1based])] = data
                except Exception:
                    continue
    return out


def _save_schedule_images(images):
    """
    清空 SCHEDULE_IMG_DIR 后保存图片,返回 {(sheet, letter): /static/... URL}
    用图片内容 hash 做文件名前缀,避免重复存
    """
    if os.path.exists(SCHEDULE_IMG_DIR):
        for f in os.listdir(SCHEDULE_IMG_DIR):
            try:
                os.remove(os.path.join(SCHEDULE_IMG_DIR, f))
            except OSError:
                pass
    os.makedirs(SCHEDULE_IMG_DIR, exist_ok=True)
    out = {}
    for (sheet, letter), data in images.items():
        h = hashlib.sha1(data).hexdigest()[:12]
        # 文件扩展名从 PNG 头判断
        ext = '.png' if data[:4] == b'\x89PNG' else '.jpg'
        fname = f'{h}_{letter}{ext}'
        path = os.path.join(SCHEDULE_IMG_DIR, fname)
        if not os.path.exists(path):
            with open(path, 'wb') as f:
                f.write(data)
        out[(sheet, letter)] = '/static/schedule_images/' + fname
    return out


def parse_schedule_workbook(file_bytes):
    """
    解析排期 xlsx,返回 (rows, stats)
    每一行:
      - 单款行(字母列 ≤ 1 个有值)→ 1 条记录,variant_letter='' 用 I 列数量
      - 拼盘行(字母列 ≥ 2 个有值)→ N 条记录,每条 variant_letter=该字母,qty=该字母列的数量
    蓝色字体的字母列也单独跳过(只完了部分款式的情况)
    """
    # data_only=True 拿公式计算结果(字母列大多是 =I/N 或 =I)
    # 但 data_only 拿不到字体颜色 / 图片,所以分两次加载:一份取值,一份取颜色和图片
    wb_vals = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    wb_fmt = openpyxl.load_workbook(io.BytesIO(file_bytes))
    aggregated = {}
    sheets_processed = []
    sheets_skipped = []
    sheet_letter_cols = {}  # 收集每个 sheet 的字母列(用于图片提取)
    rows_skipped_blue = 0
    rows_skipped_invalid = 0
    letter_rows_split = 0
    # 同货号配比回退缓存:数字前缀 sku → (normal_text, rare_text)
    sku_ratio_cache = {}

    for name in wb_vals.sheetnames:
        n = name.strip()
        if not n or not n[0].isdigit():
            sheets_skipped.append(name)
            continue
        ws_v = wb_vals[name]
        ws_f = wb_fmt[name]
        header_row, cols = _find_schedule_header(ws_v)
        if not header_row:
            sheets_skipped.append(name + '(无表头)')
            continue
        sheets_processed.append(name)

        letter_cols = _collect_letter_columns(ws_v, header_row)
        # letter_cols 空时,尝试用 R3 中文款式列(15789 这种 5"锁扣XX猫 模式)
        no_letter_style_cols = []
        if not letter_cols:
            no_letter_style_cols = _collect_no_letter_style_columns(ws_v, name_row=3, marker_row=4)
        sheet_letter_cols[name] = letter_cols

        for r in range(header_row + 1, ws_v.max_row + 1):
            po_cell_f = ws_f.cell(row=r, column=cols['po'])
            po_val = ws_v.cell(row=r, column=cols['po']).value
            if po_val is None or _is_blue_font(po_cell_f.font):
                if po_val is not None:
                    rows_skipped_blue += 1
                continue
            item_val = ws_v.cell(row=r, column=cols['item']).value
            qty_val = ws_v.cell(row=r, column=cols['qty']).value
            if item_val is None or qty_val is None:
                rows_skipped_invalid += 1
                continue
            if _is_blue_font(ws_f.cell(row=r, column=cols['item']).font) or \
               _is_blue_font(ws_f.cell(row=r, column=cols['qty']).font):
                rows_skipped_blue += 1
                continue
            try:
                qty_total = int(qty_val)
            except (ValueError, TypeError):
                rows_skipped_invalid += 1
                continue
            if qty_total <= 0:
                rows_skipped_invalid += 1
                continue

            po_no = str(po_val).strip()
            item_code = str(item_val).strip()

            def _cell_text(col_key):
                if col_key not in cols:
                    return ''
                v = ws_v.cell(row=r, column=cols[col_key]).value
                if v is None:
                    return ''
                if hasattr(v, 'strftime'):
                    return v.strftime('%Y-%m-%d')
                return str(v).strip()

            csku = _cell_text('csku')
            country = _cell_text('country')
            flag_type = _cell_text('flag_type')

            # 用户规则:布标类型按 MA布标列(AM)/ 客版布标列(AN)哪个有数判定,
            # 覆盖 AL 标签列(AL 可能空,或与 AM/AN 实际不一致)
            def _num_or_none(col_key):
                if col_key not in cols:
                    return None
                v = ws_v.cell(row=r, column=cols[col_key]).value
                try:
                    return int(round(float(v))) if v is not None and float(v) > 0 else None
                except (ValueError, TypeError):
                    return None
            ma_q = _num_or_none('ma_qty')
            kb_q = _num_or_none('kb_qty')
            if ma_q and not kb_q:
                flag_type = 'MA布标'
            elif kb_q and not ma_q:
                flag_type = '客版布标'
            # 都有数 / 都空 → 保持 AL 列原值

            # 库存布标 = 布标类型 + "-" + 国家(用户填库存时按此约定录)
            row_flag = f'{flag_type}-{country}' if (flag_type and country) else (flag_type or country or '')

            # 普通/稀有 比例 文本,空时回退同货号(数字前缀)上次的
            ratio_n_text = _cell_text('ratio_normal')
            ratio_r_text = _cell_text('ratio_rare')
            sku_prefix = _extract_sku_prefix(item_code)
            cached = sku_ratio_cache.get(sku_prefix, ('', ''))
            if not ratio_n_text:
                ratio_n_text = cached[0]
            if not ratio_r_text:
                ratio_r_text = cached[1]
            # 把这次解析到的更新进缓存(只在非空时更新)
            if ratio_n_text or ratio_r_text:
                sku_ratio_cache[sku_prefix] = (
                    ratio_n_text or cached[0],
                    ratio_r_text or cached[1],
                )

            common = {
                'series': name, 'po_no': po_no, 'item_code': item_code,
                'customer_sku': csku,
                'customer': _cell_text('customer'),
                'country': country,
                'flag_type': flag_type,
                'flag': row_flag,
                'ratio_normal_text': ratio_n_text,
                'ratio_rare_text': ratio_r_text,
                'name_cn': _cell_text('name'),
                'plan_ship_date': _cell_text('ship'),
            }

            # 收集字母列里的有效数量(非空、非零、非蓝色)
            # 字母列大多是公式 =I/N,data_only 拿到浮点;按 round 取整,
            # 不再做余数补差(避免出现文件里不存在的数字,如 19200/9 → 8×2133 + 1×2136)
            # 字母数量之和可能 ≠ I 列(行总数仍用 I 列存)
            letter_distrib = []
            for col_idx, letter in letter_cols:
                v = ws_v.cell(row=r, column=col_idx).value
                if v in (None, 0, '0'):
                    continue
                if _is_blue_font(ws_f.cell(row=r, column=col_idx).font):
                    continue
                try:
                    q = int(round(float(v)))
                except (ValueError, TypeError):
                    continue
                if q > 0:
                    letter_distrib.append((letter, q))

            # 不拆开:每条 xlsx 行 → 1 条 DB 记录,字母分布序列化进 letters_json
            if len(letter_distrib) >= 2:
                letter_rows_split += 1
                only_letter = ''
            else:
                only_letter = letter_distrib[0][0] if len(letter_distrib) == 1 else ''

            # 无字母款式列模式(15789):每列一个物料,直接读列数字,不算 ratio
            # xlsx 公式可能给小数(如 287.5),先 round,再把 sum vs PO 总数 的差额加到第一款
            letters_list = []
            if no_letter_style_cols and not letter_distrib:
                raw = []  # [[q_int, name, style], ...]
                for col_idx, mat_name, mat_style in no_letter_style_cols:
                    v = ws_v.cell(row=r, column=col_idx).value
                    if v in (None, 0, '0'):
                        continue
                    if _is_blue_font(ws_f.cell(row=r, column=col_idx).font):
                        continue
                    try:
                        q = int(round(float(v)))
                    except (ValueError, TypeError):
                        continue
                    if q <= 0:
                        continue
                    raw.append([q, mat_name, mat_style])
                if raw:
                    diff = qty_total - sum(x[0] for x in raw)
                    if diff != 0:
                        raw[0][0] += diff
                    letters_list = [{
                        'letter': '', 'qty': q, 'image_url': '',
                        'name': name, 'style': style,
                        'normal_qty': q if style == 'normal' else 0,
                        'rare_qty': q if style == 'rare' else 0,
                        'ratio_assumed': False,
                    } for q, name, style in raw]

            # 对每个字母按比例算 普通款 / 稀有款 数量(字母列模式)
            for l, q in letter_distrib:
                n_qty, n_mode = _parse_ratio(ratio_n_text, q)
                r_qty, r_mode = _parse_ratio(ratio_r_text, q)
                # "减去稀有款" 模式:先算另一项,再用 q 减
                if n_mode == 'subtract_rare' and r_qty is not None:
                    n_qty = q - r_qty
                if r_mode == 'subtract_rare' and n_qty is not None:
                    r_qty = q - n_qty
                # 兜底:N 或 R 一个有值另一个空 → 用 q 减
                if n_qty is not None and r_qty is None:
                    r_qty = q - n_qty
                elif r_qty is not None and n_qty is None:
                    n_qty = q - r_qty
                # 双兜底:两者都读不出,默认全部归普通款,加一个 ratio_assumed 标记
                ratio_assumed = False
                if n_qty is None and r_qty is None:
                    n_qty = q
                    r_qty = 0
                    ratio_assumed = True
                letters_list.append({
                    'letter': l, 'qty': q, 'image_url': '',
                    'normal_qty': n_qty, 'rare_qty': r_qty,
                    'ratio_assumed': ratio_assumed,
                })

            key = (name, po_no, item_code, csku)
            entry = aggregated.get(key)
            if entry:
                entry['qty'] += qty_total
                entry['letters_list'].extend(letters_list)
            else:
                aggregated[key] = {
                    **common,
                    'variant_letter': only_letter,
                    'qty': qty_total,
                    'letters_list': letters_list,
                }

    # 提取图片并保存,然后按 (series, letter) 回填到每条 letters_list 里
    images = _extract_letter_images(wb_fmt, sheet_letter_cols)
    img_urls = _save_schedule_images(images)
    for row in aggregated.values():
        for l in row.get('letters_list', []):
            url = img_urls.get((row['series'], l['letter']), '')
            if url:
                l['image_url'] = url
        # 单款行的 image_url(便于 SKU 视图直接显示)
        single_letter = row.get('variant_letter') or ''
        if single_letter:
            url = img_urls.get((row['series'], single_letter), '')
            if url:
                row['image_url'] = url
        # 序列化字母分布到 JSON,DB 存这个
        row['letters_json'] = json.dumps(row.get('letters_list', []), ensure_ascii=False)
        row.pop('letters_list', None)

    rows = list(aggregated.values())
    stats = {
        'sheets_processed': sheets_processed,
        'sheets_skipped': sheets_skipped,
        'rows_inserted': len(rows),
        'rows_skipped_blue': rows_skipped_blue,
        'rows_skipped_invalid': rows_skipped_invalid,
        'letter_rows_split': letter_rows_split,
        'images_extracted': len(img_urls),
    }
    return rows, stats


@app.route('/api/schedule/upload', methods=['POST'])
@role_required('admin', 'operator')
def api_schedule_upload():
    """上传排期 xlsx,整库替换"""
    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'error': '请上传 xlsx 文件'}), 400
    if not f.filename.lower().endswith('.xlsx'):
        return jsonify({'error': '只支持 .xlsx 格式'}), 400
    try:
        rows, stats = parse_schedule_workbook(f.read())
    except Exception as e:
        return jsonify({'error': f'解析失败: {e}'}), 400
    db.replace_po_schedules(rows, session.get('username'))
    return jsonify({'success': True, 'stats': stats})


@app.route('/api/schedule/info', methods=['GET'])
@login_required
def api_schedule_info():
    return jsonify(db.get_schedule_info())


def _parse_letters(letters_json):
    if not letters_json:
        return []
    try:
        return json.loads(letters_json)
    except (json.JSONDecodeError, TypeError):
        return []


def _extract_sku_prefix(item_code):
    """
    从 ITEM# 提取最前面的连续数字段作为入库货号
    例:'15758-S001' → '15758',  '15758A-S001-PKC1' → '15758',  '15704UQ1-S001' → '15704'
    """
    m = re.match(r'^(\d+)', str(item_code or '').strip())
    return m.group(1) if m else ''


# 货号别名缓存:子货号 → 父货号。CUD 后置空触发重载。
_alias_cache = None

def _invalidate_alias_cache():
    global _alias_cache
    _alias_cache = None

def _resolve_sku_prefix(prefix):
    """如果 prefix 是别名,返回 primary;否则返回原值。"""
    global _alias_cache
    if not prefix:
        return prefix
    if _alias_cache is None:
        _alias_cache = db.get_sku_alias_map()
    return _alias_cache.get(prefix, prefix)


def _match_sku_prefix(item_code):
    """提取 + 别名解析,业务匹配字母绑定/库存时用这个。"""
    return _resolve_sku_prefix(_extract_sku_prefix(item_code))


def _resolve_inventory_flag(flag_type, country, fallback):
    """
    把排期里的 (布标类型, 国家) 用 flag_mappings 翻译成入库布标
    没配置映射 → 返回 fallback(默认拼接 flag_type-country)
    """
    if flag_type and country:
        mp = db.get_flag_mapping(flag_type, country)
        if mp:
            return mp['inventory_flag']
    return fallback


def _clean_name_cn(name_cn):
    """name_cn 取开头的连续中文字符作为物料名(如 '猫系列盲盒锁扣怪 3FACING6PCS/PDQ' → '猫系列盲盒锁扣怪')"""
    if not name_cn:
        return ''
    m = re.match(r'^[一-鿿]+', name_cn.strip())
    return m.group(0) if m else name_cn.strip().split()[0]


def _enrich_letters_with_binding(item_code, letters, row_flag='', row_flag_type='', row_country=''):
    """
    给每个字母补充:
      - 绑定物料 (bound_name)
      - 该物料的总库存 (letter_stock,只按 sku+name 汇总)
      - 普通/稀有 分款库存 (normal_stock / rare_stock,按 sku+name+style+flag 精确查)
        库存布标先通过 flag_mappings 把 (flag_type, country) → 实际入库布标
      - 配套的缺口/是否够做 (针对 normal_qty / rare_qty)

    特殊情况:letter['letter'] == '' 表示"无字母款"(如 15789),不查 letter_bindings,
    直接用 letter['_name_cn'] 清洗后的中文当物料名,整 qty 当普通款。
    """
    sku_prefix = _match_sku_prefix(item_code)
    # 翻译入库布标
    inventory_flag = _resolve_inventory_flag(row_flag_type, row_country, row_flag)
    out = []
    for l in letters:
        is_no_letter = not l.get('letter')  # 空字母 = 无字母款
        info = {**l, 'bound_sku': sku_prefix,
                'flag': row_flag,           # 原始拼接(给用户看排期写的什么)
                'inventory_flag': inventory_flag,  # 翻译后实际查库存用的
                'no_letter': is_no_letter,
                }
        info.pop('_name_cn', None)  # 内部传递字段,不外泄

        if is_no_letter:
            # 优先用 parser 填入的 name(15789 这种新数据);旧合成行走 _name_cn 清洗
            name = l.get('name') or _clean_name_cn(l.get('_name_cn') or '')
            binding = None
        else:
            binding = db.get_letter_binding(sku_prefix, l['letter']) if sku_prefix else None
            name = binding['material_name'] if binding else ''

        if name:
            info['bound_name'] = name
            # 总库存(sku+name)
            total_stock = db.get_stock_by_sku_name(sku_prefix, name)
            shortage = max(0, l['qty'] - total_stock)
            info['letter_stock'] = total_stock
            info['letter_shortage'] = shortage
            info['letter_sufficient'] = shortage == 0
            # 按 style+flag 精确查 normal/rare 库存
            n_qty = l.get('normal_qty')
            r_qty = l.get('rare_qty')
            if n_qty is not None:
                n_stock = db.get_stock_by_full_dim(sku_prefix, name, 'normal', inventory_flag) if inventory_flag else None
                info['normal_stock'] = n_stock
                if n_stock is not None:
                    info['normal_shortage'] = max(0, n_qty - n_stock)
                    info['normal_sufficient'] = n_stock >= n_qty
            if r_qty is not None:
                r_stock = db.get_stock_by_full_dim(sku_prefix, name, 'rare', inventory_flag) if inventory_flag else None
                info['rare_stock'] = r_stock
                if r_stock is not None:
                    info['rare_shortage'] = max(0, r_qty - r_stock)
                    info['rare_sufficient'] = r_stock >= r_qty
        else:
            info['bound_name'] = ''
            info['letter_stock'] = None
            info['letter_shortage'] = None
            info['letter_sufficient'] = None
        out.append(info)
    return out


def _build_po_view(po, rows):
    """PO 视角:每条排期行 = 一个 ITEM# + 客户 SKU,字母分布从 letters_json 拿"""
    items = []
    for r in rows:
        letters = _parse_letters(r.get('letters_json'))
        if not letters and r.get('variant_letter'):
            letters = [{'letter': r['variant_letter'], 'qty': r['qty'],
                        'image_url': r.get('image_url') or ''}]
        elif not letters and r.get('qty', 0) > 0:
            # 无字母款(如 15789):整 qty 当普通款,不拆配比,物料名取 name_cn 的中文部分
            letters = [{
                'letter': '', 'qty': r['qty'], 'image_url': r.get('image_url') or '',
                'normal_qty': r['qty'], 'rare_qty': 0, 'ratio_assumed': False,
                '_name_cn': r.get('name_cn') or '',
            }]
        row_flag = r.get('flag') or ''
        letters = _enrich_letters_with_binding(
            r['item_code'], letters,
            row_flag=row_flag,
            row_flag_type=r.get('flag_type') or '',
            row_country=r.get('country') or '',
        )
        sku_prefix = _match_sku_prefix(r['item_code'])
        stock = db.get_sku_total_stock(sku_prefix) if sku_prefix else 0
        # 行级"是否够做":有绑定的字母都 sufficient;无绑定的字母按 total 模式判断
        bound_letters = [l for l in letters if l['bound_name']]
        unbound_letters = [l for l in letters if not l['bound_name']]
        if letters and not unbound_letters:
            # 全部字母都绑定 → 看每个字母是否各自够
            sufficient = all(l['letter_sufficient'] for l in bound_letters)
            shortage = sum(l['letter_shortage'] for l in bound_letters)
        else:
            # 有未绑定字母 → 沿用 SKU 总库存对比
            shortage = max(0, r['qty'] - stock)
            sufficient = shortage == 0
        items.append({
            'item_code': r['item_code'], 'series': r['series'],
            'customer_sku': r.get('customer_sku') or '',
            'customer': r['customer'] or '', 'country': r['country'] or '',
            'flag_type': r.get('flag_type') or '', 'flag': r.get('flag') or '',
            'ratio_normal_text': r.get('ratio_normal_text') or '',
            'ratio_rare_text': r.get('ratio_rare_text') or '',
            'name_cn': r['name_cn'] or '', 'plan_ship_date': r['plan_ship_date'] or '',
            'scheduled_qty': r['qty'],
            'letters': letters,
            'current_stock': stock,
            'shortage': shortage,
            'sufficient': sufficient,
            'has_unbound': bool(unbound_letters),
        })
    items.sort(key=lambda x: (x['item_code'], x['customer_sku']))
    return {
        'type': 'po', 'po_no': po, 'items': items,
        'all_sufficient': all(it['sufficient'] for it in items),
        'total_scheduled': sum(it['scheduled_qty'] for it in items),
        'total_stock': sum(it['current_stock'] for it in items),
        'total_shortage': sum(it['shortage'] for it in items),
    }


def _build_sku_view(query, rows, matched_by='item'):
    """
    货号视角:该货号涉及哪些 PO、各多少,库存够不够覆盖全部
    matched_by='item' 表示用户输的就是 ITEM#,直接用 query 当 SKU 去查库存;
    matched_by='customer_sku' 表示用户输的是客户 SKU(F 列),
        库存按 rows 里的 item_code 汇总
    """
    pos = []
    total_scheduled = 0
    for r in rows:
        letters = _parse_letters(r.get('letters_json'))
        if not letters and r.get('variant_letter'):
            letters = [{'letter': r['variant_letter'], 'qty': r['qty'],
                        'image_url': r.get('image_url') or ''}]
        elif not letters and r.get('qty', 0) > 0:
            # 无字母款(如 15789):整 qty 当普通款,不拆配比
            letters = [{
                'letter': '', 'qty': r['qty'], 'image_url': r.get('image_url') or '',
                'normal_qty': r['qty'], 'rare_qty': 0, 'ratio_assumed': False,
                '_name_cn': r.get('name_cn') or '',
            }]
        row_flag = r.get('flag') or ''
        letters = _enrich_letters_with_binding(
            r['item_code'], letters,
            row_flag=row_flag,
            row_flag_type=r.get('flag_type') or '',
            row_country=r.get('country') or '',
        )
        pos.append({
            'po_no': r['po_no'], 'series': r['series'],
            'item_code': r['item_code'], 'customer_sku': r.get('customer_sku') or '',
            'variant_letter': r.get('variant_letter') or '',
            'image_url': r.get('image_url') or '',
            'letters': letters,
            'customer': r['customer'] or '', 'country': r['country'] or '',
            'flag_type': r.get('flag_type') or '', 'flag': r.get('flag') or '',
            'ratio_normal_text': r.get('ratio_normal_text') or '',
            'ratio_rare_text': r.get('ratio_rare_text') or '',
            'name_cn': r['name_cn'] or '', 'plan_ship_date': r['plan_ship_date'] or '',
            'scheduled_qty': r['qty'],
        })
        total_scheduled += r['qty']
    pos.sort(key=lambda x: (x['plan_ship_date'] or '', x['po_no']))

    # 库存汇总 + 缺口汇总:都按 letter 行逐行累加,跟下面表格逐行判断口径一致。
    # 不能用 query 直接查库存表 —— query 是客户 SKU/带后缀 ITEM#,不在纯数字 sku 库存表里。
    # shortage 也必须逐行 sum:某品类不够 + 某品类过剩,不能互相抵消。
    display_code = query
    seen_keys = set()
    stock = 0
    shortage = 0
    all_letter_sufficient = True
    has_any_letter = False
    for p in pos:
        for l in p['letters']:
            if not l.get('bound_sku') or not l.get('bound_name'):
                all_letter_sufficient = False
                has_any_letter = True
                continue
            inv_flag = l.get('inventory_flag') or l.get('flag') or ''
            for style_key, stock_key, suf_key, short_key in (
                ('normal', 'normal_stock', 'normal_sufficient', 'normal_shortage'),
                ('rare', 'rare_stock', 'rare_sufficient', 'rare_shortage'),
            ):
                qty_field = 'normal_qty' if style_key == 'normal' else 'rare_qty'
                if not l.get(qty_field):
                    continue
                has_any_letter = True
                key = (l['bound_sku'], l['bound_name'], style_key, inv_flag)
                if key not in seen_keys:
                    seen_keys.add(key)
                    s = l.get(stock_key)
                    if s is not None:
                        stock += s
                shortage += l.get(short_key) or 0
                if not l.get(suf_key):
                    all_letter_sufficient = False

    sufficient = all_letter_sufficient if has_any_letter else (shortage == 0)
    return {
        'type': 'sku', 'item_code': display_code, 'matched_by': matched_by,
        'pos': pos,
        'current_stock': stock, 'total_scheduled': total_scheduled,
        'shortage': shortage, 'sufficient': sufficient,
    }


# ---- 字母绑定 ----

@app.route('/api/letter-bindings', methods=['GET'])
@login_required
def api_letter_bindings_list():
    """
    返回:
      bindings: 已配置的绑定列表 [{id, sku, letter, material_name, created_by, created_at}]
      materials_by_sku: {sku: [该 sku 在入库 / 出库出现过的物料名]} —— 新增弹窗里物料名下拉建议用
    """
    bindings = db.query_all_letter_bindings()
    skus = {b['sku'] for b in bindings}
    materials_by_sku = {sku: db.get_material_names_for_sku(sku) for sku in skus}
    return jsonify({'bindings': bindings, 'materials_by_sku': materials_by_sku})


@app.route('/api/letter-bindings', methods=['POST'])
@role_required('admin', 'operator')
def api_letter_bindings_create():
    """新增 / 更新字母绑定"""
    data = request.get_json() or {}
    sku = (data.get('sku') or '').strip()
    letter = (data.get('letter') or '').strip()
    name = (data.get('material_name') or '').strip()
    if not sku or not letter:
        return jsonify({'error': '缺少 sku 或 letter'}), 400
    if not name:
        return jsonify({'error': '请选择物料名称'}), 400
    bid = db.upsert_letter_binding(sku, letter, name, session.get('username'))
    return jsonify({'success': True, 'id': bid})


@app.route('/api/letter-bindings/<int:binding_id>', methods=['DELETE'])
@role_required('admin', 'operator')
def api_letter_bindings_delete(binding_id):
    affected = db.delete_letter_binding(binding_id)
    if affected == 0:
        return jsonify({'error': '绑定不存在'}), 404
    return jsonify({'success': True})


# ---- 货号别名 (子货号 → 父货号) ----

@app.route('/api/sku-aliases', methods=['GET'])
@login_required
def api_sku_aliases_list():
    return jsonify(db.query_all_sku_aliases())


@app.route('/api/sku-aliases', methods=['POST'])
@role_required('admin', 'operator')
def api_sku_aliases_create():
    """新增子货号→父货号映射。body: {alias_prefix, primary_prefix}"""
    data = request.get_json() or {}
    alias = (data.get('alias_prefix') or '').strip()
    primary = (data.get('primary_prefix') or '').strip()
    if not alias or not primary:
        return jsonify({'error': '缺少 alias_prefix 或 primary_prefix'}), 400
    if alias == primary:
        return jsonify({'error': '子货号不能跟父货号相同'}), 400
    # 防止链式别名(alias→primary→primary2)
    existing_map = db.get_sku_alias_map()
    if primary in existing_map:
        return jsonify({'error': f'父货号 {primary} 自己已经是子货号(映射到 {existing_map[primary]}),不能再当父货号'}), 400
    if alias in existing_map:
        return jsonify({'error': f'子货号 {alias} 已经映射到 {existing_map[alias]},先删除再加'}), 400
    try:
        aid = db.insert_sku_alias(alias, primary, session.get('username'))
    except Exception as e:
        return jsonify({'error': f'新增失败: {e}'}), 400
    _invalidate_alias_cache()
    return jsonify({'success': True, 'id': aid})


@app.route('/api/sku-aliases/<int:alias_id>', methods=['DELETE'])
@role_required('admin', 'operator')
def api_sku_aliases_delete(alias_id):
    affected = db.delete_sku_alias(alias_id)
    if affected == 0:
        return jsonify({'error': '别名不存在'}), 404
    _invalidate_alias_cache()
    return jsonify({'success': True})


# ---- 布标映射 ----

@app.route('/api/flag-mappings', methods=['GET'])
@login_required
def api_flag_mappings_list():
    """
    返回:
      mappings: 已配置的映射 [{id, flag_type, country, inventory_flag, ...}]
      unmapped: 排期里有但还没配映射的 (flag_type, country) 组合
      inventory_flags: 当前库存里所有可选的布标(来自 flags 表)
    """
    mappings = db.query_all_flag_mappings()
    sched_pairs = db.get_schedule_flag_country_pairs()
    mapped_keys = {(m['flag_type'], m['country']) for m in mappings}
    unmapped = [p for p in sched_pairs if (p['flag_type'], p['country']) not in mapped_keys]
    inventory_flags = db.query_all_flags()  # 复用现有"布标维护"里的
    return jsonify({'mappings': mappings, 'unmapped': unmapped,
                    'inventory_flags': inventory_flags})


@app.route('/api/flag-mappings', methods=['POST'])
@role_required('admin', 'operator')
def api_flag_mappings_create():
    data = request.get_json() or {}
    ft = (data.get('flag_type') or '').strip()
    cy = (data.get('country') or '').strip()
    inv = (data.get('inventory_flag') or '').strip()
    if not ft or not cy:
        return jsonify({'error': '缺少 flag_type 或 country'}), 400
    if not inv:
        return jsonify({'error': '请填写库存布标'}), 400
    mid = db.upsert_flag_mapping(ft, cy, inv, session.get('username'))
    return jsonify({'success': True, 'id': mid})


@app.route('/api/flag-mappings/<int:mapping_id>', methods=['DELETE'])
@role_required('admin', 'operator')
def api_flag_mappings_delete(mapping_id):
    affected = db.delete_flag_mapping(mapping_id)
    if affected == 0:
        return jsonify({'error': '映射不存在'}), 404
    return jsonify({'success': True})


@app.route('/api/schedule/query', methods=['GET'])
@login_required
def api_schedule_query():
    """
    搜索排期:q 依次按 PO号 → ITEM#(G) → 客户 SKU(F)三个字段尝试匹配,自动识别
    """
    q = (request.args.get('q') or request.args.get('po') or '').strip()
    if not q:
        return jsonify({'error': '请输入 PO 号或货号'}), 400
    po_rows = db.query_po_schedules_by_po(q)
    if po_rows:
        return jsonify(_build_po_view(q, po_rows))
    item_rows = db.query_po_schedules_by_item(q)
    if item_rows:
        return jsonify(_build_sku_view(q, item_rows, matched_by='item'))
    csku_rows = db.query_po_schedules_by_customer_sku(q)
    if csku_rows:
        return jsonify(_build_sku_view(q, csku_rows, matched_by='customer_sku'))
    return jsonify({'type': 'none', 'query': q, 'not_found': True})


# ---- 货号细表(xlsx,双 sheet,矩阵布局)----

_THIN = Side(border_style='thin', color='999999')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _detect_costume(records, fallback):
    rs = records or fallback
    return bool(rs) and rs[0]['category'] == 'costume'


def _build_sku_sheet(wb, sheet_name, records, sku, kind, is_costume):
    """
    构建一个 sheet:
      第 1 行:货号(合并居中)
      第 2 行:空
      第 3 行:左侧标签列空 + 各变体列的"物料名称 + 款式"
      第 4 行:左侧标签(入库日期/单号 或 出库日期/单号/PO号/领货人) + 各变体列的布标(戏服空)
      第 5 行起:每条记录一行,左侧字段 + 在匹配变体列填 qty
    """
    ws = wb.create_sheet(sheet_name)
    left_labels = ['入库日期', '单号'] if kind == 'in' else ['出库日期', '单号', 'PO号', '领货人']
    left_count = len(left_labels)

    # 变体列:毛绒 = (style, flag) 去重排序,戏服 = style 去重排序
    if is_costume:
        styles = sorted({r['style'] for r in records})
        variants = [(s, '') for s in styles]
    else:
        variants = sorted({(r['style'], r['flag']) for r in records})
    total_cols = left_count + len(variants)
    if total_cols == 0:
        total_cols = left_count or 1

    # 物料名称(取记录中第一个非空)
    name = next((r.get('name') for r in records if r.get('name')), '') or ''

    # 第 1 行:货号
    c1 = ws.cell(row=1, column=1, value=sku)
    if total_cols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c1.alignment = Alignment(horizontal='center', vertical='center')
    c1.font = Font(bold=True, size=14)

    # 第 3 行:变体列的 "物料名称 + 款式"
    for i, (style, _flag) in enumerate(variants):
        col = left_count + 1 + i
        head = f'{name} {_style_label(style)}'.strip() if name else _style_label(style)
        cell = ws.cell(row=3, column=col, value=head)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(bold=True)
        cell.border = _BORDER

    # 第 4 行:左侧标签 + 布标
    for i, lab in enumerate(left_labels):
        cell = ws.cell(row=4, column=i + 1, value=lab)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)
        cell.border = _BORDER
    if not is_costume:
        for i, (_style, flag) in enumerate(variants):
            col = left_count + 1 + i
            cell = ws.cell(row=4, column=col, value=flag)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)
            cell.border = _BORDER

    # 数据行
    variant_index = {v: i for i, v in enumerate(variants)}
    sorted_records = sorted(records, key=lambda r: (r['date'], r['id']))
    for row_i, r in enumerate(sorted_records):
        excel_row = 5 + row_i
        ws.cell(row=excel_row, column=1, value=r['date']).border = _BORDER
        ws.cell(row=excel_row, column=2, value=r['bill_no']).border = _BORDER
        if kind == 'out':
            ws.cell(row=excel_row, column=3, value=r.get('po', '') or '').border = _BORDER
            ws.cell(row=excel_row, column=4, value=r.get('picker', '') or '').border = _BORDER
        key = (r['style'], '') if is_costume else (r['style'], r['flag'])
        if key in variant_index:
            col = left_count + 1 + variant_index[key]
            cell = ws.cell(row=excel_row, column=col, value=r['qty'])
            cell.alignment = Alignment(horizontal='center')
            cell.border = _BORDER
        # 给空变体列也加边框,保持矩阵感
        for i, v in enumerate(variants):
            if v != key:
                ws.cell(row=excel_row, column=left_count + 1 + i).border = _BORDER

    # 列宽:左侧标签 12,变体列按表头长度 max 14
    for i in range(1, left_count + 1):
        ws.column_dimensions[get_column_letter(i)].width = 12
    for i in range(len(variants)):
        col_letter = get_column_letter(left_count + 1 + i)
        ws.column_dimensions[col_letter].width = 16


@app.route('/api/export/sku/<path:sku>', methods=['GET'])
@login_required
def api_export_sku_detail(sku):
    """
    导出单个货号的出入库细表(xlsx 双 sheet 矩阵布局)
    - 一个货号一个文件,入库 / 出库 各一个 sheet
    - 列 = (款式, 布标) 唯一组合;戏服没有布标,列 = 款式;戏服 sheet 跳过布标表头
    """
    in_records = [r for r in db.query_all_in_records() if r['sku'] == sku]
    out_records = [r for r in db.query_all_out_records() if r['sku'] == sku]

    if not in_records and not out_records:
        return jsonify({'error': f'货号 "{sku}" 没有任何出入库记录'}), 404

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _build_sku_sheet(wb, '入库', in_records, sku,
                     kind='in', is_costume=_detect_costume(in_records, out_records))
    _build_sku_sheet(wb, '出库', out_records, sku,
                     kind='out', is_costume=_detect_costume(out_records, in_records))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'{sku}_出入库明细_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{quote(filename)}"}
    )


# ==================== 启动 ====================

if __name__ == '__main__':
    # 开发模式:Flask 自带服务器
    print('=' * 50)
    print('华登库存管理系统(毛绒 + 戏服)')
    print('=' * 50)
    print()
    print('访问地址:')
    print('  本机:        http://localhost:5002')
    print('  局域网其他人: http://<本机IP>:5002')
    print()
    print('默认账号(请尽快改密码):')
    print('  admin / 123456 (主管)')
    print('  cg / 123456    (仓管员)')
    print('  yk / 123456    (游客)')
    print()
    print('按 Ctrl+C 停止服务')
    print('=' * 50)

    app.run(host='0.0.0.0', port=5002, debug=True)
