# core/checker.py - 核对逻辑与 Excel 生成
import re
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


def norm(val):
    """统一转成字符串，去除空格，处理整数型PO号"""
    if val is None:
        return ''
    try:
        f = float(val)
        if str(f) == 'nan':
            return ''
        if f == int(f):
            return str(int(f)).strip()
    except (ValueError, TypeError, OverflowError):
        pass
    return str(val).strip()


def hno_prefix(code):
    """提取货号前缀：取开头连续的数字部分；若货号以字母开头则保留全值"""
    s = norm(code)
    if not s:
        return ''
    m = re.match(r'^(\d+)', s)
    return m.group(1) if m else s


SKIP_SHEET_NAMES = {'JAZ销毁库存明细', '车间库存&未打发票'}


def read_main_file(filepath):
    """
    返回 list of dict，每条记录包含：
    brand, contract, hno, customer, ship_date,
    contract_qty, actual_qty, amount
    """
    xl = pd.ExcelFile(filepath, engine='xlrd')
    result = []
    skip_idx = {0, 8}
    for idx, sname in enumerate(xl.sheet_names):
        if idx in skip_idx or sname in SKIP_SHEET_NAMES:
            continue
        df = pd.read_excel(xl, sheet_name=idx, header=None)
        ncols = df.shape[1]
        if ncols >= 28:
            hno_col, cqty_col, aqty_col, amt_col, date_col = 6, 7, 9, 12, 5
        else:
            hno_col, cqty_col, aqty_col, amt_col, date_col = 5, 6, 8, 11, 4

        for i, row in df.iterrows():
            if i == 0:
                continue
            contract = norm(row.iloc[2])
            hno = hno_prefix(row.iloc[hno_col])
            brand = norm(row.iloc[0])
            if not contract or not hno:
                continue
            raw_date = row.iloc[date_col]
            if hasattr(raw_date, 'strftime'):
                ship_date = raw_date.strftime('%Y-%m-%d')
            else:
                ship_date = norm(raw_date)

            result.append({
                'brand': brand,
                'contract': contract,
                'hno': hno,
                'customer': norm(row.iloc[3]),
                'ship_date': ship_date,
                'contract_qty': norm(row.iloc[cqty_col]),
                'actual_qty': norm(row.iloc[aqty_col]),
                'amount': norm(row.iloc[amt_col]),
            })
    return result


def read_261(filepath):
    xl = pd.ExcelFile(filepath, engine='xlrd')
    pairs = set()
    skip_idx = {0, 8}
    for idx, sname in enumerate(xl.sheet_names):
        if idx in skip_idx or sname in SKIP_SHEET_NAMES:
            continue
        df = pd.read_excel(xl, sheet_name=idx, header=None)
        hno_col = 6 if df.shape[1] >= 28 else 5
        for i, row in df.iterrows():
            if i == 0:
                continue
            c, h = norm(row.iloc[2]), hno_prefix(row.iloc[hno_col])
            if c and h:
                pairs.add((c, h))
    return pairs


def read_262ck(filepath):
    xl = pd.ExcelFile(filepath, engine='openpyxl')
    pairs = set()
    sheet_map = {0: (2, 4), 2: (1, 3), 4: (2, 4), 6: (1, 2), 8: (2, 4)}
    total_sheets = len(xl.sheet_names)
    for idx, (po_col, hno_col) in sheet_map.items():
        if idx >= total_sheets:
            continue
        df = pd.read_excel(xl, sheet_name=idx, header=None)
        for i, row in df.iterrows():
            if i == 0:
                continue
            c, h = norm(row.iloc[po_col]), hno_prefix(row.iloc[hno_col])
            if c and h and not c.startswith('PO.NO'):
                pairs.add((c, h))
    return pairs


def read_zu(filepath):
    xl = pd.ExcelFile(filepath, engine='openpyxl')
    pairs = set()
    for idx in range(len(xl.sheet_names)):
        df = pd.read_excel(xl, sheet_name=idx, header=None)
        for i, row in df.iterrows():
            if i <= 1:
                continue
            c, h = norm(row.iloc[1]), hno_prefix(row.iloc[2])
            if c and h and c not in ('合同', ''):
                pairs.add((c, h))
    return pairs


def read_qty(filepath):
    xl = pd.ExcelFile(filepath, engine='openpyxl')
    pairs = set()
    for idx in range(len(xl.sheet_names)):
        df = pd.read_excel(xl, sheet_name=idx, header=None)
        ncols = df.shape[1]
        has_title_row = df.iloc[0].apply(lambda v: norm(v) != '').sum() <= 2
        if has_title_row:
            for i, row in df.iterrows():
                if i <= 1:
                    continue
                if ncols <= 11:
                    h, c = hno_prefix(row.iloc[4]), norm(row.iloc[5])
                else:
                    c, h = norm(row.iloc[4]), hno_prefix(row.iloc[6])
                if c and h:
                    pairs.add((c, h))
        else:
            for i, row in df.iterrows():
                if i == 0:
                    continue
                c, h = norm(row.iloc[3]), hno_prefix(row.iloc[5])
                if c and h:
                    pairs.add((c, h))
    return pairs


def run_check(main_path, path_261, path_262ck, path_zu, path_qty):
    """
    执行核对，返回 dict：
    {
      'total': int,
      'matched_count': int,
      'anomaly_count': int,
      'records': [...],
      'stats': {...}
    }
    """
    raw = read_main_file(main_path)
    pairs_261   = read_261(path_261)
    pairs_262ck = read_262ck(path_262ck)
    pairs_zu    = read_zu(path_zu)
    pairs_qty   = read_qty(path_qty)

    # 去重，保持顺序
    seen, unique = set(), []
    for r in raw:
        key = (r['contract'], r['hno'])
        if key not in seen:
            seen.add(key)
            unique.append(r)

    records = []
    for r in unique:
        key = (r['contract'], r['hno'])
        in_261   = key in pairs_261
        in_262ck = key in pairs_262ck
        in_zu    = key in pairs_zu
        in_qty   = key in pairs_qty
        has_match  = in_261 or in_262ck or in_zu or in_qty
        is_anomaly = not has_match
        records.append({**r,
            'in_261': in_261, 'in_262ck': in_262ck,
            'in_zu': in_zu,   'in_qty': in_qty,
            'has_match': has_match, 'is_anomaly': is_anomaly,
        })

    total = len(records)
    matched_count  = sum(1 for r in records if r['has_match'])
    anomaly_count  = sum(1 for r in records if r['is_anomaly'])

    return {
        'total': total,
        'matched_count': matched_count,
        'anomaly_count': anomaly_count,
        'records': records,
        'stats': {
            '26-1 Finished Goods':  {'found': sum(1 for r in records if r['in_261']),   'total': total},
            '26-2 Shipment Detail': {'found': sum(1 for r in records if r['in_262ck']), 'total': total},
            'ZU Shipment':          {'found': sum(1 for r in records if r['in_zu']),    'total': total},
            '26-2 Quantity':        {'found': sum(1 for r in records if r['in_qty']),   'total': total},
        }
    }


RED_FILL    = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
GREEN_FILL  = PatternFill(start_color='FF92D050', end_color='FF92D050', fill_type='solid')
HEADER_FILL = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True)
PINK_FILL   = PatternFill(start_color='FFFFC7CE', end_color='FFFFC7CE', fill_type='solid')


def build_excel(result, output_path):
    """根据 run_check 返回的 result 生成 Excel 文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Check Result'

    headers = ['Brand/Customer', 'Contract', 'Item No.', '26-1 Finished Goods', '26-2 Shipment Detail', 'ZU Shipment', '26-2 Quantity', 'Status']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for ri, r in enumerate(result['records'], 2):
        ws.cell(ri, 1, r['brand'])
        ws.cell(ri, 2, r['contract'])
        ws.cell(ri, 3, r['hno'])
        for ci, key in enumerate(['in_261', 'in_262ck', 'in_zu', 'in_qty'], 4):
            found = r[key]
            cell = ws.cell(ri, ci, 'Yes' if found else 'No')
            cell.fill = GREEN_FILL if found else RED_FILL
            cell.font = Font(color='FFFFFF', bold=True) if not found else Font()
            cell.alignment = Alignment(horizontal='center')
        sc = ws.cell(ri, 8, 'Anomaly' if r['is_anomaly'] else '')
        if r['is_anomaly']:
            sc.fill = RED_FILL
            sc.font = Font(color='FFFFFF', bold=True)
        sc.alignment = Alignment(horizontal='center')

    # 异常记录工作表
    ws2 = wb.create_sheet('Anomaly Records')
    for ci, h in enumerate(['Brand/Customer', 'Contract', 'Item No.', 'Note'], 1):
        cell = ws2.cell(1, ci, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
    for ri, r in enumerate([x for x in result['records'] if x['is_anomaly']], 2):
        ws2.cell(ri, 1, r['brand'])
        ws2.cell(ri, 2, r['contract'])
        ws2.cell(ri, 3, r['hno'])
        ws2.cell(ri, 4, 'Not found in any comparison file, needs review')
        for ci in range(1, 5):
            ws2.cell(ri, ci).fill = PINK_FILL

    for i, w in enumerate([15, 30, 20, 18, 20, 18, 15, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    for i, w in enumerate([15, 30, 20, 25], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = 'A2'

    wb.save(output_path)
