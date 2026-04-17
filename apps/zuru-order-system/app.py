# -*- coding: utf-8 -*-
"""接单表入单系统 — Flask主程序
PO Excel → 解析 → 生成新Excel下载
"""
import os
import re
import json
import time
import uuid
import logging
import logging.handlers
import threading
from datetime import datetime, timedelta

from flask import (Flask, render_template, request, jsonify,
                   send_file, Response, stream_with_context)
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import po_parser


# LOCAL_MODE 控制：本地（桌面上 Win/Mac 直接跑）= 1，cloud 部署 = 0
# 影响 /api/browse 和 /api/check_path /api/config 的任意路径访问
LOCAL_MODE = os.environ.get('LOCAL_MODE', '0') == '1'

# Flask debug：仅当显式设置时启用，避免 RCE 风险
FLASK_DEBUG = os.environ.get('FLASK_DEBUG', '0') == '1'


def _safe_filename(name):
    """清理文件名中的不安全字符，保留中文 (secure_filename 会去掉中文)"""
    if not name:
        return f'unnamed_{uuid.uuid4().hex[:8]}.xlsx'
    # 替换路径分隔符、控制字符、特殊空白
    name = name.replace('\\', '_').replace('/', '_').replace('\x00', '')
    name = name.replace('\xa0', ' ').replace('\r', '').replace('\n', '')
    # 去掉所有控制字符
    name = ''.join(c for c in name if ord(c) >= 32)
    # 去掉前导 . 防止 .. 类路径穿越（os.path.join 不会解 .. 但保险）
    name = name.lstrip('.').strip()
    if not name or name in ('.', '..'):
        return f'unnamed_{uuid.uuid4().hex[:8]}.xlsx'
    # 加 uuid 前缀避免文件名碰撞 + 防止任何残留路径技巧
    base = name[-100:]  # 限制长度
    return f'{uuid.uuid4().hex[:8]}_{base}'


APP_DIR = os.path.dirname(os.path.abspath(__file__))
# DATA_PATH: Docker部署时挂载卷，本地运行默认APP_DIR
DATA_PATH = os.environ.get('DATA_PATH', APP_DIR)

app = Flask(__name__, template_folder=os.path.join(APP_DIR, 'templates'),
            static_folder=os.path.join(APP_DIR, 'static'))
# 运行在nginx反向代理后，修正客户端IP/协议/路径前缀等header
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)

app.config['UPLOAD_FOLDER'] = os.path.join(DATA_PATH, 'uploads')
app.config['EXPORT_FOLDER'] = os.path.join(DATA_PATH, 'exports')
app.config['DATA_FOLDER'] = os.path.join(DATA_PATH, 'data')
# 实际 PO 文件 ~MB 级别，限到 32MB 防止 OOM (mem_limit 512m)
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['EXPORT_FOLDER'], exist_ok=True)
os.makedirs(app.config['DATA_FOLDER'], exist_ok=True)

# 日志：rotating 防止无限增长
LOG_FILE = os.path.join(app.config['DATA_FOLDER'], 'ops.log')
_log_handler = logging.handlers.RotatingFileHandler(
    LOG_FILE, maxBytes=10 * 1024 * 1024, backupCount=3, encoding='utf-8')
_log_handler.setFormatter(logging.Formatter('%(asctime)s %(message)s'))
logging.getLogger().setLevel(logging.INFO)
logging.getLogger().addHandler(_log_handler)


# === 加载映射数据 ===
# 映射数据（country_map/item_map）从 data/ 读取，Docker环境下可通过卷挂载覆盖
def _load_json(name):
    p = os.path.join(APP_DIR, 'data', name)
    if os.path.exists(p):
        with open(p, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}


COUNTRY_MAP = _load_json('country_map.json')
ITEM_MAP = _load_json('item_map.json')

# 任务管理：每个 process 请求一个 task_id，避免并发互相覆盖
_tasks = {}
_tasks_lock = threading.Lock()
_TASK_TTL_SECS = 600  # 10 分钟后清理已完成任务


def _new_task():
    task_id = uuid.uuid4().hex[:12]
    task = {
        'id': task_id,
        'created_at': time.time(),
        'current': 0, 'total': 0, 'msg': '排队中...',
        'done': False, 'result': None,
    }
    with _tasks_lock:
        # 清理过期任务
        now = time.time()
        stale = [tid for tid, t in _tasks.items()
                 if t['done'] and (now - t['created_at']) > _TASK_TTL_SECS]
        for tid in stale:
            _tasks.pop(tid, None)
        _tasks[task_id] = task
    return task


def _get_task(task_id):
    with _tasks_lock:
        return _tasks.get(task_id)


# 接单表路径配置（运行时产物，存在 DATA_FOLDER 下，Docker环境下卷挂载）
_CONFIG_FILE = os.path.join(app.config['DATA_FOLDER'], 'config.json')


def _load_config():
    if os.path.exists(_CONFIG_FILE):
        with open(_CONFIG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}


def _save_config(cfg):
    with open(_CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


def _read_existing_pos(filepath):
    """读取接单表B列所有PO号，返回set"""
    pos = set()
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        ws = wb.active
        for r in range(3, ws.max_row + 1):
            v = ws.cell(r, 2).value
            if v:
                s = str(v).strip().replace('.0', '')
                if s:
                    pos.add(s)
        wb.close()
    except Exception as e:
        logging.error(f'读取接单表失败: {filepath} - {e}')
    return pos


def _list_dir(path):
    """列出目录内容，返回文件夹和xlsx文件"""
    items = []
    try:
        for name in sorted(os.listdir(path)):
            full = os.path.join(path, name)
            if os.path.isdir(full):
                items.append({'name': name, 'type': 'dir'})
            elif name.lower().endswith(('.xlsx', '.xls')):
                items.append({'name': name, 'type': 'file'})
    except PermissionError:
        pass
    return items


def _translate_country(en_name):
    """英文国家名 → 中文（含洲）"""
    if not en_name:
        return ''
    cn = COUNTRY_MAP.get(en_name.strip())
    if cn:
        return cn
    for k, v in COUNTRY_MAP.items():
        if k.lower() == en_name.strip().lower():
            return v
    return en_name.strip()


def _date_serial(dt):
    """datetime → Excel序列号"""
    if not dt or not hasattr(dt, 'year'):
        return None
    return (dt - datetime(1899, 12, 30)).days


def _generate_excel(parsed_orders, custom_date=None):
    """用openpyxl生成接单表格式的新Excel

    Returns: {'ok': bool, 'filepath': str, 'filename': str, 'written': int,
              'warnings': [], 'details': [], 'revisions': []}
    """
    warnings = []
    details = []
    written = 0

    wb = openpyxl.Workbook()

    # === 排期sheet ===
    ws = wb.active
    ws.title = '排期'

    # 标题行（第1行）
    ws.merge_cells('A1:N1')
    title_cell = ws['A1']
    title_cell.value = '2026年ZURU接单金额表'
    title_cell.font = Font(name='宋体', size=18, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # 表头（第2行）
    headers = ['接单日期', '合同', '简货号', '货号', '走货期', '数量',
               '单价USD', '金额USD', '金额HK', '发票号', '出货国家',
               '生产车间', '排期品名', '备注']
    header_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    header_font = Font(name='宋体', size=11, bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 列宽
    widths = [10, 14, 8, 16, 10, 10, 10, 12, 12, 10, 14, 10, 14, 20]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    # 数据行从第3行开始
    row_idx = 3
    data_font = Font(name='宋体', size=11)
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    date_fmt = 'm"月"d"日"'

    for order in parsed_orders:
        po = order['po_number']
        po_date_serial = _date_serial(custom_date or order['po_date'])
        country_cn = _translate_country(order['destination'])
        lines = order['lines']

        if not lines:
            warnings.append(f'PO {po}: 无数据行，跳过')
            continue

        if country_cn == order['destination'] and order['destination']:
            warnings.append(f'PO {po}: 国家"{order["destination"]}"未找到中文翻译，原样写入')

        first_line = True
        skipped_items = []
        for ln in lines:
            simple_no = ln['simple_no']
            sku_full = ln['sku_spec'] or ln['sku'] or ''

            # 货号不在映射表：纯数字→跳过，M开头→照写只警告
            if simple_no and simple_no not in ITEM_MAP:
                if simple_no[0].upper() != 'M':
                    skipped_items.append(f'{simple_no}({sku_full})')
                    continue

            # 单价=0且货号不含PRODUCT → 跳过
            if not ln['price'] and 'PRODUCT' not in sku_full.upper():
                continue

            r = row_idx

            # A 接单日期（首行写）
            if first_line and po_date_serial:
                ws.cell(r, 1, po_date_serial).number_format = date_fmt

            # B 合同（首行写）
            if first_line and po:
                ws.cell(r, 2, po)

            # C 简货号
            if simple_no:
                try:
                    ws.cell(r, 3, int(simple_no))
                except (ValueError, TypeError):
                    ws.cell(r, 3, simple_no)

            # D 货号：去掉第一个-及之后的内容
            raw_item = ln['sku_spec'] or ln['sku'] or ''
            ws.cell(r, 4, raw_item.split('-')[0] if '-' in raw_item else raw_item)

            # E 走货期
            delivery_serial = _date_serial(ln['delivery'])
            if delivery_serial:
                ws.cell(r, 5, delivery_serial).number_format = date_fmt

            # F 数量
            if ln['qty']:
                ws.cell(r, 6, ln['qty'])

            # G 单价USD（美国打六折）
            if ln['price']:
                ws.cell(r, 7, round(ln['price'], 4))

            # H 金额USD = F*G
            ws.cell(r, 8).value = f'=F{r}*G{r}'

            # I 金额HK = H*7.75
            ws.cell(r, 9).value = f'=H{r}*7.75'

            # K 出货国家
            if country_cn:
                ws.cell(r, 11, country_cn)

            # 全列引用 (A:C / A:D)：避免硬编码行数限制货号表增长
            ws.cell(r, 12).value = f'=VLOOKUP(C{r},货号!A:C,3,0)'
            ws.cell(r, 13).value = f'=VLOOKUP(C{r},货号!A:D,4,0)'

            # 统一格式：绿色填充标记新数据
            for c in range(1, 15):
                cell = ws.cell(r, c)
                cell.font = data_font
                cell.border = thin_border
                cell.fill = green_fill

            row_idx += 1
            written += 1
            first_line = False

        if skipped_items:
            warnings.append(
                f'PO {po}: 以下货号不在映射表，已跳过 → {", ".join(skipped_items)}')
        details.append(f'PO {po}: {len(lines)}行')

    # 冻结表头
    ws.freeze_panes = 'A3'

    # === 货号sheet（VLOOKUP依赖）===
    ws2 = wb.create_sheet('货号')
    ws2.cell(1, 1, '货号').font = Font(bold=True)
    ws2.cell(1, 2, '中文名称').font = Font(bold=True)
    ws2.cell(1, 3, '生产车间').font = Font(bold=True)
    ws2.cell(1, 4, '排期名').font = Font(bold=True)

    for ri, (k, v) in enumerate(ITEM_MAP.items(), 2):
        try:
            ws2.cell(ri, 1, int(k))
        except (ValueError, TypeError):
            ws2.cell(ri, 1, k)
        ws2.cell(ri, 2, v.get('cn_name', ''))
        ws2.cell(ri, 3, v.get('workshop', ''))
        ws2.cell(ri, 4, v.get('schedule_name', ''))

    ws2.column_dimensions['A'].width = 10
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 10
    ws2.column_dimensions['D'].width = 16

    # 保存
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'接单表数据_{ts}.xlsx'
    filepath = os.path.join(app.config['EXPORT_FOLDER'], filename)
    wb.save(filepath)
    wb.close()

    return {
        'ok': True, 'filepath': filepath, 'filename': filename,
        'written': written, 'warnings': warnings, 'details': details,
    }


# === Flask 路由 ===

@app.route('/health')
def health():
    """Docker/nginx健康检查端点"""
    return jsonify({'status': 'ok'})


@app.route('/')
def index():
    return render_template('index.html', local_mode=LOCAL_MODE)


@app.route('/api/config', methods=['GET', 'POST'])
def config_api():
    """获取/保存配置（接单表路径）— 仅 LOCAL_MODE 下允许设置任意路径"""
    if request.method == 'GET':
        return jsonify(_load_config())
    if not LOCAL_MODE:
        return jsonify({'ok': False, 'error': '仅本机模式可设置接单表路径，请用上传'}), 403
    data = request.get_json(force=True)
    cfg = _load_config()
    if 'order_table_path' in data:
        new_path = str(data['order_table_path'] or '').strip()
        # 即使本机模式也仅允许已存在的文件
        if new_path and not os.path.isfile(new_path):
            return jsonify({'ok': False, 'error': '路径不存在'}), 400
        cfg['order_table_path'] = new_path
    _save_config(cfg)
    return jsonify({'ok': True})


@app.route('/api/browse', methods=['POST'])
def browse():
    """浏览服务器目录 — 仅本机模式开启"""
    if not LOCAL_MODE:
        return jsonify({'ok': False, 'error': '云端部署不支持浏览服务器目录，请用上传'}), 403
    data = request.get_json(force=True)
    path = data.get('path', '').strip()
    if not path:
        path = os.path.join(os.path.expanduser('~'), 'Desktop')
    if not os.path.isdir(path):
        return jsonify({'ok': False, 'error': '目录不存在'}), 400
    items = _list_dir(path)
    drives = []
    if os.name == 'nt':
        import string
        for d in string.ascii_uppercase:
            dp = f'{d}:\\'
            if os.path.isdir(dp):
                drives.append(dp)
    return jsonify({
        'ok': True,
        'current': os.path.abspath(path),
        'parent': os.path.dirname(os.path.abspath(path)),
        'items': items,
        'drives': drives,
    })


@app.route('/api/check_path', methods=['POST'])
def check_path():
    """检查接单表路径 — 仅本机模式"""
    if not LOCAL_MODE:
        return jsonify({'ok': False, 'error': '云端部署请用上传'}), 403
    data = request.get_json(force=True)
    path = data.get('path', '').strip()
    if not path or not os.path.isfile(path):
        return jsonify({'ok': False, 'error': '文件不存在'}), 400
    pos = _read_existing_pos(path)
    return jsonify({'ok': True, 'count': len(pos)})


@app.route('/api/upload_table', methods=['POST'])
def upload_table():
    """上传接单表文件（云端默认入口）"""
    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({'ok': False, 'error': '未选择文件'}), 400
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ('.xlsx', '.xls'):
        return jsonify({'ok': False, 'error': '仅支持Excel文件'}), 400
    save_path = os.path.join(app.config['DATA_FOLDER'], 'uploaded_order_table.xlsx')
    f.save(save_path)
    pos = _read_existing_pos(save_path)
    if not pos:
        return jsonify({'ok': False, 'error': '未能读取到PO数据，请确认是接单表文件'}), 400
    cfg = _load_config()
    cfg['order_table_path'] = save_path
    _save_config(cfg)
    return jsonify({'ok': True, 'count': len(pos), 'filename': f.filename})


@app.route('/api/classify', methods=['POST'])
def classify():
    """根据接单表比对，分类上传的PO文件为新单/修改单"""
    cfg = _load_config()
    table_path = cfg.get('order_table_path', '')
    if not table_path or not os.path.isfile(table_path):
        return jsonify({'error': '未设置接单表路径或文件不存在'}), 400

    existing_pos = _read_existing_pos(table_path)

    saved = []
    for f in request.files.getlist('files'):
        if not f.filename:
            continue
        ext = os.path.splitext(f.filename)[1].lower()
        if ext not in ('.xlsx', '.xls'):
            continue
        safe_name = _safe_filename(f.filename)
        path = os.path.join(app.config['UPLOAD_FOLDER'], safe_name)
        f.save(path)
        saved.append((f.filename, path))

    results = []
    for fname, path in saved:
        try:
            data = po_parser.parse(path)
            po = data['po_number']
            is_rev = po in existing_pos
            missing = []
            for ln in data['lines']:
                sn = ln['simple_no']
                if sn and sn not in ITEM_MAP:
                    sku = ln['sku_spec'] or ln['sku'] or ''
                    missing.append(f'{sn}({sku})')
            results.append({
                'filename': fname, 'po': po,
                'is_revision': is_rev, 'missing_items': missing
            })
        except Exception:
            results.append({
                'filename': fname, 'po': '',
                'is_revision': po_parser.is_revision(fname),
                'missing_items': []
            })
    po_files = {}
    for r in results:
        if r['po']:
            po_files.setdefault(r['po'], []).append(r['filename'])
    duplicate_pos = {po: fnames for po, fnames in po_files.items() if len(fnames) > 1}

    for _, path in saved:
        try:
            os.remove(path)
        except OSError:
            pass

    return jsonify({'ok': True, 'results': results, 'existing_count': len(existing_pos),
                    'duplicate_pos': duplicate_pos})


@app.route('/api/process', methods=['POST'])
def process():
    """上传PO文件 → 解析 → 生成Excel（每请求一个 task_id）"""
    saved = []
    for f in request.files.getlist('files'):
        if not f.filename:
            continue
        ext = os.path.splitext(f.filename)[1].lower()
        if ext not in ('.xlsx', '.xls'):
            continue
        safe_name = _safe_filename(f.filename)
        path = os.path.join(app.config['UPLOAD_FOLDER'], safe_name)
        f.save(path)
        saved.append((f.filename, path))

    if not saved:
        return jsonify({'error': '没有有效的Excel文件'}), 400

    cfg = _load_config()
    table_path = cfg.get('order_table_path', '')
    existing_pos = set()
    if table_path and os.path.isfile(table_path):
        existing_pos = _read_existing_pos(table_path)

    task = _new_task()
    task['total'] = len(saved)
    task['msg'] = '开始解析...'

    new_orders = []
    rev_orders = []
    parse_errors = []

    for i, (fname, path) in enumerate(saved):
        with _tasks_lock:
            task['current'] = i
            task['msg'] = f'解析 {fname}...'

        try:
            data = po_parser.parse(path)
            data['filename'] = fname
            po = data['po_number']

            if existing_pos:
                is_rev = po in existing_pos
            else:
                is_rev = po_parser.is_revision(fname)

            if is_rev:
                rev_orders.append(fname)
            else:
                new_orders.append(data)
        except Exception as e:
            parse_errors.append(f'{fname}: {str(e)[:100]}')

    with _tasks_lock:
        task['msg'] = '正在生成Excel...'
        task['current'] = task['total']

    result = {'ok': True, 'written': 0, 'warnings': [], 'details': [],
              'filename': '', 'revisions': rev_orders, 'parse_errors': parse_errors,
              'task_id': task['id']}

    custom_date_str = request.form.get('order_date', '')
    custom_date = None
    if custom_date_str:
        try:
            custom_date = datetime.strptime(custom_date_str, '%Y-%m-%d')
        except ValueError:
            pass

    if new_orders:
        gen = _generate_excel(new_orders, custom_date=custom_date)
        result['ok'] = gen['ok']
        result['written'] = gen['written']
        result['warnings'] = gen['warnings']
        result['details'] = gen['details']
        result['filename'] = gen.get('filename', '')

    with _tasks_lock:
        task['done'] = True
        task['result'] = result
        task['msg'] = '完成'

    for _, path in saved:
        try:
            os.remove(path)
        except OSError:
            pass

    return jsonify(result)


@app.route('/api/progress')
def progress():
    """SSE进度推送 — 必须带 task_id，避免 thread leak"""
    task_id = request.args.get('task_id', '').strip()

    def gen():
        if not task_id:
            yield 'data: {"error":"missing task_id","done":true}\n\n'
            return
        last_msg = ''
        deadline = time.time() + 600  # 最长 10 分钟
        while time.time() < deadline:
            t = _get_task(task_id)
            if not t:
                yield 'data: {"error":"task not found","done":true}\n\n'
                return
            with _tasks_lock:
                snapshot = {
                    'current': t['current'],
                    'total': t['total'],
                    'msg': t['msg'],
                    'done': t['done'],
                }
            msg = json.dumps(snapshot, ensure_ascii=False)
            if msg != last_msg:
                yield f'data: {msg}\n\n'
                last_msg = msg
            if snapshot['done']:
                return
            time.sleep(0.3)
        yield 'data: {"error":"timeout","done":true}\n\n'

    return Response(stream_with_context(gen()), mimetype='text/event-stream')


@app.route('/api/download/<filename>')
def download(filename):
    """下载生成的Excel"""
    # 拒绝任何包含路径分隔符或 .. 的文件名
    if '/' in filename or '\\' in filename or '..' in filename:
        return jsonify({'error': '非法文件名'}), 403
    filepath = os.path.join(app.config['EXPORT_FOLDER'], filename)
    abs_export = os.path.abspath(app.config['EXPORT_FOLDER'])
    if not os.path.abspath(filepath).startswith(abs_export + os.sep):
        return jsonify({'error': '非法路径'}), 403
    if not os.path.exists(filepath):
        return jsonify({'error': '文件不存在'}), 404
    return send_file(filepath, as_attachment=True, download_name=filename)


if __name__ == '__main__':
    port = 5005
    print('=' * 50)
    print('  ZURU 接单表入单系统')
    print(f'  http://localhost:{port}')
    print('=' * 50)
    app.run(host='0.0.0.0', port=port, debug=FLASK_DEBUG, use_reloader=False, threaded=True)
