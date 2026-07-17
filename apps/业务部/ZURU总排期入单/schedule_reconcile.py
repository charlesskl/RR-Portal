# -*- coding: utf-8 -*-
"""总排期与分排期对账。

把总排期和所有分排期扫描成同一套订单行记录，再做多重集合对比。
核心目标是核对“分排期所有有效订单行是否都汇总到了总排期”。
"""
import os
import re
import warnings
import json
from collections import Counter, defaultdict, deque
from datetime import datetime, timedelta


SCHEDULE_DIR = os.path.join(os.path.dirname(__file__), 'schedules')
EXPORT_PREFIX = '总分排期核对'
CACHE_FILE = os.path.join(os.path.dirname(__file__), 'data', 'reconcile_sub_cache.json')

SKIP_SHEET_KW = ('MA', '取消', '旧', '总排期', '汇总', 'Sheet')
SCAN_MAX_COL = 32
SCAN_HEADER_ROWS = 8


def _cell_text(value):
    if value is None:
        return ''
    if hasattr(value, 'strftime'):
        return value.strftime('%Y-%m-%d')
    s = str(value).strip()
    if s.endswith('.0'):
        s = s[:-2]
    return re.sub(r'\s+', ' ', s)


def _normalize_po(value):
    s = _cell_text(value).replace(' ', '').upper()
    if s.startswith('#'):
        s = s[1:]
    if s.endswith('.0'):
        s = s[:-2]
    return s


def _normalize_code(value):
    s = _cell_text(value).upper()
    return re.sub(r'[\s\r\n]+', '', s)


def _normalize_cpo(value):
    return _cell_text(value).upper()


def _is_valid_item(value):
    s = _normalize_code(value)
    if not s or len(s) < 3 or len(s) > 60:
        return False
    skip_kw = ('TOTAL', 'SUBTOTAL', '合计', '小计', 'ITEM', '货号',
               'N/A', 'TBD', 'NONE', '---', '合並', '合并')
    if any(kw in s for kw in skip_kw):
        return False
    return bool(re.search(r'\d', s))


def _item_base(value):
    s = _normalize_code(value)
    m = re.match(r'^(.+?)(-S\d+.*)$', s, re.I)
    return m.group(1).upper() if m else s.upper()


def _qty_number(value):
    if value is None or value == '':
        return None
    if isinstance(value, (int, float)):
        try:
            return float(value)
        except Exception:
            return None
    s = str(value).strip().replace(',', '')
    m = re.search(r'-?\d+(?:\.\d+)?', s)
    if not m:
        return None
    try:
        return float(m.group(0))
    except Exception:
        return None


def _qty_key(value):
    q = _qty_number(value)
    if q is None:
        return ''
    if abs(q - round(q)) < 0.0001:
        return str(int(round(q)))
    return f'{q:.4f}'.rstrip('0').rstrip('.')


def _date_key(value):
    if not value:
        return ''
    if hasattr(value, 'strftime'):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, (int, float)):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=int(value))).strftime('%Y-%m-%d')
        except Exception:
            return _cell_text(value)
    s = _cell_text(value)
    if not s:
        return ''
    s2 = s.replace('/', '-').replace('.', '-')
    m = re.search(r'(\d{4})-(\d{1,2})-(\d{1,2})', s2)
    if m:
        return f'{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}'
    m = re.search(r'(\d{1,2})-(\d{1,2})-(\d{4})', s2)
    if m:
        a, b, y = int(m.group(1)), int(m.group(2)), m.group(3)
        # ZURU PO 常见为 MM-DD-YYYY；如果第一段大于12，则按 DD-MM-YYYY。
        if a > 12:
            return f'{y}-{b:02d}-{a:02d}'
        return f'{y}-{a:02d}-{b:02d}'
    return s


def _should_skip_sheet(name):
    n = str(name or '').strip()
    n_upper = n.upper()
    if re.match(r'^MA(?:$|[\s_-]|排期)', n_upper):
        return True
    if re.match(r'^Sheet\d*$', n, re.I):
        return True
    return any(kw in n for kw in SKIP_SHEET_KW if kw != 'MA')


def _detect_item_col(ws):
    for row in ws.iter_rows(min_row=1, max_row=SCAN_HEADER_ROWS,
                            max_col=SCAN_MAX_COL, values_only=False):
        for cell in row:
            v = _cell_text(cell.value).upper().replace(' ', '').replace('\n', '')
            if v in ('ITEM#', 'ITEM＃', '货号', '货号#', '货号＃'):
                return cell.column, cell.row
            if '货号' in v and '系统' not in v:
                return cell.column, cell.row
            if 'ITEM' in v and ('#' in v or '＃' in v):
                return cell.column, cell.row
    return None, None


def _detect_columns(ws):
    """检测排期明细列。

    分排期与总排期的主表结构基本一致；ITEM# 列附近的相对位置最稳定。
    个别 2026 文件在左侧多一列时，相对位置仍成立。
    """
    item_col, header_row = _detect_item_col(ws)
    if not item_col:
        return None
    return {
        'header_row': header_row or 1,
        'po_date': max(item_col - 6, 1),
        'customer': max(item_col - 5, 1),
        'dest': max(item_col - 4, 1),
        'po': max(item_col - 3, 1),
        'cpo': max(item_col - 2, 1),
        'sku_line': max(item_col - 1, 1),
        'item': item_col,
        'cn_name': item_col + 1,
        'qty': item_col + 2,
        'ship_date': item_col + 6,
    }


def _record_keys(record):
    exact = (
        record['po'], record['item'], record['sku_line'], record['cpo'],
        record['qty_key'], record['ship_date_key']
    )
    if record['sku_line']:
        fuzzy = (record['po'], record['item'], record['sku_line'])
    elif record['cpo']:
        fuzzy = (record['po'], record['item'], record['cpo'])
    else:
        fuzzy = (record['po'], record['item'])
    return exact, fuzzy


def _build_record(source, filepath, sheet, row_idx, values, cols):
    po = _normalize_po(values.get('po'))
    item = _normalize_code(values.get('item'))
    qty_key = _qty_key(values.get('qty'))

    if not po or not _is_valid_item(item) or not qty_key:
        return None
    if 'PO' in po and 'NO' in po:
        return None

    rec = {
        'source': source,
        'file': os.path.basename(filepath),
        'path': filepath,
        'sheet': sheet,
        'row': row_idx,
        'po': po,
        'cpo': _normalize_cpo(values.get('cpo')),
        'sku_line': _normalize_code(values.get('sku_line')),
        'item': item,
        'item_base': _item_base(item),
        'cn_name': _cell_text(values.get('cn_name')),
        'qty': _qty_number(values.get('qty')) or 0,
        'qty_key': qty_key,
        'ship_date': _date_key(values.get('ship_date')),
        'ship_date_key': _date_key(values.get('ship_date')),
    }
    rec['exact_key'], rec['fuzzy_key'] = _record_keys(rec)
    rec['display_key'] = ' | '.join(x for x in [
        rec['po'], rec['item'], rec['sku_line'], rec['cpo'], rec['qty_key'], rec['ship_date_key']
    ] if x)
    return rec


def _scan_workbook_rows(filepath, source, master=False):
    import openpyxl
    records = []
    skipped = []

    with warnings.catch_warnings():
        warnings.simplefilter('ignore', UserWarning)
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=False, keep_links=False)
    try:
        sheet_names = wb.sheetnames
        if master:
            usable = []
            for name in sheet_names:
                if '总排期' in name and '旧' not in name and '取消' not in name and '汇总' not in name:
                    usable = [name]
                    break
            if not usable:
                usable = [sheet_names[0]]
        else:
            usable = [name for name in sheet_names if not _should_skip_sheet(name)]

        for sheet_name in usable:
            ws = wb[sheet_name]
            try:
                cols = _detect_columns(ws)
            except Exception as e:
                skipped.append({
                    'file': os.path.basename(filepath),
                    'sheet': sheet_name,
                    'reason': f'检测表头失败: {e}',
                })
                continue
            if not cols:
                skipped.append({
                    'file': os.path.basename(filepath),
                    'sheet': sheet_name,
                    'reason': '未检测到 ITEM# / 货号列',
                })
                continue
            start_row = max(int(cols['header_row']) + 1, 2)
            max_needed_col = min(max(cols.values()), SCAN_MAX_COL)
            try:
                for row in ws.iter_rows(min_row=start_row, max_col=max_needed_col, values_only=False):
                    row_idx = row[0].row if row and hasattr(row[0], 'row') else start_row

                    def val(key):
                        idx = cols[key] - 1
                        return row[idx].value if idx < len(row) else None

                    row_values = {
                        'po': val('po'),
                        'cpo': val('cpo'),
                        'sku_line': val('sku_line'),
                        'item': val('item'),
                        'cn_name': val('cn_name'),
                        'qty': val('qty'),
                        'ship_date': val('ship_date'),
                    }
                    formula_fields = [
                        key for key, value in row_values.items()
                        if isinstance(value, str) and value.startswith('=')
                    ]
                    po_value = row_values.get('po')
                    item_value = row_values.get('item')
                    has_order_identity = (
                        ('po' in formula_fields or 'item' in formula_fields)
                        or bool(_normalize_po(po_value) if 'po' not in formula_fields else '')
                        or bool(_normalize_code(item_value) if 'item' not in formula_fields else '')
                    )
                    if formula_fields and has_order_identity:
                        skipped.append({
                            'file': os.path.basename(filepath),
                            'sheet': sheet_name,
                            'reason': f'第{row_idx}行关键字段包含公式（{", ".join(formula_fields)}），无法严格核对',
                        })
                        continue
                    rec = _build_record(source, filepath, sheet_name, row_idx, row_values, cols)
                    if rec:
                        records.append(rec)
            except Exception as e:
                skipped.append({
                    'file': os.path.basename(filepath),
                    'sheet': sheet_name,
                    'reason': f'读取中断，已保留已扫描行: {e}',
                })
                continue
    finally:
        wb.close()
    return records, skipped


def _load_sub_map():
    import json
    p = os.path.join(os.path.dirname(__file__), 'data', 'sub_schedule_map.json')
    if not os.path.exists(p):
        return {}
    with open(p, 'r', encoding='utf-8') as f:
        return json.load(f)


def _locate_extra_group(record, sub_map):
    locs = sub_map.get(record['item_base'])
    if not locs:
        num_m = re.match(r'^(\d+)', record['item_base'])
        if num_m:
            locs = sub_map.get(num_m.group(1))
    if not locs:
        return ('未匹配分排期', ''), '货号未在分排期映射中'
    if len(locs) > 1:
        loc = locs[0]
        return (loc.get('file', ''), loc.get('sheet', '')), f'多归属，暂归到第1个（共{len(locs)}个）'
    loc = locs[0]
    return (loc.get('file', ''), loc.get('sheet', '')), ''


def _public_record(record, note=''):
    return {
        'file': record.get('file', ''),
        'sheet': record.get('sheet', ''),
        'row': record.get('row', ''),
        'po': record.get('po', ''),
        'cpo': record.get('cpo', ''),
        'sku_line': record.get('sku_line', ''),
        'item': record.get('item', ''),
        'qty': record.get('qty_key', ''),
        'ship_date': record.get('ship_date_key', ''),
        'key': record.get('display_key', ''),
        'note': note,
    }


def _reconcile_records(master_records, sub_records, sub_map):
    master_by_exact = defaultdict(deque)
    for rec in master_records:
        master_by_exact[rec['exact_key']].append(rec)

    group_stats = defaultdict(lambda: {
        'file': '', 'sheet': '', 'sub_count': 0, 'master_count': 0,
        'matched': 0, 'mismatch': 0, 'missing': 0, 'extra': 0,
    })

    unmatched_sub = []
    for sub in sub_records:
        gkey = (sub['file'], sub['sheet'])
        stat = group_stats[gkey]
        stat['file'], stat['sheet'] = gkey
        stat['sub_count'] += 1
        q = master_by_exact.get(sub['exact_key'])
        if q:
            q.popleft()
            stat['matched'] += 1
            stat['master_count'] += 1
        else:
            unmatched_sub.append(sub)

    extra_master = []
    for q in master_by_exact.values():
        extra_master.extend(list(q))

    missing_by_fuzzy = defaultdict(deque)
    for rec in unmatched_sub:
        missing_by_fuzzy[rec['fuzzy_key']].append(rec)
    extra_by_fuzzy = defaultdict(deque)
    for rec in extra_master:
        extra_by_fuzzy[rec['fuzzy_key']].append(rec)

    mismatches = []
    paired_sub_ids = set()
    paired_master_ids = set()
    for fkey in list(missing_by_fuzzy.keys()):
        left = missing_by_fuzzy[fkey]
        right = extra_by_fuzzy.get(fkey)
        while left and right:
            sub = left.popleft()
            master = right.popleft()
            paired_sub_ids.add(id(sub))
            paired_master_ids.add(id(master))
            gkey = (sub['file'], sub['sheet'])
            stat = group_stats[gkey]
            stat['mismatch'] += 1
            stat['master_count'] += 1
            diffs = []
            if sub['cpo'] != master['cpo']:
                diffs.append(f"客PO: 分排期={sub['cpo'] or '-'} / 总排期={master['cpo'] or '-'}")
            if sub['qty_key'] != master['qty_key']:
                diffs.append(f"数量: 分排期={sub['qty_key'] or '-'} / 总排期={master['qty_key'] or '-'}")
            if sub['ship_date_key'] != master['ship_date_key']:
                diffs.append(f"出货期: 分排期={sub['ship_date_key'] or '-'} / 总排期={master['ship_date_key'] or '-'}")
            mismatches.append({
                'sub': _public_record(sub),
                'master': _public_record(master),
                'diff': '；'.join(diffs) if diffs else '关键字段不一致',
            })

    missing = []
    for sub in unmatched_sub:
        if id(sub) in paired_sub_ids:
            continue
        stat = group_stats[(sub['file'], sub['sheet'])]
        stat['missing'] += 1
        missing.append(_public_record(sub))

    extra = []
    for master in extra_master:
        if id(master) in paired_master_ids:
            continue
        gkey, note = _locate_extra_group(master, sub_map)
        stat = group_stats[gkey]
        stat['file'], stat['sheet'] = gkey
        stat['extra'] += 1
        stat['master_count'] += 1
        extra.append(_public_record(master, note=note))

    summary_rows = []
    for stat in group_stats.values():
        stat = dict(stat)
        stat['diff'] = stat['master_count'] - stat['sub_count']
        if stat['missing'] or stat['extra'] or stat['mismatch'] or stat['diff']:
            stat['status'] = '异常'
        else:
            stat['status'] = '一致'
        summary_rows.append(stat)
    summary_rows.sort(key=lambda x: (
        0 if x['status'] == '异常' else 1,
        -abs(x['diff']), -x['missing'], -x['extra'], x['file'], x['sheet']
    ))

    return summary_rows, missing, extra, mismatches


def _list_schedule_files(schedule_dir):
    if not os.path.isdir(schedule_dir):
        raise FileNotFoundError(f'分排期目录不存在: {schedule_dir}')
    return [
        os.path.join(schedule_dir, name)
        for name in sorted(os.listdir(schedule_dir))
        if name.lower().endswith('.xlsx') and not name.startswith('~$')
    ]


def _file_signature(files):
    sig = []
    for fp in files:
        try:
            st = os.stat(fp)
            sig.append({
                'file': os.path.basename(fp),
                'size': st.st_size,
                'mtime': int(st.st_mtime),
            })
        except OSError:
            sig.append({'file': os.path.basename(fp), 'missing': True})
    return sig


def _restore_record_keys(records):
    for rec in records:
        rec['exact_key'] = tuple(rec.get('exact_key', ()))
        rec['fuzzy_key'] = tuple(rec.get('fuzzy_key', ()))
    return records


def _load_sub_cache(schedule_dir, files):
    try:
        if not os.path.exists(CACHE_FILE):
            return None
        with open(CACHE_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
        if data.get('schedule_dir') != os.path.abspath(schedule_dir):
            return None
        if data.get('signature') != _file_signature(files):
            return None
        return {
            'records': _restore_record_keys(data.get('records', [])),
            'skipped': data.get('skipped', []),
            'failed_files': data.get('failed_files', []),
        }
    except Exception:
        return None


def _save_sub_cache(schedule_dir, files, records, skipped, failed_files):
    try:
        os.makedirs(os.path.dirname(CACHE_FILE), exist_ok=True)
        payload = {
            'schedule_dir': os.path.abspath(schedule_dir),
            'signature': _file_signature(files),
            'records': records,
            'skipped': skipped,
            'failed_files': failed_files,
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        }
        tmp = CACHE_FILE + '.tmp'
        with open(tmp, 'w', encoding='utf-8') as f:
            json.dump(payload, f, ensure_ascii=False)
        os.replace(tmp, CACHE_FILE)
    except Exception:
        pass


def reconcile_schedules(master_path, schedule_dir=None, export_dir=None):
    if not os.path.exists(master_path):
        raise FileNotFoundError(f'总排期文件不存在: {master_path}')
    schedule_dir = schedule_dir or SCHEDULE_DIR
    files = _list_schedule_files(schedule_dir)
    if not files:
        raise FileNotFoundError(f'分排期目录没有 .xlsx 文件: {schedule_dir}')

    master_records, master_skipped = _scan_workbook_rows(master_path, 'master', master=True)

    cache = _load_sub_cache(schedule_dir, files)
    cache_used = bool(cache)
    if cache:
        sub_records = cache['records']
        skipped = cache['skipped']
        failed_files = cache['failed_files']
    else:
        sub_records = []
        skipped = []
        failed_files = []
        for fp in files:
            try:
                rows, sk = _scan_workbook_rows(fp, 'sub', master=False)
                sub_records.extend(rows)
                skipped.extend(sk)
            except Exception as e:
                failed_files.append({'file': os.path.basename(fp), 'error': str(e)})
        _save_sub_cache(schedule_dir, files, sub_records, skipped, failed_files)

    sub_map = _load_sub_map()
    summary_rows, missing, extra, mismatches = _reconcile_records(master_records, sub_records, sub_map)

    totals = {
        'master_rows': len(master_records),
        'sub_rows': len(sub_records),
        'matched': sum(x['matched'] for x in summary_rows),
        'mismatch': len(mismatches),
        'missing': len(missing),
        'extra': len(extra),
        'diff': len(master_records) - len(sub_records),
        'files_scanned': len(files),
        'failed_files': len(failed_files),
        'skipped_sheets': len(skipped) + len(master_skipped),
        'cache_used': cache_used,
    }
    totals['count_diff_groups'] = sum(1 for x in summary_rows if x.get('diff') != 0)
    totals['count_match'] = totals['diff'] == 0 and totals['count_diff_groups'] == 0
    totals['scan_complete'] = not failed_files and not skipped and not master_skipped
    totals['strict_match'] = (
        totals['missing'] == 0 and totals['extra'] == 0
        and totals['mismatch'] == 0 and totals['diff'] == 0
        and totals['scan_complete']
    )
    totals['ok'] = totals['strict_match']

    result = {
        'ok': True,
        'all_match': totals['strict_match'],
        'count_match': totals['count_match'],
        'strict_match': totals['strict_match'],
        'totals': totals,
        'summary': summary_rows,
        'missing': missing,
        'extra': extra,
        'mismatches': mismatches,
        'failed_files': failed_files,
        'skipped_sheets': skipped + master_skipped,
    }
    if export_dir:
        result['export_file'] = generate_reconcile_report(result, export_dir)
    return result


def _write_rows(ws, headers, rows):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(name='宋体', size=11, bold=True, color='FFFFFF')
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin
        ws.column_dimensions[c.column_letter].width = min(max(len(h) + 4, 12), 28)
    for ri, row in enumerate(rows, 2):
        for ci, key in enumerate(headers, 1):
            val = row.get(key, '')
            if isinstance(val, str) and val.startswith('='):
                val = "'" + val
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = thin
            c.alignment = Alignment(vertical='center', wrap_text=True)
            if len(str(val)) + 2 > ws.column_dimensions[c.column_letter].width:
                ws.column_dimensions[c.column_letter].width = min(len(str(val)) + 2, 45)
    ws.freeze_panes = 'A2'


def generate_reconcile_report(result, output_dir):
    import openpyxl
    os.makedirs(output_dir, exist_ok=True)
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = '核对总览'
    summary_rows = []
    for r in result.get('summary', []):
        summary_rows.append({
            '状态': r['status'], '分排期文件': r['file'], 'Sheet': r['sheet'],
            '分排期条数': r['sub_count'], '总排期条数': r['master_count'],
            '差异': r['diff'], '已匹配': r['matched'],
            '字段不一致': r['mismatch'], '分有总无': r['missing'], '总有分无': r['extra'],
        })
    _write_rows(ws, ['状态', '分排期文件', 'Sheet', '分排期条数', '总排期条数',
                     '差异', '已匹配', '字段不一致', '分有总无', '总有分无'], summary_rows)

    def flat_record(rec):
        return {
            '文件': rec.get('file', ''), 'Sheet': rec.get('sheet', ''), '行号': rec.get('row', ''),
            'PO号': rec.get('po', ''), '客PO': rec.get('cpo', ''),
            'SKU/Line': rec.get('sku_line', ''), '货号': rec.get('item', ''),
            '数量': rec.get('qty', ''), '出货期': rec.get('ship_date', ''),
            '说明': rec.get('note', ''),
        }

    ws = wb.create_sheet('分排期有_总排期无')
    _write_rows(ws, ['文件', 'Sheet', '行号', 'PO号', '客PO', 'SKU/Line',
                     '货号', '数量', '出货期', '说明'],
                [flat_record(x) for x in result.get('missing', [])])

    ws = wb.create_sheet('总排期有_分排期无')
    _write_rows(ws, ['文件', 'Sheet', '行号', 'PO号', '客PO', 'SKU/Line',
                     '货号', '数量', '出货期', '说明'],
                [flat_record(x) for x in result.get('extra', [])])

    ws = wb.create_sheet('字段不一致')
    mismatch_rows = []
    for item in result.get('mismatches', []):
        sub = item.get('sub', {})
        master = item.get('master', {})
        mismatch_rows.append({
            '差异': item.get('diff', ''),
            '分排期文件': sub.get('file', ''), '分排期Sheet': sub.get('sheet', ''),
            '分排期行号': sub.get('row', ''), '总排期行号': master.get('row', ''),
            'PO号': sub.get('po', ''), '货号': sub.get('item', ''),
            'SKU/Line': sub.get('sku_line', ''),
            '分排期客PO': sub.get('cpo', ''), '总排期客PO': master.get('cpo', ''),
            '分排期数量': sub.get('qty', ''), '总排期数量': master.get('qty', ''),
            '分排期出货期': sub.get('ship_date', ''), '总排期出货期': master.get('ship_date', ''),
        })
    _write_rows(ws, ['差异', '分排期文件', '分排期Sheet', '分排期行号', '总排期行号',
                     'PO号', '货号', 'SKU/Line', '分排期客PO', '总排期客PO',
                     '分排期数量', '总排期数量', '分排期出货期', '总排期出货期'], mismatch_rows)

    ws = wb.create_sheet('扫描说明')
    note_rows = [
        {'项目': '总排期有效行', '内容': result.get('totals', {}).get('master_rows', 0)},
        {'项目': '分排期有效行', '内容': result.get('totals', {}).get('sub_rows', 0)},
        {'项目': '条数是否完全对上', '内容': '是' if result.get('totals', {}).get('count_match') else '否'},
        {'项目': '条数异常分组数', '内容': result.get('totals', {}).get('count_diff_groups', 0)},
        {'项目': '明细是否完全对上', '内容': '是' if result.get('totals', {}).get('strict_match') else '否'},
        {'项目': '扫描分排期文件数', '内容': result.get('totals', {}).get('files_scanned', 0)},
        {'项目': '打不开的分排期文件数', '内容': result.get('totals', {}).get('failed_files', 0)},
        {'项目': '跳过sheet数', '内容': result.get('totals', {}).get('skipped_sheets', 0)},
    ]
    for f in result.get('failed_files', []):
        note_rows.append({'项目': f"打开失败: {f.get('file', '')}", '内容': f.get('error', '')})
    for s in result.get('skipped_sheets', [])[:200]:
        note_rows.append({'项目': f"跳过sheet: {s.get('file', '')}/{s.get('sheet', '')}", '内容': s.get('reason', '')})
    _write_rows(ws, ['项目', '内容'], note_rows)

    ts = datetime.now().strftime('%Y%m%d_%H%M%S_%f')
    fname = f'{EXPORT_PREFIX}_{ts}.xlsx'
    out_path = os.path.join(output_dir, fname)
    wb.save(out_path)
    wb.close()
    return fname
