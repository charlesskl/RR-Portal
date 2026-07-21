import io
import json
import os
import struct
import zipfile

import openpyxl
import pytest

import app as app_module
import master_schedule
import schedule_reconcile


def make_schedule(path, sheet_name='Schedule A', qty=100):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.cell(1, 4, 'PO')
    ws.cell(1, 5, 'CPO')
    ws.cell(1, 6, 'SKU Line')
    ws.cell(1, 7, 'ITEM#')
    ws.cell(1, 9, 'QTY')
    ws.cell(1, 13, 'Ship Date')
    ws.cell(2, 4, 'PO-1001')
    ws.cell(2, 5, 'CPO-1')
    ws.cell(2, 6, '1')
    ws.cell(2, 7, '77785-S001')
    ws.cell(2, 9, qty)
    ws.cell(2, 13, '2026-07-17')
    wb.save(path)
    wb.close()


def fake_xlsx_bytes():
    data = io.BytesIO()
    with zipfile.ZipFile(data, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr('not-an-excel-file.txt', 'hello')
    return data.getvalue()


def zip_bytes(entries):
    data = io.BytesIO()
    with zipfile.ZipFile(data, 'w', zipfile.ZIP_DEFLATED) as zf:
        for name, content in entries:
            zf.writestr(name, content)
    return data.getvalue()


@pytest.fixture
def isolated_app(tmp_path, monkeypatch):
    upload_dir = tmp_path / 'uploads'
    master_dir = upload_dir / 'master'
    export_dir = tmp_path / 'exports'
    reconcile_dir = upload_dir / 'reconcile_schedules'
    data_dir = tmp_path / 'data'
    for path in (master_dir, export_dir, reconcile_dir, data_dir):
        path.mkdir(parents=True, exist_ok=True)
    monkeypatch.setitem(app_module.app.config, 'UPLOAD_FOLDER', str(upload_dir))
    monkeypatch.setitem(app_module.app.config, 'MASTER_FOLDER', str(master_dir))
    monkeypatch.setitem(app_module.app.config, 'EXPORT_FOLDER', str(export_dir))
    monkeypatch.setitem(app_module.app.config, 'RECONCILE_FOLDER', str(reconcile_dir))
    monkeypatch.setattr(app_module, '_MASTER_STATE_FILE', str(data_dir / 'master_state.json'))
    monkeypatch.setattr(app_module, 'CACHE_FILE', str(data_dir / 'reconcile_cache.json'))
    app_module.app.config.update(TESTING=True)
    return app_module.app.test_client(), master_dir, reconcile_dir


def test_fake_xlsx_does_not_replace_existing_master(isolated_app):
    client, master_dir, _ = isolated_app
    current = master_dir / 'uploaded_master.xlsx'
    make_schedule(current, sheet_name='总排期')
    original = current.read_bytes()
    app_module._set_master_path(str(current))

    response = client.post(
        '/api/master-schedule-upload-file',
        data={'file': (io.BytesIO(fake_xlsx_bytes()), 'fake.xlsx')},
        content_type='multipart/form-data',
    )

    assert response.status_code == 400
    assert current.read_bytes() == original
    assert app_module._get_master_path() == str(current)


def test_valid_workbench_master_upload_succeeds(isolated_app, tmp_path):
    client, master_dir, _ = isolated_app
    source = tmp_path / '7.12.xlsx'
    make_schedule(source, sheet_name='总排期')

    response = client.post(
        '/api/master-schedule-upload-file',
        data={'file': (io.BytesIO(source.read_bytes()), source.name)},
        content_type='multipart/form-data',
    )

    assert response.status_code == 200
    saved = master_dir / 'uploaded_master.xlsx'
    assert saved.exists()
    workbook = openpyxl.load_workbook(saved, read_only=True)
    try:
        assert workbook.sheetnames == ['总排期']
    finally:
        workbook.close()
    assert not list(master_dir.glob('*.tmp.xlsx'))


def test_valid_legacy_master_upload_succeeds(isolated_app, tmp_path):
    client, master_dir, _ = isolated_app
    source = tmp_path / '7.12.xlsx'
    make_schedule(source, sheet_name='总排期')

    response = client.post(
        '/api/master-schedule-upload-master',
        data={'master_file': (io.BytesIO(source.read_bytes()), source.name)},
        content_type='multipart/form-data',
    )

    assert response.status_code == 200
    saved = master_dir / source.name
    assert saved.exists()
    assert not list(master_dir.glob('*.tmp.xlsx'))


def test_legacy_master_upload_rejects_fake_xlsx(isolated_app):
    client, master_dir, _ = isolated_app
    current = master_dir / 'master.xlsx'
    make_schedule(current, sheet_name='总排期')
    original = current.read_bytes()
    app_module._set_master_path(str(current))

    response = client.post(
        '/api/master-schedule-upload-master',
        data={'master_file': (io.BytesIO(fake_xlsx_bytes()), 'master.xlsx')},
        content_type='multipart/form-data',
    )

    assert response.status_code == 400
    assert current.read_bytes() == original
    assert app_module._get_master_path() == str(current)


def test_legacy_master_upload_rejects_xls(isolated_app):
    client, _, _ = isolated_app

    response = client.post(
        '/api/master-schedule-upload-master',
        data={'master_file': (io.BytesIO(b'legacy workbook'), 'master.xls')},
        content_type='multipart/form-data',
    )

    assert response.status_code == 400


def test_fake_xlsx_does_not_replace_existing_sub_schedules(isolated_app):
    client, _, reconcile_dir = isolated_app
    current = reconcile_dir / 'existing.xlsx'
    make_schedule(current)
    original = current.read_bytes()

    response = client.post(
        '/api/reconcile-schedule-files',
        data={'files': (io.BytesIO(fake_xlsx_bytes()), 'fake.xlsx')},
        content_type='multipart/form-data',
    )

    assert response.status_code == 400
    assert current.read_bytes() == original
    assert [path.name for path in reconcile_dir.iterdir()] == ['existing.xlsx']


def test_partially_invalid_sub_schedule_batch_preserves_existing_files(isolated_app, tmp_path):
    client, _, reconcile_dir = isolated_app
    current = reconcile_dir / 'existing.xlsx'
    make_schedule(current)
    valid = tmp_path / 'valid.xlsx'
    make_schedule(valid)

    response = client.post(
        '/api/reconcile-schedule-files',
        data={'files': [
            (io.BytesIO(valid.read_bytes()), 'valid.xlsx'),
            (io.BytesIO(fake_xlsx_bytes()), 'fake.xlsx'),
        ]},
        content_type='multipart/form-data',
    )

    assert response.status_code == 400
    assert [path.name for path in reconcile_dir.iterdir()] == ['existing.xlsx']


def test_zip_with_invalid_xlsx_preserves_existing_files(isolated_app, tmp_path):
    client, _, reconcile_dir = isolated_app
    current = reconcile_dir / 'existing.xlsx'
    make_schedule(current)
    valid = tmp_path / 'valid.xlsx'
    make_schedule(valid)
    archive = zip_bytes([
        ('valid.xlsx', valid.read_bytes()),
        ('fake.xlsx', fake_xlsx_bytes()),
    ])

    response = client.post(
        '/api/reconcile-schedule-files',
        data={'files': (io.BytesIO(archive), 'schedules.zip')},
        content_type='multipart/form-data',
    )

    assert response.status_code == 400
    assert [path.name for path in reconcile_dir.iterdir()] == ['existing.xlsx']


def test_outer_zip_entry_limit_is_reported_before_processing(isolated_app, monkeypatch):
    client, _, _ = isolated_app
    archive = zip_bytes([('one.xlsx', b'one'), ('two.xlsx', b'two')])
    monkeypatch.setattr(app_module, 'MAX_OUTER_ZIP_FILES', 1)

    response = client.post(
        '/api/reconcile-schedule-files',
        data={'files': (io.BytesIO(archive), 'too-many.zip')},
        content_type='multipart/form-data',
    )

    assert response.status_code == 400
    assert '超过 1 个' in ' '.join(response.get_json().get('errors', []))


def test_zip_preflight_handles_eocd_signature_inside_comment(tmp_path):
    path = tmp_path / 'commented.zip'
    with zipfile.ZipFile(path, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr('one.txt', 'one')
        zf.comment = b'comment-PK\x05\x06-marker'

    app_module._preflight_zip(str(path), 10, 1024 * 1024, '测试包')


def test_zip_preflight_rejects_forged_low_entry_count(tmp_path):
    path = tmp_path / 'forged-count.zip'
    with zipfile.ZipFile(path, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr('one.txt', 'one')
        zf.writestr('two.txt', 'two')
    data = bytearray(path.read_bytes())
    eocd = data.rfind(b'PK\x05\x06')
    struct.pack_into('<HH', data, eocd + 8, 1, 1)
    path.write_bytes(data)

    with pytest.raises(ValueError, match='条目数不一致'):
        app_module._preflight_zip(str(path), 10, 1024 * 1024, '测试包')


def test_xlsx_uncompressed_limit_is_enforced(tmp_path, monkeypatch):
    path = tmp_path / 'valid.xlsx'
    make_schedule(path)
    monkeypatch.setattr(app_module, 'MAX_XLSX_UNCOMPRESSED_BYTES', 1)

    with pytest.raises(ValueError, match='解压后'):
        app_module._validate_xlsx(str(path))


def test_reconcile_directory_swap_restores_previous_files_on_failure(isolated_app, tmp_path, monkeypatch):
    _, _, reconcile_dir = isolated_app
    old_file = reconcile_dir / 'old.xlsx'
    make_schedule(old_file)
    staging_dir = tmp_path / 'uploads' / 'stage'
    staging_dir.mkdir()
    make_schedule(staging_dir / 'new.xlsx')
    real_replace = os.replace

    def fail_activation(src, dst):
        if os.path.abspath(src) == os.path.abspath(staging_dir):
            raise OSError('simulated activation failure')
        return real_replace(src, dst)

    monkeypatch.setattr(app_module.os, 'replace', fail_activation)

    with pytest.raises(OSError, match='simulated activation failure'):
        app_module._activate_reconcile_staging(str(staging_dir))

    assert old_file.exists()
    assert not (reconcile_dir / 'new.xlsx').exists()


def test_strict_match_is_false_when_any_sub_schedule_fails(tmp_path, monkeypatch):
    master = tmp_path / 'master.xlsx'
    schedule_dir = tmp_path / 'schedules'
    schedule_dir.mkdir()
    make_schedule(master, sheet_name='总排期')
    make_schedule(schedule_dir / 'valid.xlsx')
    (schedule_dir / 'broken.xlsx').write_bytes(b'not a workbook')
    monkeypatch.setattr(schedule_reconcile, 'CACHE_FILE', str(tmp_path / 'cache.json'))

    result = schedule_reconcile.reconcile_schedules(str(master), str(schedule_dir))

    assert result['totals']['failed_files'] == 1
    assert result['strict_match'] is False
    assert result['all_match'] is False


def test_walmart_sheet_is_not_treated_as_ma_sheet():
    assert schedule_reconcile._should_skip_sheet('WALMART') is False
    assert schedule_reconcile._should_skip_sheet('K-MART') is False
    assert schedule_reconcile._should_skip_sheet('MA') is True


def test_formula_key_cell_is_reported_as_incomplete_scan(tmp_path):
    path = tmp_path / 'formula.xlsx'
    make_schedule(path)
    wb = openpyxl.load_workbook(path)
    wb.active.cell(2, 9, '=50+50')
    wb.save(path)
    wb.close()

    records, issues = schedule_reconcile._scan_workbook_rows(str(path), 'sub')

    assert records == []
    assert any('公式' in issue.get('reason', '') for issue in issues)


def test_formula_in_empty_template_row_does_not_mark_scan_incomplete(tmp_path):
    path = tmp_path / 'formula-template.xlsx'
    make_schedule(path)
    wb = openpyxl.load_workbook(path)
    wb.active.cell(3, 9, '=SUM(I4:I10)')
    wb.save(path)
    wb.close()

    records, issues = schedule_reconcile._scan_workbook_rows(str(path), 'sub')

    assert len(records) == 1
    assert issues == []


def test_dual_map_rejects_non_numeric_keys(tmp_path, monkeypatch):
    data_dir = tmp_path / 'data'
    data_dir.mkdir()
    (data_dir / 'dual_schedule_map.json').write_text(
        json.dumps({'_说明': 'metadata'}, ensure_ascii=False), encoding='utf-8'
    )
    monkeypatch.setattr(app_module, 'APP_DIR', str(tmp_path))
    monkeypatch.setattr(master_schedule, '_load_dual_map', lambda: {})
    app_module.app.config.update(TESTING=True)

    response = app_module.app.test_client().post('/api/dual-map', json={
        'items': {"');alert(1);//": {'targets': ['77711'], 'mode': 'append'}},
    })

    assert response.status_code == 400
    stored = json.loads((data_dir / 'dual_schedule_map.json').read_text(encoding='utf-8'))
    assert stored == {'_说明': 'metadata'}


def test_dual_map_rejects_non_json_form_without_clearing_config(tmp_path, monkeypatch):
    data_dir = tmp_path / 'data'
    data_dir.mkdir()
    original = {'77785': {'targets': ['77711'], 'mode': 'append'}}
    path = data_dir / 'dual_schedule_map.json'
    path.write_text(json.dumps(original), encoding='utf-8')
    monkeypatch.setattr(app_module, 'APP_DIR', str(tmp_path))
    app_module.app.config.update(TESTING=True)

    response = app_module.app.test_client().post(
        '/api/dual-map', data={'items': ''}, content_type='application/x-www-form-urlencoded'
    )

    assert response.status_code == 415
    assert json.loads(path.read_text(encoding='utf-8')) == original


def test_dual_map_reloads_after_another_worker_replaces_file(tmp_path, monkeypatch):
    dual_map_path = tmp_path / 'dual_schedule_map.json'
    dual_map_path.write_text(
        json.dumps({'77785': {'targets': ['77711'], 'mode': 'append'}}),
        encoding='utf-8',
    )
    monkeypatch.setattr(master_schedule, 'DUAL_MAP_FILE', str(dual_map_path), raising=False)
    monkeypatch.setattr(master_schedule, '_DUAL_MAP', {})
    monkeypatch.setattr(master_schedule, '_DUAL_MAP_FILE_STATE', None, raising=False)

    assert master_schedule._get_dual_map()['77785']['targets'] == ['77711']

    replacement = tmp_path / 'replacement.json'
    replacement.write_text(
        json.dumps({'77785': {'targets': ['77722'], 'mode': 'append'}}),
        encoding='utf-8',
    )
    os.replace(replacement, dual_map_path)

    assert master_schedule._get_dual_map()['77785']['targets'] == ['77722']


def test_report_cells_escape_excel_formulas():
    wb = openpyxl.Workbook()
    ws = wb.active

    schedule_reconcile._write_rows(ws, ['值'], [{'值': '=HYPERLINK("https://example.com")'}])

    assert ws['A2'].data_type == 's'
    assert ws['A2'].value.startswith("'=")
    wb.close()
