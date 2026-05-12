# -*- coding: utf-8 -*-
"""PO Excel解析器：提取头部元数据 + 商品行数据
支持宽版/窄版两种列布局（自动检测）
"""
import os
import re
import logging
from datetime import datetime, timedelta

import openpyxl

logger = logging.getLogger(__name__)


class _CellShim:
    __slots__ = ('value',)

    def __init__(self, value):
        self.value = value


_EMPTY_CELL = _CellShim(None)


class _RowGrid:
    """openpyxl worksheet 的 drop-in 替代：流式一次读完成 2D list，
    之后 ws.cell(r,c).value / ws.max_row 访问都是 O(1)。

    read_only=True 下原始 ws.cell(r,c) 每次从头 scan，循环读 N 行就是 O(N²)，
    对 200+ 行的 PO 要几十秒。iter_rows 一次流式只要不到 1s。
    """
    __slots__ = ('_rows', '_max_row', '_max_col')

    def __init__(self, ws):
        self._rows = list(ws.iter_rows(values_only=True))
        self._max_row = len(self._rows)
        self._max_col = max((len(r) for r in self._rows), default=0)

    def cell(self, r, c):
        if 1 <= r <= self._max_row:
            row = self._rows[r - 1]
            if 1 <= c <= len(row):
                return _CellShim(row[c - 1])
        return _EMPTY_CELL

    @property
    def max_row(self):
        return self._max_row

    @property
    def max_col(self):
        return self._max_col


# 修改单识别正则：文件名含Rev/R1/R2/R3/R4/Rev./.rev./Rev2.等
_REV_RE = re.compile(r'(?:Rev\d*\.?|(?:^|\W)R\d\b|\.rev\.)', re.I)


def is_revision(filename):
    """判断文件名是否为修改单"""
    return bool(_REV_RE.search(filename))


def _clean(v):
    """清理单元格值：去换行、空白、\xa0"""
    if v is None:
        return ''
    s = str(v).replace('\n', ' ').replace('\r', '').replace('\xa0', ' ').strip()
    return s


def _clean_sku(v):
    """清理货号：去掉所有空白和换行"""
    if v is None:
        return ''
    return re.sub(r'\s+', '', str(v)).strip()


def _parse_date(v):
    """解析日期 → datetime，支持多种格式"""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    s = _clean(v)
    if not s:
        return None
    # 尝试多种格式
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%m-%d-%Y', '%m/%d/%Y',
                '%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S'):
        try:
            return datetime.strptime(s[:10], fmt)
        except (ValueError, IndexError):
            continue
    # 尝试Excel序列号
    try:
        serial = float(s)
        if 40000 < serial < 60000:
            return datetime(1899, 12, 30) + timedelta(days=int(serial))
    except (ValueError, TypeError):
        pass
    return None


def _parse_number(v):
    """解析数字：去逗号、空白"""
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return v
    s = _clean(v).replace(',', '').replace(' ', '')
    try:
        return float(s) if '.' in s else int(s)
    except (ValueError, TypeError):
        return 0


def _extract_simple_no(sku_spec):
    """提取简货号：77485GQ3 → 77485，MEC421-S001 → MEC421"""
    if not sku_spec:
        return ''
    s = str(sku_spec).strip()
    if not s:
        return ''
    # M开头：字母前缀+数字（如MEC421、MB104、MTQ15）
    if s[0].upper() == 'M':
        m = re.match(r'([A-Za-z]+\d+)', s)
        return m.group(1) if m else ''
    # 普通：提取开头数字（如77485）
    m = re.match(r'(\d+)', s)
    return m.group(1) if m else ''


def _find_header_row(ws):
    """动态检测表头行：搜索'Line'+'SKU'同时出现的行
    返回表头行号，默认14
    """
    for r in range(10, 20):
        c1 = ws.cell(r, 1).value
        if not c1 or 'Line' not in str(c1):
            continue
        # SKU可能在c2~c5（不同PO布局不同）
        for c in range(2, 6):
            cv = ws.cell(r, c).value
            if cv and 'SKU' in str(cv):
                return r
    return 14  # 兜底默认


def _detect_layout(ws, header_row=14):
    """检测列布局：宽版还是窄版
    通过检查表头行中Delivery Date/Price USD的位置判断
    返回数据列偏移量（0=窄版，1=宽版）
    """
    for c in range(1, 35):
        v = ws.cell(header_row, c).value
        if v and 'Delivery' in str(v):
            # 窄版: Delivery在S(19)，宽版: 在T(20)
            if c <= 19:
                return 0  # 窄版
            else:
                return 1  # 宽版
    # 兜底：检查Name列位置
    for c in range(1, 10):
        v = ws.cell(header_row, c).value
        if v and 'Name' in str(v):
            if c <= 5:
                return 0
            else:
                return 1
    return 0  # 默认窄版


def _find_header_value(ws, row, keywords, max_col=35):
    """在指定行中找到包含关键词的单元格，返回其右侧的值"""
    for c in range(1, max_col):
        v = ws.cell(row, c).value
        if v is None:
            continue
        s = str(v).strip()
        for kw in keywords:
            if kw.lower() in s.lower():
                # 取右侧最近的非空值
                for cc in range(c + 1, min(c + 6, max_col)):
                    vv = ws.cell(row, cc).value
                    if vv is not None and str(vv).strip():
                        return vv
                return None
    return None


def _find_cell_value(ws, row, keywords, max_col=35):
    """在指定行找关键词，返回同行右侧值"""
    return _find_header_value(ws, row, keywords, max_col)


def parse(filepath):
    """解析单个PO Excel文件
    Returns: {
        'po_number': str,
        'po_date': datetime,
        'customer': str,
        'customer_po': str,
        'destination': str,  # 英文原文
        'from_person': str,
        'ship_date': datetime,
        'ship_type': str,
        'supplier': str,
        'lines': [
            {
                'line_no': str,
                'sku': str,
                'sku_spec': str,      # 完整货号
                'simple_no': str,     # 简货号（纯数字）
                'name': str,
                'inner_pcs': int,
                'outer_pcs': int,
                'barcode': str,
                'delivery': datetime,
                'price': float,
                'qty': int,
                'total_usd': float,
                'total_ctns': int,
                'customer_po': str,   # 行级客户PO
            }, ...
        ],
        'tracking_code': str,
        'packaging_info': str,
        'remark': str,
        'revision_records': str,
    }
    """
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    # 流式读成 2D 数组后立刻关闭，后续所有 ws.cell(r,c).value 访问都是 O(1)
    ws = _RowGrid(wb.active)
    wb.close()
    result = {
        'po_number': '', 'po_date': None, 'customer': '',
        'customer_po': '', 'destination': '', 'from_person': '',
        'ship_date': None, 'ship_type': '', 'supplier': '',
        'lines': [],
        'tracking_code': '', 'packaging_info': '', 'remark': '',
        'revision_records': '',
    }

    # === 头部元数据提取（动态搜索行4-13）===
    # 先找Date:所在行，以此为基准
    date_row = None
    for r in range(4, 14):
        for c in range(1, 6):
            v = ws.cell(r, c).value
            if v and 'Date' in str(v) and 'Shipment' not in str(v):
                date_row = r
                break
        if date_row:
            break
    if not date_row:
        date_row = 6  # 兜底

    result['po_date'] = _parse_date(_find_cell_value(ws, date_row, ['Date']))
    result['ship_date'] = _parse_date(_find_cell_value(ws, date_row, ['Shipment Date']))

    # PO号（Date行+1）
    po_raw = _find_cell_value(ws, date_row + 1, ['PO#', 'PO #'])
    if po_raw is not None:
        result['po_number'] = _clean(po_raw).replace('.0', '')
    if not result['po_number']:
        m = re.search(r'PO[- ]?(\d{10})', os.path.basename(filepath))
        if m:
            result['po_number'] = m.group(1)

    # 客户名（Date行+2）
    result['customer'] = _clean(_find_cell_value(ws, date_row + 2, ['Customer Name']))

    # 客户PO（Date行+3）
    cpo = _find_cell_value(ws, date_row + 3, ['Customer PO'])
    result['customer_po'] = _clean(cpo) if cpo else ''

    # 供应商（Date行+4）
    result['supplier'] = _clean(_find_cell_value(ws, date_row + 4, ['Supplier']))

    # 目的国（Date行+4~6范围搜索）
    dest = None
    for r in range(date_row + 4, date_row + 7):
        dest = _find_cell_value(ws, r, ['Destination'])
        if dest:
            break
    result['destination'] = _clean(dest) if dest else ''
    # 清理重复（如"USA,USA"）
    if result['destination'] and ',' in result['destination']:
        parts = [p.strip() for p in result['destination'].split(',') if p.strip()]
        seen = []
        for p in parts:
            if p not in seen:
                seen.append(p)
        result['destination'] = ', '.join(seen)

    # 跟单人（Date行+5~6）
    att = _find_cell_value(ws, date_row + 5, ['Att'])
    frm = _find_cell_value(ws, date_row + 6, ['From'])
    if not frm:
        frm = _find_cell_value(ws, date_row + 5, ['From'])
    result['from_person'] = _clean(frm or att or '')

    # 运输类型（Date行+1右侧）
    st = _find_cell_value(ws, date_row + 1, ['Shipment Type'])
    if not st:
        st = _find_cell_value(ws, date_row, ['Shipment Type'])
    result['ship_type'] = _clean(st) if st else ''

    # === 动态检测表头行和列布局 ===
    header_row = _find_header_row(ws)

    # 动态检测每列的实际位置（从表头关键词匹配）
    col_map = {}
    for c in range(1, 42):
        v = ws.cell(header_row, c).value
        if not v:
            continue
        s = str(v).strip().replace('\n', ' ')
        sl = s.lower()
        if 'line' in sl and c <= 2:
            col_map.setdefault('line', c)
        elif sl == 'sku':
            col_map.setdefault('sku', c)
        elif 'sku-spec' in sl or 'sku spec' in sl or sl == 'spec':
            col_map.setdefault('spec', c)
        elif 'name' in sl and 'customer' not in sl:
            col_map.setdefault('name', c)
        elif 'inner' in sl:
            col_map.setdefault('inner', c)
        elif 'outer' in sl and 'barcode' not in sl:
            col_map.setdefault('outer', c)
        elif 'barcode' in sl:
            col_map.setdefault('barcode', c)
        elif 'delivery' in sl:
            col_map.setdefault('delivery', c)
        elif 'price' in sl and 'status' not in sl:
            col_map.setdefault('price', c)
        elif 'qty' in sl:
            col_map.setdefault('qty', c)
        elif 'total' in sl and 'usd' in sl:
            col_map.setdefault('total_usd', c)
        elif 'total' in sl and 'ctn' in sl:
            col_map.setdefault('total_ctns', c)
        elif 'ship' in sl and 'type' in sl:
            col_map.setdefault('ship_type', c)
        elif 'customer' in sl and 'po' in sl:
            col_map.setdefault('cust_po', c)

    # 兜底：用offset模式（兼容旧PO）
    offset = _detect_layout(ws, header_row)
    COL_LINE = col_map.get('line', 1)
    COL_SKU = col_map.get('sku', 2)
    COL_SPEC = col_map.get('spec', 3)
    COL_NAME = col_map.get('name', 5 + offset)
    COL_INNER = col_map.get('inner', 8 + offset)
    COL_OUTER = col_map.get('outer', 12 + offset)
    COL_BARCODE = col_map.get('barcode', 17 + offset)
    COL_DELIVERY = col_map.get('delivery', 19 + offset)
    COL_PRICE = col_map.get('price', 21 + offset)
    COL_QTY = col_map.get('qty', 22 + offset)
    COL_TOTAL_USD = col_map.get('total_usd', 24 + offset)
    COL_TOTAL_CTNS = col_map.get('total_ctns', 25 + offset)
    COL_SHIP_TYPE = col_map.get('ship_type', 28 + offset)
    COL_CUST_PO = col_map.get('cust_po', 29 + offset)

    logger.info(f'[PO解析] {os.path.basename(filepath)} 表头行={header_row} 列映射={col_map}')

    # 数据起始行 = 表头行 + 2（跳过子表头行如PCS/SIZE）
    data_start_row = header_row + 2

    # === 先收集所有原始行，再做跨页合并 ===
    raw_rows = []
    totals_row = data_start_row + 100
    for r in range(data_start_row, 300):
        a_val = ws.cell(r, 1).value
        if a_val and 'Total' in str(a_val):
            totals_row = r
            break
        row_data = {
            'row': r,
            'line_no': _clean(ws.cell(r, COL_LINE).value).replace('.0', ''),
            'sku': _clean_sku(ws.cell(r, COL_SKU).value),
            'spec': _clean_sku(ws.cell(r, COL_SPEC).value),
            'name': _clean(ws.cell(r, COL_NAME).value),
            'inner_pcs': int(_parse_number(ws.cell(r, COL_INNER).value)),
            'outer_pcs': int(_parse_number(ws.cell(r, COL_OUTER).value)),
            'barcode': _clean(ws.cell(r, COL_BARCODE).value),
            'delivery': _parse_date(ws.cell(r, COL_DELIVERY).value),
            'price': float(_parse_number(ws.cell(r, COL_PRICE).value)),
            'qty': int(_parse_number(ws.cell(r, COL_QTY).value)),
            'total_usd': float(_parse_number(ws.cell(r, COL_TOTAL_USD).value)),
            'total_ctns': int(_parse_number(ws.cell(r, COL_TOTAL_CTNS).value)),
            'customer_po': _clean(ws.cell(r, COL_CUST_PO).value),
        }
        # 只要有任何有意义的数据就保留
        if row_data['sku'] or row_data['spec'] or row_data['line_no'] or row_data['qty']:
            raw_rows.append(row_data)

    # 跨页合并：当前行有line_no+数值但下一行有更完整的SKU/SPEC → 合并
    merged_rows = []
    skip_next = False
    for i, cur in enumerate(raw_rows):
        if skip_next:
            skip_next = False
            continue

        nxt = raw_rows[i + 1] if i + 1 < len(raw_rows) else None

        # 判断是否需要与下一行合并
        need_merge = False
        merge_mode = 'replace'  # replace=用下一行覆盖, append=拼接到spec末尾
        if nxt and not nxt['line_no']:
            # 情况1：当前行有line_no但无SKU/SPEC，下一行有SKU/SPEC
            if cur['line_no'] and not cur['sku'] and not cur['spec'] and (nxt['sku'] or nxt['spec']):
                need_merge = True
            # 情况2：当前行有line_no+SKU+数值但无SPEC，下一行有SPEC但无数值
            elif cur['line_no'] and cur['sku'] and not cur['spec'] and nxt['spec'] and not nxt['qty']:
                need_merge = True
            # 情况3：当前行有line_no+数值，下一行补充了SPEC或SKU
            elif cur['line_no'] and cur['qty'] and not cur['spec'] and nxt['spec']:
                need_merge = True
            # 情况4：跨单元格拆分碎片 — 当前行有完整数据，下一行只有短片段
            # 如 92146H-S00 + 1 → 92146H-S001；或 77909GQ + 1, 77909GQ1-S + 2
            elif (cur['line_no'] and cur['spec'] and cur['qty']
                  and not nxt['qty'] and not nxt['price']
                  and (nxt['spec'] and len(nxt['spec']) <= 5)
                  and (not nxt['sku'] or len(nxt['sku']) <= 3)):
                need_merge = True
                merge_mode = 'append'

        if need_merge:
            # 合并：每个字段取非空的那个，优先当前行的数值+下一行的标识
            merged = dict(cur)
            if merge_mode == 'append':
                # SPEC拼接模式：把下一行的片段追加到当前行spec
                merged['spec'] = cur['spec'] + nxt['spec']
                if nxt['sku'] and len(nxt['sku']) <= 3:
                    merged['sku'] = cur['sku'] + nxt['sku']
            else:
                merged['sku'] = cur['sku'] or nxt['sku']
                merged['spec'] = cur['spec'] or nxt['spec']
            merged['barcode'] = cur['barcode'] or nxt['barcode']
            merged['customer_po'] = cur['customer_po'] or nxt['customer_po']
            # 名称拼接
            names = [n for n in [cur['name'], nxt['name']] if n]
            merged['name'] = ' '.join(names)
            # 数值优先当前行
            if not merged['qty'] and nxt['qty']:
                merged['qty'] = nxt['qty']
            if not merged['price'] and nxt['price']:
                merged['price'] = nxt['price']
            if not merged['delivery'] and nxt['delivery']:
                merged['delivery'] = nxt['delivery']
            merged['_merged'] = True
            merged_rows.append(merged)
            skip_next = True
        else:
            merged_rows.append(cur)

    # 转换为最终lines
    for rd in merged_rows:
        if not rd['sku'] and not rd['spec']:
            continue

        line = {
            'line_no': rd['line_no'],
            'sku': rd['sku'],
            'sku_spec': rd['spec'],
            'simple_no': _extract_simple_no(rd['spec'] or rd['sku']),
            'name': rd['name'],
            'inner_pcs': rd['inner_pcs'],
            'outer_pcs': rd['outer_pcs'],
            'barcode': rd['barcode'],
            'delivery': rd['delivery'],
            'price': rd['price'],
            'qty': rd['qty'],
            'total_usd': rd['total_usd'],
            'total_ctns': rd['total_ctns'],
            'customer_po': rd['customer_po'],
        }
        if not line['delivery'] and result['ship_date']:
            line['delivery'] = result['ship_date']
        result['lines'].append(line)

    # === 备注区域（Totals行之后）===
    tc_parts, pi_parts, rm_parts, rev_parts = [], [], [], []
    current_section = None
    for r in range(totals_row + 1, totals_row + 80):
        a_val = _clean(ws.cell(r, 1).value)
        # D列(4)通常存实际内容
        d_val = _clean(ws.cell(r, 4).value)
        content = d_val or a_val

        if 'Tracking Code' in a_val:
            current_section = 'tc'
            if d_val:
                tc_parts.append(d_val)
            continue
        elif 'Packaging Info' in a_val:
            current_section = 'pi'
            if d_val:
                pi_parts.append(d_val)
            continue
        elif a_val.startswith('Remark') and 'Modifiable' not in a_val:
            current_section = 'rm'
            if d_val:
                rm_parts.append(d_val)
            continue
        elif 'Order Modifiable' in a_val or 'Revision' in a_val:
            current_section = 'rev'
            if 'Revision' in a_val and d_val:
                rev_parts.append(f'{a_val}: {d_val}')
            continue
        elif 'Special Req' in a_val or 'Additional Clause' in a_val:
            current_section = None
            continue
        elif 'PRODUCT REQ' in a_val.upper():
            current_section = None
            continue

        # 续行内容
        if current_section and content:
            if current_section == 'tc':
                tc_parts.append(content)
            elif current_section == 'pi':
                pi_parts.append(content)
            elif current_section == 'rm':
                rm_parts.append(content)
            elif current_section == 'rev':
                rev_parts.append(content)

    result['tracking_code'] = '\n'.join(tc_parts)
    result['packaging_info'] = '\n'.join(pi_parts)
    result['remark'] = '\n'.join(rm_parts)
    result['revision_records'] = '\n'.join(rev_parts)

    logger.info(f'[PO解析] PO={result["po_number"]} 客户={result["customer"]} '
                f'{len(result["lines"])}行数据 目的国={result["destination"]}')
    return result
