# -*- coding: utf-8 -*-
"""翻译排期模块：MA扣数表格式的排期文件检测、扫描、导出

支持的文件格式（如 ZURU-MA 扣数表 RR新格式）：
- 每个sheet对应一个ZURU货号（sheet名=货号或货号前缀）
- 上半部分为MA区域（Row1-2表头, Row3 MA Balance, MA数据, MA Total行）
- 下半部分为PO排期区域（订单号4500xxx格式）
- 列布局：A版本 B客户 C下单日期 D货期 E订单号 F跟单 G货号 H数量 I扣单编号 J+扣数 ...WB列

导出时只填6列（C-H），WB列复制公式=$H{row}，其余留空手动填写。
"""
import os
import re
import logging
from collections import OrderedDict
from datetime import datetime

import openpyxl
import openpyxl.utils
import openpyxl.descriptors.base as _db

# AutoFilter补丁（与hy_schedule.py一致）
_orig_mp_set = _db.MatchPattern.__set__
def _lenient_mp_set(self, instance, value):
    try:
        _orig_mp_set(self, instance, value)
    except ValueError:
        instance.__dict__[self.name] = None
_db.MatchPattern.__set__ = _lenient_mp_set


# 翻译排期数据列号
FY_COL = {
    'po_date': 3,       # C: 下单日期
    'ship_date': 4,     # D: 货期
    'po': 5,            # E: 订单号
    'from_person': 6,   # F: 跟单
    'item': 7,          # G: ZURU货号
    'qty': 8,           # H: 订单数量
}

# 辅助sheet黑名单
_FY_SKIP_SHEETS = {'汇总表', '汇总', 'Sheet1', 'sheet1', 'Sheet'}


def detect_fy_file(filepath):
    """检测文件是否为翻译排期（MA扣数表）格式。

    判断逻辑：
    1. 文件名含"扣数" → True
    2. 第一个sheet的A3含"MA Balance" → True
    """
    fn = os.path.basename(filepath)
    if '扣数' in fn:
        return True

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        if not wb.sheetnames:
            wb.close()
            return False
        ws = wb[wb.sheetnames[0]]
        a3 = str(ws.cell(3, 1).value or '').upper()
        wb.close()
        return 'MA' in a3 and 'BALANCE' in a3
    except Exception:
        return False


def _find_ma_total_row(ws, max_scan=500):
    """在sheet中查找 'MA Total' 或 'MA - total' 行号"""
    for r in range(1, max_scan):
        v = str(ws.cell(r, 1).value or '').strip()
        if re.match(r'MA\s*[-–]?\s*[Tt]otal', v):
            return r
    return None


def scan_fy_items(filepath):
    """扫描翻译排期文件，构建 item_map 条目。

    对每个非辅助sheet：
    1. sheet名的数字前缀作为主key
    2. sheet全名（去空格大写）也作为key
    3. PO区域G列的具体货号也加入映射

    Returns: {item_key_upper: [{file, sheet, type: "fy"}]}
    """
    fn = os.path.basename(filepath)
    mapping = {}

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        logging.warning(f'[翻译排期] 打开失败 {fn}: {e}')
        return {}

    for sn in wb.sheetnames:
        sn_stripped = sn.strip()
        if sn_stripped in _FY_SKIP_SHEETS:
            continue

        entry = {'file': fn, 'sheet': sn, 'type': 'fy'}
        sn_upper = re.sub(r'[\s\n]+', '', sn_stripped).upper()

        # 1) 数字前缀作为主key（如 "10097-B" → "10097"）
        num_m = re.match(r'(\d+)', sn_upper)
        if num_m:
            num_key = num_m.group(1)
            mapping.setdefault(num_key, [])
            if entry not in mapping[num_key]:
                mapping[num_key].append(entry)

        # 2) sheet全名也作为key
        if sn_upper:
            for variant in {sn_upper, sn_upper.replace('-', '')}:
                mapping.setdefault(variant, [])
                if entry not in mapping[variant]:
                    mapping[variant].append(entry)

        # 3) 扫描PO区域G列的具体货号
        try:
            ws = wb[sn]
            ma_total_row = _find_ma_total_row(ws)
            po_start = (ma_total_row + 1) if ma_total_row else 4

            empty = 0
            for r in range(po_start, 5000):
                v = ws.cell(r, 7).value  # G列
                if not v:
                    empty += 1
                    if empty > 30:
                        break
                    continue
                empty = 0
                s = re.sub(r'[\s\n]+', '', str(v)).strip().upper()
                if not s or not re.match(r'\d', s):
                    continue
                mapping.setdefault(s, [])
                if entry not in mapping[s]:
                    mapping[s].append(entry)
        except Exception:
            pass

    total_items = len(mapping)
    total_sheets = sum(1 for sn in wb.sheetnames if sn.strip() not in _FY_SKIP_SHEETS)
    wb.close()
    logging.info(f'[翻译排期] 扫描完成 {fn}: {total_sheets}个sheet, {total_items}个货号key')
    return mapping


def _read_fy_sheet_headers(schedule_dir, source_file, sheet_name):
    """从源排期文件读取表头行和WB列位置。

    Returns: (headers_r1, headers_r2, wb_col_indices, max_col)
    """
    fpath = os.path.join(schedule_dir, source_file)
    if not os.path.exists(fpath):
        return [], [], [], 8

    try:
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    except Exception:
        return [], [], [], 8

    # 匹配sheet（处理尾部空格）
    ws = None
    target = sheet_name.strip()
    for sn in wb.sheetnames:
        if sn.strip() == target:
            ws = wb[sn]
            break
    if not ws:
        wb.close()
        return [], [], [], 8

    max_col = min(ws.max_column or 8, 50)

    headers_r1 = []
    headers_r2 = []
    wb_cols = []

    for c in range(1, max_col + 1):
        h1 = str(ws.cell(1, c).value or '')
        h2 = str(ws.cell(2, c).value or '')
        headers_r1.append(h1)
        headers_r2.append(h2)

        # WB列检测：表头含 "WB"
        if 'WB' in (h1 + h2).upper():
            wb_cols.append(c)

    wb.close()
    return headers_r1, headers_r2, wb_cols, max_col


def generate_fy_export(fy_rows, output_dir, schedule_dir):
    """生成翻译排期导出Excel。

    每个目标sheet一个Excel页签，列布局与源文件一致：
    - C-H列填入PO数据
    - WB列写入公式 =$H{row}
    - 其余列留空

    Args:
        fy_rows: [{target_file, target_sheet, po_date, ship_date, po, from_person, item, qty}]
        output_dir: 导出目录
        schedule_dir: 排期文件目录

    Returns: 导出文件名，或 None
    """
    if not fy_rows:
        return None

    # 按target_sheet分组，保持首次出现顺序
    rows_by_sheet = OrderedDict()
    for row in fy_rows:
        ts = row.get('target_sheet') or row.get('target_file', '未知')
        rows_by_sheet.setdefault(ts, []).append(row)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    existing_names = set()

    for target_sheet, rows in rows_by_sheet.items():
        source_file = rows[0].get('target_file', '')

        # 读取源文件表头和WB列
        headers_r1, headers_r2, wb_cols, max_col = _read_fy_sheet_headers(
            schedule_dir, source_file, target_sheet)

        # 生成sheet名
        sheet_name = re.sub(r'[\\/?*\[\]:]', '', str(target_sheet).strip())
        if not sheet_name:
            sheet_name = 'Sheet'
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
        base = sheet_name
        idx = 2
        while sheet_name in existing_names:
            suffix = f'({idx})'
            sheet_name = base[:31 - len(suffix)] + suffix
            idx += 1
        existing_names.add(sheet_name)

        ws = wb.create_sheet(title=sheet_name)

        # 写表头
        data_start = 2
        if headers_r1:
            for ci, h in enumerate(headers_r1, 1):
                if h:
                    ws.cell(1, ci, h)
            if any(headers_r2):
                for ci, h in enumerate(headers_r2, 1):
                    if h:
                        ws.cell(2, ci, h)
                data_start = 3
        else:
            # 无源文件时用默认表头
            labels = {3: '下单日期', 4: '货期', 5: '订单号',
                      6: '跟单', 7: 'ZURU货号', 8: '订单数量'}
            for col, label in labels.items():
                ws.cell(1, col, label)

        # 写数据行
        for ri_offset, row in enumerate(rows):
            ri = data_start + ri_offset

            # 日期字段防御：过滤 pandas NaT / NaN / None
            for date_key in ('po_date', 'ship_date'):
                val = row.get(date_key)
                col = FY_COL[date_key]
                if val is not None and hasattr(val, 'year'):
                    ws.cell(ri, col, val)
                elif isinstance(val, str) and val.strip():
                    ws.cell(ri, col, val)
            ws.cell(ri, FY_COL['po'], row.get('po'))
            ws.cell(ri, FY_COL['from_person'], row.get('from_person'))
            ws.cell(ri, FY_COL['item'], row.get('item'))
            qty = row.get('qty')
            if qty:
                ws.cell(ri, FY_COL['qty'], qty)

            # WB公式列
            for col in wb_cols:
                ws.cell(ri, col, f'=$H{ri}')

        # 列宽
        col_widths = {3: 12, 4: 12, 5: 15, 6: 14, 7: 15, 8: 12}
        for c, w in col_widths.items():
            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

    # 保存
    os.makedirs(output_dir, exist_ok=True)
    filename = f'翻译排期导出_{datetime.now().strftime("%m%d_%H%M%S")}.xlsx'
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)

    logging.info(f'[翻译排期] 导出: {len(fy_rows)}行, {len(rows_by_sheet)}个sheet, 文件={filename}')
    return filename
