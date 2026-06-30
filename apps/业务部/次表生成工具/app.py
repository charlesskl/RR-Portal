"""
单价表次表生成工具 - Flask主应用
端口: 5007
"""
import os
import json
import logging
import uuid
from flask import Flask, request, jsonify, render_template, send_file
from formula_engine import calc_all
from excel_engine import load_excel, scan_sheet, generate_subtable, get_sheet_info, find_append_position, process_upload, process_upload_multi, extract_formula_numbers, extract_columns

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
logger = logging.getLogger(__name__)


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/calculate', methods=['POST'])
def api_calculate():
    """公式计算器API"""
    try:
        params = request.get_json()
        result = calc_all(params)
        return jsonify({'success': True, 'data': result})
    except Exception as e:
        logger.exception("计算出错")
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/upload', methods=['POST'])
def api_upload():
    """上传Excel文件，返回sheet列表"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '未选择文件'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'success': False, 'error': '未选择文件'}), 400

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ('.xls', '.xlsx'):
        return jsonify({'success': False, 'error': '仅支持 .xls 和 .xlsx 文件'}), 400

    # 保存文件
    file_id = str(uuid.uuid4())[:8]
    safe_name = f"{file_id}{ext}"
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], safe_name)
    file.save(filepath)

    try:
        wb, actual_path, converted = load_excel(filepath, read_only=True)
        sheets = get_sheet_info(wb)
        wb.close()

        return jsonify({
            'success': True,
            'file_id': file_id,
            'filename': file.filename,
            'filepath': actual_path,
            'converted': converted,
            'sheets': sheets,
        })
    except Exception as e:
        logger.exception("上传文件解析出错")
        return jsonify({'success': False, 'error': f'文件解析失败: {str(e)}'}), 400


@app.route('/api/scan', methods=['POST'])
def api_scan():
    """扫描指定sheet，识别关键行"""
    data = request.get_json()
    filepath = data.get('filepath')
    sheet_name = data.get('sheet_name')

    if not filepath or not sheet_name:
        return jsonify({'success': False, 'error': '缺少参数'}), 400

    # 路径安全检查
    real_path = os.path.realpath(filepath)
    upload_dir = os.path.realpath(app.config['UPLOAD_FOLDER'])
    if not real_path.startswith(upload_dir):
        return jsonify({'success': False, 'error': '非法路径'}), 403

    try:
        wb, actual_path, _ = load_excel(filepath, read_only=True)
        ws = wb[sheet_name]
        found = scan_sheet(ws)

        # read_only模式不支持ws.cell()随机访问，用iter_rows批量读取所需行
        needed_rows = set()
        for k, v in found.items():
            if isinstance(v, int):
                needed_rows.add(v)
        data_col_end = found.get('data_col_end', found.get('data_col', 3))
        max_col_needed = data_col_end + 2  # 多读1列用于备注

        # PP到CC之间的中间行也需要读取（前半段行跳过功能）
        pp_r = found.get('product_price')
        cc_r = found.get('carton_cost')
        if pp_r and cc_r and cc_r > pp_r + 1:
            for mid_r in range(pp_r + 1, cc_r):
                needed_rows.add(mid_r)

        # 一次性读取所有需要的行 -> cell_cache[(row, col)] = value
        cell_cache = {}
        if needed_rows:
            min_r, max_r = min(needed_rows), max(needed_rows)
            current_r = min_r - 1
            for row in ws.iter_rows(min_row=min_r, max_row=max_r, min_col=1, max_col=max_col_needed):
                current_r += 1
                if current_r in needed_rows:
                    for col_offset, cell in enumerate(row):
                        cell_cache[(current_r, 1 + col_offset)] = cell.value

        # 读取关键行的实际文本，便于前端显示
        row_details = {}
        for key, row_num in found.items():
            if isinstance(row_num, int):
                cell_val = cell_cache.get((row_num, 1))
                row_details[key] = {
                    'row': row_num,
                    'text': str(cell_val)[:80] if cell_val else '',
                }

        # 自动读取主表关键数值
        auto_values = {}
        def _safe_float(v):
            try:
                return float(v) if v is not None else None
            except (ValueError, TypeError):
                return None

        for av_key, av_row_key in [
            ('exchange_rate', 'exchange_rate'),
            ('original_exchange_rate', 'original_exchange_rate'),
            ('waste_pct', 'waste_pct'),
            ('markup_pct', 'markup_pct'),
        ]:
            row = found.get(av_row_key)
            if row:
                val = _safe_float(cell_cache.get((row, 2)))
                if val is not None:
                    auto_values[av_key] = val
        # 主表pcs数
        carton_row = found.get('carton_cost')
        if carton_row:
            import re as _re
            ct = cell_cache.get((carton_row, 1)) or ''
            m = _re.search(r'(\d+)\s*pcs', str(ct), _re.IGNORECASE)
            if m: auto_values['main_pcs'] = int(m.group(1))
        # 主表纸箱尺寸备注（从Carton Cost行右侧读）
        desc_cell = cell_cache.get((carton_row, data_col_end + 1)) if carton_row else None
        if desc_cell:
            auto_values['carton_desc'] = str(desc_cell)[:100]

        # 读取主表产品货号（通常在B2或B1）
        product_item = None
        for r_idx in range(1, 4):
            for row in ws.iter_rows(min_row=r_idx, max_row=r_idx, min_col=1, max_col=2):
                a_val = row[0].value
                b_val = row[1].value if len(row) > 1 else None
                if a_val and isinstance(a_val, str) and 'ITEM' in a_val.upper():
                    if b_val:
                        product_item = str(b_val).strip()
                    break
            if product_item:
                break
        if product_item:
            auto_values['product_item'] = product_item

        # 读取表头行（row 1 ~ 主表末尾，供new_sheet模式选择复制）
        header_end = found.get('main_table_end', 12)
        header_end = min(header_end, 40)  # 安全上限
        header_rows = []
        formula_cache = {}  # {row_num: (a_text, formula_str)}
        data_col = found.get('data_col', 3)
        current_hr = 0
        for row in ws.iter_rows(min_row=1, max_row=header_end, min_col=1, max_col=max_col_needed):
            current_hr += 1
            a_val = row[0].value
            b_val = row[1].value if len(row) > 1 else None
            # 读取data_col公式
            dc_val = row[data_col - 1].value if len(row) >= data_col else None
            if dc_val and isinstance(dc_val, str) and dc_val.startswith('='):
                formula_cache[current_hr] = (
                    str(a_val)[:60] if a_val else '',
                    dc_val
                )
            # 跳过完全空的行
            if a_val is None and b_val is None:
                continue
            header_rows.append({
                'row': current_hr,
                'a': str(a_val)[:60] if a_val else '',
                'b': str(b_val)[:60] if b_val else '',
            })

        # 构建公式列表（供clone_sheet模式的公式编辑器使用）
        # 排除已由clone_sheet参数控制的行（carton_cost, tape_cost, packing_labor）
        managed_rows = {found.get('carton_cost'), found.get('tape_cost'), found.get('packing_labor')}
        managed_rows.discard(None)
        formulas = []
        for row_num, (a_text, formula_str) in sorted(formula_cache.items()):
            if row_num in managed_rows:
                continue
            nums = extract_formula_numbers(formula_str)
            if not nums:
                continue
            formulas.append({
                'row': row_num,
                'label': a_text,
                'formula': formula_str,
                'numbers': nums,
            })

        # 检测PP到CC之间的中间行（前半段物料行）
        mid_rows = []
        if pp_r and cc_r and cc_r > pp_r + 1:
            for mid_r in range(pp_r + 1, cc_r):
                a_text = cell_cache.get((mid_r, 1))
                if a_text and str(a_text).strip():
                    mid_rows.append({'row': mid_r, 'text': str(a_text).strip()[:60]})

        wb.close()

        return jsonify({
            'success': True,
            'found': {k: v for k, v in found.items() if isinstance(v, int)},
            'row_details': row_details,
            'header_rows': header_rows,
            'append_position': found.get('main_table_end', 0),
            'auto_values': auto_values,
            'mid_rows': mid_rows,
            'formulas': formulas,
            'product_columns': found.get('product_columns', []),
        })
    except Exception as e:
        logger.exception("扫描sheet出错")
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/generate', methods=['POST'])
def api_generate():
    """生成次表并返回下载链接（支持多Sheet多组）"""
    data = request.get_json()

    filepath = data.get('filepath')
    output_mode = data.get('output_mode', 'same_sheet')
    include_header_rows = data.get('include_header_rows', [])  # new_sheet模式的表头行
    formula_overrides_raw = data.get('formula_overrides', {})  # clone_sheet模式的公式覆盖
    # 键从字符串转为int
    formula_overrides = {int(k): v for k, v in formula_overrides_raw.items()} if formula_overrides_raw else {}
    clone_rows = data.get('clone_rows', [])  # clone_sheet模式的行选择
    name_overrides_raw = data.get('name_overrides', {})  # clone_sheet模式的名称覆盖
    name_overrides = {int(k): v for k, v in name_overrides_raw.items()} if name_overrides_raw else {}
    sheets_raw = data.get('sheets')  # 新：多Sheet模式
    sheet_name = data.get('sheet_name')  # 旧：单Sheet兼容
    params_raw = data.get('params', {})  # 旧：单Sheet兼容

    if not filepath:
        return jsonify({'success': False, 'error': '缺少参数'}), 400

    # 路径安全检查
    real_path = os.path.realpath(filepath)
    upload_dir = os.path.realpath(app.config['UPLOAD_FOLDER'])
    if not real_path.startswith(upload_dir):
        return jsonify({'success': False, 'error': '非法路径'}), 403

    # extract_columns模式：直接复制+删列，不需要纸箱参数
    if output_mode == 'extract_columns':
        extract_config = data.get('extract_config', [])
        if not extract_config:
            return jsonify({'success': False, 'error': '缺少提取列配置'}), 400
        extract_item_name = data.get('extract_item_name', '').strip() or None
        extract_formula_overrides = data.get('extract_formula_overrides', {}) or None
        try:
            output_path = extract_columns(filepath, extract_config,
                                          item_name=extract_item_name,
                                          formula_overrides=extract_formula_overrides)
            return jsonify({
                'success': True,
                'output_path': output_path,
                'download_url': f'/api/download?path={output_path}',
            })
        except Exception as e:
            logger.exception("提取列出错")
            return jsonify({'success': False, 'error': str(e)}), 400

    required_fields = ['item_name', 'pcs_per_carton', 'carton_cost_rmb',
                       'tape_cost_rmb', 'packing_labor_rmb', 'cbm']

    # 统一为 sheets_list 格式: [{ sheet_name, params: [...] }]
    if sheets_raw and isinstance(sheets_raw, list):
        sheets_list = sheets_raw
    elif sheet_name:
        # 向下兼容单Sheet模式
        if isinstance(params_raw, dict):
            params_raw = [params_raw]
        sheets_list = [{'sheet_name': sheet_name, 'params': params_raw}]
    else:
        return jsonify({'success': False, 'error': '缺少sheet参数'}), 400

    # 校验每个Sheet的每组参数
    for sheet_entry in sheets_list:
        sn = sheet_entry.get('sheet_name', '?')
        params_list = sheet_entry.get('params', [])
        if isinstance(params_list, dict):
            params_list = [params_list]
            sheet_entry['params'] = params_list
        for idx, params in enumerate(params_list):
            for field in required_fields:
                if field not in params:
                    return jsonify({'success': False, 'error': f'[{sn}] 组{idx+1}缺少参数: {field}'}), 400

    try:
        output_path = process_upload_multi(filepath, sheets_list, output_mode, include_header_rows, formula_overrides,
                                          clone_rows=clone_rows, name_overrides=name_overrides)

        return jsonify({
            'success': True,
            'output_path': output_path,
            'download_url': f'/api/download?path={output_path}',
        })
    except Exception as e:
        logger.exception("生成次表出错")
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/download')
def api_download():
    """下载生成的Excel文件"""
    filepath = request.args.get('path')
    if not filepath or not os.path.exists(filepath):
        return jsonify({'success': False, 'error': '文件不存在'}), 404

    # 路径安全检查：只允许下载uploads目录内的文件
    real_path = os.path.realpath(filepath)
    upload_dir = os.path.realpath(app.config['UPLOAD_FOLDER'])
    if not real_path.startswith(upload_dir):
        return jsonify({'success': False, 'error': '非法路径'}), 403

    return send_file(
        filepath,
        as_attachment=True,
        download_name=os.path.basename(filepath),
    )


@app.route('/api/paper_prices')
def api_paper_prices():
    """读取公式表中的纸价数据"""
    formula_file = os.path.join(os.path.dirname(__file__), '..', '公式表', '纸箱计算公式（最新）5.29.xlsx')
    formula_file = os.path.normpath(formula_file)
    if not os.path.exists(formula_file):
        # 尝试桌面路径
        formula_file = r'C:\Users\Administrator\Desktop\欣姐项目\公式表\纸箱计算公式（最新）5.29.xlsx'
    if not os.path.exists(formula_file):
        return jsonify({'success': False, 'error': '公式表文件不存在'}), 404

    try:
        from openpyxl import load_workbook as _lwb
        wb = _lwb(formula_file, data_only=True, read_only=True)
        ws = wb.active

        prices = []
        # 英寸体系纸价 (E1~E11区域, F=含税, G=不含税)
        for r in range(2, 12):
            name = ws.cell(row=r, column=5).value  # E列=纸种名
            tax_price = ws.cell(row=r, column=6).value  # F列=含税价
            no_tax_price = ws.cell(row=r, column=7).value  # G列=不含税价
            if name and (tax_price or no_tax_price):
                prices.append({
                    'name': str(name).strip(),
                    'tax_price': round(float(tax_price), 4) if tax_price else None,
                    'no_tax_price': round(float(no_tax_price), 4) if no_tax_price else None,
                })
        wb.close()

        return jsonify({'success': True, 'prices': prices})
    except Exception as e:
        logger.exception("读取纸价出错")
        return jsonify({'success': False, 'error': str(e)}), 400


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5007, debug=True)
