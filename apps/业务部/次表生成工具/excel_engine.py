"""
Excel引擎 - 读取主表、识别关键行、生成次表
支持 .xls（转.xlsx）和 .xlsx 格式
"""
import os
import re
import copy
import logging
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# 关键行匹配模式
ROW_PATTERNS = {
    'product_price': re.compile(r'Product\s*Price.*(?:exclude|Packaging)', re.IGNORECASE),
    'carton_cost': re.compile(r'Carton\s*Cost', re.IGNORECASE),
    'tape_cost': re.compile(r'(?:Carton\s*)?Tape\s*Cost', re.IGNORECASE),
    'packing_labor': re.compile(r'(?:Carton\s*)?Packing\s*labor', re.IGNORECASE),
    'ex_factory': re.compile(r'Ex[\-\s]*Factory', re.IGNORECASE),
    'fob_40f': re.compile(r'40F\s*FCL\s*FOB', re.IGNORECASE),
    'shipping_40f': re.compile(r'Shipping\s*40F', re.IGNORECASE),
    'fob_20f': re.compile(r'20F[\s\-]*FCL\s*FOB', re.IGNORECASE),
    'shipping_20f': re.compile(r'Shipping\s*20F', re.IGNORECASE),
    'shipping_lcl': re.compile(r'(?:Shipping\s*)?LCL', re.IGNORECASE),
    'exchange_rate': re.compile(r'(?:USD|Exchange).*Rate', re.IGNORECASE),
    'original_exchange_rate': re.compile(r'Original\s*Exchange\s*Rate', re.IGNORECASE),
    'waste_pct': re.compile(r'%\s*Waste', re.IGNORECASE),
    'markup_pct': re.compile(r'%\s*Mark\s*up', re.IGNORECASE),
}


def _convert_xls_via_com(xls_path, xlsx_path):
    """用 WPS/Excel COM 自动化转换 .xls → .xlsx，完整保留格式"""
    import win32com.client
    import pythoncom
    pythoncom.CoInitialize()
    app = None
    wb = None
    try:
        # 尝试 WPS表格(ket) → Excel，kwps是WPS文字不是表格
        for prog_id in ('ket.Application', 'Excel.Application'):
            try:
                app = win32com.client.Dispatch(prog_id)
                break
            except Exception:
                continue
        if app is None:
            raise RuntimeError('无法启动 WPS 或 Excel')
        app.Visible = False
        app.DisplayAlerts = False
        wb = app.Workbooks.Open(os.path.abspath(xls_path))
        # 51 = xlOpenXMLWorkbook (.xlsx)
        wb.SaveAs(os.path.abspath(xlsx_path), FileFormat=51)
        wb.Close(False)
    finally:
        if wb is not None:
            try:
                wb.Close(False)
            except Exception:
                pass
        if app is not None:
            app.Quit()
        pythoncom.CoUninitialize()


def _com_resave(xlsx_path):
    """用 WPS/Excel COM 重新打开并保存 xlsx 文件，修复 openpyxl 破坏的主题字体引用"""
    import win32com.client
    import pythoncom
    pythoncom.CoInitialize()
    app = None
    wb = None
    try:
        for prog_id in ('ket.Application', 'Excel.Application'):
            try:
                app = win32com.client.Dispatch(prog_id)
                break
            except Exception:
                continue
        if app is None:
            return
        app.Visible = False
        app.DisplayAlerts = False
        wb = app.Workbooks.Open(os.path.abspath(xlsx_path))
        wb.Save()
        wb.Close(False)
        logger.info(f'COM重新保存成功: {xlsx_path}')
    except Exception as e:
        logger.warning(f'COM重新保存失败({e})，跳过')
    finally:
        if wb is not None:
            try:
                wb.Close(False)
            except Exception:
                pass
        if app is not None:
            app.Quit()
        pythoncom.CoUninitialize()


def _convert_xls_via_xlrd(xls_path, xlsx_path):
    """用 xlrd+openpyxl 转换 .xls → .xlsx（回退方案，格式会丢失）"""
    import xlrd
    xls_wb = xlrd.open_workbook(xls_path, formatting_info=True)
    xlsx_wb = Workbook()
    xlsx_wb.remove(xlsx_wb.active)

    for sheet_name in xls_wb.sheet_names():
        xls_ws = xls_wb.sheet_by_name(sheet_name)
        xlsx_ws = xlsx_wb.create_sheet(title=sheet_name)

        for row_idx in range(xls_ws.nrows):
            for col_idx in range(xls_ws.ncols):
                cell = xls_ws.cell(row_idx, col_idx)
                xlsx_cell = xlsx_ws.cell(row=row_idx + 1, column=col_idx + 1)
                if cell.ctype == 0:
                    continue
                elif cell.ctype == 1:
                    xlsx_cell.value = cell.value
                elif cell.ctype == 2:
                    xlsx_cell.value = cell.value
                elif cell.ctype == 3:
                    try:
                        date_tuple = xlrd.xldate_as_tuple(cell.value, xls_wb.datemode)
                        from datetime import datetime
                        xlsx_cell.value = datetime(*date_tuple)
                        xlsx_cell.number_format = 'YYYY-MM-DD'
                    except Exception:
                        xlsx_cell.value = cell.value
                elif cell.ctype == 4:
                    xlsx_cell.value = bool(cell.value)
                else:
                    xlsx_cell.value = cell.value

        for col_idx in range(xls_ws.ncols):
            col_letter = get_column_letter(col_idx + 1)
            max_len = 0
            for row_idx in range(min(20, xls_ws.nrows)):
                val = xls_ws.cell_value(row_idx, col_idx)
                if val:
                    max_len = max(max_len, len(str(val)))
            if max_len > 0:
                xlsx_ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    xlsx_wb.save(xlsx_path)


def convert_xls_to_xlsx(xls_path):
    """
    将.xls转为.xlsx
    优先使用 WPS/Excel COM 自动化（完整保留格式），
    失败时回退到 xlrd 方案（格式会丢失）。
    """
    xlsx_path = xls_path.rsplit('.', 1)[0] + '_converted.xlsx'
    try:
        _convert_xls_via_com(xls_path, xlsx_path)
        logger.info(f'COM转换成功: {xls_path} → {xlsx_path}')
    except Exception as e:
        logger.warning(f'COM转换失败({e})，回退到xlrd方案')
        _convert_xls_via_xlrd(xls_path, xlsx_path)
    return xlsx_path


def load_excel(filepath, read_only=False):
    """
    加载Excel文件，返回workbook和是否转换标记
    .xls文件会先转换为.xlsx
    read_only=True 用于上传/扫描场景，大幅提升大文件加载速度
    """
    ext = os.path.splitext(filepath)[1].lower()
    converted = False

    if ext == '.xls':
        filepath = convert_xls_to_xlsx(filepath)
        converted = True

    wb = load_workbook(filepath, data_only=False, read_only=read_only)
    return wb, filepath, converted


def scan_sheet(ws):
    """
    扫描worksheet，识别关键行位置
    返回字典：{key: row_number, ...}
    兼容read_only模式（使用iter_rows而非ws.max_row）
    """
    found = {}

    # 用iter_rows兼容read_only模式，只读A列
    row_idx = 0
    for row in ws.iter_rows(min_col=1, max_col=1):
        row_idx += 1
        cell = row[0]
        cell_a = cell.value
        if not cell_a or not isinstance(cell_a, str):
            continue

        cell_text = cell_a.strip()

        for key, pattern in ROW_PATTERNS.items():
            if pattern.search(cell_text):
                if key not in found:
                    found[key] = row_idx

    # 查找数据列（均价或第一个数量档位列）
    found['data_col'] = find_data_column(ws, found)

    # 查找数据列结束位置（多档区间表的最后一个数据列）
    found['data_col_end'] = find_data_column_end(ws, found)

    # 查找主表最后一行
    found['main_table_end'] = find_main_table_end(ws, found)

    # 检测多列sheet的产品列（各列的产品名称和列号）
    data_col = found.get('data_col', 3)
    data_col_end = found.get('data_col_end', data_col)
    if data_col_end > data_col:
        found['product_columns'] = _detect_product_columns(ws, found)

    return found


def _detect_product_columns(ws, found):
    """
    检测多列sheet中的产品列。
    在exchange_rate行到product_price行之间，找到有多个纯文本值的header行，
    提取产品名和列号。
    返回: [{'col': 列号, 'name': 产品名}, ...]
    """
    data_col = found.get('data_col', 3)
    data_col_end = found.get('data_col_end', data_col)
    pp_row = found.get('product_price')
    er_row = found.get('exchange_rate')
    if not pp_row:
        return []

    # 搜索范围：exchange_rate行到product_price行之间
    search_start = max(1, (er_row or 1))
    search_end = pp_row

    best_cols = []
    for row in ws.iter_rows(min_row=search_start, max_row=search_end,
                            min_col=data_col, max_col=data_col_end):
        cols = []
        for col_offset, cell in enumerate(row):
            val = cell.value
            col_num = data_col + col_offset
            # 只接受纯文本（非公式、非数字、非标签文本）
            if val and isinstance(val, str) and not val.startswith('=') and val.strip() not in ('－', '-', '') \
                    and '\n' not in val and 'Product' not in val and 'Carton Size' not in val:
                cols.append({'col': col_num, 'name': val.strip()[:40]})
        # 选择有最多产品名的行
        if len(cols) >= 2 and len(cols) > len(best_cols):
            best_cols = cols

    return best_cols


def find_data_column(ws, found):
    """找到数据列的列号（通常是C列=3）"""
    pp_row = found.get('product_price')
    if pp_row:
        for row in ws.iter_rows(min_row=pp_row, max_row=pp_row, min_col=2, max_col=20):
            for col_offset, cell in enumerate(row):
                val = cell.value
                if val is not None and val != '－' and val != '-':
                    try:
                        float(val)
                        return 2 + col_offset  # min_col=2, 所以列号=2+offset
                    except (ValueError, TypeError):
                        continue
    return 3  # 默认C列


def find_data_column_end(ws, found):
    """找到最后一个有数据的列号（多档区间表为N列=14，单列表等于data_col）"""
    data_col = found.get('data_col', 3)
    check_row = found.get('ex_factory') or found.get('carton_cost') or found.get('product_price')
    if not check_row:
        return data_col

    # 用iter_rows读取该行，兼容read_only模式，扫描到30列足够
    last_col = data_col
    cells = {}
    for row in ws.iter_rows(min_row=check_row, max_row=check_row, min_col=data_col, max_col=30):
        for col_offset, cell in enumerate(row):
            cells[data_col + col_offset] = cell.value

    for col in range(data_col, 31):
        val = cells.get(col)
        if val is not None and val != '' and val != '－' and val != '-':
            last_col = col
        else:
            next_val = cells.get(col + 1)
            if next_val is None or next_val == '' or next_val == '－':
                break
    return last_col


def find_main_table_end(ws, found):
    """找到主表的最后一行（用于确定追加位置），兼容read_only模式"""
    candidates = [v for k, v in found.items() if isinstance(v, int)]
    if candidates:
        max_row = max(candidates)
        # 从max_row往下扫描找空白区域，用iter_rows兼容read_only
        scan_end = max_row + 20
        current_r = max_row - 1
        for row in ws.iter_rows(min_row=max_row, max_row=scan_end, min_col=1, max_col=9):
            current_r += 1
            all_empty = all(cell.value is None for cell in row)
            cell_a = row[0].value
            if cell_a is None or (isinstance(cell_a, str) and cell_a.strip() == ''):
                if all_empty:
                    return current_r
        return max_row + 1
    # 没有识别到任何关键行，返回一个安全值
    row_count = 0
    for _ in ws.iter_rows(min_col=1, max_col=1):
        row_count += 1
    return row_count + 1


def find_append_position(ws):
    """
    找到追加次表的位置
    如果已有次表，在最后一个次表后追加
    """
    # 从底部往上找最后一个有内容的行
    last_content_row = 1
    for r in range(ws.max_row, 0, -1):
        for c in range(1, min(ws.max_column + 1, 10)):
            if ws.cell(row=r, column=c).value is not None:
                last_content_row = r
                break
        if last_content_row > 1:
            break

    return last_content_row + 2  # 空一行再追加


def _collect_main_formulas(ws, found):
    """
    收集主表中 Carton Cost 到 LCL 之间所有行的公式和B列值。
    返回: [(row_num, a_text, b_value_or_formula, c_formula), ...]
    """
    data_col = found.get('data_col', 3)
    start_keys = ['carton_cost', 'tape_cost', 'packing_labor',
                  'ex_factory', 'fob_40f', 'shipping_40f',
                  'fob_20f', 'shipping_20f', 'shipping_lcl']

    rows_info = []
    for key in start_keys:
        row = found.get(key)
        if not row:
            continue
        a_val = ws.cell(row=row, column=1).value or ''
        b_val = ws.cell(row=row, column=2).value
        c_val = ws.cell(row=row, column=data_col).value
        rows_info.append({
            'key': key,
            'src_row': row,
            'a_text': a_val,
            'b_value': b_val,
            'c_value': c_val,
        })
    return rows_info


def _remap_formula(formula, row_map, pcs_old, pcs_new, remove_rows=None):
    """
    将主表公式中的行号引用替换为次表的新行号。
    同时替换pcs数（如12→6）。

    row_map: {旧行号: 新行号, ...}
    remove_rows: 需要从公式中移除引用的主表行号集合（前半段行跳过）
    """
    if not formula or not isinstance(formula, str) or not formula.startswith('='):
        return formula

    result = formula

    # 先移除被跳过的行引用（在remap之前处理，因为行号还是原始值）
    if remove_rows:
        for rm_row in sorted(remove_rows, reverse=True):
            # 移除 +$C$9 或 +C9 格式（中间或末尾的项）
            for pat in [r'\+\$?[A-Z]+\$?' + str(rm_row) + r'(?!\d)',
                        r'\$?[A-Z]+\$?' + str(rm_row) + r'(?!\d)\+']:
                result = re.sub(pat, '', result)

    # 按行号从大到小替换，避免 $B$2 误替换 $B$22 中的 2
    sorted_rows = sorted(row_map.keys(), reverse=True)
    for old_row in sorted_rows:
        new_row = row_map[old_row]
        # 替换各种引用格式: $B$22, B22, $C$22, C22 等
        # (?!\d) 负向前瞻：确保行号后面不跟数字，防止C5匹配到C50
        for pattern_fmt in [r'\$([A-Z])\${}(?!\d)', r'([A-Z]){}(?!\d)']:
            pattern = pattern_fmt.format(old_row)
            replacement_fmt = '${}${}' if '$' in pattern_fmt else '{}{}'
            result = re.sub(
                pattern,
                lambda m, nr=new_row, fmt=replacement_fmt: fmt.format(m.group(1), nr),
                result
            )

    # 替换pcs数：只替换被 * 或 / 包围的独立数字
    if pcs_old != pcs_new and pcs_old > 0:
        # 匹配 *12) 或 /12+  或 /12) 等场景中的pcs数
        result = re.sub(
            r'(?<=[*/])' + str(pcs_old) + r'(?=[)\s/+\-]|$)',
            str(pcs_new),
            result
        )

    return result


def _fill_chain_columns(ws_target, row, data_col, data_col_end, number_format='$#,##0.000'):
    """对于固定成本行，D-N列写链式传递公式: =前列同行号"""
    if data_col_end <= data_col:
        return
    for col in range(data_col + 1, data_col_end + 1):
        prev_letter = get_column_letter(col - 1)
        ws_target.cell(row=row, column=col).value = f'={prev_letter}{row}'
        ws_target.cell(row=row, column=col).number_format = number_format


def _resolve_material_position(position, found):
    """
    将 after_row_N 格式的位置转换为5个zone之一。
    根据N与主表关键行的位置关系:
      N < carton_cost → product
      N = carton_cost 或 CC~Tape之间 → carton
      N = tape_cost 或 Tape~Labor之间 → after_tape
      N = packing_labor 或 Labor~ExFactory之间 → after_labor
      N >= ex_factory → after_exfactory
    返回: (zone_key, row_n)
    """
    # 兼容旧格式（直接是zone名）
    if position in ('product', 'carton', 'after_tape', 'after_labor', 'after_exfactory'):
        return position, 0

    m = re.match(r'after_row_(\d+)', position or '')
    if not m:
        return 'carton', 0  # 默认carton区

    row_n = int(m.group(1))

    cc = found.get('carton_cost', 0)
    tc = found.get('tape_cost', 0)
    pl = found.get('packing_labor', 0)
    ef = found.get('ex_factory', 0)

    if cc and row_n < cc:
        return 'product', row_n
    if ef and row_n >= ef:
        return 'after_exfactory', row_n
    if pl and row_n >= pl:
        return 'after_labor', row_n
    if tc and row_n >= tc:
        return 'after_tape', row_n
    # cc <= row_n < tc (or tc not found)
    return 'carton', row_n


# ===================== 公式数字提取（供克隆表编辑器使用） =====================

# 匹配单元格引用（含$锚定、跨sheet引用）
_CELL_REF_RE = re.compile(r"\$?[A-Z]{1,3}\$?\d+")
# 匹配方括号内容（外部工作簿引用 [1]Sheet!A1）
_BRACKET_RE = re.compile(r"\[[^\]]*\]")
# 匹配数字字面量（整数或小数）
_NUMBER_RE = re.compile(r"\d+\.?\d*")


def extract_formula_numbers(formula):
    """
    从Excel公式中提取可编辑的数字字面量，排除单元格引用中的数字。

    参数:
        formula: 公式字符串，如 '=0.01*2/1.08/B5'
    返回:
        list of {'value': str, 'start': int, 'end': int}
        start/end 是相对于完整公式字符串的字符位置（左闭右开）
    """
    if not formula or not isinstance(formula, str) or not formula.startswith('='):
        return []

    # 收集需要排除的字符位置（单元格引用 + 方括号内容）
    excluded = set()
    for m in _CELL_REF_RE.finditer(formula):
        excluded.update(range(m.start(), m.end()))
    for m in _BRACKET_RE.finditer(formula):
        excluded.update(range(m.start(), m.end()))

    # 提取不在排除区的数字
    numbers = []
    for m in _NUMBER_RE.finditer(formula):
        if not any(i in excluded for i in range(m.start(), m.end())):
            numbers.append({
                'value': m.group(),
                'start': m.start(),
                'end': m.end(),
            })
    return numbers


def clone_sheet(wb, sheet_name, found, params, formula_overrides=None,
                clone_rows=None, name_overrides=None):
    """
    克隆主表到新sheet，替换成本参数（Carton Cost/Tape/Labor等）。
    公式保留原结构，只替换参数值。

    适用场景：基于同一主表结构，用不同纸箱参数生成新货号的完整表。
    formula_overrides: {src_row: new_formula_string} 用户在前端编辑的公式覆盖
    """
    ws = wb[sheet_name]
    data_col = found.get('data_col', 3)
    data_col_end = found.get('data_col_end', data_col)
    exchange_row = found.get('exchange_rate')
    waste_row = found.get('waste_pct')
    markup_row = found.get('markup_pct')
    original_rate_row = found.get('original_exchange_rate')
    is_multi_col = data_col_end > data_col

    pcs = params['pcs_per_carton']

    # 预处理内联物料
    for mat in params.get('materials', []):
        zone, row_n = _resolve_material_position(mat.get('position', 'carton'), found)
        mat['position'] = zone
        mat['_sort_row'] = row_n
    if params.get('materials'):
        params['materials'].sort(key=lambda m: (m['position'], m.get('_sort_row', 0)))

    # 分离内联物料（<carton_cost且有_sort_row）和zone物料
    cc_row = found.get('carton_cost', 0)
    main_end = found.get('main_table_end', ws.max_row)
    name_overrides = name_overrides or {}
    if clone_rows:
        # 用户选择的行（保持排序）
        all_rows = sorted(set(clone_rows) & set(range(1, main_end + 1)))
    else:
        all_rows = list(range(1, main_end + 1))
    inline_mats = {}
    zone_mats = []
    for mat in params.get('materials', []):
        sr = mat.get('_sort_row', 0)
        if sr > 0 and sr in all_rows:
            inline_mats.setdefault(sr, []).append(mat)
        else:
            zone_mats.append(mat)

    # 预先计算 row_map（原始行 → 克隆表中的行），内联物料插入导致行号偏移
    row_map = {}
    tgt = 1
    for src_row in all_rows:
        row_map[src_row] = tgt
        tgt += 1
        tgt += len(inline_mats.get(src_row, []))

    # 汇率/乘数引用 — 使用 remapped 行号，避免内联物料插入导致引用错位
    mapped_exchange = row_map.get(exchange_row, exchange_row) if exchange_row else None
    rate_ref = f'$B${mapped_exchange}' if mapped_exchange else '7.08'
    if waste_row and markup_row:
        mapped_waste = row_map.get(waste_row, waste_row)
        mapped_markup = row_map.get(markup_row, markup_row)
        multiplier_ref = f'(1+$B${mapped_waste}+$B${mapped_markup})'
    else:
        multiplier_ref = '1.12'

    # 创建新sheet
    new_name = params['item_name'][:31]
    if new_name in wb.sheetnames:
        new_name = new_name[:28] + '_01'
    ws_target = wb.create_sheet(title=new_name)

    # 复制全部行（含内联物料插入+公式remap+公式覆盖）
    total_written, inline_material_rows, _clone_row_map = _copy_header_rows(
        ws, ws_target, all_rows, max(data_col_end, ws.max_column or 1),
        inline_materials=inline_mats,
        data_col=data_col, data_col_end=data_col_end,
        multiplier_ref=multiplier_ref, rate_ref=rate_ref,
        formula_overrides=formula_overrides, pcs=pcs,
        name_overrides=name_overrides,
    )

    # 替换 PRODUCT ITEM#（源表第2行B列 → 用 row_map 映射目标行）
    ws_target.cell(row=row_map.get(2, 2), column=2).value = params.get('item_name', '')

    # 兜底：data列公式/数值cell如果仍为General格式，设为 $#,##0.000
    for r in range(1, total_written + 1):
        for c in range(data_col, data_col_end + 1):
            cell = ws_target.cell(row=r, column=c)
            if cell.number_format == 'General' and cell.value is not None:
                if isinstance(cell.value, (int, float)) or (
                        isinstance(cell.value, str) and cell.value.startswith('=')):
                    cell.number_format = '$#,##0.000'

    # 提取主表pcs数（从Carton Cost行A列文字提取）
    main_pcs = 0
    carton_text = ws.cell(row=cc_row, column=1).value or '' if cc_row else ''
    pcs_match = re.search(r'(\d+)\s*pcs', carton_text, re.IGNORECASE)
    if pcs_match:
        main_pcs = int(pcs_match.group(1))

    # ===== 覆盖参数 =====
    actual_desc_col = data_col_end + 1 if data_col_end > data_col else data_col + 1

    # Carton Cost 行
    if cc_row and cc_row in row_map:
        r = row_map[cc_row]
        # A列: 替换pcs数
        a_val = ws_target.cell(row=r, column=1).value or ''
        if main_pcs > 0 and isinstance(a_val, str):
            ws_target.cell(row=r, column=1).value = a_val.replace(str(main_pcs), str(pcs))
        ws_target.cell(row=r, column=data_col).value = (
            f'={round(params["carton_cost_rmb"], 3)}*{multiplier_ref}/{rate_ref}/{pcs}')
        _fill_chain_columns(ws_target, r, data_col, data_col_end)
        # 备注
        if params.get('carton_desc'):
            ws_target.cell(row=r, column=actual_desc_col).value = f'纸箱尺寸：{params["carton_desc"]}'

    # Carton Tape Cost 行
    tape_row = found.get('tape_cost')
    if tape_row and tape_row in row_map and not params.get('skip_rows', {}).get('tape'):
        r = row_map[tape_row]
        ws_target.cell(row=r, column=data_col).value = (
            f'={round(params["tape_cost_rmb"], 3)}*{multiplier_ref}/{rate_ref}/{pcs}')
        _fill_chain_columns(ws_target, r, data_col, data_col_end)

    # Packing Labor 行
    labor_row = found.get('packing_labor')
    if labor_row and labor_row in row_map and not params.get('skip_rows', {}).get('labor'):
        r = row_map[labor_row]
        if original_rate_row and is_multi_col:
            ws_target.cell(row=r, column=data_col).value = (
                f'={round(params["packing_labor_rmb"], 3)}*$B${row_map.get(original_rate_row, original_rate_row)}/{rate_ref}')
        else:
            ws_target.cell(row=r, column=data_col).value = (
                f'={round(params["packing_labor_rmb"], 3)}*{multiplier_ref}/{rate_ref}/{pcs}')
        _fill_chain_columns(ws_target, r, data_col, data_col_end)

    # zone物料写入（carton区物料插在Carton Cost下面）
    if zone_mats:
        # 需要在克隆表中找到carton_cost后面的空间插入
        # 但克隆表已经全部复制了，zone物料应该通过内联方式已经处理
        # 如果还有zone_mats（不在表头范围内的），追加到Ex-Factory之前
        pass  # zone物料在克隆模式下应通过内联物料方式处理

    # A列中的pcs数替换（Ex-Factory、FOB、Shipping等行）
    pcs_replace_keys = ['ex_factory', 'fob_40f', 'shipping_40f',
                        'fob_20f', 'shipping_20f', 'shipping_lcl',
                        'packing_labor']
    for key in pcs_replace_keys:
        src_row = found.get(key)
        if src_row and src_row in row_map:
            r = row_map[src_row]
            a_val = ws_target.cell(row=r, column=1).value
            if main_pcs > 0 and isinstance(a_val, str) and str(main_pcs) in a_val:
                ws_target.cell(row=r, column=1).value = a_val.replace(str(main_pcs), str(pcs))

    # FOB行的B列(CBM)替换
    L_cm = params.get('length_cm', 0)
    W_cm = params.get('width_cm', 0)
    H_cm = params.get('height_cm', 0)
    cbm_formula = f'={L_cm}*{W_cm}*{H_cm}/1000000' if (L_cm and W_cm and H_cm) else None
    if cbm_formula:
        for key in ['fob_40f', 'fob_20f', 'shipping_lcl']:
            src_row = found.get(key)
            if src_row and src_row in row_map:
                r = row_map[src_row]
                ws_target.cell(row=r, column=2).value = cbm_formula
                ws_target.cell(row=r, column=2).number_format = '0.000'

    # ===== 公式重建兜底（.xls转换后公式丢失的场景）=====
    col_letter = get_column_letter(data_col)

    def _is_formula(row_key):
        """检查目标sheet中某关键行的data_col是否为公式"""
        src = found.get(row_key)
        if not src or src not in row_map:
            return True  # 行不存在，视为不需要重建
        cell = ws_target.cell(row=row_map[src], column=data_col)
        return isinstance(cell.value, str) and str(cell.value).startswith('=')

    # 1) Product Price in Packaging: 如果不是公式，重建为上方行求和
    pp_src = found.get('product_price')
    if pp_src and pp_src in row_map and not _is_formula('product_price'):
        pp_tgt = row_map[pp_src]
        # 找第一个材料行：PP行上方、header区之下的第一个有值的行
        first_mat_tgt = None
        search_start = max(row_map.get(found.get('original_exchange_rate', 0), 0),
                           row_map.get(found.get('exchange_rate', 0), 0)) + 1
        if search_start <= 1:
            search_start = 8  # 安全默认：跳过前面的header行
        # 跳过标题行（如"Product Material Break Down"等纯文字行）
        # 从search_start扫描到PP行前，找第一个data_col有数值/公式的行
        for check_r in range(search_start, pp_tgt):
            v = ws_target.cell(row=check_r, column=data_col).value
            if v is not None:
                first_mat_tgt = check_r
                break
        if first_mat_tgt and first_mat_tgt < pp_tgt:
            ws_target.cell(row=pp_tgt, column=data_col).value = (
                f'=SUM({col_letter}{first_mat_tgt}:{col_letter}{pp_tgt - 1})')

    # 2) Ex-Factory: 如果不是公式，重建为 PP到EF-1行的求和
    ef_src = found.get('ex_factory')
    if ef_src and ef_src in row_map and not _is_formula('ex_factory'):
        ef_tgt = row_map[ef_src]
        if pp_src and pp_src in row_map:
            pp_tgt = row_map[pp_src]
            ws_target.cell(row=ef_tgt, column=data_col).value = (
                f'=SUM({col_letter}{pp_tgt}:{col_letter}{ef_tgt - 1})')

    # 3) FOB/Shipping: 如果不是公式，重建标准运费公式
    ef_tgt = row_map.get(ef_src) if ef_src else None
    if ef_tgt:
        fob_40f_src = found.get('fob_40f')
        if fob_40f_src and fob_40f_src in row_map and not _is_formula('fob_40f'):
            r = row_map[fob_40f_src]
            ws_target.cell(row=r, column=data_col).value = (
                f'=6500/((58/B{r}*{pcs}))/{rate_ref}')

        ship_40f_src = found.get('shipping_40f')
        if ship_40f_src and ship_40f_src in row_map and not _is_formula('shipping_40f'):
            r = row_map[ship_40f_src]
            fob_40f_tgt = row_map.get(fob_40f_src)
            if fob_40f_tgt:
                ws_target.cell(row=r, column=data_col).value = (
                    f'={col_letter}{ef_tgt}+{col_letter}{fob_40f_tgt}')

        fob_20f_src = found.get('fob_20f')
        if fob_20f_src and fob_20f_src in row_map and not _is_formula('fob_20f'):
            r = row_map[fob_20f_src]
            ws_target.cell(row=r, column=data_col).value = (
                f'=4000/((26/B{r}*{pcs}))/{rate_ref}')

        ship_20f_src = found.get('shipping_20f')
        if ship_20f_src and ship_20f_src in row_map and not _is_formula('shipping_20f'):
            r = row_map[ship_20f_src]
            fob_20f_tgt = row_map.get(fob_20f_src)
            if fob_20f_tgt:
                ws_target.cell(row=r, column=data_col).value = (
                    f'={col_letter}{ef_tgt}+{col_letter}{fob_20f_tgt}')

        lcl_src = found.get('shipping_lcl')
        if lcl_src and lcl_src in row_map and not _is_formula('shipping_lcl'):
            r = row_map[lcl_src]
            # LCL = CBM * 350 / 汇率 / pcs + Ex-Factory（用自身行的B列CBM）
            ws_target.cell(row=r, column=data_col).value = (
                f'=B{r}*350/{rate_ref}/{pcs}+{col_letter}{ef_tgt}')

    return wb


def _material_formula(mat, multiplier_ref, rate_ref, pcs):
    """根据物料类型生成对应的Excel公式"""
    rmb = round(mat.get('price_rmb', 0), 3)
    mat_type = mat.get('type', '')

    # 自定义物料：可选是否除加成和汇率
    if mat_type == 'custom':
        no_markup = mat.get('no_markup', False)
        no_rate = mat.get('no_rate', False)
        if no_markup and no_rate:
            return f'={rmb}'
        elif no_markup:
            return f'={rmb}/{rate_ref}'
        elif no_rate:
            return f'={rmb}*{multiplier_ref}'
        else:
            return f'={rmb}*{multiplier_ref}/{rate_ref}'

    # 平卡/邮包盒：除PCS
    if mat_type in ('card', 'mailbox'):
        return f'={rmb}*{multiplier_ref}/{rate_ref}/{pcs}'

    # 默认公式（护角、打包带、纸滑板等）
    return f'={rmb}*{multiplier_ref}/{rate_ref}'


def generate_subtable(wb, sheet_name, found, params):
    """
    在指定sheet追加次表

    参数:
        wb: openpyxl workbook
        sheet_name: 目标sheet名
        found: scan_sheet返回的关键行字典
        params: {
            'item_name': '15780-S002 6PCS/CTN',
            'pcs_per_carton': 6,
            'carton_desc': '46.5*28.5*32.5cm, K3K',
            'carton_cost_rmb': 2.846,     # 纸箱成本RMB/箱
            'tape_cost_rmb': 0.104,       # 胶纸RMB/箱
            'packing_labor_rmb': 0.188,   # 人工RMB/箱
            'cbm': 0.0431,               # 立方米
            'output_mode': 'same_sheet' / 'new_sheet' / 'new_file'
        }
    """
    ws = wb[sheet_name]
    data_col = found.get('data_col', 3)
    data_col_end = found.get('data_col_end', data_col)
    exchange_row = found.get('exchange_rate')
    pp_row = found.get('product_price')

    # new_sheet模式需要跨sheet引用主表单元格
    is_new_sheet = params.get('output_mode') == 'new_sheet'
    # Excel跨sheet引用前缀: '主表名'!
    sheet_prefix = f"'{sheet_name}'!" if is_new_sheet else ''

    # 汇率单元格引用
    rate_ref = f'{sheet_prefix}$B${exchange_row}' if exchange_row else '7.08'

    # Product Price单元格引用
    pp_col_letter = get_column_letter(data_col)
    pp_ref = f'{sheet_prefix}${pp_col_letter}${pp_row}' if pp_row else '0'

    # Waste% 和 Markup% 单元格引用
    waste_row = found.get('waste_pct')
    markup_row = found.get('markup_pct')
    # 构建乘数引用: (1 + Waste% + Markup%)
    if waste_row and markup_row:
        multiplier_ref = f'(1+{sheet_prefix}$B${waste_row}+{sheet_prefix}$B${markup_row})'
    else:
        multiplier_ref = '1.12'  # 默认 1 + 0.02 + 0.10

    pcs = params['pcs_per_carton']

    # 预处理物料位置: after_row_N → zone key + 原始行号
    for mat in params.get('materials', []):
        zone, row_n = _resolve_material_position(mat.get('position', 'carton'), found)
        mat['position'] = zone
        mat['_sort_row'] = row_n
    # 同zone内按原始行号排序
    if params.get('materials'):
        params['materials'].sort(key=lambda m: (m['position'], m.get('_sort_row', 0)))

    # 用于追踪内联到表头区的物料行号（供Ex-Factory公式引用）
    inline_material_rows = []

    header_row_map = {}  # 表头行映射: 源行号→目标行号（new_sheet模式下使用）

    if params.get('output_mode') == 'new_sheet':
        # 创建新sheet
        new_name = params['item_name'][:31]  # Excel sheet名最长31字符
        if new_name in wb.sheetnames:
            new_name = new_name[:28] + '_01'
        ws_target = wb.create_sheet(title=new_name)
        # 从主表复制用户选中的表头行
        header_row_nums = params.get('include_header_rows', [])
        if header_row_nums:
            # 分离内联物料（_sort_row在表头范围且<carton_cost）和zone物料
            sorted_headers = sorted(header_row_nums)
            cc_row = found.get('carton_cost', 0)
            inline_mats = {}  # {src_row: [mat, ...]}
            zone_mats = []
            for mat in params.get('materials', []):
                sr = mat.get('_sort_row', 0)
                if sr > 0 and (cc_row == 0 or sr < cc_row):
                    # 找到最近的 <= sr 的表头行
                    anchor = None
                    for h in sorted_headers:
                        if h <= sr:
                            anchor = h
                        else:
                            break
                    if anchor is not None:
                        inline_mats.setdefault(anchor, []).append(mat)
                    else:
                        zone_mats.append(mat)
                else:
                    zone_mats.append(mat)
            params['materials'] = zone_mats

            # 预计算row_map: 与_copy_header_rows Phase1逻辑一致
            pre_sorted = sorted(header_row_nums)
            pre_row_map = {}
            pre_tgt = 1
            for src in pre_sorted:
                pre_row_map[src] = pre_tgt
                pre_tgt += 1
                pre_tgt += len(inline_mats.get(src, []))

            # 如果表头包含汇率/waste/markup行，在调用前就切换为本sheet引用
            header_set = set(header_row_nums)
            if exchange_row and exchange_row in header_set:
                rate_ref = f'$B${pre_row_map.get(exchange_row, exchange_row)}'
            if waste_row and markup_row and waste_row in header_set and markup_row in header_set:
                mw = pre_row_map.get(waste_row, waste_row)
                mm = pre_row_map.get(markup_row, markup_row)
                multiplier_ref = f'(1+$B${mw}+$B${mm})'
            if pp_row and pp_row in header_set:
                pp_ref = f'${pp_col_letter}${pre_row_map.get(pp_row, pp_row)}'
            if exchange_row and exchange_row in header_set:
                sheet_prefix = ''  # 关键行在本sheet，不需要跨sheet前缀

            total_written, inline_material_rows, header_row_map = _copy_header_rows(
                ws, ws_target, header_row_nums, data_col_end,
                inline_materials=inline_mats,
                data_col=data_col, data_col_end=data_col_end,
                multiplier_ref=multiplier_ref, rate_ref=rate_ref,
                pcs=pcs,
            )
            start_row = total_written + 1  # 紧接表头+内联物料
        else:
            # 没有选中任何表头行，退化为旧行为
            _write_subtable_header(ws_target, params, sheet_name, exchange_row)
            start_row = 3
    else:
        ws_target = ws
        start_row = find_append_position(ws)

    # === 写入次表内容 ===
    col_a = 1  # A列
    col_b = 2  # B列
    desc_col = data_col + 1  # 备注列
    actual_desc_col = data_col_end + 1 if data_col_end > data_col else data_col + 1  # 实际备注列

    # 标题样式
    title_fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
    title_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    value_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    r = start_row

    # 标题行: new_sheet模式有header时跳过（header已提供上下文）
    has_headers = is_new_sheet and params.get('include_header_rows')
    if not has_headers:
        title_cell = ws_target.cell(row=r, column=col_a)
        title_cell.value = params['item_name']
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal='center')
        merge_end = data_col_end if data_col_end > data_col else data_col
        for c in range(col_a, merge_end + 1):
            ws_target.cell(row=r, column=c).fill = title_fill
            ws_target.cell(row=r, column=c).border = thin_border
        ws_target.merge_cells(
            start_row=r, start_column=col_a,
            end_row=r, end_column=merge_end
        )
        r += 1

    # === Product区物料行（写在Title和Carton Cost之间）===
    product_material_rows = []
    prod_mat_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
    for mat in params.get('materials', []):
        if mat.get('position') != 'product':
            continue
        mat_desc = mat.get('desc', '')
        mat_rmb = mat.get('price_rmb', 0)
        if not mat_desc or mat_rmb <= 0:
            continue

        ws_target.cell(row=r, column=col_a).value = mat_desc
        ws_target.cell(row=r, column=col_a).fill = prod_mat_fill
        ws_target.cell(row=r, column=col_a).border = thin_border
        ws_target.cell(row=r, column=col_b).value = '－'
        ws_target.cell(row=r, column=col_b).border = thin_border
        # 物料是per piece价格: =RMB*(1+Waste%+Markup%)/汇率
        ws_target.cell(row=r, column=data_col).value = f'={round(mat_rmb, 3)}*{multiplier_ref}/{rate_ref}'
        ws_target.cell(row=r, column=data_col).number_format = '$#,##0.000'
        ws_target.cell(row=r, column=data_col).fill = prod_mat_fill
        ws_target.cell(row=r, column=data_col).border = thin_border
        _fill_chain_columns(ws_target, r, data_col, data_col_end)
        mat_remark = mat.get('remark', '')
        if mat_remark:
            ws_target.cell(row=r, column=actual_desc_col).value = mat_remark
        product_material_rows.append(r)
        r += 1

    # Carton Cost 行
    ws_target.cell(row=r, column=col_a).value = f'Carton Cost ({pcs}pcs)'
    ws_target.cell(row=r, column=col_a).border = thin_border
    ws_target.cell(row=r, column=col_b).value = '－'
    ws_target.cell(row=r, column=col_b).border = thin_border
    # Carton Cost/pcs USD = RMB * (1+Waste%+Markup%) / 汇率 / pcs
    ws_target.cell(row=r, column=data_col).value = f'={round(params["carton_cost_rmb"], 3)}*{multiplier_ref}/{rate_ref}/{pcs}'
    ws_target.cell(row=r, column=data_col).number_format = '$#,##0.000'
    ws_target.cell(row=r, column=data_col).border = thin_border
    _fill_chain_columns(ws_target, r, data_col, data_col_end)
    # 备注栏写纸箱描述
    if params.get('carton_desc'):
        ws_target.cell(row=r, column=actual_desc_col).value = f"纸箱尺寸：{params['carton_desc']}"
    carton_row = r
    r += 1

    # === 附加物料行（Carton区，插在 Carton Cost 和 Tape 之间）===
    material_rows = []
    mat_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    for mat in params.get('materials', []):
        # 只处理carton区物料，其他zone由各自的_write_materials_at处理
        if mat.get('position') != 'carton':
            continue
        mat_desc = mat.get('desc', '')
        mat_rmb = mat.get('price_rmb', 0)
        if not mat_desc or mat_rmb <= 0:
            continue

        ws_target.cell(row=r, column=col_a).value = mat_desc
        ws_target.cell(row=r, column=col_a).fill = mat_fill
        ws_target.cell(row=r, column=col_a).border = thin_border
        ws_target.cell(row=r, column=col_b).value = '－'
        ws_target.cell(row=r, column=col_b).border = thin_border
        # 物料公式：根据type决定是否除PCS、是否去掉加成/汇率
        ws_target.cell(row=r, column=data_col).value = _material_formula(mat, multiplier_ref, rate_ref, pcs)
        ws_target.cell(row=r, column=data_col).number_format = '$#,##0.000'
        ws_target.cell(row=r, column=data_col).fill = mat_fill
        ws_target.cell(row=r, column=data_col).border = thin_border
        _fill_chain_columns(ws_target, r, data_col, data_col_end)
        mat_remark = mat.get('remark', '')
        if mat_remark:
            ws_target.cell(row=r, column=actual_desc_col).value = mat_remark
        material_rows.append(r)
        r += 1

    # 跳过行配置
    skip = params.get('skip_rows', {})
    # 前半段行跳过：主表PP到CC之间需要跳过的行号
    mid_row_remove = set(skip.get('mid_row_indices', []))

    # 辅助函数：在指定位置写入物料行，返回写入的行号列表
    def _write_materials_at(position_key, fill_color='E8F5E9'):
        nonlocal r
        rows_written = []
        mat_f = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        for mat in params.get('materials', []):
            if mat.get('position') != position_key:
                continue
            mat_desc = mat.get('desc', '')
            mat_rmb = mat.get('price_rmb', 0)
            if not mat_desc or mat_rmb <= 0:
                continue
            ws_target.cell(row=r, column=col_a).value = mat_desc
            ws_target.cell(row=r, column=col_a).fill = mat_f
            ws_target.cell(row=r, column=col_a).border = thin_border
            ws_target.cell(row=r, column=col_b).value = '－'
            ws_target.cell(row=r, column=col_b).border = thin_border
            ws_target.cell(row=r, column=data_col).value = _material_formula(mat, multiplier_ref, rate_ref, pcs)
            ws_target.cell(row=r, column=data_col).number_format = '$#,##0.000'
            ws_target.cell(row=r, column=data_col).fill = mat_f
            ws_target.cell(row=r, column=data_col).border = thin_border
            _fill_chain_columns(ws_target, r, data_col, data_col_end)
            mat_remark = mat.get('remark', '')
            if mat_remark:
                ws_target.cell(row=r, column=actual_desc_col).value = mat_remark
            rows_written.append(r)
            r += 1
        return rows_written

    # Carton Tape Cost 行
    tape_row = None
    if not skip.get('tape'):
        ws_target.cell(row=r, column=col_a).value = 'Carton Tape Cost'
        ws_target.cell(row=r, column=col_a).border = thin_border
        ws_target.cell(row=r, column=col_b).value = '－'
        ws_target.cell(row=r, column=col_b).border = thin_border
        ws_target.cell(row=r, column=data_col).value = f'={round(params["tape_cost_rmb"], 3)}*{multiplier_ref}/{rate_ref}/{pcs}'
        ws_target.cell(row=r, column=data_col).number_format = '$#,##0.000'
        ws_target.cell(row=r, column=data_col).border = thin_border
        _fill_chain_columns(ws_target, r, data_col, data_col_end)
        tape_row = r
        r += 1

    # === after_tape 位置物料 ===
    after_tape_rows = _write_materials_at('after_tape', 'E3F2FD')

    # Carton Packing labor 行
    labor_row = None
    if not skip.get('labor'):
        ws_target.cell(row=r, column=col_a).value = f'Carton Packing labor({pcs}pcs)'
        ws_target.cell(row=r, column=col_a).border = thin_border
        ws_target.cell(row=r, column=col_b).value = '－'
        ws_target.cell(row=r, column=col_b).border = thin_border
        # Labor公式：检测是否有Original Exchange Rate行 且 是多列结构
        original_rate_row = found.get('original_exchange_rate')
        is_multi_col = data_col_end > data_col
        if original_rate_row and is_multi_col:
            mapped_orig = header_row_map.get(original_rate_row, original_rate_row)
            labor_formula = f'={round(params["packing_labor_rmb"], 3)}*{sheet_prefix}$B${mapped_orig}/{rate_ref}'
        else:
            labor_formula = f'={round(params["packing_labor_rmb"], 3)}*{multiplier_ref}/{rate_ref}/{pcs}'
        ws_target.cell(row=r, column=data_col).value = labor_formula
        ws_target.cell(row=r, column=data_col).number_format = '$#,##0.000'
        ws_target.cell(row=r, column=data_col).border = thin_border
        _fill_chain_columns(ws_target, r, data_col, data_col_end)
        labor_row = r
        r += 1

    # === after_labor 位置物料 ===
    after_labor_rows = _write_materials_at('after_labor', 'E3F2FD')

    # === Ex-Factory 到 LCL: 从主表复制公式 + 替换行号 ===
    col_letter = pp_col_letter

    # 收集主表中 Ex-Factory 到 LCL 的所有行
    copy_keys = ['ex_factory', 'fob_40f', 'shipping_40f',
                 'fob_20f', 'shipping_20f', 'shipping_lcl']
    main_rows = _collect_main_formulas(ws, found)

    # 提取主表的pcs数（从Carton Cost行的A列文字中提取）
    main_pcs = 0
    carton_text = ws.cell(row=found.get('carton_cost', 1), column=1).value or ''
    pcs_match = re.search(r'(\d+)\s*pcs', carton_text, re.IGNORECASE)
    if pcs_match:
        main_pcs = int(pcs_match.group(1))

    # 构建行号映射: 主表行号 → 次表行号
    row_map = {}
    # Product Price 行映射（保持不变，因为引用主表的PP）
    if pp_row:
        row_map[pp_row] = pp_row  # PP行不动，保留绝对引用

    # 先预分配次表行号给所有需要复制的行
    # Ex-Factory 行（FOB/Shipping的映射延迟到after_exfactory物料写入后）
    ex_factory_row = r
    row_map[found.get('ex_factory', 0)] = r

    # 也把 carton/tape/labor 映射加入（因为Ex-Factory公式可能引用它们，跳过的行不映射）
    if found.get('carton_cost'):
        row_map[found['carton_cost']] = carton_row
    if found.get('tape_cost') and tape_row:
        row_map[found['tape_cost']] = tape_row
    if found.get('packing_labor') and labor_row:
        row_map[found['packing_labor']] = labor_row

    # CBM公式
    L_cm = params.get('length_cm', 0)
    W_cm = params.get('width_cm', 0)
    H_cm = params.get('height_cm', 0)
    cbm_formula = f'={L_cm}*{W_cm}*{H_cm}/1000000' if (L_cm and W_cm and H_cm) else params.get('cbm', 0)

    # 写入 Ex-Factory 行
    ws_target.cell(row=r, column=col_a).value = f'Ex-Factory {pcs}pcs/ctn\n(Factroy provides Shipping Documents)'
    ws_target.cell(row=r, column=col_a).font = Font(bold=True)
    ws_target.cell(row=r, column=col_a).alignment = Alignment(wrap_text=True)
    ws_target.cell(row=r, column=col_a).border = thin_border
    ws_target.cell(row=r, column=col_b).value = '－'
    ws_target.cell(row=r, column=col_b).border = thin_border
    # Ex-Factory公式: same_sheet优先从主表复制+替换行号，new_sheet或公式丢失时生成新公式
    src_ex = found.get('ex_factory')
    src_formula = ws.cell(row=src_ex, column=data_col).value if src_ex else None
    if not is_new_sheet and src_formula and isinstance(src_formula, str) and src_formula.startswith('='):
        ex_formula = _remap_formula(src_formula, row_map, main_pcs, pcs, remove_rows=mid_row_remove)
    else:
        # 公式丢失（如.xls转换）或Ex-Factory行不存在 → 生成新公式
        ex_parts = [pp_ref]
        # inline_material_rows 已在PP的SUM范围内，不再重复添加
        for mr in product_material_rows:
            ex_parts.append(f'{col_letter}{mr}')
        ex_parts.append(f'{col_letter}{carton_row}')
        if tape_row: ex_parts.append(f'{col_letter}{tape_row}')
        for mr in after_tape_rows:
            ex_parts.append(f'{col_letter}{mr}')
        if labor_row: ex_parts.append(f'{col_letter}{labor_row}')
        for mr in after_labor_rows:
            ex_parts.append(f'{col_letter}{mr}')
        for mr in material_rows:
            ex_parts.append(f'{col_letter}{mr}')
        ex_formula = '=' + '+'.join(ex_parts)
    ws_target.cell(row=r, column=data_col).value = ex_formula
    ws_target.cell(row=r, column=data_col).number_format = '$#,##0.000'
    ws_target.cell(row=r, column=data_col).fill = value_fill
    ws_target.cell(row=r, column=data_col).font = Font(bold=True)
    ws_target.cell(row=r, column=data_col).border = thin_border
    # 多列: Ex-Factory是汇总行，每列从主表对应列复制公式+替换行号
    for col in range(data_col + 1, data_col_end + 1):
        src_f = ws.cell(row=src_ex, column=col).value if src_ex else None
        if not is_new_sheet and src_f and isinstance(src_f, str) and src_f.startswith('='):
            ws_target.cell(row=r, column=col).value = _remap_formula(src_f, row_map, main_pcs, pcs, remove_rows=mid_row_remove)
        else:
            # new_sheet模式或公式丢失 → 生成新公式（pp_ref已包含跨sheet前缀）
            cl = get_column_letter(col)
            mapped_pp = header_row_map.get(pp_row, pp_row) if pp_row else None
            pp_cross = f"{sheet_prefix}{cl}${mapped_pp}" if mapped_pp else '0'
            ex_parts_c = [pp_cross]
            # inline_material_rows 已在PP的SUM范围内，不再重复添加
            for mr in product_material_rows:
                ex_parts_c.append(f'{cl}{mr}')
            ex_parts_c.append(f'{cl}{carton_row}')
            if tape_row: ex_parts_c.append(f'{cl}{tape_row}')
            for mr in after_tape_rows:
                ex_parts_c.append(f'{cl}{mr}')
            if labor_row: ex_parts_c.append(f'{cl}{labor_row}')
            for mr in after_labor_rows:
                ex_parts_c.append(f'{cl}{mr}')
            for mr in material_rows:
                ex_parts_c.append(f'{cl}{mr}')
            ws_target.cell(row=r, column=col).value = '=' + '+'.join(ex_parts_c)
        ws_target.cell(row=r, column=col).number_format = '$#,##0.000'
        ws_target.cell(row=r, column=col).fill = value_fill
        ws_target.cell(row=r, column=col).font = Font(bold=True)
    r += 1

    # === after_exfactory 位置物料（不参与Ex-Factory公式） ===
    after_exfactory_rows = _write_materials_at('after_exfactory', 'FCE4EC')

    # FOB/Shipping 行号映射（在after_exfactory物料后预分配，确保行号正确）
    next_r = r
    for key in ['fob_40f', 'shipping_40f', 'fob_20f', 'shipping_20f', 'shipping_lcl']:
        src_row = found.get(key, 0)
        if src_row and not skip.get(key):
            row_map[src_row] = next_r
            next_r += 1

    # 写入 FOB / Shipping 行: 优先从主表复制公式，公式丢失时生成标准运费公式
    # 标准运费常量（行业标准值，所有Fuggler/Rainbocorn文件一致）
    FREIGHT_40F_COST = 6500   # 40尺柜费用
    FREIGHT_40F_DIM = 58      # 40尺柜内高(cm)
    FREIGHT_20F_COST = 4000   # 20尺柜费用
    FREIGHT_20F_DIM = 26      # 20尺柜内高(cm)
    FREIGHT_LCL_RATE = 350    # LCL散货费率(per CBM)

    # 跟踪已写入的行号，用于 Shipping = ExFactory + FOB 引用
    written_rows = {'ex_factory': ex_factory_row}  # ex_factory_row 是上面写入的行

    for key in ['fob_40f', 'shipping_40f', 'fob_20f', 'shipping_20f', 'shipping_lcl']:
        src_row = found.get(key)
        if not src_row:
            continue
        # 检查是否跳过该行
        if skip.get(key):
            continue

        new_row = row_map.get(src_row, r)
        written_rows[key] = r

        # A列: 从主表复制文字，替换pcs数
        src_a = ws.cell(row=src_row, column=1).value or ''
        if main_pcs > 0:
            new_a = src_a.replace(str(main_pcs), str(pcs))
        else:
            new_a = src_a
        ws_target.cell(row=r, column=col_a).value = new_a
        ws_target.cell(row=r, column=col_a).border = thin_border

        # B列: 从主表复制公式/值+替换行号
        src_b = ws.cell(row=src_row, column=2).value
        if src_b is not None:
            if src_b == '－' or src_b == '-':
                ws_target.cell(row=r, column=col_b).value = '－'
            elif key == 'fob_40f':
                ws_target.cell(row=r, column=col_b).value = cbm_formula
            elif isinstance(src_b, str) and src_b.startswith('='):
                ws_target.cell(row=r, column=col_b).value = _remap_formula(src_b, row_map, main_pcs, pcs)
            else:
                if key in ('fob_20f', 'shipping_lcl') and isinstance(src_b, (int, float)):
                    ws_target.cell(row=r, column=col_b).value = cbm_formula
                else:
                    ws_target.cell(row=r, column=col_b).value = src_b
        ws_target.cell(row=r, column=col_b).number_format = '0.000'
        ws_target.cell(row=r, column=col_b).border = thin_border

        # C列: same_sheet优先复制公式+替换行号，new_sheet或公式丢失时生成标准运费公式
        src_c = ws.cell(row=src_row, column=data_col).value
        if not is_new_sheet and src_c and isinstance(src_c, str) and src_c.startswith('='):
            # same_sheet + .xlsx文件：公式完整，正常remap
            ws_target.cell(row=r, column=data_col).value = _remap_formula(src_c, row_map, main_pcs, pcs)
        else:
            # 公式丢失（.xls转换）→ 生成标准运费公式
            ex_r = written_rows.get('ex_factory', r - 1)
            b_ref = f'$B${r}'
            if key == 'fob_40f':
                # 40F运费 = 6500 / (58 / CBM * pcs) / 汇率
                ws_target.cell(row=r, column=data_col).value = f'=({FREIGHT_40F_COST}/(({FREIGHT_40F_DIM}/{b_ref}*{pcs}))/{rate_ref})'
            elif key == 'shipping_40f':
                # Shipping 40F = Ex-Factory + FOB 40F freight
                fob40_r = written_rows.get('fob_40f', r - 1)
                ws_target.cell(row=r, column=data_col).value = f'={col_letter}{ex_r}+{col_letter}{fob40_r}'
            elif key == 'fob_20f':
                # 20F运费 = 4000 / (26 / CBM * pcs) / 汇率
                ws_target.cell(row=r, column=data_col).value = f'=({FREIGHT_20F_COST}/(({FREIGHT_20F_DIM}/{b_ref}*{pcs}))/{rate_ref})'
            elif key == 'shipping_20f':
                # Shipping 20F = Ex-Factory + FOB 20F freight
                fob20_r = written_rows.get('fob_20f', r - 1)
                ws_target.cell(row=r, column=data_col).value = f'={col_letter}{ex_r}+{col_letter}{fob20_r}'
            elif key == 'shipping_lcl':
                # LCL = CBM * 350 / 汇率 / pcs + Ex-Factory
                ws_target.cell(row=r, column=data_col).value = f'={b_ref}*{FREIGHT_LCL_RATE}/{rate_ref}/{pcs}+{col_letter}{ex_r}'
        ws_target.cell(row=r, column=data_col).number_format = '$#,##0.000'
        ws_target.cell(row=r, column=data_col).border = thin_border

        # 多列: same_sheet优先复制公式，new_sheet或公式丢失时用链式传递
        for col in range(data_col + 1, data_col_end + 1):
            src_val = ws.cell(row=src_row, column=col).value
            if not is_new_sheet and src_val and isinstance(src_val, str) and src_val.startswith('='):
                ws_target.cell(row=r, column=col).value = _remap_formula(src_val, row_map, main_pcs, pcs)
            else:
                if key in ('shipping_40f', 'shipping_20f', 'shipping_lcl'):
                    # Shipping行: 用同列的Ex-Factory + 同列的FOB
                    cl = get_column_letter(col)
                    if key == 'shipping_40f':
                        fob_r = written_rows.get('fob_40f', r - 1)
                    elif key == 'shipping_20f':
                        fob_r = written_rows.get('fob_20f', r - 1)
                    else:
                        fob_r = r  # LCL自身就是freight+ExFactory
                    if key == 'shipping_lcl':
                        ws_target.cell(row=r, column=col).value = f'={b_ref}*{FREIGHT_LCL_RATE}/{rate_ref}/{pcs}+{cl}{ex_r}'
                    else:
                        ws_target.cell(row=r, column=col).value = f'={cl}{ex_r}+{cl}{fob_r}'
                else:
                    # FOB freight行: 链式传递（各列CBM相同，运费相同）
                    prev = get_column_letter(col - 1)
                    ws_target.cell(row=r, column=col).value = f'={prev}{r}'
            ws_target.cell(row=r, column=col).number_format = '$#,##0.000'

        # 样式: Shipping行加粗+黄底, FOB行普通
        if key.startswith('shipping'):
            ws_target.cell(row=r, column=col_a).font = Font(bold=True)
            for sc in range(data_col, data_col_end + 1):
                ws_target.cell(row=r, column=sc).fill = value_fill
                ws_target.cell(row=r, column=sc).font = Font(bold=True)

        r += 1

    return wb


def _copy_header_rows(ws_source, ws_target, row_nums, max_col,
                      inline_materials=None, data_col=3, data_col_end=3,
                      multiplier_ref='1.12', rate_ref='7.08',
                      formula_overrides=None, pcs=1, name_overrides=None):
    """
    从主表复制指定行到新sheet，保留格式。
    可在指定行后插入物料行。

    row_nums: 要复制的主表行号列表 [1, 2, 3, 5, 6]
    inline_materials: {src_row: [mat_dict, ...]} 在该src_row后插入物料
    返回: (total_rows_written, [插入的物料行在目标sheet中的行号])
    """
    inline_materials = inline_materials or {}
    formula_overrides = formula_overrides or {}
    name_overrides = name_overrides or {}
    inserted_material_rows = []

    mat_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    actual_desc_col = data_col_end + 1 if data_col_end > data_col else data_col + 1

    # Phase 1: 先计算 row_map（源行→目标行），考虑内联插入导致的偏移
    sorted_rows = sorted(row_nums)
    row_map = {}  # {源行号: 目标行号}
    target_row = 1
    for src_row in sorted_rows:
        row_map[src_row] = target_row
        target_row += 1
        target_row += len(inline_materials.get(src_row, []))
    has_offset = any(src != tgt for src, tgt in row_map.items())

    # Phase 2: 复制行 + 用 row_map 重写公式
    target_row = 1
    for src_row in sorted_rows:
        for col in range(1, max_col + 2):  # 多读1列用于备注
            src_cell = ws_source.cell(row=src_row, column=col)
            value = src_cell.value
            # 名称覆盖：如果用户修改了A列名称
            if col == 1 and src_row in name_overrides:
                value = name_overrides[src_row]
            # 公式覆盖：如果用户在前端编辑了该行data_col的公式，使用覆盖值
            if col == data_col and src_row in formula_overrides:
                value = formula_overrides[src_row]
            # 如果有偏移且是公式，用row_map重写行号引用
            if has_offset and isinstance(value, str) and value.startswith('='):
                value = _remap_formula(value, row_map, 0, 0)
            tgt_cell = ws_target.cell(row=target_row, column=col)
            tgt_cell.value = value
            # 复制格式（始终复制，不依赖 has_style — 行/列级继承样式时 has_style 可能为 False）
            tgt_cell.font = copy.copy(src_cell.font)
            tgt_cell.fill = copy.copy(src_cell.fill)
            tgt_cell.alignment = copy.copy(src_cell.alignment)
            tgt_cell.border = copy.copy(src_cell.border)
            tgt_cell.number_format = src_cell.number_format
        target_row += 1

        # 在该行后插入物料
        for mat in inline_materials.get(src_row, []):
            ws_target.cell(row=target_row, column=1).value = mat.get('desc', '')
            ws_target.cell(row=target_row, column=1).fill = mat_fill
            ws_target.cell(row=target_row, column=1).border = thin_border
            ws_target.cell(row=target_row, column=2).value = '－'
            ws_target.cell(row=target_row, column=2).border = thin_border
            ws_target.cell(row=target_row, column=data_col).value = (
                _material_formula(mat, multiplier_ref, rate_ref, pcs))
            ws_target.cell(row=target_row, column=data_col).number_format = '$#,##0.000'
            ws_target.cell(row=target_row, column=data_col).fill = mat_fill
            ws_target.cell(row=target_row, column=data_col).border = thin_border
            _fill_chain_columns(ws_target, target_row, data_col, data_col_end)
            remark = mat.get('remark', '')
            if remark:
                ws_target.cell(row=target_row, column=actual_desc_col).value = remark
            inserted_material_rows.append(target_row)
            target_row += 1

    # 复制行高（用 row_map 映射源行→目标行）
    for src_row, tgt_row in row_map.items():
        if src_row in ws_source.row_dimensions:
            src_rd = ws_source.row_dimensions[src_row]
            if src_rd.height is not None:
                ws_target.row_dimensions[tgt_row].height = src_rd.height

    # 复制合并单元格（只复制涉及已复制行的合并区域，用 row_map 映射行号）
    src_row_set = set(row_map.keys())
    for merged_range in ws_source.merged_cells.ranges:
        min_r, max_r = merged_range.min_row, merged_range.max_row
        if min_r in src_row_set and max_r in src_row_set:
            ws_target.merge_cells(
                start_row=row_map[min_r], start_column=merged_range.min_col,
                end_row=row_map[max_r], end_column=merged_range.max_col
            )

    # 复制列宽
    for col in range(1, max_col + 2):
        col_letter = get_column_letter(col)
        if col_letter in ws_source.column_dimensions:
            ws_target.column_dimensions[col_letter].width = ws_source.column_dimensions[col_letter].width

    return target_row - 1, inserted_material_rows, row_map


def _write_subtable_header(ws, params, source_sheet, exchange_row):
    """为新sheet写入头部信息"""
    ws.cell(row=1, column=1).value = 'CUSTOMER :'
    ws.cell(row=1, column=2).value = 'ZURU'
    ws.cell(row=2, column=1).value = 'PRODUCT  ITEM#: '
    ws.cell(row=2, column=2).value = params.get('item_name', '')


def get_sheet_info(wb):
    """获取workbook中所有sheet的基本信息（兼容read_only模式）"""
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        if wb.read_only:
            # read_only模式下用迭代计数（只扫前几行估算即可）
            row_count = 0
            col_count = 0
            for row in ws.iter_rows(min_col=1, max_col=1):
                row_count += 1
            # 列数从第1行取
            for row in ws.iter_rows(min_row=1, max_row=1):
                col_count = len(row)
            sheets.append({'name': name, 'rows': row_count, 'cols': col_count})
        else:
            sheets.append({'name': name, 'rows': ws.max_row, 'cols': ws.max_column})
    return sheets


def process_upload(filepath, sheet_name, params_list, output_mode='same_sheet'):
    """
    完整处理流程：加载 → 扫描 → 生成次表(多组) → 保存

    params_list: 单个dict或list[dict]，向下兼容
    返回: (output_filepath, scan_result)
    """
    # 兼容单个dict
    if isinstance(params_list, dict):
        params_list = [params_list]

    wb, actual_path, converted = load_excel(filepath)

    ws = wb[sheet_name]
    found = scan_sheet(ws)

    if output_mode == 'new_file':
        # 复制原sheet到新workbook
        new_wb = Workbook()
        new_wb.remove(new_wb.active)
        source_ws = wb[sheet_name]
        new_ws = new_wb.create_sheet(title=sheet_name)

        # 复制所有数据和格式
        for row in source_ws.iter_rows():
            for cell in row:
                new_cell = new_ws.cell(row=cell.row, column=cell.column)
                new_cell.value = cell.value
                new_cell.font = copy.copy(cell.font)
                new_cell.fill = copy.copy(cell.fill)
                new_cell.alignment = copy.copy(cell.alignment)
                new_cell.border = copy.copy(cell.border)
                new_cell.number_format = cell.number_format

        # 复制列宽
        for col_letter, dim in source_ws.column_dimensions.items():
            new_ws.column_dimensions[col_letter].width = dim.width

        wb = new_wb
        actual_output_mode = 'same_sheet'
    else:
        actual_output_mode = output_mode

    # 循环生成每组次表
    for params in params_list:
        params['output_mode'] = actual_output_mode
        generate_subtable(wb, sheet_name, found, params)
        # 重新扫描以获取更新后的 main_table_end
        ws = wb[sheet_name]
        found = scan_sheet(ws)

    # 保存
    base_name = os.path.splitext(os.path.basename(filepath))[0]
    if converted:
        base_name = base_name.replace('_converted', '')
    output_name = f"{base_name}_次表.xlsx"
    output_dir = os.path.dirname(filepath)
    output_path = os.path.join(output_dir, output_name)
    wb.save(output_path)
    _com_resave(output_path)

    return output_path, found


def process_upload_multi(filepath, sheets_list, output_mode='same_sheet', include_header_rows=None, formula_overrides=None,
                         clone_rows=None, name_overrides=None):
    """
    多Sheet批量生成：加载文件一次，对每个Sheet分别扫描+生成次表，保存一次。

    sheets_list: [{ sheet_name: str, params: list[dict] }]
    include_header_rows: new_sheet模式下要复制的主表行号列表
    返回: output_filepath
    """
    wb, actual_path, converted = load_excel(filepath)

    # new_file模式：创建含目标sheet及其引用sheet的新workbook
    if output_mode == 'new_file':
        target_sheets = {e['sheet_name'] for e in sheets_list}

        # 扫描目标sheet公式中引用的其他sheet，一并复制（防止#REF!）
        ref_sheets = set()
        all_sheet_set = set(wb.sheetnames)
        for sn in target_sheets:
            source_ws = wb[sn]
            for row in source_ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        # 匹配带引号的跨sheet引用: 'sheet名'!
                        for m in re.finditer(r"'([^']+)'!", cell.value):
                            ref_name = m.group(1)
                            if ref_name in all_sheet_set and ref_name not in target_sheets:
                                ref_sheets.add(ref_name)
                        # 匹配不带引号的跨sheet引用: sheet名!（如中文sheet名）
                        for m in re.finditer(r'(?<!\')(\w+)!', cell.value):
                            ref_name = m.group(1)
                            if ref_name in all_sheet_set and ref_name not in target_sheets:
                                ref_sheets.add(ref_name)
        all_sheets_to_copy = list(target_sheets) + sorted(ref_sheets)
        if ref_sheets:
            logger.info(f'new_file模式: 额外复制被引用的sheet: {ref_sheets}')

        new_wb = Workbook()
        new_wb.remove(new_wb.active)
        for sn in all_sheets_to_copy:
            source_ws = wb[sn]
            new_ws = new_wb.create_sheet(title=sn)
            # 复制数据和格式
            for row in source_ws.iter_rows():
                for cell in row:
                    new_cell = new_ws.cell(row=cell.row, column=cell.column)
                    new_cell.value = cell.value
                    new_cell.font = copy.copy(cell.font)
                    new_cell.fill = copy.copy(cell.fill)
                    new_cell.alignment = copy.copy(cell.alignment)
                    new_cell.border = copy.copy(cell.border)
                    new_cell.number_format = cell.number_format
            # 复制列宽
            for col_letter, dim in source_ws.column_dimensions.items():
                new_ws.column_dimensions[col_letter].width = dim.width
            # 复制行高
            for r in range(1, source_ws.max_row + 1):
                if source_ws.row_dimensions[r].height is not None:
                    new_ws.row_dimensions[r].height = source_ws.row_dimensions[r].height
            # 复制合并单元格
            for merged_range in source_ws.merged_cells.ranges:
                new_ws.merge_cells(str(merged_range))
        wb = new_wb
        actual_mode = 'same_sheet'  # new_file内部用same_sheet追加
    else:
        actual_mode = output_mode

    for sheet_entry in sheets_list:
        sheet_name = sheet_entry['sheet_name']
        params_list = sheet_entry.get('params', [])
        if isinstance(params_list, dict):
            params_list = [params_list]

        ws = wb[sheet_name]
        found = scan_sheet(ws)

        for params in params_list:
            params['output_mode'] = actual_mode
            if include_header_rows:
                params['include_header_rows'] = include_header_rows
            if actual_mode == 'clone_sheet':
                clone_sheet(wb, sheet_name, found, params, formula_overrides=formula_overrides,
                           clone_rows=clone_rows, name_overrides=name_overrides)
            else:
                generate_subtable(wb, sheet_name, found, params)
            # 重新扫描以获取更新后的追加位置
            ws = wb[sheet_name]
            found = scan_sheet(ws)

    # 保存
    base_name = os.path.splitext(os.path.basename(filepath))[0]
    if converted:
        base_name = base_name.replace('_converted', '')
    output_name = f"{base_name}_次表.xlsx"
    output_dir = os.path.dirname(filepath)
    output_path = os.path.join(output_dir, output_name)
    wb.save(output_path)
    _com_resave(output_path)

    return output_path


def extract_columns(filepath, sheets_config, item_name=None, formula_overrides=None):
    """
    提取多列sheet中的选中产品列。
    用COM打开文件，删除未选中的产品列，可选修改ITEM名称和公式参数，保存为新文件。

    sheets_config: [{
        'sheet_name': str,
        'keep_cols': [int, ...],       # 要保留的产品列号
        'all_product_cols': [int, ...], # 所有产品列号
    }]
    item_name: str or None - 修改R2的PRODUCT ITEM#名称
    formula_overrides: {row_num: {old_num: new_num, ...}, ...} - 公式中数字替换
    返回: output_filepath
    """
    import win32com.client
    import pythoncom
    pythoncom.CoInitialize()

    # 先用openpyxl加载（处理xls转换）
    _, actual_path, converted = load_excel(filepath)

    # 准备输出路径
    base_name = os.path.splitext(os.path.basename(filepath))[0]
    if converted:
        base_name = base_name.replace('_converted', '')
    output_name = f"{base_name}_提取.xlsx"
    output_dir = os.path.dirname(actual_path)
    output_path = os.path.join(output_dir, output_name)

    # 复制文件
    import shutil
    shutil.copy2(actual_path, output_path)

    # 用COM打开并删列
    app = None
    try:
        app = win32com.client.Dispatch('ket.Application')
        app.Visible = False
        app.DisplayAlerts = False

        wb = app.Workbooks.Open(os.path.abspath(output_path))

        for cfg in sheets_config:
            sheet_name = cfg['sheet_name']
            keep_cols = set(cfg['keep_cols'])
            all_product_cols = sorted(cfg['all_product_cols'])

            try:
                ws = wb.Sheets(sheet_name)
            except Exception:
                logger.warning(f'extract_columns: sheet "{sheet_name}" not found, skip')
                continue

            # 删列前：修复链式引用（如D列=E11引用B列，B被删后会#REF!）
            # 对每个保留列的每个单元格，如果公式引用了要删除的列，追踪链到保留列或根公式
            delete_cols = set(all_product_cols) - keep_cols
            max_row = ws.UsedRange.Rows.Count + ws.UsedRange.Row - 1
            _chain_re = re.compile(r'^=([A-Z]+)(\d+)$')
            from openpyxl.utils import column_index_from_string as _col_idx
            for col in sorted(keep_cols):
                for r in range(1, max_row + 1):
                    cell = ws.Cells(r, col)
                    if not cell.HasFormula:
                        continue
                    formula = cell.Formula
                    # 检查是否是简单的同行引用 =X<row>（链式引用的典型格式）
                    m = _chain_re.match(formula)
                    if not m:
                        continue
                    ref_col_letter = m.group(1)
                    ref_row = int(m.group(2))
                    ref_col = _col_idx(ref_col_letter)
                    if ref_col not in delete_cols:
                        continue
                    # 追踪链：ref_col被删，找它引用的列，直到找到保留列或非链式公式
                    visited = set()
                    cur_col = ref_col
                    resolved = None
                    while cur_col in delete_cols and cur_col not in visited:
                        visited.add(cur_col)
                        chain_cell = ws.Cells(ref_row, cur_col)
                        if not chain_cell.HasFormula:
                            # 不是公式，用值替代
                            resolved = ('value', chain_cell.Value)
                            break
                        chain_formula = chain_cell.Formula
                        cm = _chain_re.match(chain_formula)
                        if not cm:
                            # 不是简单引用，直接复制这个公式
                            resolved = ('formula', chain_formula)
                            break
                        next_col = _col_idx(cm.group(1))
                        if next_col not in delete_cols:
                            # 找到了保留列，用它的引用
                            resolved = ('formula', chain_formula)
                            break
                        cur_col = next_col
                    if resolved:
                        if resolved[0] == 'formula':
                            cell.Formula = resolved[1]
                        else:
                            cell.Value = resolved[1]

            # 从右到左删除未选中的列（避免索引偏移）
            for col in reversed(all_product_cols):
                if col not in keep_cols:
                    ws.Columns(col).Delete()

        # 修改ITEM名称（遍历所有sheet的R1~R4找PRODUCT ITEM#）
        if item_name:
            for cfg in sheets_config:
                try:
                    ws = wb.Sheets(cfg['sheet_name'])
                    for r in range(1, 5):
                        a_val = ws.Cells(r, 1).Value
                        if a_val and isinstance(a_val, str) and 'ITEM' in a_val.upper():
                            ws.Cells(r, 2).Value = item_name
                            break
                except Exception:
                    pass

        # 公式参数替换（用COM直接改公式字符串中的数字）
        if formula_overrides:
            for cfg in sheets_config:
                try:
                    ws = wb.Sheets(cfg['sheet_name'])
                except Exception:
                    continue
                for row_str, replacements in formula_overrides.items():
                    row_num = int(row_str)
                    if not replacements:
                        continue
                    # 遍历该行所有有公式的列
                    max_col = ws.UsedRange.Columns.Count + ws.UsedRange.Column - 1
                    for c in range(1, max_col + 1):
                        cell = ws.Cells(row_num, c)
                        if not cell.HasFormula:
                            continue
                        formula = cell.Formula
                        new_formula = formula
                        for old_val, new_val in replacements.items():
                            new_formula = new_formula.replace(str(old_val), str(new_val))
                        if new_formula != formula:
                            cell.Formula = new_formula

        wb.Save()
        wb.Close(False)
        logger.info(f'extract_columns: saved to {output_path}')

    except Exception as e:
        logger.exception("extract_columns COM error")
        raise
    finally:
        if app:
            try:
                app.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()

    return output_path
