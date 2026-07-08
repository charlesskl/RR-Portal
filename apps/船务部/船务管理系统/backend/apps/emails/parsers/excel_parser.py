"""Packing List Excel 解析器 — 从 .xlsx 文件中提取装箱单数据。"""

import re

import openpyxl


# CBM 相关关键词
_CBM_KEYWORDS = ['cbm', "measm't", 'measmt', 'measurement', 'meas']


def _normalize(text: str) -> str:
    """将文本转为小写并去除多余空白和标点。"""
    if text is None:
        return ''
    text = str(text).strip().lower()
    # 去除尾部的 . 和 #
    text = re.sub(r'[.#]+$', '', text)
    return text


def _match_field(header_raw: str) -> str | None:
    """基于关键词匹配，将列头映射到字段名。

    按照规范的关键词匹配规则（优先级从高到低）：
    - Contains "SKU" → product_code
    - "RETAIL UNIT" → quantity
    - "PC PER CARTON" or ("PC" and "CARTON") → spec
    - "NO OF CARTON" or ("NO" and "CARTON") → pieces
    - "NO OF PALLET" or ("PALLET" and "NO") → pallet_count
    - "CUSTOMER PO ITEM" → customer_po_item_no  (必须在 customer_po 之前检查)
    - "CUSTOMER PO" (but not "ITEM") → customer_po
    - "ZURU PO" or "PO NO" → contract_number
    - "ACTUAL FACTORY" or "ACTUAL" → factory_english
    - "SUPPLIER" → supplier
    """
    if header_raw is None:
        return None
    h = _normalize(str(header_raw))
    if not h:
        return None

    # SKU / Supplier Style No. / Style number / CONTACT (FR格式) → product_code
    if 'sku' in h:
        return 'product_code'
    if h == 'contact' or h.startswith('contact'):
        return 'product_code'
    if 'supplier' in h and 'style' in h:
        return 'product_code'
    if 'style' in h and ('no' in h.split() or 'number' in h or 'num' in h):
        return 'product_code'

    # Retail Unit → quantity（最高优先级）
    if 'retail' in h and 'unit' in h:
        return 'retail_unit'   # 先用独立 key，_build_column_map 里再处理优先级
    # Shipping QUANTITY → shipping_quantity（可能是卡板数，也可能是数量）
    if 'shipping' in h and 'quantity' in h:
        return 'shipping_quantity'
    # ORDER QTY / Order QTY → quantity（ZURU 多 BKG PL 格式）
    if 'order' in h and ('qty' in h.split() or h.endswith('qty')):
        return 'quantity'
    if h == 'quantity' or 'quantity' in h:
        return 'quantity'

    # PC PER CARTON → spec (必须在 pieces 之前)
    if ('pc' in h.split() or 'pcs' in h.split()) and 'carton' in h:
        return 'spec'
    # CASE PACK → spec（ZURU 多 BKG PL 格式：每箱个数）
    if 'case' in h and 'pack' in h:
        return 'spec'
    # INNERS → spec（裕泰仓 Kmart 箱单：每箱内盒数）
    if 'inners' in h or h == 'inner':
        return 'spec'

    # NO OF CARTON / CTNS / Cartons → pieces
    if 'no' in h and 'carton' in h:
        return 'pieces'
    # TOTAL CARTON → pieces（ZURU 多 BKG PL 格式：箱数）
    if 'total' in h and 'carton' in h:
        return 'pieces'
    if h == 'ctns' or h.startswith('ctns') or (h == 'cartons') or (h == 'carton'):
        return 'pieces'

    # NO OF PALLET → pallet_count
    if 'pallet' in h and 'no' in h:
        return 'pallet_count'

    # CUSTOMER PO ITEM → customer_po_item_no (先于 customer_po)
    if 'customer' in h and 'po' in h and 'item' in h:
        return 'customer_po_item_no'

    # CUSTOMER PO (but not ITEM) → customer_po
    if 'customer' in h and 'po' in h and 'item' not in h:
        return 'customer_po'

    # Order no. / Order No → customer_po（裕泰仓 Kmart 箱单：客户订单号）
    # 'qty' 已在前面被 'ORDER QTY' → quantity 规则截走，到这里 'order' + 'no' 安全
    if 'order' in h and 'no' in h.split() and 'qty' not in h:
        return 'customer_po'

    # ZURU PO or PO NO → contract_number
    if 'zuru' in h and 'po' in h:
        return 'contract_number'
    # 泛规则 PO+NO：排除 "PO Request No" 等含 'request' 的列（这些不是合同号）
    if 'po' in h and 'no' in h and 'item' not in h and 'customer' not in h and 'request' not in h:
        return 'contract_number'

    # MAIN ASSEMBLY FACTORY → main_factory (优先于 actual factory)
    if 'main' in h and 'factory' in h:
        return 'main_factory'

    # ACTUAL FACTORY → factory_english（要求同时含 factory，避免命中 "ACTUAL CBM/ONE CARTON"）
    if 'actual' in h and 'factory' in h:
        return 'factory_english'

    # FACTORY (单独) → 如果还没匹配到 factory 相关字段
    if h == 'factory' or (h.endswith('factory') and 'actual' not in h and 'main' not in h):
        return 'factory_short'

    # Factoy Name (typo of Factory) → factory_short（客户 Excel 模板常见拼写错误）
    if 'factoy' in h:
        return 'factory_short'

    # 做柜工厂 / 装柜工厂 → _zuogui_factory（独立字段，避免被通用 '工厂' 规则吞掉）
    if header_raw and ('做柜工厂' in str(header_raw) or '装柜工厂' in str(header_raw)):
        return '_zuogui_factory'

    # 中文"工厂"列头
    if header_raw and '工厂' in str(header_raw):
        return 'factory_short'

    # 中文"订单号" → contract_number（裕泰仓 Kmart 箱单 ZURU PO 列名）
    if header_raw and '订单号' in str(header_raw):
        return 'contract_number'

    # 中文"港口" → destination_port（裕泰仓 Kmart 箱单 per-row 卸货港，含目的国信息）
    if header_raw and '港口' in str(header_raw):
        return 'destination_port'

    # N.W / NW → nw_per_box（每箱净重）
    # 精确匹配 "n.w" 或 "n/w"，避免把 "n.w.t"、total 等误识别
    if h in ('n.w', 'n/w', 'nw'):
        return 'nw_per_box'

    # G.W / GW → gw_per_box（每箱毛重）
    if h in ('g.w', 'g/w', 'gw'):
        return 'gw_per_box'

    # SUPPLIER → supplier
    if 'supplier' in h:
        return 'supplier'

    # REMARK → remark
    if h == 'remark' or 'remark' in h:
        return 'remark'

    # CDS BKG# / BKG → _cds_bkg_number（ZURU 多 BKG PL 用 BKG，CDS PL 用 CDS BKG#）
    if 'cds' in h and 'bkg' in h:
        return '_cds_bkg_number'
    if h == 'bkg' or h.startswith('bkg'):
        return '_cds_bkg_number'
    # Booking Key / Booking Key# → _cds_bkg_number（ZURU 总表分柜明细 sheet 的 BKG 列名）
    if 'booking' in h and 'key' in h:
        return '_cds_bkg_number'

    # BOOKING → _booking_line
    if h == 'booking':
        return '_booking_line'

    # CARRIER SO NUMBER → carrier_so
    if 'carrier' in h and 'so' in h:
        return 'carrier_so'

    # SO# / SO NO. / SO N° → _so_column（ZURU 多 BKG PL 列，值如 "35624147/5"）
    if h in ('so#', 'so no', 'so no.', 'so number', 'so n°'):
        return '_so_column'

    # CONTAINER NO. LOADING → container_no_loading（区别于 container_type_loding）
    # "Container No. loading" 含 container + no + loading，但不含 type
    if 'container' in h and 'no' in h and 'loading' in h and 'type' not in h:
        return 'container_no_loading'

    # CONTAINER TYPE LODING → container_type_loding
    if 'container' in h and 'type' in h:
        return 'container_type_loding'

    return None


def _find_sheet(wb: openpyxl.Workbook):
    """查找 Packing list 工作表（大小写不敏感），如找不到则返回活动工作表。

    优先级：
    1. 含 "分柜"/"明细"/"detail" 的 packing 相关 sheet
    2. 含 "<工厂>做柜\\n<柜型>" 合并单元格分柜标签的 packing sheet（如 "Packing list (2)"）
    3. 精确名称 "packing list"
    4. 其他含 "packing" 的 sheet
    """
    packing_sheets = []
    plain_match = None
    for name in wb.sheetnames:
        nl = name.lower().strip()
        if nl == 'packing list' and plain_match is None:
            plain_match = name
        if 'packing' in nl:
            packing_sheets.append(name)

    # 0. sheet 名含"N个<柜型>"（如"4个40HQ" / "2个45HC"）视为多柜分柜表，最高优先级
    for name in wb.sheetnames:
        if re.search(r'\d+\s*个\s*\d+\s*(?:HQ|HC|GP)', name, re.IGNORECASE):
            return wb[name]

    # 1. 先找分柜/明细 sheet（即使存在普通 packing list 也优先）
    for name in packing_sheets:
        if '分柜' in name or '明细' in name or 'detail' in name.lower():
            return wb[name]

    # 2. 找含 "<工厂>做柜\n<柜型>" 合并单元格分柜标签的 packing sheet
    _zg_re = re.compile(r'[一-鿿]+[做装拼]柜\s*[\r\n]+\s*\d+\s*(?:HQ|GP|HC)', re.IGNORECASE)
    for name in packing_sheets:
        ws = wb[name]
        for mr in ws.merged_cells.ranges:
            v = ws.cell(mr.min_row, mr.min_col).value
            if v and isinstance(v, str) and _zg_re.search(v):
                return ws

    if plain_match:
        return wb[plain_match]
    if packing_sheets:
        return wb[packing_sheets[0]]
    return wb.active


def _find_header_row(ws, max_rows: int = 30) -> tuple[int, list]:
    """在前 max_rows 行中查找含有 SKU 和 CARTON 关键词的表头行。

    支持双行表头（如 ZURU 多 BKG PL：R21 含 SKU / R22 含 Carton）：
    单行命中后会与下一行合并比较，取识别字段更多者。
    """
    def _merge(upper, lower):
        out = []
        for cu, cl in zip(upper, lower):
            su = '' if cu is None else str(cu).strip()
            sl = '' if cl is None else str(cl).strip()
            if su and sl:
                out.append(f'{su} {sl}')
            elif su:
                out.append(su)
            elif sl:
                out.append(sl)
            else:
                out.append(None)
        return out

    def _score(cells):
        """识别字段数：_match_field 命中即记 1，去重。"""
        seen = set()
        for c in cells:
            f = _match_field(c)
            if f and f not in seen:
                seen.add(f)
        return len(seen)

    rows_cache = []
    for row_idx in range(1, max_rows + 1):
        cells = []
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            cells.append(val)
        rows_cache.append(cells)

        text = ' '.join(str(c).lower() for c in cells if c is not None)
        # 排除明显的业务文档头行（这些行可能恰好含 cbm/ctns 摘要而误命中）
        # 注：不能用 'invoice no' / 'date:' 排除——PL 真表头也含这些作为列名
        if any(kw in text for kw in ('messrs:', 'messrs：', 'address:', 'address：',
                                      'packing list no', 'booking notice',
                                      'contact:', 'contact：')):
            continue
        is_match = (
            ('sku' in text and 'carton' in text)
            or ((re.search(r'\bstyle\b', text) or 'keycode' in text) and ('carton' in text or 'cbm' in text or 'ctns' in text))
            or ('description' in text and 'carton' in text)
        )
        if is_match:
            # 与下一行合并比较，取识别字段更多者
            if row_idx < max_rows:
                next_cells = [ws.cell(row=row_idx + 1, column=c).value for c in range(1, ws.max_column + 1)]
                merged = _merge(cells, next_cells)
                if _score(merged) > _score(cells):
                    # 用合并表头，数据从 row_idx + 2 行开始
                    return row_idx + 1, merged
            return row_idx, cells

    # 单行未匹配：尝试合并相邻两行
    for i in range(len(rows_cache) - 1):
        merged = _merge(rows_cache[i], rows_cache[i + 1])
        text = ' '.join(str(c).lower() for c in merged if c is not None)
        if 'sku' in text and 'carton' in text:
            return i + 2, merged

    # 未找到时返回第 1 行
    cells = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    return 1, cells


def _build_column_map(headers: list) -> dict:
    """构建列索引映射。

    Returns:
        dict: {field_name: column_index, ...} 以及可能的
              '_cbm_columns': [col_indices] 用于 CBM/尺寸列
    """
    col_map = {}

    # 映射标准字段（每个字段只取第一个匹配）
    for idx, header in enumerate(headers):
        field = _match_field(header)
        if field is not None and field not in col_map:
            col_map[field] = idx

    # 处理 Retail Unit 与 Shipping QUANTITY 的优先级：
    # 如果同时存在，Retail Unit → quantity，Shipping QUANTITY → pallet_count（卡板数）
    # 如果只有 Shipping QUANTITY，则作为 quantity
    if 'retail_unit' in col_map and 'shipping_quantity' in col_map:
        col_map['quantity'] = col_map.pop('retail_unit')
        if 'pallet_count' not in col_map:
            col_map['pallet_count'] = col_map.pop('shipping_quantity')
        else:
            col_map.pop('shipping_quantity')
    elif 'retail_unit' in col_map:
        col_map['quantity'] = col_map.pop('retail_unit')
    elif 'shipping_quantity' in col_map:
        col_map['quantity'] = col_map.pop('shipping_quantity')

    # 处理 CBM 相关列
    cbm_indices = []
    measm_col = None
    for idx, header in enumerate(headers):
        if header is None:
            continue
        h = _normalize(str(header))
        # 检查是否是 CBM/MEASM 关键词
        if any(kw in h for kw in _CBM_KEYWORDS):
            cbm_indices.append(idx)
            if measm_col is None:
                measm_col = idx
        # 检查单字符 L/W/H（通常紧挨 CBM 列）
        elif h in ('l', 'w', 'h'):
            cbm_indices.append(idx)

    # 如果只找到一个 MEASM'T(CBM) 列，检查后续列是否为 x,数值,x,数值,CBM 的模式
    # 实际格式：O=MEASM'T(CBM), 数据行中 O=长, P=x, Q=宽, R=x, S=高, T=CBM
    if measm_col is not None and len(cbm_indices) == 1:
        # 标记为 MEASM 模式，后续解析时从 measm_col 开始取连续6列
        col_map['_measm_col'] = measm_col

        # 检测 MEASM 头是否是"每箱 CBM"格式（如 "ACTUAL CBM/ONE CARTON"）
        # 若是，则 measm_col+5 是每箱 CBM，需要另寻 TOTAL/总 CBM 列作为行总 CBM
        _mh = _normalize(str(headers[measm_col]))
        if 'one carton' in _mh or 'per carton' in _mh:
            # 在 measm_col+6 之后 5 列内寻找 'total'/'cbm 合计' 等 → 用作每行总 CBM
            for _ti in range(measm_col + 6, min(measm_col + 12, len(headers))):
                _th = _normalize(str(headers[_ti] or ''))
                if not _th:
                    continue
                if 'total' in _th or 'cbm 合计' in _th or '总 cbm' in _th:
                    col_map['_total_cbm_col'] = _ti
                    break

    # 按索引排序
    cbm_indices.sort()
    col_map['_cbm_columns'] = cbm_indices

    return col_map


def _extract_cbm_values(row_values: list, cbm_columns: list, measm_col: int | None = None, total_cbm_col: int | None = None) -> tuple:
    """从 CBM 相关列提取体积和箱规。

    Args:
        total_cbm_col: 若 MEASM 头是"每箱 CBM"（如 ACTUAL CBM/ONE CARTON），
                       则用此列作为每行总 CBM（取代 measm_col+5）

    Returns:
        (volume, box_dimensions)
    """
    # MEASM 模式：MEASM'T(CBM) 列后面跟尺寸和 CBM
    if measm_col is not None:
        l_val = _safe_get(row_values, measm_col)
        sep = _safe_get(row_values, measm_col + 1)

        # 格式A：单元格独立，用 'x' 分隔（L | x | W | x | H | CBM）
        if sep is not None and str(sep).strip().lower() == 'x':
            w_val = _safe_get(row_values, measm_col + 2)
            h_val = _safe_get(row_values, measm_col + 4)
            # 优先用 total_cbm_col（"ACTUAL CBM/ONE CARTON" 格式下 measm_col+5 是每箱 CBM）
            if total_cbm_col is not None:
                cbm_val = _safe_get(row_values, total_cbm_col)
            else:
                cbm_val = _safe_get(row_values, measm_col + 5)
            dims = []
            for v in (l_val, w_val, h_val):
                if v is not None and v != '' and str(v).strip().lower() != 'x':
                    dims.append(str(v))
            box_dimensions = '*'.join(dims) if len(dims) == 3 else ''
            volume = cbm_val if cbm_val is not None else ''
            return (volume, box_dimensions)

        # 格式B：单格内 "35.3 x 26.3 x 33.5"，下一格为 CBM
        if l_val is not None and isinstance(l_val, str) and 'x' in l_val.lower():
            parts = re.split(r'\s*[xX×]\s*', l_val.strip())
            nums = [p.strip() for p in parts if re.match(r'[\d.]+', p.strip())]
            box_dimensions = '*'.join(nums) if len(nums) == 3 else ''
            cbm_val = _safe_get(row_values, measm_col + 1)
            volume = cbm_val if cbm_val is not None else ''
            return (volume, box_dimensions)

        # 格式C：measm_col 直接就是 CBM 数值
        # 但若 l_val>80（合理柜单 CBM 上限约 76）且 measm_col+5 为 None，
        # 说明可能是合并单元格丢了真实 CBM，l_val 是误读的长度数字 → 标记为缺失
        if isinstance(l_val, (int, float)) and l_val > 80:
            val_at_5 = _safe_get(row_values, measm_col + 5)
            if val_at_5 is None:
                return ('', '')
        volume = l_val if l_val is not None else ''
        return (volume, '')

    if not cbm_columns:
        return ('', '')

    if len(cbm_columns) >= 4:
        # 前 3 列为 L/W/H，最后 1 列为总 CBM
        l_val = _safe_get(row_values, cbm_columns[0])
        w_val = _safe_get(row_values, cbm_columns[1])
        h_val = _safe_get(row_values, cbm_columns[2])
        cbm_val = _safe_get(row_values, cbm_columns[-1])

        dims = []
        for v in (l_val, w_val, h_val):
            if v is not None and v != '':
                dims.append(str(v))
        box_dimensions = '*'.join(dims) if len(dims) == 3 else ''

        volume = cbm_val if cbm_val is not None else ''
        return (volume, box_dimensions)
    else:
        cbm_val = _safe_get(row_values, cbm_columns[-1])
        volume = cbm_val if cbm_val is not None else ''
        return (volume, '')


def _safe_get(lst, idx, default=None):
    """安全获取列表元素。"""
    if idx is not None and 0 <= idx < len(lst):
        return lst[idx]
    return default


def _convert_xls_to_xlsx(xls_path: str) -> str:
    """将 .xls 文件转换为临时 .xlsx 文件路径。"""
    import xlrd
    import tempfile
    import os
    wb = xlrd.open_workbook(xls_path)
    new_wb = openpyxl.Workbook()
    new_wb.remove(new_wb.active)
    for sn in wb.sheet_names():
        ws = wb.sheet_by_name(sn)
        new_ws = new_wb.create_sheet(sn[:31] or 'Sheet')
        for r in range(ws.nrows):
            for c in range(ws.ncols):
                v = ws.cell_value(r, c)
                new_ws.cell(row=r + 1, column=c + 1, value=v)
    fd, tmp_path = tempfile.mkstemp(suffix='.xlsx', prefix='xls_conv_')
    os.close(fd)
    new_wb.save(tmp_path)
    return tmp_path


def parse_po_extract_xls(file_path: str) -> list[dict]:
    """解析 ZURU "Purchase Order Extract" 格式的交仓 PL（.xls）。

    特征列：PO Num, Open Date (PO), Item ID, Pieces Ordered, CTN, CBM,
            Booking#, PO#, Remark, DC

    返回值字段：product_code(Item ID), customer_po(PO Num), contract_number(PO#),
                spec/pieces (CTN), quantity (Pieces Ordered), volume (CBM),
                booking_number (Booking#), remark (BX 号),
                _delivery_date (Open Date PO 转 'M/D' 格式),
                _dc, nw_per_box, gw_per_box, box_dimensions
    """
    try:
        import xlrd
    except ImportError:
        return []
    from datetime import datetime, timedelta

    try:
        wb = xlrd.open_workbook(file_path)
    except Exception:
        return []

    # 优先找 'Purchase Order Extract' sheet
    ws = None
    for sn in wb.sheet_names():
        if 'purchase' in sn.lower() and 'extract' in sn.lower():
            ws = wb.sheet_by_name(sn)
            break
    if ws is None:
        ws = wb.sheet_by_index(0)

    # 查找表头行（必须含 PO Num, CBM, Remark 这几个标志列）
    header_row = -1
    headers = []
    for r in range(min(ws.nrows, 10)):
        row = [str(ws.cell_value(r, c) or '').strip() for c in range(ws.ncols)]
        if 'PO Num' in row and 'CBM' in row and 'Remark' in row:
            header_row = r
            headers = row
            break
    if header_row < 0:
        return []

    col = {h: i for i, h in enumerate(headers) if h}

    def _v(r, key, default=''):
        i = col.get(key)
        if i is None or i >= ws.ncols:
            return default
        v = ws.cell_value(r, i)
        return v if v not in (None, '') else default

    def _num(r, key):
        v = _v(r, key, 0)
        try:
            return float(v)
        except (ValueError, TypeError):
            return 0.0

    def _excel_date(serial):
        try:
            d = datetime(1899, 12, 30) + timedelta(days=int(float(serial)))
            return f'{d.month}/{d.day}'
        except (ValueError, TypeError):
            return ''

    def _str(v):
        if v is None or v == '':
            return ''
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v).strip()

    # 找紧邻 CBM 右边、无表头的尺寸列（如 '36.5*21.5*27'）
    cbm_idx = col.get('CBM')
    dim_idx = None
    if cbm_idx is not None and cbm_idx + 1 < ws.ncols:
        if not headers[cbm_idx + 1]:
            dim_idx = cbm_idx + 1

    items = []
    for r in range(header_row + 1, ws.nrows):
        po_num = _v(r, 'PO Num')
        item_id = _v(r, 'Item ID')
        if not po_num or not item_id:
            continue

        ctn = int(_num(r, 'CTN'))
        nw_total = _num(r, 'N W')
        gw_total = _num(r, 'G W')
        box_dim = ''
        if dim_idx is not None:
            box_dim = _str(ws.cell_value(r, dim_idx))

        items.append({
            'product_code': _str(item_id),
            'customer_po': _str(po_num),
            'contract_number': _str(_v(r, 'PO#')),
            'spec': ctn,
            'pieces': ctn,
            'quantity': int(_num(r, 'Pieces Ordered')),
            'volume': round(_num(r, 'CBM'), 3),
            'nw_per_box': round(nw_total / ctn, 3) if ctn else 0,
            'gw_per_box': round(gw_total / ctn, 3) if ctn else 0,
            'box_dimensions': box_dim,
            'booking_number': _str(_v(r, 'Booking#')),
            'remark': _str(_v(r, 'Remark')),
            'factory_english': '兴信',  # 交仓邮件默认兴信本厂
            'factory_remark': '兴信',
            'supplier': '',
            'main_factory': '',
            'factory_short': '',
            '_delivery_date': _excel_date(_num(r, 'Open Date (PO)')),
            '_dc': _str(_v(r, 'DC')),
            '_invoice': _str(_v(r, 'INVOICE#')),
        })
    return items


def parse_packing_list(file_path: str) -> list[dict]:
    """解析 Packing List Excel 文件。

    Args:
        file_path: .xlsx 文件路径

    Returns:
        list[dict]: 每行数据的字典列表，包含以下键：
            product_code, quantity, spec, pieces, pallet_count,
            customer_po, customer_po_item_no, contract_number,
            factory_english, volume, box_dimensions
    """
    # .xls 老格式：先尝试用专用解析器（PO Extract 等），失败再转换为 .xlsx
    if file_path.lower().endswith('.xls'):
        po_items = parse_po_extract_xls(file_path)
        if po_items:
            return po_items
        # 其他 .xls 格式：转换为 .xlsx 再走标准管道
        file_path = _convert_xls_to_xlsx(file_path)

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = _find_sheet(wb)

    header_row, headers = _find_header_row(ws)
    col_map = _build_column_map(headers)

    cbm_columns = col_map.pop('_cbm_columns', [])
    measm_col = col_map.pop('_measm_col', None)
    total_cbm_col = col_map.pop('_total_cbm_col', None)

    # 一次性读取所有单元格值（避免逐 cell 访问，性能 10x+）
    # _all_rows[r-1] 对应行 r 的 [c1, c2, ...] 值列表
    _all_rows = list(ws.iter_rows(values_only=True))
    _max_row = len(_all_rows)
    _max_col = ws.max_column

    def _get(r_idx, c_idx):
        """1-based 行列号取值（同 ws.cell(r,c).value 但走缓存）。"""
        if 1 <= r_idx <= _max_row and 1 <= c_idx <= _max_col:
            row = _all_rows[r_idx - 1]
            if c_idx - 1 < len(row):
                return row[c_idx - 1]
        return None

    # 展开合并单元格：{(row, col): value}
    _merged_values = {}
    for _mr in ws.merged_cells.ranges:
        _v = _get(_mr.min_row, _mr.min_col)
        if _v is None:
            continue
        for _r in range(_mr.min_row, _mr.max_row + 1):
            for _c in range(_mr.min_col, _mr.max_col + 1):
                _merged_values[(_r, _c)] = _v

    # 车次列识别（如 ZURU 客上车 PL 的 CUSTOMER ITEM DESCRIPTION 列）
    # 模式：N*车型/M.D收（如 "1*45'自备柜/5.13收"、"1*8.5米车/5.19收"、"1*8T车/6.5收"）
    _truck_label_re = re.compile(
        r'\d+\s*\*\s*\S+?[/／]\s*\d+(?:[.\d]+)?\s*收'
    )
    _truck_col = None
    _truck_match_count = 0
    # 性能：仅当合并单元格中有命中模式时才扫描全列；大表（>500 行）跳过全列扫描
    _has_merged_truck = any(
        isinstance(v, str) and _truck_label_re.search(v) for v in _merged_values.values()
    )
    if _has_merged_truck or _max_row <= 500:
        for _c in range(1, _max_col + 1):
            _hits = 0
            for _r in range(header_row + 1, _max_row + 1):
                _v = _merged_values.get((_r, _c))
                if _v is None:
                    _v = _get(_r, _c)
                if _v and isinstance(_v, str) and _truck_label_re.search(_v):
                    _hits += 1
            if _hits > _truck_match_count:
                _truck_match_count = _hits
                _truck_col = _c
    if _truck_match_count < 2:
        _truck_col = None

    results = []
    # 维护已设置过 _container_assignment 的 carrier_so 集合
    # 用途：同 SO 的后续行不再走格式4 fallback（避免打散柜组）
    _carrier_so_with_asn = set()
    # container_type_loding 列上一行的合并单元格值（如 "丹尼做柜\n40GP"）
    # 用途：连续相同合并值的行只在首行设 asn（让 views cur_key 沿用，避免后续行 fmt4 误设默认兴信）
    _ctl_col_idx = col_map.get('container_type_loding')  # 1-based 转 0-based 在用时减 1
    _prev_ctl_merged = None
    standard_fields = [
        'product_code', 'quantity', 'spec', 'pieces', 'pallet_count',
        'customer_po', 'customer_po_item_no', 'contract_number',
        'factory_english', 'supplier', 'main_factory', 'factory_short', 'remark', 'carrier_so', 'container_type_loding', 'container_no_loading', '_cds_bkg_number', '_booking_line', '_zuogui_factory', '_so_column',
        'nw_per_box', 'gw_per_box', 'destination_port',
    ]

    # 检测是否为多柜分组格式（SUB TOTAL 行分隔）
    # 性能：限制扫描前 8 列（SUB TOTAL 一般在前几列），并跳过空行
    _container_group = 0
    _sub_total_cols = []
    _max_st_col = min(_max_col, 8)
    for _r_idx, _row in enumerate(_all_rows, 1):
        for _c_idx, _v in enumerate(_row[:_max_st_col], 1):
            if _v is None:
                continue
            _s = str(_v).strip().upper()
            if _s in ('SUB TOTAL:', 'SUB TOTAL', 'SUBTOTAL'):
                _sub_total_cols.append((_r_idx, _c_idx))
                break
    _is_multi_container = len(_sub_total_cols) > 1

    for row_idx in range(header_row + 1, _max_row + 1):
        # 用缓存的 _all_rows 避免逐 cell 访问；按 _max_col 补齐到固定长度
        _row_tuple = _all_rows[row_idx - 1]
        row_values = list(_row_tuple) + [None] * max(0, _max_col - len(_row_tuple))

        # 多柜格式：优先检测任意列SUB TOTAL
        if _is_multi_container:
            _row_strs = [str(v or '').strip().upper() for v in row_values]
            if any(v in ('SUB TOTAL:', 'SUB TOTAL', 'SUBTOTAL') for v in _row_strs):
                _container_group += 1
                continue

        # 检查 SKU 是否为空 — 跳过空行
        sku_col = col_map.get('product_code')
        if sku_col is not None:
            sku_val = row_values[sku_col]
            if sku_val is None or str(sku_val).strip() == '':
                continue
            sku_str = str(sku_val).strip().upper()
            if sku_str in ('TOTAL', 'TOTAL:', '合计', 'GRAND TOTAL', 'GRAND TOTAL:', 'SKU NO.', 'SKU NO', 'SKU'):
                continue

        record = {}
        for field in standard_fields:
            idx = col_map.get(field)
            if idx is not None and idx < len(row_values):
                val = row_values[idx]
                record[field] = val if val is not None else ''
            else:
                record[field] = ''

        # 提取 CBM 相关值
        volume, box_dimensions = _extract_cbm_values(row_values, cbm_columns, measm_col, total_cbm_col)
        record['volume'] = volume
        record['box_dimensions'] = box_dimensions

        # container_type_loding 列合并值检查：
        # 若本行该列(合并扩展后)与上一行相同 → 这是合并范围非首行，不需要重新设 asn
        _ctl_merged_now = None
        if _ctl_col_idx is not None:
            _ctl_r1c = (row_idx, _ctl_col_idx + 1)
            _ctl_merged_now = _merged_values.get(_ctl_r1c)
            if _ctl_merged_now is None:
                _ctl_merged_now = row_values[_ctl_col_idx] if _ctl_col_idx < len(row_values) else None
        _is_ctl_continuation = (
            _ctl_merged_now is not None
            and _prev_ctl_merged is not None
            and _ctl_merged_now == _prev_ctl_merged
        )

        # 检测做柜分配列 — 支持四种格式：
        # 格式1: 无表头列有 "SO#XXX\n雨禾做柜\n40HQ"
        # 格式2: Remarks列有 "40HC\nKoyo装柜" + 旁边列有 "柜N" + SO列
        # 格式3: container_type_loding 列有值（含柜型，如 "1*40HQ\n兴信拼柜\n柜2"），用于分组触发
        # 格式5: 出货清单(分柜) Actual Factory 列合并单元格 "40HQ/兴信 SO#269850035"
        # 注：合并范围内非首行 row_values[ci] 是 None，跳过 → 让 views 的 cur_key 沿用首行的 assignment
        for ci in range(len(row_values)):
            cv = row_values[ci]  # 只看原始值，合并范围只首行非 None
            if not cv or not isinstance(cv, str):
                continue
            cv_s = cv.strip()
            if '做柜' in cv_s and 'SO#' in cv_s:
                record['_container_assignment'] = cv_s
                break
            if ('装柜' in cv_s or '做柜' in cv_s or '拼柜' in cv_s) and re.search(r'\d+(?:H[QC]|GP)', cv_s, re.IGNORECASE):
                # 排除中文说明性长文本（如 "1*40HQ，由兴信安排装柜，其他工厂将货物送到兴信拼柜"）
                # 真实分柜标签通常很短：'40HC\nKoyo装柜\n柜2' / '1*40HQ\n兴信做柜'
                # 含中文标点（，。、；）或长度过长（>40）的视为说明文字，跳过
                if len(cv_s) > 40 or any(p in cv_s for p in '，。、；！？'):
                    continue
                # 格式2：找旁边的柜号和SO号
                _gui = ''
                _so = ''
                for _ci2 in range(max(0, ci-2), min(len(row_values), ci+3)):
                    _v2 = row_values[_ci2]
                    if _v2 is None:
                        continue
                    if isinstance(_v2, str):
                        _v2s = _v2.strip()
                        if re.match(r'^柜\d+$', _v2s):
                            _gui = _v2s
                        elif re.match(r'^[A-Z]{3,}[\d]+$', _v2s):
                            _so = _v2s
                        elif re.match(r'^\d{6,12}$', _v2s):
                            # 纯数字 SO（如 ZURU 出货 SO NO 列：17481605）
                            _so = _v2s
                    elif isinstance(_v2, (int, float)):
                        # 数字型 SO（来自 Excel 数字单元格）
                        _v2s = str(int(_v2)) if isinstance(_v2, float) and _v2.is_integer() else str(_v2)
                        if re.match(r'^\d{6,12}$', _v2s):
                            _so = _v2s
                record['_container_assignment'] = f'SO#{_so}\n{cv_s}\n{_gui}'.strip()
                break
            # 格式5: "40HQ/兴信 SO#269850035" 或 "40HQ/泰亨，注意XXX SO#269850065"
            #         典型出现在 出货清单(分柜) 的 Actual Factory 列合并单元格
            #         归一化为 "<柜型>\n<工厂>做柜\nSO#<so>"，让下游 views 的 factory 提取逻辑命中
            _fmt5_m = re.match(
                r'\s*(\d+\s*(?:HQ|GP|HC))\s*[/／]\s*([一-鿿A-Za-z]+)',
                cv_s, re.IGNORECASE,
            )
            if _fmt5_m and 'SO#' in cv_s:
                _ct5 = _fmt5_m.group(1).strip().upper()
                _fac5 = _fmt5_m.group(2).strip()
                _so5_m = re.search(r'SO#\s*([A-Z]*\d{6,})', cv_s)
                _so5 = _so5_m.group(1) if _so5_m else ''
                record['_container_assignment'] = (
                    f'{_ct5}\n{_fac5}做柜\nSO#{_so5}' if _so5 else f'{_ct5}\n{_fac5}做柜'
                )
                if _so5:
                    record['_container_so'] = _so5
                break

        # 格式3：container_type_loding 含"N*柜型"（如 1*40HQ，2*40HQ）才触发多柜分组
        # 注意：必须有乘号(*)，避免把普通柜型记录"40HQ"误触发多柜分支
        if not record.get('_container_assignment'):
            ctl = str(record.get('container_type_loding') or '').strip()
            if ctl and re.search(r'\d+\s*\*\s*\d+\s*(?:HQ|GP|HC)', ctl, re.IGNORECASE):
                record['_container_assignment'] = ctl
        # 格式4: Container No. loading 列（如 C260304887）→ 多柜分组
        # 有明确柜号列时，以该列值作为分组标识
        # 若同时有 main_factory 列，用该工厂名作为做柜工厂标识；否则默认兴信做柜
        # 注：本行或同 carrier_so 的前面行已通过格式1/2/5设置 _container_assignment 时，
        #     **不覆盖** _container_assignment（避免打散同一 SO 的柜组），
        #     但 _container_number 仍设置（让柜号模式正确分组，含领头行）
        _cso_now = str(record.get('carrier_so') or '').strip()
        _has_explicit_asn = bool(record.get('_container_assignment'))
        # 抑制条件：仅当本行已通过格式1/2/5 设过 _container_assignment 时不覆盖
        # 注：曾用 _is_ctl_continuation / _carrier_so_with_asn 抑制——但同一 SO 下多柜场景
        # （如 PL260500989：同 SO 含 兴信柜 + 博锐柜）会让第二个柜的首行被错误抑制，
        # 导致 _process_tjx_groups 无法识别新柜的做柜工厂
        _suppress_asn_fmt4 = _has_explicit_asn
        if not record.get('_container_number') and record.get('container_no_loading'):
            cno = str(record['container_no_loading']).strip()
            ctl = str(record.get('container_type_loding') or '').strip()
            mf = str(record.get('main_factory') or '').strip()
            if cno and ctl.upper() != 'CFS':  # CFS=拼柜中转站，不用于单柜分组
                record['_container_number'] = cno
                if not _suppress_asn_fmt4:
                    # main_factory 列被填合并单元格的多行分柜标签（如 "1*40HQ\n兴信做柜\nSO号"）时，
                    # 提取里面的工厂名，避免整段被拼成 "XX做柜" 误识为外厂
                    if '\n' in mf or any(k in mf for k in ('做柜', '装柜', '拼柜')):
                        m_fac = re.search(r'([一-鿿A-Za-z]+)[做装拼]柜', mf)
                        mf = m_fac.group(1).strip() if m_fac else ''
                    zuogui = mf if mf else '兴信'
                    # 英文公司全名（含 ≥2 空格 或 长度>25）当工厂名拼"做柜"会让本行 _ca 与
                    # 前置 SO#\n<柜型>\n<中文工厂>做柜 形式的 _ca 不一致，被 _process_tjx_groups
                    # 错误切成多组。这种情况下跳过 _ca 设置，让本行沿用前一行 cur_key。
                    if zuogui and (zuogui.count(' ') >= 2 or len(zuogui) > 25):
                        zuogui = ''
                    if zuogui and not record.get('_container_assignment'):
                        record['_container_assignment'] = f'{ctl}\n{zuogui}做柜' if ctl else f'{zuogui}做柜'

        # 记录已设置 _container_assignment 的 carrier_so，让后续相同 SO 的行跳过格式4 的 assignment 部分
        if _cso_now and record.get('_container_assignment'):
            _carrier_so_with_asn.add(_cso_now)

        # 更新 prev ctl 合并值（用于下一行检测）
        _prev_ctl_merged = _ctl_merged_now
        # 柜号列检测（所有行都检测，包括有装柜标记的行）
        for ci in range(len(row_values)):
            cv = row_values[ci]
            if cv and isinstance(cv, str) and re.match(r'^柜\d+$', cv.strip()):
                record['_container_number'] = cv.strip()
                # 找SO号
                for _ci2 in range(ci-2, min(len(row_values), ci+3)):
                    _v2 = row_values[_ci2] if 0 <= _ci2 < len(row_values) else None
                    if _v2 and isinstance(_v2, str) and re.match(r'^[A-Z]{3,}[\d]+$', _v2.strip()):
                        record['_container_so'] = _v2.strip()
                break

        if _is_multi_container:
            record['_container_group'] = _container_group

        # 车次标签（来自 F 列等含合并单元格的车次列）
        if _truck_col is not None:
            _tv = _merged_values.get((row_idx, _truck_col))
            if _tv is None:
                _tv = _get(row_idx, _truck_col)
            if _tv and isinstance(_tv, str) and _truck_label_re.search(_tv):
                record['_truck_label'] = _tv.strip()

        results.append(record)

    wb.close()

    # 把 per-row destination_port 字符串（如 'SYDNEY, AUSTRALIA'）转换成中文国家
    # 写到 record['country']，供 ShipmentItem.country 使用（覆盖邮件级 country）
    if any(r.get('destination_port') for r in results):
        from .pdf_parser import _port_to_country
        for r in results:
            dp = (r.get('destination_port') or '').strip()
            if dp:
                # 优先用 "," 后的国家名（如 'SYDNEY, AUSTRALIA' → 'AUSTRALIA'）
                tail = dp.rsplit(',', 1)[-1].strip()
                country = _port_to_country(tail) or _port_to_country(dp)
                if country:
                    r['country'] = country

    return results


def parse_warehouse_receipt(file_path: str) -> list[dict]:
    """解析兴信入仓单（落货纸）Excel — 中文表头 Sheet1。

    典型表头（约 r7）：
      报关单对应项号 | 报关品名 | 商品编码 | PO No. | Vendor Part Number |
      数量QTY | 箱数 | 包装规格(CM)长 | 包装规格(CM)宽 | 包装规格(CM)高 | 体积CBM | 净重(KG) | 毛重(KG)
    数据行从表头下一行开始，至遇到 "Total:" / "合计" / 空 SKU 行结束。

    返回与 parse_packing_list 兼容的 item 字典列表，自动标记为兴信货。
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception:
        return []

    items = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        # 在前 15 行找含 'Vendor Part Number' 或 '数量QTY' + '体积CBM' 的表头
        header_row = None
        col_sku = col_qty = col_ctn = col_cbm = col_po = None
        col_l = col_w = col_h = None
        for r in range(1, min(ws.max_row + 1, 15)):
            row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            row_text = ' '.join(str(v or '') for v in row_vals)
            if ('Vendor Part Number' in row_text or 'SKU' in row_text.upper()) and ('数量' in row_text or 'QTY' in row_text.upper()) and ('CBM' in row_text.upper() or '体积' in row_text):
                header_row = r
                for c, v in enumerate(row_vals, 1):
                    vs = str(v or '').strip()
                    if not vs:
                        continue
                    if 'Vendor Part Number' in vs or vs.upper() == 'SKU':
                        col_sku = c
                    elif vs.startswith('数量') or 'QTY' in vs.upper():
                        col_qty = c
                    elif vs == '箱数' or 'CTN' in vs.upper() or 'CARTON' in vs.upper():
                        col_ctn = c
                    elif '体积' in vs or 'CBM' in vs.upper():
                        col_cbm = c
                    elif vs.startswith('PO No') or vs == 'PO No.':
                        col_po = c
                    elif '包装' in vs and '长' in vs:
                        col_l = c
                    elif '包装' in vs and '宽' in vs:
                        col_w = c
                    elif '包装' in vs and '高' in vs:
                        col_h = c
                break
        if not header_row or not col_sku or not col_cbm:
            continue

        for r in range(header_row + 1, ws.max_row + 1):
            sku = ws.cell(r, col_sku).value
            if sku is None:
                continue
            sku_s = str(sku).strip()
            if not sku_s:
                continue
            # 跳过合计行
            row_text = ' '.join(str(ws.cell(r, c).value or '') for c in range(1, ws.max_column + 1))
            if 'total' in row_text.lower() or '合计' in row_text or '总计' in row_text:
                continue

            try:
                qty = int(ws.cell(r, col_qty).value or 0) if col_qty else 0
            except (ValueError, TypeError):
                qty = 0
            try:
                ctn = int(ws.cell(r, col_ctn).value or 0) if col_ctn else 0
            except (ValueError, TypeError):
                ctn = 0
            try:
                cbm = round(float(ws.cell(r, col_cbm).value or 0), 3)
            except (ValueError, TypeError):
                cbm = 0
            po_val = ws.cell(r, col_po).value if col_po else ''
            po = str(po_val).strip() if po_val is not None else ''

            box_dim = ''
            if col_l and col_w and col_h:
                _l = ws.cell(r, col_l).value
                _w = ws.cell(r, col_w).value
                _h = ws.cell(r, col_h).value
                if _l and _w and _h:
                    box_dim = f'{_l}*{_w}*{_h}'

            items.append({
                'product_code': sku_s,
                'quantity': qty,
                'pieces': ctn,
                'spec': (qty // ctn) if ctn else 0,
                'pallet_count': 0,
                'volume': cbm,
                'box_dimensions': box_dim,
                'customer_po': po,
                'customer_po_item_no': '',
                'contract_number': '',
                'factory_english': 'Dong Guan Hanson Plastic Product Ltd',
                'factory_remark': '兴信',
                'supplier': '',
                'main_factory': '',
                'factory_short': '',
                'remark': '',
                'carrier_so': '',
                'container_type_loding': '',
                'container_no_loading': '',
                'nw_per_box': 0,
                'gw_per_box': 0,
                '_source_file': '',
            })
        if items:
            break  # 只取第一个含表头的 sheet
    return items


def parse_cds_bkg_mapping(file_path: str) -> list:
    """读取 ZURU 邮件 Excel 中的 Sheet2，返回每个 Booking 的记录列表。

    每条记录：{'bkg': 'AMZ...', 'booking': 'Booking 8-1*40HQ-1', 'factory': '兴信', 'contracts': {'4500195273', ...}}

    Sheet2 结构（含合并单元格）：
      A: CDS BKG#  B: Booking  C: 做柜工厂  D: Supplier Name  E: ZURU PO No.  F: SO#  G: CBM
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception:
        return []

    # 找含 CDS BKG# 的 sheet
    target_ws = None
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(min_row=1, max_row=10):
            for cell in row:
                if str(cell.value or '').strip().upper().startswith('CDS'):
                    target_ws = ws
                    break
            if target_ws:
                break
        if target_ws:
            break

    if not target_ws:
        return []

    # 找表头行
    header_row = None
    col_bkg = col_booking = col_factory = col_contract = None
    for i, row in enumerate(target_ws.iter_rows(min_row=1, max_row=15), 1):
        raw_vals = [str(c.value or '') for c in row]
        vals = [v.strip().lower() for v in raw_vals]
        if any('cds' in v and 'bkg' in v for v in vals):
            header_row = i
            for j, v in enumerate(vals):
                if 'cds' in v and 'bkg' in v:
                    col_bkg = j
                elif 'booking' in v:
                    col_booking = j
                elif '做柜' in raw_vals[j] or ('factory' in v and 'supplier' not in v):
                    col_factory = j
                elif 'po' in v and ('zuru' in v or 'no' in v):
                    col_contract = j
            break

    if header_row is None or col_bkg is None:
        return []

    records = []  # list of {'bkg', 'booking', 'factory', 'contracts': set}
    cur_bkg = ''
    cur_booking = ''
    cur_factory = ''

    for row in target_ws.iter_rows(min_row=header_row + 1):
        cells = [c.value for c in row]
        if not any(cells):
            continue

        def _s(idx):
            if idx is None or idx >= len(cells):
                return ''
            return str(cells[idx] or '').strip()

        bkg = _s(col_bkg)
        booking = _s(col_booking) if col_booking is not None else ''
        factory = _s(col_factory) if col_factory is not None else ''
        contract = _s(col_contract) if col_contract is not None else ''

        if bkg:
            cur_bkg = bkg
        if booking:
            cur_booking = booking
        if factory:
            cur_factory = factory
            # 新 Booking 开始 → 新建记录
            records.append({'bkg': cur_bkg, 'booking': cur_booking, 'factory': cur_factory, 'contracts': set()})

        if records and contract:
            records[-1]['contracts'].add(contract)

    return records


def parse_yax_excel(file_path: str) -> dict:
    """解析 YAX 拖车通知单 Excel（马士基FCA并柜拖车通知单）。

    Args:
        file_path: .xlsx 文件路径

    Returns:
        dict: {
            'booking_number': str,        # 船公司订舱单号（YAX开头）
            'container_type': str,        # 柜型
            'customs_cutoff': str,        # 截关时间
            'si_deadline': str,           # 截补料时间
            'customs_broker': str,        # 报关行
            'po_so_mapping': list,        # 所有PO条目（含落货纸号码）
            'containers': list,           # 按柜分组结果，每柜一个dict：
                                          #   {'container_index': int,   # 第几柜（从1开始）
                                          #    'total_cbm': float,       # 该柜累计CBM
                                          #    'entries': list}          # 该柜的PO条目
        }
    """
    # 整柜标准体积，用于多柜分组（超过此值认为开始新的一柜）
    FULL_CONTAINER_CBM = 67.426

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    result = {
        'booking_number': '',
        'container_type': '',
        'customs_cutoff': '',
        'si_deadline': '',
        'customs_broker': '',
        'po_so_mapping': [],
        'containers': [],
    }

    po_table_start = None

    # 扫描所有行提取头部信息和定位 PO 表格
    for row_idx in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row_idx, column=1).value

        if cell_a is None:
            continue

        cell_a_str = str(cell_a).strip()
        cell_a_lower = cell_a_str.lower()

        # 提取头部字段（值可能在B列或C列）
        def _get_val(r=row_idx):
            for col in (2, 3, 4):
                v = ws.cell(row=r, column=col).value
                if v is not None and str(v).strip():
                    return str(v).strip()
            return ''

        if '订舱' in cell_a_str:
            v = _get_val()
            if v:
                result['booking_number'] = v
        elif '柜型' in cell_a_str:
            v = _get_val()
            if v:
                result['container_type'] = v
        elif '截关' in cell_a_str and '补料' not in cell_a_str:
            v = _get_val()
            if v:
                result['customs_cutoff'] = v
        elif '截补料' in cell_a_str or '补料' in cell_a_str:
            v = _get_val()
            if v:
                result['si_deadline'] = v

        # 提取报关行（从"报关员电话"列或包含"报关行"的单元格）
        if not result['customs_broker']:
            for col in range(1, min(ws.max_column + 1, 15)):
                cv = ws.cell(row=row_idx, column=col).value
                if cv and '报关行' in str(cv):
                    broker = str(cv).split(',')[0].split('，')[0].strip()
                    if broker and '报关方式' not in broker:
                        result['customs_broker'] = broker
                        break

        # 查找 PO 表格起始行
        if 'po' in cell_a_lower and po_table_start is None:
            row_text = ''
            for col in range(1, ws.max_column + 1):
                v = ws.cell(row=row_idx, column=col).value
                if v is not None:
                    row_text += str(v).lower() + ' '
            if 'po' in row_text and ('yat' in row_text or 'yax' in row_text
                                     or 'cbm' in row_text):
                po_table_start = row_idx

    # 解析 PO 表格
    if po_table_start is not None:
        header_row = po_table_start
        headers = []
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=header_row, column=col).value
            headers.append(str(v).strip().lower() if v is not None else '')

        po_col = None
        so_col = None
        cbm_col = None
        factory_col = None
        cargo_receipt_col = None  # 落货纸号码列

        for idx, h in enumerate(headers):
            if 'po' in h and po_col is None:
                po_col = idx
            elif h in ('yat', 'yax') or 'so' in h:
                so_col = idx
            elif h == 'cbm':
                cbm_col = idx
            elif 'factory' in h or 'supplier' in h:
                factory_col = idx
            elif any(kw in h for kw in ('cargo receipt', 'receipt no', '落货纸', 'receipt')):
                cargo_receipt_col = idx

        for row_idx in range(header_row + 1, ws.max_row + 1):
            row_values = []
            for col in range(1, ws.max_column + 1):
                row_values.append(ws.cell(row=row_idx, column=col).value)

            po_val = row_values[po_col] if po_col is not None and po_col < len(row_values) else None
            if po_val is None or str(po_val).strip() == '':
                continue

            entry = {
                'po': str(po_val).strip(),
                'so': '',
                'cbm': None,
                'factory': '',
                'cargo_receipt': '',   # 落货纸号码
            }

            # SO 号：优先找 YAX 开头的值
            if so_col is not None and so_col < len(row_values):
                so_val = row_values[so_col]
                if so_val is not None:
                    entry['so'] = str(so_val).strip()

            if not entry['so'].upper().startswith('YAX'):
                for idx, v in enumerate(row_values):
                    if v is not None and str(v).strip().upper().startswith('YAX'):
                        entry['so'] = str(v).strip()
                        break

            if cbm_col is not None and cbm_col < len(row_values):
                cbm_val = row_values[cbm_col]
                if cbm_val is not None:
                    try:
                        entry['cbm'] = float(cbm_val)
                    except (ValueError, TypeError):
                        entry['cbm'] = cbm_val

            if factory_col is not None and factory_col < len(row_values):
                f_val = row_values[factory_col]
                if f_val is not None:
                    entry['factory'] = str(f_val).strip()

            # 落货纸号码：从专属列取，没有则扫描每行所有单元格找以字母开头+纯数字的编号
            if cargo_receipt_col is not None and cargo_receipt_col < len(row_values):
                rv = row_values[cargo_receipt_col]
                if rv is not None:
                    entry['cargo_receipt'] = str(rv).strip()
            if not entry['cargo_receipt']:
                import re as _re
                for v in row_values:
                    if v is not None:
                        sv = str(v).strip()
                        # 落货纸号码格式：字母前缀+6位以上数字，如 CR2604001234、HK26040012
                        if _re.match(r'^[A-Za-z]{1,4}\d{6,}$', sv):
                            entry['cargo_receipt'] = sv
                            break

            result['po_so_mapping'].append(entry)

    # 按柜分组：累计CBM，超过 FULL_CONTAINER_CBM 时开新柜
    # 同时处理同一个YAX Excel有多柜的情况
    if result['po_so_mapping']:
        containers = []
        current_cbm = 0.0
        current_entries = []

        for entry in result['po_so_mapping']:
            cbm = entry.get('cbm') or 0.0
            try:
                cbm = float(cbm)
            except (ValueError, TypeError):
                cbm = 0.0

            # 若加入当前柜会超出整柜体积，且当前柜已有数据，则开新柜
            if current_entries and (current_cbm + cbm) > FULL_CONTAINER_CBM * 1.05:
                containers.append({
                    'container_index': len(containers) + 1,
                    'total_cbm': round(current_cbm, 3),
                    'entries': current_entries,
                })
                current_cbm = 0.0
                current_entries = []

            current_cbm += cbm
            current_entries.append(entry)

        if current_entries:
            containers.append({
                'container_index': len(containers) + 1,
                'total_cbm': round(current_cbm, 3),
                'entries': current_entries,
            })

        result['containers'] = containers

    wb.close()
    return result
