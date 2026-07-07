"""
柜单 Excel 生成引擎
"""
import os
import re
import tempfile
from datetime import datetime
from decimal import Decimal
from zoneinfo import ZoneInfo

CST = ZoneInfo('Asia/Shanghai')

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from apps.shipments.models import Shipment, ShipmentItem, ShipmentSubItem
from apps.master_data.models import FactoryMapping, ProductMapping

# ---------------------------------------------------------------------------
# 样式
# ---------------------------------------------------------------------------
THIN = Side(style='thin')
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NO_BORDER = Border()

FNT_TO = Font(name='Times New Roman', size=14)
FNT_TITLE = Font(name='Times New Roman', size=23)
FNT_SUB = Font(name='Times New Roman', size=19)
FNT_HDR = Font(name='Times New Roman', size=12)
FNT = Font(name='Times New Roman', size=13)
FNT_BOLD = Font(name='Times New Roman', size=12)
FNT_GREEN = Font(name='Times New Roman', size=11, color='008000')
FNT_RED = Font(name='Times New Roman', size=14, color='FF0000')
FNT_YLW = Font(name='Times New Roman', size=14)
FNT_FOOTER = Font(name='Times New Roman', size=13)

AC = Alignment(horizontal='center', vertical='center', wrap_text=True)
AL = Alignment(horizontal='left', vertical='center', wrap_text=True)
AR = Alignment(horizontal='right', vertical='center', wrap_text=True)
ALN = Alignment(horizontal='left', vertical='center', wrap_text=False)  # 不换行

YELLOW = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# 列定义
COLS = [
    ('备注', 8), ('序号', 8), ('洋行', 8), ('合同', 16), ('货号', 14),
    ('货名', 24), ('国家', 10), ('玩具\n类别', 9), ('数量', 10), ('件数', 10),
    ('毛重', 10), ('净重', 10), ('体积', 10), ('客人PO#', 16),
    ('每单\n总件数', 10), ('每箱\n毛重', 9), ('每箱\n净重', 9),
]
EXTRA_COL = ('Customer PO item No.', 14)


def _c(ws, r, c, v, font=FNT, align=AC, border=None, fill=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = font
    cell.alignment = align
    if border:
        cell.border = border
    if fill:
        cell.fill = fill
    return cell


def _merge(ws, r, c1, c2, v, font, align=AC):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    _c(ws, r, c1, v, font, align)


def _row_border(ws, r, n, font=FNT):
    for c in range(1, n + 1):
        cell = ws.cell(row=r, column=c)
        cell.border = THIN_BORDER
        cell.font = font
        cell.alignment = AC


def _extract_qty_per_box(product_name):
    """从货名中提取每箱个数，如 '迷你冰箱75个/箱' → 75"""
    m = re.search(r'(\d+)\s*个/箱', product_name or '')
    return int(m.group(1)) if m else 0


def _factory_remark(item):
    """获取工厂备注：优先用 factory_remark，为空时默认'兴信'（本厂）"""
    remark = (item.factory_remark or '').strip()
    if remark:
        return remark
    return '兴信'


def _has_en_po(items):
    return any(item.customer_po and re.search(r'[a-zA-Z]', item.customer_po) for item in items)


def _is_cab(s):
    return s.shipment_type in ('normal', 'customer_load', 'fcl')


def _row_country(item, s):
    """柜单"国家"列：优先 item.country（per-row 解析），回退 Shipment.delivery_address（用户手填收货地）。"""
    return item.country or s.delivery_address or ''


def _fd(dt):
    if not dt:
        return ''
    if hasattr(dt, 'tzinfo') and dt.tzinfo:
        dt = dt.astimezone(CST)
    return f'{dt.month}月{dt.day}日'


def _fdt(dt):
    if not dt:
        return ''
    if hasattr(dt, 'tzinfo') and dt.tzinfo:
        dt = dt.astimezone(CST)
    return f'{dt.month}月{dt.day}日 {dt.hour:02d}:{dt.minute:02d}'


def _dec(v):
    return Decimal(str(v)) if v is not None else Decimal('0')


def _title(s):
    if s.remarks and '柜' in s.remarks:
        return s.remarks
    cn = s.customer.name if s.customer else ''
    return f'{s.port or ""} {cn} {s.so_number or ""} 柜'


# ---------------------------------------------------------------------------
# 表头
# ---------------------------------------------------------------------------

def _hdr_normal(ws, s, nc):
    # Row 1: TO（合并A1:C1）
    _merge(ws, 1, 1, 3, 'TO: 廖平/王军', FNT_TO, AL)

    # 兴信做柜且有外厂货时，在王军行右侧显示"兴信拖柜，外厂列表送兴信拼柜，深圳报关"
    mf = s.main_factory or ''
    _ext_factories = []
    if mf and '兴信' in mf:
        _seen = set()
        for _it in s.items.all():
            _fr = (_it.factory_remark or '').strip()
            if _fr and _fr != '兴信' and _fr not in _seen:
                _seen.add(_fr)
                _ext_factories.append(_fr)
    if _ext_factories:
        _ext_txt = '兴信拖柜，' + '，'.join(_ext_factories) + '送兴信拼柜，深圳报关'
        # 按文字宽度动态计算合并列数，然后居中放置（不整行合并）
        _need_w = len(_ext_txt) * 2.2 + 8
        _col_w = [COLS[i][1] if i < len(COLS) else 9 for i in range(nc)]
        _total_w = sum(_col_w)
        # 居中起始偏移 = (总宽 - 文字宽) / 2，但不得进入 A-C（TO区域）
        _center_offset = (_total_w - _need_w) / 2
        _acc, _start_c = 0, 4
        for _ci in range(nc):
            if _acc >= _center_offset:
                _start_c = max(4, _ci + 1)
                break
            _acc += _col_w[_ci]
        _acc2, _end_c = 0, _start_c
        for _ci in range(_start_c - 1, nc):
            _acc2 += _col_w[_ci]
            _end_c = _ci + 1
            if _acc2 >= _need_w:
                break
        _end_c = min(_end_c, nc)
        _merge(ws, 1, _start_c, _end_c, _ext_txt, FNT_YLW, AC)
        for c in range(_start_c, _end_c + 1):
            ws.cell(row=1, column=c).fill = YELLOW

    # Row 2: 标题（合并F2:H2，居中显示）
    _merge(ws, 2, 6, 8, _title(s), FNT_TITLE, AC)

    if mf and '兴信' not in mf:
        # 去掉 HQ/CY 等仓库前缀，只保留工厂名
        _mf_clean = re.sub(r'^(HQ|CY|QX|SZ|GZ)\s*', '', mf).strip()
        txt = f'送{_mf_clean}拼柜，深圳报关'
        _merge(ws, 2, 13, 16, txt, FNT_YLW, AC)
        for c in range(13, 17):
            ws.cell(row=2, column=c).fill = YELLOW

    # Row 3: 副标题（合并E3:K3，不整行合并）
    _merge(ws, 3, 5, 11, f'{_fd(s.ship_date)} 出    {s.container_type or ""}  SO # {s.so_number or ""}', FNT_SUB, AC)


def _hdr_warehouse(ws, s, nc):
    _merge(ws, 1, 1, 3, 'TO：廖平/王军', FNT_TO, AL)
    cn = s.customer.name if s.customer else ''
    # 标题居中 F2:H2
    _merge(ws, 2, 6, 8, f'{s.port or "盐田"}{cn}{s.remarks.split("柜")[0].split("车")[0].split(cn)[-1] if s.remarks and cn in s.remarks else ""}车', FNT_TITLE, AC)
    # 副标题居中 E3:K3
    _merge(ws, 3, 5, 11, f'{_fd(s.ship_date)} 入{s.warehouse or ""}  {s.container_type or ""}  吨车如下：各', FNT_SUB, AC)


def _hdr_cl(ws, s, nc):
    _merge(ws, 1, 1, 3, 'TO: 廖平/王军', FNT_TO, AL)
    # 标题居中 F2:H2
    _merge(ws, 2, 6, 8, _title(s), FNT_TITLE, AC)
    # 右上角：报关行名称 + 客上柜备注（整柜不带"拼柜"）
    broker = s.customs_broker or ''
    mf = (s.main_factory or '').strip()
    special_req = (s.special_requirements or '').strip()
    is_zheng = not mf or mf == '兴信'  # 无主拼工厂或兴信=整柜
    # 昱升拼柜：special_requirements 已含完整描述，直接用作第二行
    if '昱升' in special_req:
        remark_text = f'{broker}\n{special_req}'
    elif is_zheng:
        remark_text = f'{broker}\n**客上柜**深圳报关'
    else:
        remark_text = f'{broker}\n**客上柜**拼柜，深圳报关'
    ws.merge_cells(start_row=2, start_column=13, end_row=2, end_column=16)
    _c(ws, 2, 13, remark_text, FNT_YLW, AC)
    for c in range(13, 17):
        ws.cell(row=2, column=c).fill = YELLOW
    cn = f' ({s.container_number})' if s.container_number else ''
    # 副标题居中 E3:K3
    _merge(ws, 3, 5, 11, f'{_fd(s.ship_date)} 出    {s.container_type or ""} SO #{s.so_number or ""}{cn}', FNT_SUB, AC)


def _hdr_qx(ws, s, nc):
    _merge(ws, 1, 1, 3, 'TO: 廖平/王军', FNT_TO, AL)
    cn = s.customer.name if s.customer else ''
    # 标题居中 F2:H2
    _merge(ws, 2, 6, 8, f'{cn}{s.so_number or ""}车，入清溪物流园', FNT_TITLE, AC)
    _merge(ws, 2, 13, 16, '**客上车**', FNT_RED, AC)
    for c in range(13, 17):
        ws.cell(row=2, column=c).fill = YELLOW
    # 副标题居中 E3:K3
    _merge(ws, 3, 5, 11, f'{_fd(s.ship_date)} 入清溪物流园  {s.container_type or ""}', FNT_SUB, AC)


def _hdr_customer_truck(ws, s, nc):
    """客上车：格式同交仓，右上角显示'客上车，送库有，深圳报关'"""
    _merge(ws, 1, 1, 3, 'TO：廖平/王军', FNT_TO, AL)
    cn = s.customer.name if s.customer else ''
    _merge(ws, 2, 6, 8, f'{s.port or "深圳"}{cn}车', FNT_TITLE, AC)
    _merge(ws, 2, 13, 16, '客上车，送库有，深圳报关', FNT_YLW, AC)
    for c in range(13, 17):
        ws.cell(row=2, column=c).fill = YELLOW
    _merge(ws, 3, 5, 11, f'{_fd(s.ship_date)} 送库有  {s.container_type or ""}  吨车如下：各', FNT_SUB, AC)


HDRS = {'normal': _hdr_normal, 'warehouse': _hdr_warehouse, 'customer_load': _hdr_cl, 'qingxi': _hdr_qx, 'customer_truck': _hdr_customer_truck}


# ---------------------------------------------------------------------------
# 主函数
# ---------------------------------------------------------------------------

def generate_container_sheet(shipment_id):
    s = Shipment.objects.select_related('customer', 'created_by').get(pk=shipment_id)
    items = list(ShipmentItem.objects.filter(shipment=s).prefetch_related('sub_items').order_by('seq_number'))

    # 预加载 ProductMapping.factory_short，用于自动填充备注列
    # 优先级：ShipmentItem.factory_remark > ProductMapping.factory_short > '兴信'
    _pm_remark_cache = {}
    _codes = [i.product_code for i in items if i.product_code]
    if _codes:
        for _pm in ProductMapping.objects.filter(product_code__in=_codes).values('product_code', 'factory_short'):
            if _pm['factory_short'] and _pm['product_code'] not in _pm_remark_cache:
                _pm_remark_cache[_pm['product_code']] = _pm['factory_short']

    def _factory_remark(item):
        remark = (item.factory_remark or '').strip()
        if remark:
            return remark
        cached = _pm_remark_cache.get(item.product_code or '', '')
        if cached:
            return cached
        return '兴信'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '柜单'

    has_ex = _has_en_po(items)
    cols = list(COLS)
    if has_ex:
        cols.append(EXTRA_COL)
    nc = len(cols)

    # 列宽
    for i, (_, w) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    cab = _is_cab(s)
    is_wh = s.shipment_type in ('warehouse', 'customer_truck')

    # === Row 1-3: 表头（无框线） ===
    HDRS.get(s.shipment_type, _hdr_normal)(ws, s, nc)
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 45
    ws.row_dimensions[3].height = 35
    ws.row_dimensions[4].height = 5

    # === Row 5: 列标题（有框线） ===
    ws.row_dimensions[5].height = 35
    for i, (name, _) in enumerate(cols, 1):
        _c(ws, 5, i, name, FNT_HDR, AC, THIN_BORDER)

    # === Row 6+: 数据 ===
    r = 6
    sq, sp = 0, 0
    sgw, snw, sv = Decimal('0'), Decimal('0'), Decimal('0')

    # 判断是否有 SLB 货号 → 决定品牌显示方式
    has_slb = any('SLB' in (item.product_code or '').upper() for item in items)
    # 收集所有品牌用于底部汇总
    all_brands = set()

    for item in items:
        sub_items = list(item.sub_items.all()) if hasattr(item, 'sub_items') else []
        has_mix = len(sub_items) > 0

        sq += item.quantity or 0
        sp += item.pieces or 0
        sgw += _dec(item.gross_weight)
        snw += _dec(item.net_weight)
        sv += _dec(item.volume)
        # 子行的体积也计入总和
        for sub in sub_items:
            sv += _dec(sub.volume)
            sq += sub.quantity or 0

        fn = item.product_name or ''

        if has_mix:
            # --- 混合装：主行 + 子行各占一行，共用字段合并单元格 ---
            total_rows = 1 + len(sub_items)  # 主行 + 子行数
            r1 = r  # 起始行
            r2 = r + total_rows - 1  # 结束行

            # 先给所有行加框线
            for rr in range(r1, r2 + 1):
                ws.row_dimensions[rr].height = 35
                _row_border(ws, rr, nc)

            # --- 主行（r1）：各自填写的列 ---
            _c(ws, r1, 1, _factory_remark(item), FNT, AC, THIN_BORDER)  # A: 备注
            _c(ws, r1, 2, item.seq_number, FNT, AC, THIN_BORDER)  # B: 序号
            _c(ws, r1, 3, item.trading_company or '', FNT, AC, THIN_BORDER)  # C: 洋行
            _c(ws, r1, 6, fn, FNT, AC, THIN_BORDER)  # F: 货名
            _c(ws, r1, 7, _row_country(item, s), FNT, AC, THIN_BORDER)  # G: 国家
            _c(ws, r1, 8, item.toy_category or '', FNT, AC, THIN_BORDER)  # H: 玩具类别
            # I: 数量 = 件数 × 货名中的每箱个数
            main_qty_per_box = _extract_qty_per_box(fn)
            main_qty = (item.pieces or 0) * main_qty_per_box if main_qty_per_box else item.quantity
            _c(ws, r1, 9, main_qty, FNT, AC, THIN_BORDER)  # I: 数量
            # L: 净重（取整）
            _c(ws, r1, 12, int(round(float(item.net_weight))) if item.net_weight else '', FNT, AC, THIN_BORDER)
            # Q: 每箱净重（两位小数）
            _c(ws, r1, 17, round(float(item.net_weight_per_box), 2) if item.net_weight_per_box else '', FNT, AC, THIN_BORDER)

            # --- 子行：各自填写的列 ---
            for si, sub in enumerate(sub_items):
                sr = r1 + 1 + si
                sub_fn = sub.product_name or ''
                _c(ws, sr, 1, _factory_remark(item), FNT, AC, THIN_BORDER)  # A: 备注
                _c(ws, sr, 2, item.seq_number + 1 + si, FNT, AC, THIN_BORDER)  # B: 序号递增
                _c(ws, sr, 3, item.trading_company or '', FNT, AC, THIN_BORDER)  # C: 洋行
                _c(ws, sr, 6, sub_fn, FNT, AC, THIN_BORDER)  # F: 货名
                _c(ws, sr, 7, sub.country or _row_country(item, s), FNT, AC, THIN_BORDER)  # G: 国家
                _c(ws, sr, 8, sub.toy_category or '', FNT, AC, THIN_BORDER)  # H: 玩具类别
                # I: 数量 = 件数 × 货名中的每箱个数
                sub_qty_per_box = _extract_qty_per_box(sub_fn)
                sub_qty = (item.pieces or 0) * sub_qty_per_box if sub_qty_per_box else sub.quantity
                _c(ws, sr, 9, sub_qty, FNT, AC, THIN_BORDER)  # I: 数量
                # L: 净重 = 件数 × 每箱净重（子行volume存的是每箱净重）
                sub_nw_per_box = float(sub.volume) if sub.volume else 0
                sub_net_weight = int(round((item.pieces or 0) * sub_nw_per_box)) if sub_nw_per_box else ''
                _c(ws, sr, 12, sub_net_weight, FNT, AC, THIN_BORDER)  # L: 净重
                # Q: 每箱净重
                _c(ws, sr, 17, round(sub_nw_per_box, 2) if sub_nw_per_box else '', FNT, AC, THIN_BORDER)

            # --- 合并单元格的列（共用数据） ---
            # O列：客上柜用SUMIFS公式汇总同货号+同合同+同客PO的总件数，其他类型用=J{r1}
            is_cl = s.shipment_type == 'customer_load'
            if _factory_remark(item) == '兴信':
                if is_cl:
                    o_formula = f'=SUMIFS(J$6:J$9999,E$6:E$9999,E{r1},D$6:D$9999,D{r1},N$6:N$9999,N{r1})'
                else:
                    o_formula = f'=J{r1}'
            else:
                o_formula = ''
            merge_cols = {
                4: item.contract_number or '',    # D: 合同
                5: item.product_code or '',        # E: 货号
                10: item.pieces,                   # J: 件数
                11: int(round(float(item.gross_weight))) if item.gross_weight else '',  # K: 毛重（取整）
                13: item.volume,                   # M: 体积
                14: item.customer_po or '',        # N: 客PO
                15: o_formula,                     # O: 每单总件数
                16: round(float(item.gross_weight_per_box), 2) if item.gross_weight_per_box else '',  # P: 每箱毛重（两位小数）
            }
            for col, val in merge_cols.items():
                if total_rows > 1:
                    ws.merge_cells(start_row=r1, start_column=col, end_row=r2, end_column=col)
                _c(ws, r1, col, val, FNT, AC, THIN_BORDER)

            r = r2 + 1

        else:
            # --- 普通行（无混合装） ---
            gw_per = round(float(item.gross_weight_per_box), 2) if item.gross_weight_per_box else ''
            nw_per = round(float(item.net_weight_per_box), 2) if item.net_weight_per_box else ''
            vals = [
                _factory_remark(item), item.seq_number, item.trading_company or '',
                item.contract_number or '', item.product_code or '', fn,
                _row_country(item, s), item.toy_category or '',
                item.quantity, item.pieces,
                None,  # K: 毛重（用公式）
                None,  # L: 净重（用公式）
                item.volume, item.customer_po or '',
                None,  # O: 每单总件数（用公式 =J{r}）
                gw_per, nw_per,
            ]
            if has_ex:
                vals.append(item.customer_po_item_no or '')

            ws.row_dimensions[r].height = 35
            for ci, v in enumerate(vals, 1):
                _c(ws, r, ci, v, FNT, AC, THIN_BORDER)
            # 毛重公式：=INT(件数 × 每箱毛重)，净重公式：=INT(件数 × 每箱净重)
            # SLB货号的 gw_per_box/nw_per_box 已在入库时存为每箱小数值（来自 ProductMapping）
            if gw_per:
                ws.cell(row=r, column=11).value = f'=INT(J{r}*P{r})'
            if nw_per:
                ws.cell(row=r, column=12).value = f'=INT(J{r}*Q{r})'
            # O列：客上柜用SUMIFS公式汇总同货号+同合同+同客PO的总件数，其他类型用=J{r}
            if _factory_remark(item) == '兴信':
                if s.shipment_type == 'customer_load':
                    ws.cell(row=r, column=15).value = (
                        f'=SUMIFS(J$6:J$9999,E$6:E$9999,E{r},D$6:D$9999,D{r},N$6:N$9999,N{r})'
                    )
                else:
                    ws.cell(row=r, column=15).value = f'=J{r}'
            r += 1

        # 收集品牌
        if item.brand:
            all_brands.add(item.brand)

        # 品牌+卡板行
        pl = item.pallet_count or 0
        if has_slb:
            # 有SLB货号：品牌跟在每个货号后面（含卡板信息）
            if item.brand or pl > 0:
                ws.row_dimensions[r].height = 40
                _row_border(ws, r, nc)
                if item.brand:
                    _c(ws, r, 6, f'品牌: {item.brand}', FNT, AL, THIN_BORDER)
                if pl > 0:
                    _c(ws, r, 10, f'{pl}卡板', FNT, AC, THIN_BORDER)
                    _c(ws, r, 11, pl, FNT, AC, THIN_BORDER)
                r += 1
        else:
            # 无SLB货号：品牌统一放底部，这里只显示卡板
            if pl > 0:
                ws.row_dimensions[r].height = 40
                _row_border(ws, r, nc)
                _c(ws, r, 10, f'{pl}卡板', FNT, AC, THIN_BORDER)
                _c(ws, r, 11, pl, FNT, AC, THIN_BORDER)
                r += 1

        # 箱规行（整柜不显示，交仓不显示（放底部），特定仓库不显示）
        show = not cab and not is_wh
        if show:
            wh = s.warehouse or ''
            for skip in ['盐田2号仓', '清溪物流园', '自提', '送外厂']:
                if skip in wh:
                    show = False
                    break
        if show and item.box_dimensions:
            ws.row_dimensions[r].height = 22
            _row_border(ws, r, nc)
            _c(ws, r, 6, f':{item.box_dimensions}', FNT, AL, THIN_BORDER)
            r += 1

    # === 品牌汇总行（无SLB时，统一放底部） ===
    if not has_slb and all_brands:
        # 按货号前缀分组品牌
        brand_lines = []
        brand_by_prefix = {}
        for item in items:
            if item.brand:
                code = item.product_code or ''
                # 精确前缀优先（从长到短）
                if code.startswith('15789'):
                    brand_by_prefix.setdefault('15789开头', set()).add(item.brand)
                elif code.startswith('15783'):
                    brand_by_prefix.setdefault('15783开头', set()).add(item.brand)
                elif code.startswith('15756'):
                    brand_by_prefix.setdefault('15756开头', set()).add(item.brand)
                elif code.startswith('92'):
                    brand_by_prefix.setdefault('92开头', set()).add(item.brand)
                elif code.startswith('157'):
                    brand_by_prefix.setdefault('157开头', set()).add(item.brand)
                elif code.startswith('95'):
                    brand_by_prefix.setdefault('95开头', set()).add(item.brand)
                elif code[0:1].isdigit() and code[0] == '7':
                    brand_by_prefix.setdefault('7字开头', set()).add(item.brand)
                else:
                    brand_by_prefix.setdefault('其他', set()).add(item.brand)
        for prefix, brands in brand_by_prefix.items():
            brand_lines.append(f'{prefix}品牌：{"／".join(sorted(brands))}')

        # 品牌汇总合并为一个大单元格（E:F合并，多行显示）
        brand_text = '\n'.join(brand_lines)
        line_count = len(brand_lines)
        ws.row_dimensions[r].height = max(40, line_count * 18)
        _row_border(ws, r, nc)
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        _c(ws, r, 5, brand_text, FNT, AL, THIN_BORDER)
        r += 1

    # === 合计行（用SUM公式）=== SI 内容放在合计行左侧
    ws.row_dimensions[r].height = 35
    _c(ws, r, 8, '合计', FNT_BOLD, AC, THIN_BORDER)
    _c(ws, r, 9, f'=SUM(I6:I{r-1})', FNT_BOLD, AC, THIN_BORDER)   # 数量
    _c(ws, r, 10, f'=SUM(J6:J{r-1})', FNT_BOLD, AC, THIN_BORDER)  # 件数
    _c(ws, r, 11, f'=SUM(K6:K{r-1})', FNT_BOLD, AC, THIN_BORDER)  # 毛重
    _c(ws, r, 12, f'=SUM(L6:L{r-1})', FNT_BOLD, AC, THIN_BORDER)  # 净重
    _c(ws, r, 13, f'=SUM(M6:M{r-1})', FNT_BOLD, AC, THIN_BORDER)  # 体积
    _c(ws, r, 14, 'CBM', FNT_BOLD, AL, THIN_BORDER)
    # 非交仓单：SI 显示在合计行左侧（A-D）
    if s.shipment_type != 'warehouse':
        _si_txt = f'SI: {_fdt(s.si_deadline)}' if s.si_deadline else ''
        _merge(ws, r, 1, 4, _si_txt, FNT_FOOTER, ALN)
    r += 1

    # === 底部信息 ===
    FH = 22
    is_wh = s.shipment_type in ('warehouse', 'customer_truck')
    cr = ''
    if s.created_by:
        cr = getattr(s.created_by, 'display_name', '') or s.created_by.username
    now = datetime.now()

    if is_wh:
        # 交仓底部：只有截数期、制表、收货人
        r += 1  # 空行

        # 截数期
        ws.row_dimensions[r].height = FH
        cd = f'截数期：{_fd(s.cutoff_date)}' if s.cutoff_date else ''
        _merge(ws, r, 1, 4, cd, FNT_FOOTER, ALN)
        r += 1

        # 制表
        ws.row_dimensions[r].height = FH
        _merge(ws, r, 1, 4, f'制表：{cr}  {now.year}/{now.month}/{now.day}', FNT_FOOTER, ALN)
        con = s.customer.consignee if s.customer else ''
        _merge(ws, r, 5, 8, f'收货人：{con}', FNT_FOOTER, ALN)
        code = s.customer.consignee_code if s.customer else ''
        _merge(ws, r, 12, nc, f'收货人代码：{code}', FNT_FOOTER, ALN)
        r += 1

        # 每个货号的长宽高（一行显示，多个用空格分隔）
        _skip_dims = False
        wh = s.warehouse or ''
        for skip in ['盐田2号仓', '清溪物流园', '自提', '送外厂']:
            if skip in wh:
                _skip_dims = True
                break
        if not _skip_dims:
            _dims = []
            for item in items:
                if item.box_dimensions:
                    _dims.append(f'{item.product_code}:#{item.box_dimensions}')
            if _dims:
                ws.row_dimensions[r].height = FH
                # 每3个一行
                for i in range(0, len(_dims), 3):
                    _line = '          '.join(_dims[i:i+3])
                    ws.row_dimensions[r].height = FH
                    _merge(ws, r, 1, nc, _line, FNT_FOOTER, ALN)
                    r += 1
    else:
        # 正常/客上柜底部

        def _weight_box(row, label):
            """在 H-I 写标签、J-K 写空值框（带边框）"""
            ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)
            _c(ws, row, 8, label, FNT_FOOTER, ALN, THIN_BORDER)
            ws.cell(row=row, column=9).border = THIN_BORDER
            ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=11)
            ws.cell(row=row, column=10).border = THIN_BORDER
            ws.cell(row=row, column=11).border = THIN_BORDER

        # 截数期行 + 填充物重量 + 柜号（同行右侧）
        ws.row_dimensions[r].height = FH
        cd = f'截数期：{_fdt(s.cutoff_date)}' if s.cutoff_date else ''
        _merge(ws, r, 1, 4, cd, FNT_FOOTER, ALN)
        _weight_box(r, '填充物重量')
        _merge(ws, r, 14, nc, f'柜号：{s.container_number or ""}', FNT_FOOTER, ALN)
        r += 1

        # 制表行 + 柜重 + 船封（同行右侧）
        ws.row_dimensions[r].height = FH
        _merge(ws, r, 1, 4, f'制表：{cr}  {now.year}/{now.month}/{now.day}', FNT_FOOTER, ALN)
        _weight_box(r, '柜重')
        _merge(ws, r, 14, nc, f'船封：{s.seal_number or ""}', FNT_FOOTER, ALN)
        r += 1

        # 整个集装箱重量（左右无内容）
        ws.row_dimensions[r].height = FH
        _weight_box(r, '整个集装箱重量')
        r += 1

        # 贸易方式 + 发货人
        ws.row_dimensions[r].height = FH
        _merge(ws, r, 1, 3, '贸易方式：进料对口', FNT_FOOTER, ALN)
        _merge(ws, r, 5, 10, '发货人：东莞兴信塑胶制品有限公司 (4419946995)', FNT_FOOTER, ALN)
        r += 1

        # 拼箱 + 收货人 + 收货人代码
        ws.row_dimensions[r].height = FH
        con = s.customer.consignee if s.customer else ''
        code = s.customer.consignee_code if s.customer else ''
        _merge(ws, r, 1, 3, '拼箱：是', FNT_FOOTER, ALN)
        _merge(ws, r, 5, 8, f'收货人：{con}', FNT_FOOTER, ALN)
        _merge(ws, r, 12, nc, f'收货人代码：{code}', FNT_FOOTER, ALN)
    r += 1

    # 送博锐或库有拼柜时，添加"带卡板送货"
    mf = (s.main_factory or '').strip()
    if any(k in mf for k in ('博锐', '搏锐', '库有')):
        ws.row_dimensions[r].height = 30
        FNT_KBD = Font(name='Times New Roman', size=14)
        _merge(ws, r, 14, nc, '带卡板送货', FNT_KBD, AC)
        for c in range(14, nc + 1):
            ws.cell(row=r, column=c).fill = YELLOW
        r += 1

    # 特殊要求/装柜备注 — 只在兴信做柜时显示（外厂做柜不显示）
    _mf = (s.main_factory or '').strip()
    _is_xingxin_maker = not _mf or '兴信' in _mf
    _remarks = (s.special_requirements or '').strip() if _is_xingxin_maker else ''
    if _remarks:
        r += 1
        FNT_REMARK = Font(name='Times New Roman', size=11, color='000000')
        # 右下角紧凑展示：从右侧往左按文字宽度计算起始列，不整行合并
        # 中文字符11pt约占1.8单位，加8单位留白；最左不超过第14列(N)
        _need_w = len(_remarks) * 1.8 + 8
        _acc_w, _start_c = 0, nc
        for _ci in range(nc - 1, 12, -1):   # 最左到第14列(index 13)
            _acc_w += COLS[_ci][1] if _ci < len(COLS) else 9
            _start_c = _ci + 1
            if _acc_w >= _need_w:
                break
        _start_c = max(_start_c, 14)        # 确保在右侧区域
        _merge(ws, r, _start_c, nc, _remarks, FNT_REMARK,
               Alignment(horizontal='center', vertical='center', wrap_text=True))
        for c in range(_start_c, nc + 1):
            ws.cell(row=r, column=c).fill = YELLOW
        # 根据文字长度和可用列宽自动调整行高
        _avail_w = sum(COLS[i][1] for i in range(_start_c - 1, nc) if i < len(COLS))
        _lines = max(1, len(_remarks) * 1.8 // max(_avail_w, 1)) + 1
        ws.row_dimensions[r].height = max(30, int(_lines) * 18)

    # === 页面设置（A4横向，缩放至1页宽）===
    ws.page_setup.paperSize = ws.PAPERSIZE_A4   # A4纸
    ws.page_setup.orientation = 'landscape'      # 横向
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # 保存 — 文件名格式：ZURU合同号港口编号货号（如 ZURU4500194240盐田1441柜77673TQ1）
    _items_qs = s.items.all()
    _contract = next((i.contract_number for i in _items_qs if i.contract_number), '') or ''
    _port = s.port or ''
    # 柜号：从 remarks 提取数字，如"盐田 ZURU 1441 柜" → "1441柜"
    _cab_m = re.search(r'(\d+)\s*柜', s.remarks or '')
    _cab = f'{_cab_m.group(1)}柜' if _cab_m else ''
    _codes = ''.join(i.product_code for i in _items_qs if i.product_code)
    _raw_name = f'ZURU{_contract}{_port}{_cab}{_codes}'
    safe_name = re.sub(r'[<>:"/\\|?*\s]', '', _raw_name)[:80]  # 限制文件名长度避免超出系统限制
    path = os.path.join(tempfile.gettempdir(), f'{safe_name}.xlsx')
    wb.save(path)
    wb.close()
    return path
