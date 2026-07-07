"""
扫描X盘船务部共享的柜单文件，按客户分类导入货号映射。
Usage: python manage.py import_product_mappings
"""
import os
import re
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand
from apps.master_data.models import ProductMapping

SKIP_CN = re.compile(r'[\u4e00-\u9fff]')
SKIP_LONG = re.compile(r'^\d{11,}$')

# 已知客户名（用于从目录名匹配）
KNOWN_CUSTOMERS = ['ZURU', 'MOOSE', 'TIG', 'TOMY', 'CEPIA', 'AZAD', 'ZANZOON']

# 扫描目录配置
BASE = r'X:\船务部共享'
SCAN_DIRS = [
    os.path.join(BASE, '2020-2023-2024年出货明细', '2020年出货明细'),
    os.path.join(BASE, '2020-2023-2024年出货明细', '2021年出货明细'),
    os.path.join(BASE, '2020-2023-2024年出货明细', '2022年出货明细'),
    os.path.join(BASE, '2020-2023-2024年出货明细', '2023年出货资料'),
    os.path.join(BASE, '2020-2023-2024年出货明细', '2024年出货资料'),
    os.path.join(BASE, '2025年出货资料'),
    os.path.join(BASE, '2026-1月'),
    os.path.join(BASE, '2026-2月'),
    os.path.join(BASE, '2026-3月'),
]


def _guess_customer(filepath):
    """从文件路径推断客户名"""
    parts = filepath.upper().replace('\\', '/').split('/')
    for part in parts:
        for c in KNOWN_CUSTOMERS:
            if c.upper() == part.strip().upper():
                return c
    # 从文件名前缀推断
    fname = os.path.basename(filepath).upper()
    if fname.startswith('ZUR'):
        return 'ZURU'
    for c in KNOWN_CUSTOMERS:
        if c.upper() in fname:
            return c
    return ''


def _valid_code(code):
    """检查是否有效货号"""
    if not code or len(code) < 2 or len(code) > 30:
        return False
    if SKIP_CN.search(code):
        return False
    if SKIP_LONG.match(code):
        return False
    if code in ('0', '0.0', 'None', 'nan'):
        return False
    return True


def _to_dec(val):
    """转Decimal"""
    if val is None or val == '' or val == 0:
        return None
    try:
        d = Decimal(str(val)).quantize(Decimal('0.001'))
        return d if d > 0 else None
    except (InvalidOperation, ValueError):
        return None


def _find_header_row(ws_func, max_rows, max_cols):
    """查找表头行（含'货号'的行）"""
    for r in range(max_rows):
        for c in range(min(max_cols, 10)):
            val = ws_func(r, c)
            if val and '货号' in str(val):
                return r
    return None


def _parse_xls(filepath, customer):
    """解析.xls文件"""
    import xlrd
    try:
        wb = xlrd.open_workbook(filepath)
    except Exception:
        return []

    results = []
    ws = wb.sheet_by_index(0)
    hr = _find_header_row(lambda r, c: ws.cell_value(r, c) if c < ws.ncols else '', min(ws.nrows, 10), ws.ncols)
    if hr is None:
        return []

    # 找货号列位置
    code_col = None
    name_col = None
    cat_col = None
    gw_col = None
    nw_col = None
    for c in range(ws.ncols):
        val = str(ws.cell_value(hr, c)).strip()
        if '货号' in val and code_col is None:
            code_col = c
        elif '货名' in val and name_col is None:
            name_col = c
        elif '类别' in val and cat_col is None:
            cat_col = c
        elif '每箱' in val and '毛' in val and gw_col is None:
            gw_col = c
        elif '每箱' in val and '净' in val and nw_col is None:
            nw_col = c

    if code_col is None:
        return []

    for r in range(hr + 1, ws.nrows):
        code = str(ws.cell_value(r, code_col)).strip()
        # 去掉.0后缀
        if code.endswith('.0'):
            code = code[:-2]
        if not _valid_code(code):
            continue

        name = str(ws.cell_value(r, name_col)).strip() if name_col and name_col < ws.ncols else ''
        cat = str(ws.cell_value(r, cat_col)).strip() if cat_col and cat_col < ws.ncols else ''
        gw = _to_dec(ws.cell_value(r, gw_col)) if gw_col and gw_col < ws.ncols else None
        nw = _to_dec(ws.cell_value(r, nw_col)) if nw_col and nw_col < ws.ncols else None

        results.append({
            'code': code, 'name': name, 'cat': cat,
            'gw': gw, 'nw': nw, 'customer': customer,
        })
    return results


def _parse_xlsx(filepath, customer):
    """解析.xlsx文件"""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    except Exception:
        return []

    results = []
    ws = wb.active

    def cell_val(r, c):
        try:
            row_data = list(ws.iter_rows(min_row=r+1, max_row=r+1, values_only=True))
            if row_data and c < len(row_data[0]):
                return row_data[0][c]
        except Exception:
            pass
        return None

    # 读取前10行找表头
    rows_cache = []
    for i, row in enumerate(ws.iter_rows(max_row=10, values_only=True)):
        rows_cache.append(row)

    hr = None
    for r, row in enumerate(rows_cache):
        for val in row:
            if val and '货号' in str(val):
                hr = r
                break
        if hr is not None:
            break

    if hr is None:
        wb.close()
        return []

    header = rows_cache[hr]
    code_col = name_col = cat_col = gw_col = nw_col = None
    for c, val in enumerate(header):
        v = str(val or '').strip()
        if '货号' in v and code_col is None:
            code_col = c
        elif '货名' in v and name_col is None:
            name_col = c
        elif '类别' in v and cat_col is None:
            cat_col = c
        elif '每箱' in v and '毛' in v and gw_col is None:
            gw_col = c
        elif '每箱' in v and '净' in v and nw_col is None:
            nw_col = c

    if code_col is None:
        wb.close()
        return []

    for i, row in enumerate(ws.iter_rows(min_row=hr+2, values_only=True)):
        if code_col >= len(row):
            continue
        code = str(row[code_col] or '').strip()
        if code.endswith('.0'):
            code = code[:-2]
        if not _valid_code(code):
            continue

        name = str(row[name_col] or '').strip() if name_col and name_col < len(row) else ''
        cat = str(row[cat_col] or '').strip() if cat_col and cat_col < len(row) else ''
        gw = _to_dec(row[gw_col]) if gw_col and gw_col < len(row) else None
        nw = _to_dec(row[nw_col]) if nw_col and nw_col < len(row) else None

        results.append({
            'code': code, 'name': name, 'cat': cat,
            'gw': gw, 'nw': nw, 'customer': customer,
        })

    wb.close()
    return results


class Command(BaseCommand):
    help = '从X盘导入货号映射（按客户分类）'

    def handle(self, *args, **options):
        self.stdout.write('开始扫描...')
        all_items = []
        file_count = 0
        skip_count = 0

        for scan_dir in SCAN_DIRS:
            if not os.path.exists(scan_dir):
                self.stdout.write(f'  跳过不存在: {scan_dir}')
                continue
            self.stdout.write(f'  扫描: {scan_dir}')

            for root, dirs, files in os.walk(scan_dir):
                for fname in files:
                    # 跳过临时文件
                    if fname.startswith('~$'):
                        continue
                    ext = os.path.splitext(fname)[1].lower()
                    if ext not in ('.xls', '.xlsx'):
                        continue

                    filepath = os.path.join(root, fname)

                    # 文件大小限制
                    try:
                        size = os.path.getsize(filepath)
                    except OSError:
                        continue
                    if ext == '.xls' and size > 2 * 1024 * 1024:
                        skip_count += 1
                        continue
                    if ext == '.xlsx' and size > 1 * 1024 * 1024:
                        skip_count += 1
                        continue

                    customer = _guess_customer(filepath)
                    file_count += 1

                    try:
                        if ext == '.xls':
                            items = _parse_xls(filepath, customer)
                        else:
                            items = _parse_xlsx(filepath, customer)
                        all_items.extend(items)
                    except Exception:
                        pass

                    if file_count % 500 == 0:
                        self.stdout.write(f'    已扫描 {file_count} 文件, {len(all_items)} 条数据...')

        self.stdout.write(f'\n扫描完成: {file_count} 文件, {len(all_items)} 条原始数据, 跳过 {skip_count} 个大文件')

        # 去重并导入（同一客户同一货号只保留最新的有重量数据的）
        self.stdout.write('正在导入...')
        merged = {}
        for item in all_items:
            key = (item['code'], item['customer'])
            existing = merged.get(key)
            if existing is None:
                merged[key] = item
            else:
                # 有重量数据的优先
                if item['gw'] and not existing['gw']:
                    merged[key] = item
                elif item['name'] and not existing['name']:
                    existing['name'] = item['name']
                    existing['cat'] = item['cat'] or existing['cat']

        created = 0
        updated = 0
        for (code, customer), item in merged.items():
            obj, is_new = ProductMapping.objects.update_or_create(
                product_code=code,
                customer_name=customer,
                defaults={
                    'product_name': item['name'] or '',
                    'toy_category': item['cat'] or '',
                    'gross_weight_per_box': item['gw'],
                    'net_weight_per_box': item['nw'],
                    'source': 'X盘',
                },
            )
            if is_new:
                created += 1
            else:
                updated += 1

        self.stdout.write(self.style.SUCCESS(
            f'\n导入完成: 新增 {created}, 更新 {updated}, 合计 {created + updated}'
        ))

        # 按客户统计
        from django.db.models import Count
        stats = ProductMapping.objects.values('customer_name').annotate(cnt=Count('id')).order_by('-cnt')
        self.stdout.write('\n按客户统计:')
        for s in stats:
            self.stdout.write(f'  {s["customer_name"] or "(未知)"}: {s["cnt"]}')
