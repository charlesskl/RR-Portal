import re
import tempfile
from decimal import Decimal, InvalidOperation

from rest_framework import viewsets, filters, status
from rest_framework.decorators import api_view, parser_classes
from rest_framework.parsers import MultiPartParser
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response

from rest_framework.permissions import IsAuthenticated

from .models import Customer, TransportCompany, FactoryMapping, ProductMapping, DestinationPortMapping
from .serializers import (
    CustomerSerializer,
    TransportCompanySerializer,
    FactoryMappingSerializer,
    ProductMappingSerializer,
    DestinationPortMappingSerializer,
)


class StandardPagination(PageNumberPagination):
    page_size = 50
    page_size_query_param = 'page_size'
    max_page_size = 200


class CustomerViewSet(viewsets.ModelViewSet):
    queryset = Customer.objects.all()
    serializer_class = CustomerSerializer


class TransportCompanyViewSet(viewsets.ModelViewSet):
    queryset = TransportCompany.objects.all()
    serializer_class = TransportCompanySerializer


class FactoryMappingViewSet(viewsets.ModelViewSet):
    queryset = FactoryMapping.objects.all()
    serializer_class = FactoryMappingSerializer


class ProductMappingViewSet(viewsets.ModelViewSet):
    serializer_class = ProductMappingSerializer
    pagination_class = StandardPagination
    filter_backends = [filters.SearchFilter]
    search_fields = ['product_code', 'product_name']

    def get_queryset(self):
        from django.db.models import Case, When, Value, IntegerField, Q
        from django.db.models import F
        qs = ProductMapping.objects.annotate(
            _sort_priority=Case(
                # 无货名排后面
                When(product_name='', then=Value(1)),
                # 假货名排后面：N卡板、立放、品牌:、装柜、纯数字等
                When(product_name__regex=r'^\d+\s*卡板$', then=Value(1)),
                When(product_name__in=['立放', '立装放柜', '立放装柜'], then=Value(1)),
                When(product_name__startswith='品牌:', then=Value(1)),
                When(product_name__regex=r'^在.+[装做]柜$', then=Value(1)),
                When(product_name__contains='要求立放', then=Value(1)),
                When(product_name__regex=r'^第.+层要求', then=Value(1)),
                When(product_name__regex=r'^\d+(\.\d+)?$', then=Value(1)),
                # 真正有货名的排前面
                default=Value(0),
                output_field=IntegerField(),
            )
        ).order_by(
            '_sort_priority', 'customer_name', 'product_code',
            # 同货号多规格时：无个数装（默认规格）排最前，再按个数升序
            F('qty_per_box').asc(nulls_first=True)
        )
        customer = self.request.query_params.get('customer')
        if customer:
            qs = qs.filter(customer_name__iexact=customer)
        source = self.request.query_params.get('source')
        if source:
            qs = qs.filter(source=source)
        return qs


SKIP_CN = re.compile(r'[\u4e00-\u9fff]')


def _parse_qty_per_box(name: str):
    """
    从货名中解析每箱个数，作为多规格区分键。
    支持格式：
      "卡哇伊DIY球25个/箱"  → 25
      "狗蛋四代6个/箱"      → 6
      "12pcs/ctn"           → 12（英文格式）
    找不到则返回 None（视为默认规格）。
    """
    if not name:
        return None
    # 中文格式：数字 + 个/箱
    m = re.search(r'(\d+)\s*个\s*[/／]\s*箱', name)
    if m:
        return int(m.group(1))
    # 英文格式：数字 + pcs/ctn 或 pc/ctn
    m = re.search(r'(\d+)\s*pcs?\s*/\s*ctn', name, re.IGNORECASE)
    if m:
        return int(m.group(1))
    return None


def _to_dec(val):
    if val is None or val == '' or val == 0:
        return None
    try:
        d = Decimal(str(val)).quantize(Decimal('0.001'))
        return d if d > 0 else None
    except (InvalidOperation, ValueError):
        return None


def _parse_upload_excel(filepath):
    """解析上传的柜单Excel，提取货号映射"""
    ext = filepath.rsplit('.', 1)[-1].lower()
    items = []

    if ext == 'xls':
        import xlrd
        try:
            wb = xlrd.open_workbook(filepath)
        except Exception:
            return []
        ws = wb.sheet_by_index(0)
        # 找表头
        hr = None
        code_col = name_col = cat_col = gw_col = nw_col = rmk_col = None
        for r in range(min(ws.nrows, 10)):
            for c in range(ws.ncols):
                if ws.cell_value(r, c) and '货号' in str(ws.cell_value(r, c)):
                    hr = r
                    break
            if hr is not None:
                break
        if hr is None:
            return []
        for c in range(ws.ncols):
            v = str(ws.cell_value(hr, c)).strip()
            if '货号' in v and code_col is None: code_col = c
            elif '货名' in v and name_col is None: name_col = c
            elif '类别' in v and cat_col is None: cat_col = c
            elif '每箱' in v and '毛' in v and gw_col is None: gw_col = c
            elif '每箱' in v and '净' in v and nw_col is None: nw_col = c
            elif '备注' in v and rmk_col is None: rmk_col = c
        if code_col is None:
            return []
        for r in range(hr + 1, ws.nrows):
            code = str(ws.cell_value(r, code_col)).strip()
            if code.endswith('.0'): code = code[:-2]
            if not code or len(code) < 2 or SKIP_CN.search(code): continue
            items.append({
                'product_code': code,
                'product_name': str(ws.cell_value(r, name_col)).strip() if name_col and name_col < ws.ncols else '',
                'toy_category': str(ws.cell_value(r, cat_col)).strip() if cat_col and cat_col < ws.ncols else '',
                'gross_weight_per_box': _to_dec(ws.cell_value(r, gw_col)) if gw_col and gw_col < ws.ncols else None,
                'net_weight_per_box': _to_dec(ws.cell_value(r, nw_col)) if nw_col and nw_col < ws.ncols else None,
                'factory_short': str(ws.cell_value(r, rmk_col)).strip() if rmk_col is not None and rmk_col < ws.ncols else '',
            })
    else:
        import openpyxl
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        except Exception:
            return []
        ws = wb.active
        rows = list(ws.iter_rows(max_row=min(ws.max_row or 100, 100), values_only=True))
        wb.close()
        hr = None
        code_col = name_col = cat_col = gw_col = nw_col = rmk_col = None
        for r, row in enumerate(rows[:10]):
            for val in row:
                if val and '货号' in str(val):
                    hr = r
                    break
            if hr is not None:
                break
        if hr is None:
            return []
        for c, val in enumerate(rows[hr]):
            v = str(val or '').strip()
            if '货号' in v and code_col is None: code_col = c
            elif '货名' in v and name_col is None: name_col = c
            elif '类别' in v and cat_col is None: cat_col = c
            elif '每箱' in v and '毛' in v and gw_col is None: gw_col = c
            elif '每箱' in v and '净' in v and nw_col is None: nw_col = c
            elif '备注' in v and rmk_col is None: rmk_col = c
        if code_col is None:
            return []
        for row in rows[hr + 1:]:
            if code_col >= len(row): continue
            code = str(row[code_col] or '').strip()
            if code.endswith('.0'): code = code[:-2]
            if not code or len(code) < 2 or SKIP_CN.search(code): continue
            items.append({
                'product_code': code,
                'product_name': str(row[name_col] or '').strip() if name_col and name_col < len(row) else '',
                'toy_category': str(row[cat_col] or '').strip() if cat_col and cat_col < len(row) else '',
                'gross_weight_per_box': _to_dec(row[gw_col]) if gw_col and gw_col < len(row) else None,
                'net_weight_per_box': _to_dec(row[nw_col]) if nw_col and nw_col < len(row) else None,
                'factory_short': str(row[rmk_col] or '').strip() if rmk_col is not None and rmk_col < len(row) else '',
            })
    return items


@api_view(['POST'])
@parser_classes([MultiPartParser])
def import_daily(request):
    """每日新增导入柜单Excel到货号映射"""
    files = request.FILES.getlist('files')
    customer = request.data.get('customer', '')
    if not files:
        return Response({'error': '请上传文件'}, status=status.HTTP_400_BAD_REQUEST)

    import os
    all_items = []
    errors = []
    for f in files:
        # 保存临时文件
        ext = f.name.rsplit('.', 1)[-1] if '.' in f.name else 'xls'
        tmp = tempfile.NamedTemporaryFile(suffix=f'.{ext}', delete=False)
        for chunk in f.chunks():
            tmp.write(chunk)
        tmp.close()
        try:
            items = _parse_upload_excel(tmp.name)
            all_items.extend(items)
        except Exception as e:
            errors.append(f'{f.name}: {str(e)}')
        finally:
            os.unlink(tmp.name)

    created = 0
    updated = 0
    result_items = []
    for item in all_items:
        # 从货名解析每箱个数，作为独立规格的区分键：
        # "25个/箱" → qty=25，"12个/箱" → qty=12，找不到 → qty=None（默认规格）
        # 相同货号 + 不同 qty_per_box → 各自独立保存，不覆盖
        qty = _parse_qty_per_box(item.get('product_name', ''))
        defaults = {
            'product_name': item['product_name'] or '',
            'toy_category': item['toy_category'] or '',
            'gross_weight_per_box': item['gross_weight_per_box'],
            'net_weight_per_box': item['net_weight_per_box'],
            'source': 'daily',
        }
        # 备注列（做货工厂）：有值且不是默认的"兴信"才写入，避免覆盖手动设置的值
        fs = (item.get('factory_short') or '').strip()
        if fs and fs != '兴信':
            defaults['factory_short'] = fs
        obj, is_new = ProductMapping.objects.update_or_create(
            product_code=item['product_code'],
            customer_name=customer,
            qty_per_box=qty,
            defaults=defaults,
        )
        if is_new:
            created += 1
            item['status'] = 'new'
        else:
            updated += 1
            item['status'] = 'updated'
        result_items.append(item)

    return Response({
        'created': created,
        'updated': updated,
        'errors': errors,
        'total_parsed': len(all_items),
    })


class DestinationPortMappingViewSet(viewsets.ModelViewSet):
    queryset = DestinationPortMapping.objects.all().order_by('port_name')
    serializer_class = DestinationPortMappingSerializer
    permission_classes = [IsAuthenticated]
