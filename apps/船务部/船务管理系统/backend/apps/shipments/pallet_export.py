"""卡板数报表 Excel 生成器。"""
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .pallet_report import group_by_factory_so

# 样式
_TITLE_FONT = Font(name='微软雅黑', size=16, bold=True)
_HEADER_FONT = Font(name='微软雅黑', size=11, bold=True)
_BOLD_FONT = Font(bold=True)
_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
_LEFT = Alignment(horizontal='left', vertical='center')

_TITLE_FILL = PatternFill(start_color='D9E8F5', end_color='D9E8F5', fill_type='solid')
_CATEGORY_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
_FACTORY_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

_DOUBLE_TOP = Border(top=Side(style='double'))


def generate_xlsx(data: dict) -> bytes:
    """生成单 sheet 卡板数报表 xlsx。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '卡板数报表'

    widths = [12, 16, 14, 28, 14, 14, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    row = 1

    # 标题
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=1, value='卡板数月度统计报表')
    cell.font = _TITLE_FONT
    cell.alignment = _CENTER
    cell.fill = _TITLE_FILL
    row += 1

    ws.cell(row=row, column=1, value=f'统计期间: {data["period_start"]} ~ {data["period_end"]}')
    row += 1
    ws.cell(row=row, column=1, value=f'筛选工厂: {data["factories_filter"]}    筛选分类: {data["categories_filter"]}')
    row += 1
    ws.cell(row=row, column=1, value=f'导出时间: {datetime.now().strftime("%Y-%m-%d %H:%M")}')
    row += 2

    grand_total_pallets = 0

    if data.get('self_items'):
        row, sub = _write_self_section(ws, row, data['self_items'])
        grand_total_pallets += sub

    if data.get('local_items'):
        row, sub = _write_grouped_section(ws, row, '送外厂卡板（兴信→外厂拼柜）',
                                          data['local_items'], factory_field='zuogui_factory')
        grand_total_pallets += sub

    if data.get('external_items'):
        row, sub = _write_grouped_section(ws, row, '外厂送来卡板（外厂→兴信拼柜）',
                                          data['external_items'], factory_field='factory_remark')
        grand_total_pallets += sub

    if data.get('manual_borui'):
        row, sub = _write_manual_section(ws, row, '送博锐手填', data['manual_borui'])
        grand_total_pallets += sub

    if data.get('manual_kuyou'):
        row, sub = _write_manual_section(ws, row, '送库有手填', data['manual_kuyou'])
        grand_total_pallets += sub

    row += 1
    cell = ws.cell(row=row, column=1, value=f'总计: {grand_total_pallets} 卡板')
    cell.font = _BOLD_FONT
    cell.border = _DOUBLE_TOP

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_self_section(ws, row, items):
    """本厂做柜：按 SO 分组，2 层"""
    subtotal = sum(int(it.get('pallet_count') or 0) for it in items)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row=row, column=1, value=f'【本厂做柜卡板（兴信做柜）】  小计: {subtotal} 卡板')
    c.font = _HEADER_FONT; c.fill = _CATEGORY_FILL
    row += 1
    headers = ['时间', 'SO号', '货号', '货名', '合同号', '客PO', '件数', '卡板数']
    for i, h in enumerate(headers, 1):
        cc = ws.cell(row=row, column=i, value=h); cc.font = _BOLD_FONT; cc.alignment = _CENTER
    row += 1
    so_groups = {}
    for it in items:
        so = it.get('so_number') or '-'
        so_groups.setdefault(so, []).append(it)
    for so, lst in so_groups.items():
        so_pallets = sum(int(it.get('pallet_count') or 0) for it in lst)
        so_pieces = sum(int(it.get('pieces') or 0) for it in lst)
        ws.cell(row=row, column=1, value=lst[0].get('ship_date', '')).font = _BOLD_FONT
        ws.cell(row=row, column=2, value=so).font = _BOLD_FONT
        ws.cell(row=row, column=7, value=so_pieces).font = _BOLD_FONT
        ws.cell(row=row, column=8, value=so_pallets).font = _BOLD_FONT
        row += 1
        for it in lst:
            ws.cell(row=row, column=3, value=it.get('product_code', ''))
            ws.cell(row=row, column=4, value=it.get('product_name', ''))
            ws.cell(row=row, column=5, value=it.get('contract_number', ''))
            ws.cell(row=row, column=6, value=it.get('customer_po', ''))
            ws.cell(row=row, column=7, value=int(it.get('pieces') or 0))
            ws.cell(row=row, column=8, value=int(it.get('pallet_count') or 0))
            row += 1
    cell = ws.cell(row=row, column=8, value=f'合计: {subtotal}')
    cell.font = _BOLD_FONT
    row += 2
    return row, subtotal


def _write_grouped_section(ws, row, title, items, factory_field):
    """送外厂/外厂送来：3 层（工厂→SO→货号）"""
    subtotal = sum(int(it.get('pallet_count') or 0) for it in items)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row=row, column=1, value=f'【{title}】  小计: {subtotal} 卡板')
    c.font = _HEADER_FONT; c.fill = _CATEGORY_FILL
    row += 1

    grouped = group_by_factory_so(items, factory_field)
    for factory, so_dict in grouped.items():
        factory_pallets = sum(int(it.get('pallet_count') or 0)
                              for lst in so_dict.values() for it in lst)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        c = ws.cell(row=row, column=1, value=f'  【{factory}】小计: {factory_pallets}')
        c.font = _BOLD_FONT; c.fill = _FACTORY_FILL
        row += 1
        headers = ['时间', 'SO号', '货号', '货名', '合同号', '客PO', '件数', '卡板数']
        for i, h in enumerate(headers, 1):
            cc = ws.cell(row=row, column=i, value=h); cc.font = _BOLD_FONT; cc.alignment = _CENTER
        row += 1
        for so, lst in so_dict.items():
            so_pallets = sum(int(it.get('pallet_count') or 0) for it in lst)
            so_pieces = sum(int(it.get('pieces') or 0) for it in lst)
            ws.cell(row=row, column=1, value=lst[0].get('ship_date', '')).font = _BOLD_FONT
            ws.cell(row=row, column=2, value=so).font = _BOLD_FONT
            ws.cell(row=row, column=7, value=so_pieces).font = _BOLD_FONT
            ws.cell(row=row, column=8, value=so_pallets).font = _BOLD_FONT
            row += 1
            for it in lst:
                ws.cell(row=row, column=3, value=it.get('product_code', ''))
                ws.cell(row=row, column=4, value=it.get('product_name', ''))
                ws.cell(row=row, column=5, value=it.get('contract_number', ''))
                ws.cell(row=row, column=6, value=it.get('customer_po', ''))
                ws.cell(row=row, column=7, value=int(it.get('pieces') or 0))
                ws.cell(row=row, column=8, value=int(it.get('pallet_count') or 0))
                row += 1
    row += 1
    return row, subtotal


def _write_manual_section(ws, row, title, items):
    """送博锐/送库有手填：2 层（SO→货号）"""
    subtotal = sum(int(it.get('pallet_count') or 0) for it in items)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row=row, column=1, value=f'【{title}】  小计: {subtotal} 卡板')
    c.font = _HEADER_FONT; c.fill = _CATEGORY_FILL
    row += 1
    headers = ['日期', 'SO号', '货号', '货名', '合同号', '', '件数', '卡板数']
    for i, h in enumerate(headers, 1):
        cc = ws.cell(row=row, column=i, value=h); cc.font = _BOLD_FONT; cc.alignment = _CENTER
    row += 1
    so_groups = {}
    for it in items:
        so = it.get('so_number') or '-'
        so_groups.setdefault(so, []).append(it)
    for so, lst in so_groups.items():
        so_pallets = sum(int(it.get('pallet_count') or 0) for it in lst)
        so_pieces = sum(int(it.get('pieces') or 0) for it in lst)
        ws.cell(row=row, column=1, value=lst[0].get('date', '')).font = _BOLD_FONT
        ws.cell(row=row, column=2, value=so).font = _BOLD_FONT
        ws.cell(row=row, column=7, value=so_pieces).font = _BOLD_FONT
        ws.cell(row=row, column=8, value=so_pallets).font = _BOLD_FONT
        row += 1
        for it in lst:
            ws.cell(row=row, column=3, value=it.get('product_code', ''))
            ws.cell(row=row, column=4, value=it.get('product_name', ''))
            ws.cell(row=row, column=5, value=it.get('contract_number', ''))
            ws.cell(row=row, column=7, value=int(it.get('pieces') or 0))
            ws.cell(row=row, column=8, value=int(it.get('pallet_count') or 0))
            row += 1
    row += 1
    return row, subtotal
