"""Excel 解析器单元测试。"""

import os
import tempfile

import openpyxl
from django.test import TestCase

from apps.emails.parsers.excel_parser import _find_header_row, _match_field, parse_packing_list, parse_yax_excel


class FindHeaderRowTest(TestCase):
    """_find_header_row 不应把 MESSRS / ADDRESS / BOOKING NOTICE 等业务文档头行误识为表头。"""

    def _make_ws_with_messrs_then_real_header(self):
        # 复现 PL260600451 那封交仓邮件：R7=MESSRS 行 + 末尾 cell 含「兴信:44CTNS/...CBM」
        # R10=真表头（SKU NO. ... NO of CARTON ...）
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Packing list'
        ws.cell(row=5, column=1, value='Packing / Weight List')
        ws.cell(row=7, column=1, value='MESSRS: BRANDS IN STYLE (PVT) LTD')
        ws.cell(row=7, column=28, value='DATE:')
        ws.cell(row=7, column=31, value='兴信：44CTNS/251.03KGS/2.854CBM\n泰亨：40CTNS/271.2KGS/2.354CBM')
        ws.cell(row=8, column=1, value='address: WAIDYA MAWATHA')
        ws.cell(row=8, column=28, value='PACKING LIST NO:')
        ws.cell(row=9, column=28, value='BOOKING NOTICE NO:')
        # R10 真表头
        headers = ['SKU NO.', 'DESCRIPTION', 'HS code', 'Shipping QUANTITY', 'Retail Unit',
                   'PC PER\nCARTON', 'NO of\nCARTON', 'PC PER\nPALLET', 'NO of\nPALLET']
        for i, h in enumerate(headers, 1):
            ws.cell(row=10, column=i, value=h)
        return ws

    def test_messrs_row_with_ctns_cbm_summary_not_header(self):
        ws = self._make_ws_with_messrs_then_real_header()
        row_idx, _ = _find_header_row(ws)
        assert row_idx == 10, f'expected R10 真表头，但选了 R{row_idx}'


class MatchFieldTest(TestCase):
    """_match_field 字段名映射规则。"""

    def test_booking_key_header_maps_to_cds_bkg(self):
        # ZURU 总表分柜明细 sheet 的 BKG 列叫 "Booking Key#" 而不是 "BKG"
        assert _match_field('Booking Key#') == '_cds_bkg_number'
        assert _match_field('Booking Key') == '_cds_bkg_number'
        assert _match_field('BookingKey') == '_cds_bkg_number'

    def test_booking_column_still_maps_to_booking_line(self):
        # 不能把 Booking 列误吞到 _cds_bkg_number
        assert _match_field('Booking') == '_booking_line'

    def test_cds_bkg_and_bare_bkg_still_work(self):
        assert _match_field('CDS BKG#') == '_cds_bkg_number'
        assert _match_field('BKG') == '_cds_bkg_number'

    def test_kmart_yutai_warehouse_headers(self):
        # 裕泰仓 Kmart 箱单格式：Order no. = 客 PO，订单号 = 合同号，INNERS = 规格/每箱
        assert _match_field('Order no.') == 'customer_po'
        assert _match_field('Order No') == 'customer_po'
        assert _match_field('订单号') == 'contract_number'
        assert _match_field('INNERS') == 'spec'
        assert _match_field('Inners') == 'spec'

    def test_existing_rules_not_broken_by_new_headers(self):
        assert _match_field('ORDER QTY') == 'quantity'   # 不能被 Order no. 规则吞
        assert _match_field('Customer PO#') == 'customer_po'
        assert _match_field('ZURU PO No.') == 'contract_number'
        assert _match_field('PC PER CARTON') == 'spec'


class ExcelParserTest(TestCase):
    """Packing List Excel 解析测试。"""

    def _create_test_packing_list(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Packing list'
        headers = [
            'SKU NO.', 'Shipping QUANTITY', 'Retail Unit', 'PC PER CARTON',
            'NO of CARTON', 'PC PER PALLET', 'NO of PALLET',
            'L', 'W', 'H', 'CBM',
            'Customer PO#', 'Customer PO item No.',
            'ZURU PO No.', 'Actual factory', 'Supplier Name',
        ]
        ws.append(headers)
        ws.append([
            '15754', 1, 4068, 4, 1017, None, None,
            44, 36.5, 27.5, 15.5,
            '0958543467', '', '4500189918',
            'Dong Guan Hanson Plastic Product Ltd', 'Hanson',
        ])
        ws.append([
            '77794UQ1', 1, 10960, 8, 1370, None, 18,
            50, 40, 30, 25.864,
            '10001643491-3890', 'B0DRBC1FN2', '4500189603',
            'Dong Guan Hanson Plastic Product Ltd', 'Hanson',
        ])
        path = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False).name
        wb.save(path)
        return path

    def test_parse_basic_fields(self):
        path = self._create_test_packing_list()
        try:
            items = parse_packing_list(path)
            assert len(items) == 2
            assert items[0]['product_code'] == '15754'
            assert items[0]['quantity'] == 4068
            assert items[0]['pieces'] == 1017
            assert items[0]['contract_number'] == '4500189918'
            assert items[0]['customer_po'] == '0958543467'
            assert items[0]['volume'] == 15.5
            assert items[0]['box_dimensions'] == '44*36.5*27.5'
        finally:
            os.unlink(path)

    def test_parse_customer_po_item_no(self):
        path = self._create_test_packing_list()
        try:
            items = parse_packing_list(path)
            assert items[1]['customer_po_item_no'] == 'B0DRBC1FN2'
            assert items[0]['customer_po_item_no'] == ''
        finally:
            os.unlink(path)

    def test_parse_supplier(self):
        path = self._create_test_packing_list()
        try:
            items = parse_packing_list(path)
            assert items[0]['supplier'] == 'Hanson'
        finally:
            os.unlink(path)

    def test_parse_second_row(self):
        path = self._create_test_packing_list()
        try:
            items = parse_packing_list(path)
            assert items[1]['product_code'] == '77794UQ1'
            assert items[1]['quantity'] == 10960
            assert items[1]['pallet_count'] == 18
            assert items[1]['volume'] == 25.864
            assert items[1]['box_dimensions'] == '50*40*30'
        finally:
            os.unlink(path)


class YaxExcelParserTest(TestCase):
    """YAX 拖车通知单 Excel 解析测试。"""

    def _create_test_yax_excel(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        # Header area
        ws['A1'] = '马士基FCA并柜拖车通知单'
        ws['A2'] = '船公司订舱单号'
        ws['B2'] = '149601102372'
        ws['A3'] = '柜型'
        ws['B3'] = '40GP(40HQ SUB)'
        ws['A4'] = '截关时间'
        ws['B4'] = '2026/3/7 12:00'
        ws['A5'] = '截补料时间'
        ws['B5'] = '2026/3/3 15:00'
        # PO-SO mapping table
        row = 20
        ws.cell(row=row, column=1, value='PO 号码')
        ws.cell(row=row, column=2, value='')
        ws.cell(row=row, column=3, value='YAT')
        ws.cell(row=row, column=4, value='CBM')
        ws.cell(row=row, column=5, value='KGS')
        ws.cell(row=row, column=6, value='Cartons')
        # Data rows
        row = 21
        ws.cell(row=row, column=1, value=10001519913)
        ws.cell(row=row, column=2, value='YAT')
        ws.cell(row=row, column=3, value='YAX5634105')
        ws.cell(row=row, column=4, value=13.935)
        row = 22
        ws.cell(row=row, column=1, value=10001519921)
        ws.cell(row=row, column=2, value='YAT')
        ws.cell(row=row, column=3, value='YAX5634172')
        ws.cell(row=row, column=4, value=2.806)
        path = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False).name
        wb.save(path)
        return path

    def test_parse_header_fields(self):
        path = self._create_test_yax_excel()
        try:
            result = parse_yax_excel(path)
            assert result['booking_number'] == '149601102372'
            assert '40GP' in result['container_type']
            assert result['customs_cutoff'] == '2026/3/7 12:00'
            assert result['si_deadline'] == '2026/3/3 15:00'
        finally:
            os.unlink(path)

    def test_parse_po_so_mapping(self):
        path = self._create_test_yax_excel()
        try:
            result = parse_yax_excel(path)
            assert len(result['po_so_mapping']) >= 2
            first = result['po_so_mapping'][0]
            assert first['so'] == 'YAX5634105'
            assert first['cbm'] == 13.935
        finally:
            os.unlink(path)

    def test_parse_po_values(self):
        path = self._create_test_yax_excel()
        try:
            result = parse_yax_excel(path)
            pos = [item['po'] for item in result['po_so_mapping']]
            assert '10001519913' in pos
            assert '10001519921' in pos
        finally:
            os.unlink(path)
