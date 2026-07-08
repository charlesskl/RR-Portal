"""
柜单 Excel 生成引擎测试
"""
import os

import openpyxl
from django.test import TestCase

from apps.accounts.models import User
from apps.generator.base_generator import generate_container_sheet
from apps.master_data.models import Customer
from apps.shipments.models import Shipment, ShipmentItem


class GeneratorTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username='ship1', password='test', role='shipping',
        )
        self.customer = Customer.objects.create(
            name='ZURU',
            consignee='ZURU INC.',
            consignee_code='66934704',
            is_brand_auto=True,
        )
        self.shipment = Shipment.objects.create(
            shipment_type='normal',
            customer=self.customer,
            so_number='SHZ7952680',
            container_type="1*40'HQ",
            port='盐田',
            created_by=self.user,
        )
        ShipmentItem.objects.create(
            shipment=self.shipment,
            factory_remark='兴信',
            seq_number=1,
            trading_company='ZURU',
            contract_number='4500189918',
            product_code='15754',
            product_name='明星系列',
            spec='4个/箱',
            country='美国',
            toy_category='塑胶',
            quantity=4068,
            pieces=1017,
            gross_weight_per_box=1.03,
            net_weight_per_box=0.50,
            volume=15.5,
            customer_po='0958543467',
            brand='ZURU/SPIN MASTER',
            total_pieces_per_order=1017,
        )

    def test_generate_normal_sheet(self):
        """测试正常柜单生成"""
        output_path = generate_container_sheet(self.shipment.id)
        self.assertTrue(os.path.exists(output_path))

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        # 检查标题包含 ZURU
        title_cell = ws.cell(row=2, column=6).value
        self.assertIn('ZURU', title_cell)

        # 检查列标题行（第5行）
        self.assertEqual(ws.cell(row=5, column=1).value, '备注')
        self.assertEqual(ws.cell(row=5, column=2).value, '序号')
        self.assertEqual(ws.cell(row=5, column=5).value, '货号')
        self.assertEqual(ws.cell(row=5, column=6).value, '货名')

        # 检查数据包含产品代码 15754
        data_row = 6
        self.assertEqual(ws.cell(row=data_row, column=5).value, '15754')

        # 检查货名+规格 格式
        self.assertEqual(ws.cell(row=data_row, column=6).value, '明星系列')

        # 检查品牌行
        brand_row = data_row + 1
        self.assertIn('ZURU/SPIN MASTER', ws.cell(row=brand_row, column=5).value)

        # 检查收货人信息（在E列=column 5）
        found_consignee = False
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value and 'ZURU INC.' in str(cell.value):
                    found_consignee = True
                    break
        self.assertTrue(found_consignee, '应包含收货人信息')

        wb.close()
        os.unlink(output_path)

    def test_generate_warehouse_sheet(self):
        """测试入仓/车柜单生成"""
        self.shipment.shipment_type = 'warehouse'
        self.shipment.warehouse = '盐田1号仓'
        self.shipment.save()

        # 添加箱规用于测试
        item = ShipmentItem.objects.filter(shipment=self.shipment).first()
        item.box_dimensions = '44*36.5*27.5'
        item.save()

        output_path = generate_container_sheet(self.shipment.id)
        self.assertTrue(os.path.exists(output_path))

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        # 入仓类型标题应包含"车"
        title_cell = ws.cell(row=2, column=6).value
        self.assertIn('车', title_cell)

        # 非排除仓库应该显示箱规
        found_box_dim = False
        for row in ws.iter_rows(min_row=6, max_row=ws.max_row):
            for cell in row:
                if cell.value and '44*36.5*27.5' in str(cell.value):
                    found_box_dim = True
                    break
        self.assertTrue(found_box_dim, '非排除仓库应显示箱规')

        wb.close()
        os.unlink(output_path)

    def test_no_box_dim_for_excluded_warehouse(self):
        """测试盐田2号仓不显示箱规"""
        self.shipment.shipment_type = 'warehouse'
        self.shipment.warehouse = '盐田2号仓'
        self.shipment.save()

        item = ShipmentItem.objects.filter(shipment=self.shipment).first()
        item.box_dimensions = '44*36.5*27.5'
        item.save()

        output_path = generate_container_sheet(self.shipment.id)
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        found_box_dim = False
        for row in ws.iter_rows(min_row=6, max_row=ws.max_row):
            for cell in row:
                if cell.value and '44*36.5*27.5' in str(cell.value):
                    found_box_dim = True
                    break
        self.assertFalse(found_box_dim, '盐田2号仓不应显示箱规')

        wb.close()
        os.unlink(output_path)

    def test_extra_po_column_with_english_letters(self):
        """测试客户PO包含英文字母时添加额外列"""
        item = ShipmentItem.objects.filter(shipment=self.shipment).first()
        item.customer_po = 'ABC123'
        item.customer_po_item_no = '10'
        item.save()

        output_path = generate_container_sheet(self.shipment.id)
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        # 第18列应该是 Customer PO item No.
        self.assertEqual(ws.cell(row=5, column=18).value, 'Customer PO item No.')

        wb.close()
        os.unlink(output_path)

    def test_generate_customer_load_sheet(self):
        """测试客上柜柜单生成"""
        self.shipment.shipment_type = 'customer_load'
        self.shipment.customs_broker = '华展报关'
        self.shipment.save()

        output_path = generate_container_sheet(self.shipment.id)
        self.assertTrue(os.path.exists(output_path))

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        title_cell = ws.cell(row=2, column=6).value
        self.assertIn('ZURU', title_cell)

        wb.close()
        os.unlink(output_path)

    def test_generate_qingxi_sheet(self):
        """测试清溪物流园柜单生成"""
        self.shipment.shipment_type = 'qingxi'
        self.shipment.save()

        output_path = generate_container_sheet(self.shipment.id)
        self.assertTrue(os.path.exists(output_path))

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        title_cell = ws.cell(row=2, column=6).value
        self.assertIn('清溪物流园', title_cell)

        wb.close()
        os.unlink(output_path)

    def test_summary_row_totals(self):
        """测试合计行数据正确"""
        # 添加第二个明细项
        ShipmentItem.objects.create(
            shipment=self.shipment,
            seq_number=2,
            trading_company='ZURU',
            contract_number='4500189919',
            product_code='15755',
            product_name='另一个产品',
            quantity=2000,
            pieces=500,
            gross_weight_per_box=1.00,
            net_weight_per_box=0.48,
            volume=10.0,
            customer_po='0958543468',
            total_pieces_per_order=500,
        )

        output_path = generate_container_sheet(self.shipment.id)
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        # 找合计行（合计在H列=column 8）
        summary_row = None
        for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=8, max_col=8):
            for cell in row:
                if cell.value == '合计':
                    summary_row = cell.row
                    break

        self.assertIsNotNone(summary_row, '应存在合计行')
        # 数量合计 = 4068 + 2000
        self.assertEqual(ws.cell(row=summary_row, column=9).value, f'=SUM(I6:I{summary_row - 1})')
        # 件数合计 = 1017 + 500
        self.assertEqual(ws.cell(row=summary_row, column=10).value, f'=SUM(J6:J{summary_row - 1})')

        wb.close()
        os.unlink(output_path)
