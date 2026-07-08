"""Tests for pallet_report aggregation helper."""
from datetime import date
from django.test import TestCase
from apps.shipments.pallet_report import (
    classify_pallet_items,
    group_by_factory_so,
    apply_filters,
)


class ClassifyPalletItemsTest(TestCase):
    """把每条 PL item 分类到 self/local/external 之一"""

    def test_xingxin_zuogui_own_goods_classified_as_self(self):
        items = [{
            'zuogui_factory': '兴信',
            'product_code': '15756',
            'factory_remark': '兴信',
            'pallet_count': 10,
        }]
        result = classify_pallet_items(items)
        assert result['self'] == items
        assert result['local'] == []
        assert result['external'] == []

    def test_xingxin_zuogui_external_goods_classified_as_external(self):
        items = [{
            'zuogui_factory': '兴信',
            'product_code': '15786',
            'factory_remark': '华康',
            'pallet_count': 5,
        }]
        result = classify_pallet_items(items)
        assert result['self'] == []
        assert result['external'] == items

    def test_external_zuogui_classified_as_local(self):
        items = [{
            'zuogui_factory': '博锐',
            'product_code': '92146',
            'factory_remark': '兴信',
            'pallet_count': 3,
        }]
        result = classify_pallet_items(items)
        assert result['local'] == items

    def test_zero_pallet_count_excluded(self):
        items = [{
            'zuogui_factory': '兴信',
            'product_code': '15756',
            'factory_remark': '兴信',
            'pallet_count': 0,
        }]
        result = classify_pallet_items(items)
        assert result['self'] == []


class GroupByFactorySoTest(TestCase):
    def test_groups_by_factory_then_so(self):
        items = [
            {'factory_remark': '博锐', 'so_number': 'SO1', 'pallet_count': 1, 'product_code': 'A'},
            {'factory_remark': '博锐', 'so_number': 'SO1', 'pallet_count': 2, 'product_code': 'B'},
            {'factory_remark': '博锐', 'so_number': 'SO2', 'pallet_count': 3, 'product_code': 'C'},
            {'factory_remark': '库有', 'so_number': 'SO3', 'pallet_count': 4, 'product_code': 'D'},
        ]
        result = group_by_factory_so(items, 'factory_remark')
        assert list(result.keys()) == ['博锐', '库有']
        assert list(result['博锐'].keys()) == ['SO1', 'SO2']
        assert len(result['博锐']['SO1']) == 2
        assert len(result['博锐']['SO2']) == 1
        assert len(result['库有']['SO3']) == 1

    def test_empty_factory_uses_unknown(self):
        items = [{'factory_remark': '', 'so_number': 'SO1', 'pallet_count': 1}]
        result = group_by_factory_so(items, 'factory_remark')
        assert '未知' in result


class ApplyFiltersTest(TestCase):
    def test_date_range_filter(self):
        # 用 YYYY-MM-DD 完整日期避免依赖系统当前年份（_parse_ship_date 对 M/D 用 today().year 兜底）
        items = [
            {'ship_date': '2026-05-15', 'factory_remark': '兴信', 'pallet_count': 1},
            {'ship_date': '2026-06-10', 'factory_remark': '兴信', 'pallet_count': 2},
            {'ship_date': '2026-04-30', 'factory_remark': '兴信', 'pallet_count': 3},
        ]
        result = apply_filters(items, start=date(2026, 5, 1), end=date(2026, 5, 31))
        assert len(result) == 1
        assert result[0]['ship_date'] == '2026-05-15'

    def test_factory_filter(self):
        items = [
            {'factory_remark': '博锐', 'pallet_count': 1},
            {'factory_remark': '库有', 'pallet_count': 2},
            {'factory_remark': '华康', 'pallet_count': 3},
        ]
        result = apply_filters(items, factories=['博锐', '库有'])
        assert len(result) == 2

    def test_no_filter_returns_all(self):
        items = [{'factory_remark': '兴信', 'pallet_count': 1}]
        result = apply_filters(items)
        assert len(result) == 1


import io
import openpyxl


class GenerateXlsxTest(TestCase):
    def test_generate_xlsx_returns_valid_bytes(self):
        from apps.shipments.pallet_export import generate_xlsx
        data = {
            'period_start': '2026-04-01',
            'period_end': '2026-04-30',
            'factories_filter': '全部',
            'categories_filter': '全部',
            'self_items': [
                {'ship_date': '4/5', 'so_number': 'SO1', 'product_code': '15756',
                 'product_name': '货A', 'contract_number': 'C1', 'customer_po': 'PO1',
                 'pieces': 100, 'pallet_count': 10, 'factory_remark': '兴信'},
            ],
            'local_items': [],
            'external_items': [],
            'manual_borui': [],
            'manual_kuyou': [],
        }
        blob = generate_xlsx(data)
        assert isinstance(blob, bytes)
        wb = openpyxl.load_workbook(io.BytesIO(blob))
        ws = wb.active
        cell_values = [str(c.value or '') for row in ws.iter_rows() for c in row]
        text = ' '.join(cell_values)
        assert '卡板数月度统计报表' in text
        assert '15756' in text
        assert '10' in text  # 卡板数
        assert '本厂做柜' in text

    def test_generate_xlsx_with_manual_data(self):
        from apps.shipments.pallet_export import generate_xlsx
        data = {
            'period_start': '2026-04-01',
            'period_end': '2026-04-30',
            'factories_filter': '全部',
            'categories_filter': '全部',
            'self_items': [], 'local_items': [], 'external_items': [],
            'manual_borui': [
                {'date': '4/6', 'so_number': 'SOB1', 'product_code': '7152',
                 'product_name': '', 'contract_number': '', 'pieces': 50, 'pallet_count': 5},
            ],
            'manual_kuyou': [],
        }
        blob = generate_xlsx(data)
        wb = openpyxl.load_workbook(io.BytesIO(blob))
        ws = wb.active
        text = ' '.join(str(c.value or '') for row in ws.iter_rows() for c in row)
        assert '送博锐手填' in text
        assert 'SOB1' in text
        assert '7152' in text


from django.contrib.auth import get_user_model
from rest_framework.test import APIClient


class PalletExportAPITest(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(username='exp', password='x')

    def setUp(self):
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)

    def test_export_returns_xlsx(self):
        payload = {
            'start': '2026-04-01',
            'end': '2026-04-30',
            'factories': [],
            'categories': [],
            'self_items': [
                {'ship_date': '4/5', 'so_number': 'SO1', 'product_code': '15756',
                 'product_name': '货A', 'contract_number': 'C1', 'customer_po': 'PO1',
                 'pieces': 100, 'pallet_count': 10, 'factory_remark': '兴信',
                 'zuogui_factory': '兴信'},
            ],
            'local_items': [],
            'external_items': [],
            'manual_borui': [],
            'manual_kuyou': [],
        }
        resp = self.client.post('/api/pallets/export/', payload, format='json')
        assert resp.status_code == 200
        assert 'spreadsheetml' in resp['Content-Type']
        assert 'attachment' in resp['Content-Disposition']
        # 校验返回的 xlsx 实际可解析且含期望内容
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        text = ' '.join(str(c.value or '') for row in ws.iter_rows() for c in row)
        assert '卡板数月度统计报表' in text
        assert '15756' in text
        assert '本厂做柜' in text
        assert 'SO1' in text

    def test_export_rejects_invalid_dates(self):
        # 缺日期 → 400
        resp = self.client.post('/api/pallets/export/', {
            'start': '', 'end': '', 'self_items': [], 'local_items': [],
            'external_items': [], 'manual_borui': [], 'manual_kuyou': [],
        }, format='json')
        assert resp.status_code == 400
        # 格式错 → 400
        resp = self.client.post('/api/pallets/export/', {
            'start': '2026/04/01', 'end': '2026-04-30', 'self_items': [], 'local_items': [],
            'external_items': [], 'manual_borui': [], 'manual_kuyou': [],
        }, format='json')
        assert resp.status_code == 400
