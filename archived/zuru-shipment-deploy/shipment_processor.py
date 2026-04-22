# -*- coding: utf-8 -*-
"""
出货标记核心处理程序

用法: python shipment_processor.py <接单表.xlsx> <出货文件夹>

规则：
1. 匹配：合同号 + 简货号（开头连续数字）
2. 可标记：备注列为空 + 数量>0
3. 顺序匹配可用行，标记出货日期到备注列
4. XML手术式写入，100%保留原文件格式
5. 输出：接单表_更新.xlsx
"""

import os
import re
import copy
import shutil
import logging
import zipfile
from io import BytesIO
from datetime import datetime

from lxml import etree

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

logger = logging.getLogger(__name__)

# ============================================================
# 常量
# ============================================================

# 卡板单后缀
SET_SUFFIX_RE = re.compile(r'^(.+?)(SLB|SLD|SLT|SK)(.*)?$', re.IGNORECASE)

# 提取开头连续数字作为简货号
_LEADING_DIGITS_RE = re.compile(r'^(\d+)')

# 出货日期提取（如 "3月9日" 或 "3月9出"）
_DATE_RE = re.compile(r'(\d{1,2}月\d{1,2})[日出]')

# xlsx XML命名空间
_XLSX_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
_REL_NS = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'

# 单元格引用正则（如 A10, AA123）
_CELL_REF_RE = re.compile(r'^([A-Z]+)(\d+)$')

# 混装货号映射
MIXED_MAP = {
    '7153': ['7149', '7150'],
    '7154': ['7151', '7152'],
    '25257': ['25251', '25252', '25253'],
}

# 产品组辅助行关键词
AUX_KEYWORDS = ('收缩指商', '收缩膜', '收藏指南', 'PDQ')

def _today_serial():
    """返回今天的Excel日期序列号"""
    return (datetime.now() - datetime(1899, 12, 31)).days


# 需要清除的隐藏/特殊字符
_HIDDEN_CHARS = re.compile(
    r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f'
    r'\u00a0'
    r'\u200b\u200c\u200d\u200e\u200f'
    r'\u2028\u2029'
    r'\u202a-\u202e'
    r'\ufeff'
    r'\u3000'
    r']'
)


# ============================================================
# 工具函数
# ============================================================

def _normalize(val):
    """单元格值 → 干净字符串（用于匹配键）"""
    if val is None:
        return ''
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    if isinstance(val, int):
        return str(val)
    s = str(val)
    s = _HIDDEN_CHARS.sub('', s)
    s = s.strip()
    s = s.upper()
    return s


def _extract_simple_no(item_no):
    """提取开头连续数字作为简货号。15750A→15750, 92148B→92148"""
    m = _LEADING_DIGITS_RE.match(item_no)
    return m.group(1) if m else item_no


def _to_simple_key(item_no):
    """出货货号→简货号：保留SLB/SLD/SLT/SK后缀，取开头连续数字+后缀。
    15750A→15750, 77711SLB→77711SLB, 15779SLD→15779SLD"""
    m = SET_SUFFIX_RE.match(item_no)
    if m:
        base = _extract_simple_no(m.group(1))
        return base + m.group(2).upper()
    return _extract_simple_no(item_no)


def _is_set_item(item_no):
    """判断货号是否为卡板单（含SLB/SLD/SLT/SK后缀）"""
    return bool(SET_SUFFIX_RE.match(item_no))


def _strip_set_suffix(item_no):
    """去掉SLB/SLD/SLT/SK后缀"""
    m = SET_SUFFIX_RE.match(item_no)
    return m.group(1) if m else item_no


def _find_slb_pair(ws, slb_row, col_item, col_contract, col_qty, max_row):
    """找SLB行的配对产品行（SLB行紧接的下一个同基础货号非SLB行）。
    SLB结构: SLB行(卡板数) → 产品行(产品数量) [→ 辅助行(收缩膜/PDQ等)]
    返回: (product_row, product_qty, aux_rows) 或 (None, 0, [])
    """
    slb_item = _normalize(ws.cell(row=slb_row, column=col_item).value)
    base = _extract_simple_no(_strip_set_suffix(slb_item))
    product_row = None
    product_qty = 0
    aux_rows = []

    for r in range(slb_row + 1, min(slb_row + 20, max_row + 1)):
        rc = _normalize(ws.cell(row=r, column=col_contract).value)
        if rc:  # 遇到新合同号行，停止
            break
        item = _normalize(ws.cell(row=r, column=col_item).value)
        if _is_set_item(item):
            # 遇到下一个SLB行，停止
            break
        if product_row is None:
            # 第一个非SLB行 = 产品行
            item_base = _extract_simple_no(item)
            if item_base == base:
                product_row = r
                try:
                    product_qty = int(float(
                        ws.cell(row=r, column=col_qty).value or 0))
                except (ValueError, TypeError):
                    product_qty = 0
            else:
                aux_rows.append(r)
        else:
            # 产品行之后的辅助行
            aux_rows.append(r)

    return product_row, product_qty, aux_rows


def _is_auxiliary_row(ws, row_idx, col_item, max_col):
    """判断是否为产品组辅助行。
    规则1: 货号列字体非黑色（红/紫/灰等）→ 辅助行
    规则2: 货号或任意单元格含关键词（收缩指商/收缩膜/收藏指南/PDQ）→ 辅助行
    """
    # 规则1: 检查货号列字体颜色，非黑色即辅助行
    item_cell = ws.cell(row=row_idx, column=col_item)
    font_color = item_cell.font.color
    if font_color is not None:
        if font_color.type == 'rgb' and font_color.rgb is not None:
            rgb = str(font_color.rgb)
            # 黑色: 00000000 或 FF000000，其他颜色视为辅助行
            if rgb not in ('00000000', 'FF000000'):
                return True
        elif font_color.type == 'indexed' and font_color.indexed is not None:
            # indexed=0 通常是黑色，其他颜色视为辅助行
            if font_color.indexed not in (0, 1, 8, 64):
                return True
    # 规则2: 关键词匹配
    item_val = str(item_cell.value or '').strip()
    for kw in AUX_KEYWORDS:
        if kw in item_val:
            return True
    for c in range(1, max_col + 1):
        val = ws.cell(row=row_idx, column=c).value
        if val is None:
            continue
        s = str(val).strip()
        for kw in AUX_KEYWORDS:
            if kw == s:
                return True
    return False


def _collect_available(ws, candidates, col_qty, col_beizhu, col_contract,
                       col_item, max_col, is_slb=False, log_func=None):
    """从候选行中筛选可标记行，返回 [(row, qty, needs_contract_fill), ...]
    普通货号：只标合同列不为空的行
    SLB货号：合同列为空但货号含SLB后缀也可标记（需补填合同号）
    """
    available = []
    for r in candidates:
        contract_val = ws.cell(row=r, column=col_contract).value
        has_contract = contract_val is not None and str(contract_val).strip()
        needs_fill = False

        if not has_contract:
            if is_slb:
                # SLB货号：检查该行货号是否也含SLB后缀
                full_item = _normalize(ws.cell(row=r, column=col_item).value)
                if _is_set_item(full_item):
                    needs_fill = True  # 需要补填合同号
                else:
                    if log_func:
                        log_func(f"    跳过行{r}: 非SLB子行不标记")
                    continue
            else:
                # 非SLB出货：检查最近B-filled行类型
                # SLB/SET类型行的正下方B=None行 → SLB产品子行，跳过（由SLB配对逻辑处理）
                # 普通行下方的B=None行 → 多item续行或同货号重复订单行，允许
                nearest_above_is_set = None
                for prev_r in range(r - 1, max(r - 30, 0), -1):
                    prev_b = ws.cell(row=prev_r, column=col_contract).value
                    prev_item = _normalize(ws.cell(row=prev_r, column=col_item).value)
                    if not prev_item:
                        continue
                    if any(kw in prev_item for kw in AUX_KEYWORDS):
                        continue
                    if prev_b is not None and str(prev_b).strip():
                        nearest_above_is_set = _is_set_item(prev_item)
                        break
                    # B=None行继续往上找
                if nearest_above_is_set is not False:
                    # SLB/SET类行 或 未找到B-filled行 → 跳过
                    if log_func:
                        log_func(f"    跳过行{r}: SLB产品子行或无合同行（上方B-filled={nearest_above_is_set}）")
                    continue
                needs_fill = True
        if col_beizhu:
            bz = ws.cell(row=r, column=col_beizhu).value
            if bz is not None and str(bz).strip():
                bz_str = str(bz).strip()
                # 允许覆盖"额外费用"
                if bz_str != '额外费用':
                    if log_func:
                        log_func(f"    排除行{r}: 备注不为空='{bz_str[:30]}'")
                    continue
        try:
            qty = int(float(ws.cell(row=r, column=col_qty).value))
        except (ValueError, TypeError):
            if log_func:
                log_func(f"    排除行{r}: 数量无效={ws.cell(row=r, column=col_qty).value}")
            continue
        if qty > 0:
            available.append((r, qty, needs_fill))
        else:
            if log_func:
                log_func(f"    排除行{r}: 数量={qty}≤0")
    return available


def _find_sub_rows(ws, main_row, col_contract, max_row, col_item=None):
    """找主行下方的子行（合同列为空的连续行，直到下一个合同号行）。
    若提供 col_item，遇到与主行货号相同的B=None行时停止（独立订单行，非辅助子行）。
    返回: [row_idx, ...] 不含主行本身"""
    subs = []
    main_item = None
    if col_item is not None:
        main_item = _normalize(ws.cell(row=main_row, column=col_item).value)
    for r in range(main_row + 1, max_row + 1):
        rc = _normalize(ws.cell(row=r, column=col_contract).value)
        if rc:
            break
        if col_item is not None and main_item:
            row_item = _normalize(ws.cell(row=r, column=col_item).value)
            if row_item and row_item == main_item:
                break  # 同货号独立订单行，停止
        subs.append(r)
    return subs


def _find_blue_sub_rows(ws, main_row, col_contract, col_item, max_row):
    """蓝色填充专用子行查找：遇到与主行货号相同的B=None行时停止（不上色独立订单行）。
    只返回真正的辅助子行（不同货号，如SLB组件、备注行等）。"""
    subs = []
    main_item = _normalize(ws.cell(row=main_row, column=col_item).value)
    for r in range(main_row + 1, max_row + 1):
        rc = _normalize(ws.cell(row=r, column=col_contract).value)
        if rc:
            break  # 遇到B-filled行停止
        if main_item:
            row_item = _normalize(ws.cell(row=r, column=col_item).value)
            if row_item and row_item == main_item:
                break  # 同货号独立订单行，不加入蓝色范围
        subs.append(r)
    return subs


def _get_sub_qtys(ws, sub_rows, col_qty):
    """读取子行数量，返回 {row: qty} 字典"""
    result = {}
    for sr in sub_rows:
        try:
            sq = int(float(ws.cell(row=sr, column=col_qty).value or 0))
        except (ValueError, TypeError):
            sq = 0
        result[sr] = sq
    return result


def _select_match(available, ship_qty):
    """从可用行中选择最佳匹配。
    优先级: 精确匹配 → 最接近且≥出货量的行 → None
    返回: (row, qty, needs_fill, match_type)
        match_type: 'exact' | 'partial' | None
    """
    # 1. 精确匹配
    for row, qty, needs_fill in available:
        if qty == ship_qty:
            return row, qty, needs_fill, 'exact'
    # 2. 最接近且≥出货量的行（ceiling匹配，避免选过大的行）
    partial = [(r, q, n) for r, q, n in available if q > ship_qty]
    if partial:
        best = min(partial, key=lambda x: x[1])  # 选最小的≥ship_qty
        return best[0], best[1], best[2], 'partial'
    return None, 0, False, None


def _match_group(ws, available, col_qty, col_contract, ship_qty,
                 used_rows, max_row):
    """当单行不够扣时，检查主行+子行的组总量是否匹配。
    返回: (row_group, group_total, 'group_exact'|'group_partial') 或 None
    row_group = [主行, 子行1, 子行2, ...], group_total = 组内总数量
    注意：计算时排除已在 used_rows 中的子行（避免把已被拆行的子行计入总量）"""
    for main_r, main_qty, needs_fill in available:
        if main_r in used_rows:
            continue
        sub_rows = _find_sub_rows(ws, main_r, col_contract, max_row)
        # 排除已被使用的子行
        valid_sub_rows = [sr for sr in sub_rows if sr not in used_rows]
        valid_sub_qtys = _get_sub_qtys(ws, valid_sub_rows, col_qty)
        group_total = main_qty + sum(valid_sub_qtys.values())
        if group_total == ship_qty:
            return [main_r] + valid_sub_rows, group_total, 'group_exact'
        elif group_total > ship_qty:
            return [main_r] + valid_sub_rows, group_total, 'group_partial'
    return None


def _match_sub_rows(ws, candidates, col_qty, col_beizhu, col_contract,
                    ship_qty, used_rows):
    """当主行数量不够时，在子行（合同列为空的候选行）中查找匹配。
    优先精确匹配，其次够扣。
    返回: (row, qty, 'exact'|'partial') 或 None"""
    sub_candidates = []
    for r in candidates:
        if r in used_rows:
            continue
        contract_val = ws.cell(row=r, column=col_contract).value
        if contract_val is not None and str(contract_val).strip():
            continue  # 跳过主行（已在available中检查过）
        if col_beizhu:
            bz = ws.cell(row=r, column=col_beizhu).value
            if bz is not None and str(bz).strip():
                if str(bz).strip() != '额外费用':
                    continue
        try:
            qty = int(float(ws.cell(row=r, column=col_qty).value))
        except (ValueError, TypeError):
            continue
        if qty > 0:
            sub_candidates.append((r, qty))
    # 优先精确匹配
    for r, qty in sub_candidates:
        if qty == ship_qty:
            return r, qty, 'exact'
    # 够扣的行
    for r, qty in sub_candidates:
        if qty > ship_qty:
            return r, qty, 'partial'
    return None



# ============================================================
# XML手术式写入（100%保留格式）
# ============================================================

def _col_letter(col):
    """1-based列号 → Excel列字母。1→A, 14→N, 27→AA"""
    result = ''
    while col > 0:
        col, rem = divmod(col - 1, 26)
        result = chr(65 + rem) + result
    return result


def _find_sheet_xml_path(zf, sheet_title):
    """在xlsx ZIP中查找指定sheet名称对应的XML路径"""
    ns_wb = _XLSX_NS
    ns_r = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'

    try:
        wb_xml = zf.read('xl/workbook.xml')
        wb_root = etree.fromstring(wb_xml)

        rid = None
        for sheet_elem in wb_root.iter(f'{{{ns_wb}}}sheet'):
            if sheet_elem.get('name') == sheet_title:
                rid = sheet_elem.get(f'{{{ns_r}}}id')
                break

        if not rid:
            # 找不到匹配名称，取第一个sheet
            for sheet_elem in wb_root.iter(f'{{{ns_wb}}}sheet'):
                rid = sheet_elem.get(f'{{{ns_r}}}id')
                break

        if rid:
            rels_xml = zf.read('xl/_rels/workbook.xml.rels')
            rels_root = etree.fromstring(rels_xml)
            for rel in rels_root.iter(f'{{{_REL_NS}}}Relationship'):
                if rel.get('Id') == rid:
                    target = rel.get('Target')
                    if not target.startswith('/'):
                        return f'xl/{target}'
                    return target.lstrip('/')
    except Exception as e:
        logger.warning(f"查找sheet路径失败: {e}")

    return 'xl/worksheets/sheet1.xml'


def _write_cell_value(cell_elem, value, ns):
    """将单元格设置为数字或inline string，保留样式属性"""
    for child in list(cell_elem):
        cell_elem.remove(child)
    # 清理可能的公式相关属性
    for attr in ('cm',):
        if attr in cell_elem.attrib:
            del cell_elem.attrib[attr]

    if isinstance(value, (int, float)):
        # 数字值：<v>46094</v>
        if 't' in cell_elem.attrib:
            del cell_elem.attrib['t']
        v_elem = etree.SubElement(cell_elem, f'{{{ns}}}v')
        v_elem.text = str(value)
    else:
        # 文本值：inline string
        cell_elem.set('t', 'inlineStr')
        is_elem = etree.SubElement(cell_elem, f'{{{ns}}}is')
        t_elem = etree.SubElement(is_elem, f'{{{ns}}}t')
        t_elem.text = str(value)


_FORMULA_REF_RE = re.compile(r'(\$?[A-Za-z]+)(\$?)(\d+)')


def _shift_formula_row(formula, old_row, offset):
    """将公式中相对行引用为 old_row 的单元格行号偏移 offset。
    绝对行引用（$行号）不变。"""
    def _replace(m):
        col_part, abs_flag, row_str = m.group(1), m.group(2), m.group(3)
        if abs_flag == '' and int(row_str) == old_row:
            return f'{col_part}{int(row_str) + offset}'
        return m.group(0)
    return _FORMULA_REF_RE.sub(_replace, formula)


def _shift_row_elem(row_elem, offset, ns):
    """将 <row> 元素的行号、所有子 <c> 的引用、以及 <f> 公式中的行引用偏移 offset 行"""
    old_r = int(row_elem.get('r'))
    new_r = old_r + offset
    row_elem.set('r', str(new_r))
    for cell in row_elem.findall(f'{{{ns}}}c'):
        ref = cell.get('r', '')
        m = _CELL_REF_RE.match(ref)
        if m:
            cell.set('r', f'{m.group(1)}{new_r}')
        # 同步更新公式中的相对行引用
        f_elem = cell.find(f'{{{ns}}}f')
        if f_elem is not None and f_elem.text:
            f_elem.text = _shift_formula_row(f_elem.text, old_r, offset)


def _set_cell_value_xml(row_elem, col_letter, row_num, value, ns):
    """在 <row> 元素中设置指定列的单元格值"""
    ref = f'{col_letter}{row_num}'
    for cell in row_elem.findall(f'{{{ns}}}c'):
        if cell.get('r') == ref:
            _write_cell_value(cell, value, ns)
            return
    # 单元格不存在，创建
    cell = etree.SubElement(row_elem, f'{{{ns}}}c')
    cell.set('r', ref)
    _write_cell_value(cell, value, ns)


def _xml_split_rows(sheet_data, split_info, ns, get_blue_func=None):
    """在XML sheetData中执行拆行操作。

    1. 找到原行组的 <row> 元素
    2. 深拷贝为剩余行组
    3. 修改原行组数量 → 出货数量
    4. 修改克隆行组数量 → 剩余数量
    5. 将克隆行组插入原行组后面
    6. 所有后续行的行号 + group_size
    7. 若split_info含余量标记(remainder_mark)，将日期/蓝色写入第一个克隆行
    """
    from copy import deepcopy

    row_group = split_info['row_group']
    shipped_qtys = split_info['shipped_qtys']
    remain_qtys = split_info['remain_qtys']
    qty_col = split_info['qty_col']
    qty_col_letter = _col_letter(qty_col)
    group_size = len(row_group)

    # 找所有 <row> 元素，按行号索引
    all_rows = sheet_data.findall(f'{{{ns}}}row')
    row_elems = {}
    for elem in all_rows:
        row_elems[int(elem.get('r'))] = elem

    # 找原行组的元素
    group_elems = []
    for r in row_group:
        if r in row_elems:
            group_elems.append((r, row_elems[r]))

    if not group_elems:
        return

    # 1. 后续行全部下移 group_size 行（从下往上，避免冲突）
    last_group_row = max(r for r, _ in group_elems)
    rows_to_shift = [(int(e.get('r')), e) for e in all_rows
                     if int(e.get('r')) > last_group_row]
    rows_to_shift.sort(key=lambda x: x[0], reverse=True)
    for _, elem in rows_to_shift:
        _shift_row_elem(elem, group_size, ns)

    # 2. 深拷贝原行组 → 剩余行组（插入到原行组后面）
    insert_after = group_elems[-1][1]
    first_clone = None
    for r, orig_elem in reversed(group_elems):
        clone = deepcopy(orig_elem)
        _shift_row_elem(clone, group_size, ns)
        # 修改克隆行的数量为剩余数量
        if r in remain_qtys:
            _set_cell_value_xml(clone, qty_col_letter,
                                int(clone.get('r')),
                                remain_qtys[r], ns)
        insert_after.addnext(clone)
        if r == row_group[0]:
            first_clone = clone  # 记录第一个克隆行（主行克隆=余量行）

    # 3. 修改原行组数量为出货数量（余量已被覆盖则保留原始数量，主行保持全量显示）
    if not split_info.get('remainder_mark'):
        for r, orig_elem in group_elems:
            if r in shipped_qtys:
                _set_cell_value_xml(orig_elem, qty_col_letter, r,
                                    shipped_qtys[r], ns)

    # 4. 若有余量标记，余量克隆行只标蓝色（不填合同，不写日期）
    remainder_mark = split_info.get('remainder_mark')
    if remainder_mark and first_clone is not None and get_blue_func is not None:
        # 蓝色填充整行（不填合同，不写日期）
        for cell_elem in first_clone.findall(f'{{{ns}}}c'):
            orig_s = cell_elem.get('s', '0')
            cell_elem.set('s', get_blue_func(orig_s))


def _xml_swap_adjacent_groups(sheet_data, group_a, group_b, ns):
    """
    在XML中交换两个相邻行组的顺序。
    group_a 当前在前（未选中有合同行组），group_b 紧跟其后（选中的needs_fill行组）。
    交换后 group_b 在前，group_a 在后。
    返回: {old_row: new_row} 行号映射表（供更新 edits_map / blue_rows 使用）
    """
    sorted_a = sorted(group_a)
    sorted_b = sorted(group_b)
    size_a = len(sorted_a)
    size_b = len(sorted_b)
    start_a = sorted_a[0]

    # 安全检查：两组必须相邻
    if min(sorted_b) != max(sorted_a) + 1:
        logger.warning(
            f"行组不相邻，跳过交换: group_a结束={max(sorted_a)}, "
            f"group_b开始={min(sorted_b)}")
        return {}

    all_rows = sheet_data.findall(f'{{{ns}}}row')
    row_elems = {int(e.get('r')): e for e in all_rows}

    elems_a = [row_elems[r] for r in sorted_a if r in row_elems]
    elems_b = [row_elems[r] for r in sorted_b if r in row_elems]

    if not elems_a or not elems_b:
        return {}

    # group_b 上移至 start_a（负偏移）
    b_offset = start_a - sorted_b[0]
    for elem in elems_b:
        _shift_row_elem(elem, b_offset, ns)

    # group_a 下移 size_b（正偏移）
    for elem in elems_a:
        _shift_row_elem(elem, size_b, ns)

    # XML重排：正序将 group_b 各元素插到 group_a 第一个元素之前
    first_a = elems_a[0]
    for elem_b in elems_b:
        first_a.addprevious(elem_b)

    # 构建行号映射
    row_map = {}
    for i, r in enumerate(sorted_b):
        row_map[r] = start_a + i
    for i, r in enumerate(sorted_a):
        row_map[r] = start_a + size_b + i
    return row_map


def _xml_move_rows_before(sheet_data, rows_to_move, before_row, ns):
    """
    将 rows_to_move 中的行（必须全部在 before_row 之后）移到 before_row 之前。
    中间夹杂的未选中行跟在原有合同行后面。
    返回 {old_row: new_row} 映射。
    """
    rows_to_move = sorted(rows_to_move)
    if not rows_to_move or any(r <= before_row for r in rows_to_move):
        logger.warning(
            f"_xml_move_rows_before: rows_to_move不合法，跳过 "
            f"before_row={before_row}, rows={rows_to_move}")
        return {}

    max_r = max(rows_to_move)
    all_affected = list(range(before_row, max_r + 1))
    move_set = set(rows_to_move)
    stay_rows = [r for r in all_affected if r not in move_set]
    new_order = rows_to_move + stay_rows

    # 构建行号映射 old→new
    row_map = {old: before_row + i for i, old in enumerate(new_order)}

    # 收集受影响范围内的 XML 行元素（以当前 r 属性为键）
    row_elems = {}
    for elem in sheet_data.findall(f'{{{ns}}}row'):
        r = int(elem.get('r'))
        if before_row <= r <= max_r:
            row_elems[r] = elem

    if before_row not in row_elems:
        return {}

    before_elem = row_elems[before_row]

    # 物理重排：依次将 rows_to_move 各元素插到 before_elem 之前
    for old_r in rows_to_move:
        if old_r in row_elems:
            before_elem.addprevious(row_elems[old_r])

    # 更新行号和单元格引用
    for old_r, elem in row_elems.items():
        new_r = row_map[old_r]
        elem.set('r', str(new_r))
        for cell in elem.findall(f'{{{ns}}}c'):
            ref = cell.get('r', '')
            m = _CELL_REF_RE.match(ref)
            if m:
                cell.set('r', f'{m.group(1)}{new_r}')
            f_elem = cell.find(f'{{{ns}}}f')
            if f_elem is not None and f_elem.text:
                f_elem.text = _shift_formula_row(f_elem.text, old_r, new_r - old_r)

    return row_map


def _surgical_xlsx_write(src_path, dst_path, sheet_title, cell_edits,
                         blue_rows=None, row_splits=None, row_reorders=None):
    """
    复制src到dst，然后在dst中手术式修改指定单元格。

    cell_edits: [(row, col, value), ...] — 1-based行列号
    blue_rows: set of row numbers — 这些行整行蓝色填充
    row_splits: [{'row_group': [int], 'shipped_qtys': {row: qty},
                  'remain_qtys': {row: qty}, 'qty_col': int, ...}]
                — 拆行操作清单，克隆行组插入到原行组下方
    """
    shutil.copy2(src_path, dst_path)

    if not cell_edits:
        return

    ns = _XLSX_NS

    # 按行分组编辑
    edits_map = {}
    for row_num, col_num, value in cell_edits:
        ref = f'{_col_letter(col_num)}{row_num}'
        edits_map.setdefault(row_num, {})[ref] = value

    temp_path = dst_path + '.tmp'

    with zipfile.ZipFile(dst_path, 'r') as zin:
        sheet_path = _find_sheet_xml_path(zin, sheet_title)
        sheet_xml = zin.read(sheet_path)

        # 解析XML（保留空白和格式）
        parser = etree.XMLParser(remove_blank_text=False)
        root = etree.fromstring(sheet_xml, parser)

        sheet_data = root.find(f'{{{ns}}}sheetData')
        if sheet_data is None:
            logger.warning("sheetData未找到，跳过写入")
            return

        # --- 在styles.xml中添加蓝色填充 ---
        styles_xml = zin.read('xl/styles.xml')
        styles_root = etree.fromstring(styles_xml, parser)

        # 添加蓝色填充 #00B0F0
        fills_elem = styles_root.find(f'{{{ns}}}fills')
        fill_wrapper = etree.SubElement(fills_elem, f'{{{ns}}}fill')
        pf = etree.SubElement(fill_wrapper, f'{{{ns}}}patternFill')
        pf.set('patternType', 'solid')
        fg = etree.SubElement(pf, f'{{{ns}}}fgColor')
        fg.set('rgb', 'FF00B0F0')
        bg = etree.SubElement(pf, f'{{{ns}}}bgColor')
        bg.set('indexed', '64')
        blue_fill_id = str(len(fills_elem) - 1)
        fills_elem.set('count', str(len(fills_elem)))

        # 为每种原始样式创建蓝色填充克隆
        cell_xfs = styles_root.find(f'{{{ns}}}cellXfs')
        blue_clone_map = {}  # 原样式idx → 蓝色克隆idx

        def _get_blue_clone(orig_s, left_align=False):
            """获取原样式的蓝色填充版本，缓存避免重复创建"""
            cache_key = (orig_s, left_align)
            if cache_key in blue_clone_map:
                return blue_clone_map[cache_key]
            orig_xf = cell_xfs[int(orig_s)]
            new_xf = etree.SubElement(cell_xfs, f'{{{ns}}}xf')
            # 复制原样式的所有属性
            for attr, val in orig_xf.attrib.items():
                new_xf.set(attr, val)
            # 改fillId为蓝色
            new_xf.set('fillId', blue_fill_id)
            new_xf.set('applyFill', '1')
            # 左对齐（写入的内容）
            if left_align:
                new_xf.set('applyAlignment', '1')
                align = new_xf.find(f'{{{ns}}}alignment')
                if align is None:
                    align = etree.SubElement(new_xf, f'{{{ns}}}alignment')
                align.set('horizontal', 'left')
            new_idx = str(len(cell_xfs) - 1)
            blue_clone_map[cache_key] = new_idx
            return new_idx

        # 默认蓝色样式（无原始样式的单元格用）
        default_blue = _get_blue_clone('0')
        # 默认蓝色+左对齐样式（写入内容的单元格用）
        default_blue_left = _get_blue_clone('0', left_align=True)

        # --- 行组重排序（将选中行移到未选中有合同行之前）---
        if row_reorders:
            reorder_map = {}
            for rd in row_reorders:
                if rd.get('type') == 'move_before':
                    mapping = _xml_move_rows_before(
                        sheet_data, rd['rows_to_move'], rd['before_row'], ns)
                else:
                    mapping = _xml_swap_adjacent_groups(
                        sheet_data, rd['group_a'], rd['group_b'], ns)
                # 正确组合映射：reorder_map[k] → mapping[reorder_map[k]]
                all_keys = set(reorder_map) | set(mapping)
                composed = {}
                for orig_r in all_keys:
                    mid = reorder_map.get(orig_r, orig_r)
                    final = mapping.get(mid, mid)
                    if final != orig_r:
                        composed[orig_r] = final
                reorder_map = composed
            if reorder_map:
                # 更新 edits_map 行号
                new_edits = {}
                for row_num, cells in edits_map.items():
                    new_row = reorder_map.get(row_num, row_num)
                    new_cells = {}
                    for ref, val in cells.items():
                        m = _CELL_REF_RE.match(ref)
                        new_ref = f'{m.group(1)}{new_row}' if m else ref
                        new_cells[new_ref] = val
                    new_edits[new_row] = new_cells
                edits_map = new_edits
                # 更新 blue_rows 行号
                if blue_rows:
                    blue_rows = {reorder_map.get(r, r) for r in blue_rows}
                # 更新 row_splits 行号（重排序后行号变化，split必须用新行号）
                if row_splits:
                    for sp in row_splits:
                        sp['start_row'] = reorder_map.get(sp['start_row'], sp['start_row'])
                        sp['row_group'] = [reorder_map.get(r, r) for r in sp['row_group']]
                        sp['shipped_qtys'] = {reorder_map.get(r, r): q for r, q in sp['shipped_qtys'].items()}
                        sp['remain_qtys'] = {reorder_map.get(r, r): q for r, q in sp['remain_qtys'].items()}

        # --- 拆行处理（从下往上，避免行号偏移）---
        if row_splits:
            # 按start_row升序排列，用于计算行号偏移
            splits_asc = sorted(row_splits, key=lambda s: s['start_row'])
            # 从下往上执行XML拆行
            for sp in reversed(splits_asc):
                _xml_split_rows(sheet_data, sp, ns,
                                get_blue_func=_get_blue_clone)

            # 更新 sheet dimension
            total_added = sum(len(sp['row_group']) for sp in row_splits)
            dim = root.find(f'{{{ns}}}dimension')
            if dim is not None:
                ref = dim.get('ref', '')
                m = re.match(r'([A-Z]+\d+):([A-Z]+)(\d+)', ref)
                if m:
                    new_last = int(m.group(3)) + total_added
                    dim.set('ref', f'{m.group(1)}:{m.group(2)}{new_last}')

            # 修正cell_edits和blue_rows的行号偏移
            # 每个拆行在group最后一行后面插入group_size行，
            # 导致后续行号+group_size
            # 预缓存每个拆行组的最后一行和组大小
            split_boundaries = [
                (max(sp['row_group']), len(sp['row_group']))
                for sp in splits_asc
            ]

            def _adjust_row(orig_row):
                """计算拆行后的新行号"""
                offset = 0
                for last_in_group, group_size in split_boundaries:
                    if orig_row > last_in_group:
                        offset += group_size
                return orig_row + offset

            # 重建edits_map用偏移后的行号
            new_edits = {}
            for row_num, cells in edits_map.items():
                new_row = _adjust_row(row_num)
                new_cells = {}
                for ref, val in cells.items():
                    col_part = _CELL_REF_RE.match(ref)
                    if col_part:
                        new_ref = f'{col_part.group(1)}{new_row}'
                    else:
                        new_ref = ref
                    new_cells[new_ref] = val
                new_edits[new_row] = new_cells
            edits_map = new_edits

            # 重建blue_rows用偏移后的行号
            if blue_rows:
                blue_rows = {_adjust_row(r) for r in blue_rows}

        # --- 修改sheet数据 ---
        blue_set = blue_rows or set()
        target_rows = set(edits_map.keys()) | blue_set

        for row_elem in sheet_data.findall(f'{{{ns}}}row'):
            rn = int(row_elem.get('r'))
            if rn not in target_rows:
                continue

            pending = dict(edits_map.get(rn, {}))
            need_blue = rn in blue_set

            for cell_elem in row_elem.findall(f'{{{ns}}}c'):
                cell_ref = cell_elem.get('r')
                is_edited = cell_ref in pending
                # 蓝色填充整行（保留原字体/数字格式）
                if need_blue:
                    orig_s = cell_elem.get('s', '0')
                    # 被写入的单元格用左对齐样式
                    cell_elem.set('s', _get_blue_clone(orig_s, left_align=is_edited))
                # 写入数据
                if is_edited:
                    _write_cell_value(cell_elem, pending.pop(cell_ref), ns)

            # 新建不存在的单元格
            for cell_ref, value in pending.items():
                cell_elem = etree.SubElement(row_elem, f'{{{ns}}}c')
                cell_elem.set('r', cell_ref)
                if need_blue:
                    cell_elem.set('s', default_blue_left)
                _write_cell_value(cell_elem, value, ns)

        cell_xfs.set('count', str(len(cell_xfs)))
        modified_styles = etree.tostring(
            styles_root, xml_declaration=True, encoding='UTF-8',
            standalone=True)

        # 序列化
        modified_bytes = etree.tostring(
            root, xml_declaration=True, encoding='UTF-8', standalone=True)

        # 写入新ZIP（含修改后的styles.xml）
        with zipfile.ZipFile(temp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                if item.filename == sheet_path:
                    zout.writestr(item, modified_bytes)
                elif item.filename == 'xl/styles.xml':
                    zout.writestr(item, modified_styles)
                else:
                    zout.writestr(item, zin.read(item.filename))

    os.replace(temp_path, dst_path)


# ============================================================
# 库存错误报告
# ============================================================

def _generate_error_report(errors, output_dir):
    """生成库存错误报告Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "库存错误报告"

    header_font = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(patternType='solid', fgColor='C0392B')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    headers = ['合同号', '出货货号', '简货号', '出货数量', '可用库存', '缺口数量', '原因']
    col_widths = [18, 18, 14, 14, 14, 14, 32]

    for c, (title, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=c, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        if c <= 26:
            ws.column_dimensions[chr(64 + c)].width = width

    data_font = Font(name='Microsoft YaHei', size=10)
    red_font = Font(name='Microsoft YaHei', size=10, color='C0392B', bold=True)
    center_align = Alignment(horizontal='center')

    for i, err in enumerate(errors, 2):
        gap = err['ship_qty'] - err['available_qty']
        row_data = [
            err['contract'], err['item_no'], err.get('simple_no', ''),
            err['ship_qty'], err['available_qty'], gap, err['reason'],
        ]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.font = red_font if c == 6 else data_font
            cell.border = thin_border
            if c >= 4:
                cell.alignment = center_align

    report_path = os.path.join(output_dir, "库存错误报告.xlsx")
    wb.save(report_path)
    wb.close()
    return report_path


# ============================================================
# 读取接单表
# ============================================================

def detect_columns(ws, keywords):
    """动态检测表头列。额外检测'简货号'列。"""
    for r in range(1, min(11, ws.max_row + 1)):
        col_map = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val is None:
                continue
            s = str(val).strip().replace('\n', '')
            if '简货号' in s and '简货号' not in col_map:
                col_map['简货号'] = c
                continue
            for kw in keywords:
                if kw in s and kw not in col_map:
                    col_map[kw] = c
                    break
        if all(k in col_map for k in ['合同', '货号', '数量']):
            return r, col_map
    return None, None


# ============================================================
# 读取出货资料
# ============================================================

def _read_sheet_data(filepath):
    """读取Excel首个Sheet全部数据，支持 .xls、.xlsx 和 .et（WPS格式）"""
    ext = os.path.splitext(filepath)[1].lower()
    if ext == '.xls':
        import xlrd
        wb = xlrd.open_workbook(filepath)
        ws = wb.sheet_by_index(0)
        return [[ws.cell_value(i, j) for j in range(ws.ncols)]
                for i in range(ws.nrows)]
    if ext == '.xlsx':
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        data = [list(row) for row in ws.iter_rows(values_only=True)]
        wb.close()
        return data
    if ext == '.et':  # WPS私有格式：先尝试OOXML，再回退xlrd
        try:
            from io import BytesIO
            with open(filepath, 'rb') as _f:
                raw = BytesIO(_f.read())
            wb = load_workbook(raw, read_only=True, data_only=True)
            ws = wb.active
            data = [list(row) for row in ws.iter_rows(values_only=True)]
            wb.close()
            return data
        except Exception:
            try:
                import xlrd
                wb = xlrd.open_workbook(filepath)
                ws = wb.sheet_by_index(0)
                return [[ws.cell_value(i, j) for j in range(ws.ncols)]
                        for i in range(ws.nrows)]
            except Exception:
                return []
    return []


def _extract_file_date(rows):
    """从出货文件前5行提取出货日期（如 "3月9日" 或 "3月9出"），统一返回 "X月X日" """
    for i in range(min(5, len(rows))):
        for val in rows[i]:
            if not isinstance(val, str):
                continue
            m = _DATE_RE.search(val)
            if m:
                return m.group(1) + '日'
    return None


def _extract_shipping_method(rows, fname=''):
    """从出货文件G2单元格（行索引1，列索引6）提取出货方式，如"ZURU342车"/"盐田ZURU1045柜"。
    G2为空时从文件名提取兜底。"""
    if len(rows) > 1 and len(rows[1]) > 6:
        v = rows[1][6]
        if v is not None:
            s = str(v).strip()
            if s:
                return s
    # 兜底：从文件名提取出货方式（如"盐田313车"、"盐田1045柜"）
    if fname:
        m = re.search(r'(盐田|蛇口|南沙|黄埔)?\.?\s*(ZURU\s*)?\d+\s*(车|柜)', fname)
        if m:
            return m.group(0).strip()
    return ''


def read_shipment_folder(folder):
    """
    读取出货文件夹内所有Excel，筛选备注=华登/兴信。
    返回: [(合同, 货号, 数量, 日期, 件数), ...] 保留每条独立记录（不合并）
    """
    records = []

    if not os.path.isdir(folder):
        logger.error(f"文件夹不存在: {folder}")
        return records

    import hashlib
    seen_hashes = set()       # 用于跳过内容完全相同的重复文件
    cross_file_records = set()  # 跨文件去重：(合同, 货号, 数量, 件数, 日期)

    for fname in os.listdir(folder):
        if fname.startswith('~$'):
            continue
        if not fname.lower().endswith(('.xls', '.xlsx', '.et')):
            continue

        fpath = os.path.join(folder, fname)

        # 跳过内容重复的文件（如"盐田821柜71126 .xls"和"盐田821柜71126 (1).xls"）
        try:
            with open(fpath, 'rb') as _f:
                file_hash = hashlib.md5(_f.read()).hexdigest()
            if file_hash in seen_hashes:
                logger.info(f"  跳过重复文件(内容相同): {fname}")
                continue
            seen_hashes.add(file_hash)
        except Exception:
            pass
        try:
            rows = _read_sheet_data(fpath)
        except Exception as e:
            logger.error(f"读取失败 {fname}: {e}")
            continue

        if not rows:
            continue

        file_date = _extract_file_date(rows)
        shipping_method = _extract_shipping_method(rows, fname)
        if file_date:
            logger.info(f"  {fname}: 出货日期={file_date}")

        target = ['备注', '合同', '货号', '数量', '件数']
        header_idx = None
        col_map = {}
        for i in range(min(10, len(rows))):
            vals = [str(v).strip().replace('\n', '') if v else ''
                    for v in rows[i]]
            found = {}
            for kw in target:
                for j, v in enumerate(vals):
                    if kw in v and kw not in found:
                        found[kw] = j
                        break
            # 件数列可选，其余四列为必须
            if all(k in found for k in ['备注', '合同', '货号', '数量']):
                header_idx = i
                col_map = found
                break

        if header_idx is None:
            logger.warning(f"  {fname}: 未找到表头，跳过")
            continue

        def _get(row, idx):
            return row[idx] if idx < len(row) else None

        # 先收集本文件的所有记录（不做去重，保留同文件内重复行如25282两条3408）
        file_records = []
        for i in range(header_idx + 1, len(rows)):
            row = rows[i]
            factory = str(_get(row, col_map['备注']) or '').strip()
            if factory not in ('华登', '兴信'):
                continue

            contract = _normalize(_get(row, col_map['合同']))
            item_no = _normalize(_get(row, col_map['货号']))
            if not contract or not item_no:
                continue

            try:
                qty = int(float(_get(row, col_map['数量'])))
            except (ValueError, TypeError):
                continue
            if qty <= 0:
                continue

            # 提取件数（可选列）
            cases = 0
            if '件数' in col_map:
                try:
                    cases = int(float(_get(row, col_map['件数']) or 0))
                except (ValueError, TypeError):
                    cases = 0

            file_records.append(
                (contract, item_no, qty, file_date or '', cases, shipping_method))

        # 跨文件去重：同一(合同+货号+数量+件数+日期)如已从其他文件添加过则跳过
        # 同文件内的重复行（如25282两条相同qty）全部保留
        count = 0
        for rec in file_records:
            dedup_key = (rec[0], rec[1], rec[2], rec[4], rec[3])
            if dedup_key not in cross_file_records:
                records.append(rec)
                count += 1
        # 本文件所有记录的key加入跨文件seen集合（去重key不含文件内重复区分）
        cross_file_records.update(
            (r[0], r[1], r[2], r[4], r[3]) for r in file_records)

        logger.info(f"  {fname}: {count} 条有效出货记录")

    logger.info(f"出货记录合计: {len(records)} 条")
    return records


# ============================================================
# 核心处理：匹配出货 → 标记备注
# ============================================================

def process(order_path, shipment_folder, output_path=None,
            mark_date='', log_callback=None, amount_table_path=None):
    """
    出货标记主函数。

    参数:
        order_path: 接单表Excel路径
        shipment_folder: 出货资料文件夹路径
        output_path: 输出路径（默认 接单表_更新.xlsx）
        mark_date: 统一标记日期（如 "3月14日"），为空则用出货文件日期
        log_callback: 日志回调函数（用于GUI实时显示）

    返回: (output_path, stats) 成功 或 (None, error_msg) 失败
    """
    def log(msg):
        logger.info(msg)
        if log_callback:
            log_callback(msg)

    log("=" * 50)
    log("步骤1: 读取出货资料")
    log("=" * 50)

    records = read_shipment_folder(shipment_folder)
    if not records:
        log("没有有效的出货数据")
        return None, "没有有效的出货数据"

    # 统一日期：如果指定了mark_date，所有条目用同一个日期
    if mark_date:
        log(f"  统一备注日期: {mark_date}")

    for idx, (c, i, qty, rec_date, cases, _sm) in enumerate(records, 1):
        simple = _to_simple_key(i)
        if _is_set_item(i):
            tag = " [卡板单]"
        elif _extract_simple_no(i) in MIXED_MAP:
            tag = " [混装]"
        else:
            tag = ""
        d = mark_date or rec_date
        log(f"  [{idx}] 合同={c}, 货号={i}→简={simple}{tag}, 数量={qty}, 件数={cases}, 日期={d}")

    # ---- 读取接单表（只读模式，不修改原文件）----
    log("")
    log("=" * 50)
    log("步骤2: 读取接单表")
    log("=" * 50)

    try:
        wb = load_workbook(order_path, data_only=True)
    except Exception as e:
        log(f"无法打开接单表: {e}")
        return None, f"无法打开接单表: {e}"

    # 优先找"排期"工作表，找不到则逐个sheet检测表头
    ws = None
    target_cols = ['合同', '货号', '数量', '备注']
    _SCHEDULE_NAMES = ('排期',)
    for name in _SCHEDULE_NAMES:
        if name in wb.sheetnames:
            ws = wb[name]
            log(f"  定位到工作表: {name}")
            break
    if ws is None:
        # 没有名为"排期"的sheet，逐个尝试找有正确表头的
        for _ws in wb.worksheets:
            _hr, _cm = detect_columns(_ws, target_cols)
            if _hr is not None:
                ws = _ws
                log(f"  自动检测到工作表: {_ws.title}")
                break
    if ws is None:
        ws = wb.active
        log(f"  未找到排期工作表，使用活动表: {ws.title}")

    sheet_title = ws.title
    header_row, col_map = detect_columns(
        ws, target_cols)

    if header_row is None:
        log("接单表未找到有效表头（至少需要: 合同、货号、数量）")
        wb.close()
        return None, "接单表未找到有效表头"

    col_contract = col_map['合同']
    col_item = col_map['货号']
    col_qty = col_map['数量']
    col_beizhu = col_map.get('备注')
    col_simple = col_map.get('简货号', col_item)
    max_col = ws.max_column

    has_simple_col = '简货号' in col_map
    log(f"接单表: {ws.max_row} 行 × {max_col} 列, Sheet='{sheet_title}'")
    if has_simple_col:
        log(f"  ✓ 检测到简货号列: 第{col_simple}列")
    else:
        log(f"  ✗ 未检测到简货号列，将从货号列自动提取简货号")
    log(f"  表头行={header_row}, 合同列={col_contract}, "
        f"简货号列={col_simple}, 货号列={col_item}, "
        f"数量列={col_qty}, 备注列={col_beizhu}")

    # 出货金额表所需额外列（走货期/单价/金额/国家/车间）
    col_delivery = col_price = col_amount_usd = col_amount_hk = None
    col_country = col_factory = None
    if amount_table_path and header_row:
        for _c in range(1, max_col + 1):
            _v = ws.cell(row=header_row, column=_c).value
            if not _v:
                continue
            _s = str(_v).strip()
            if '走货期' in _s and col_delivery is None:
                col_delivery = _c
            elif '单价' in _s and 'USD' in _s and col_price is None:
                col_price = _c
            elif '金额' in _s and 'USD' in _s and col_amount_usd is None:
                col_amount_usd = _c
            elif '金额' in _s and 'HK' in _s and col_amount_hk is None:
                col_amount_hk = _c
            elif '出货国家' in _s and col_country is None:
                col_country = _c
            elif '生产车间' in _s and col_factory is None:
                col_factory = _c

    # 建立行索引: (合同, 简货号) → [row_idx, ...]
    row_index = {}
    aux_skipped = 0
    last_contract = ''
    for r in range(header_row + 1, ws.max_row + 1):
        rc = _normalize(ws.cell(row=r, column=col_contract).value)
        if rc:
            last_contract = rc
        effective_contract = rc or last_contract

        ri_raw = _normalize(ws.cell(row=r, column=col_simple).value)
        if not effective_contract or not ri_raw:
            continue

        if _is_auxiliary_row(ws, r, col_item, max_col):
            aux_skipped += 1
            continue

        full_item = _normalize(ws.cell(row=r, column=col_item).value)

        # SLB/SLD/SLT/SK行也参与索引，key带后缀以区分
        ri = _extract_simple_no(ri_raw)
        m = SET_SUFFIX_RE.match(full_item)
        if m:
            ri = ri + m.group(2).upper()
        row_index.setdefault((effective_contract, ri), []).append(r)

    row_to_simple_key = {r: sk for (rc, sk), rows in row_index.items() for r in rows}
    log(f"  行索引: {len(row_index)} 个合同+简货号组合")
    if aux_skipped:
        log(f"  跳过辅助行(收缩指商/PDQ): {aux_skipped} 行")

    # ---- 匹配出货 → 收集标记操作 ----
    log("")
    log("=" * 50)
    log("步骤3: 匹配出货数据")
    log("=" * 50)

    stats = {
        'processed': 0,
        'not_found': 0,
        'rows_marked': 0,
    }

    actions = []  # [(row_idx, date_str, needs_fill, contract), ...]
    row_splits = []  # 部分扣减时的拆行指令
    row_reorders = []  # needs_fill精确匹配时，将选中行移到有合同未选中行之前
    used_rows = set()  # 已匹配的行，防止同一行被重复匹配
    sub_rows_cache = {}  # {main_row: [sub_rows]}，避免 blue_rows 阶段重复扫描
    inventory_errors = []

    # 找到接单日期列（A列=1）
    col_order_date = 1

    # 收集每个合同的首行接单日期（用于补填SLB行）
    contract_order_dates = {}
    for r in range(header_row + 1, ws.max_row + 1):
        rc = _normalize(ws.cell(row=r, column=col_contract).value)
        if rc and rc not in contract_order_dates:
            od = ws.cell(row=r, column=col_order_date).value
            if od is not None:
                contract_order_dates[rc] = od

    merged_indices = set()  # 已被look-ahead合并处理的记录索引
    for rec_idx, (contract, item_no, ship_qty,
                  rec_date, ship_cases, _sm_rec) in enumerate(records):
        if rec_idx in merged_indices:
            continue
        simple_key = _to_simple_key(item_no)
        ship_date = mark_date or rec_date
        slb = _is_set_item(item_no)
        is_group = slb or (_extract_simple_no(item_no) in MIXED_MAP)

        candidates = row_index.get((contract, simple_key), [])
        available = _collect_available(
            ws, candidates, col_qty, col_beizhu, col_contract,
            col_item, max_col, is_slb=slb)
        # 过滤掉已匹配的行
        available = [(r, q, n) for r, q, n in available
                     if r not in used_rows]

        if is_group:
            tag = '[卡板单]' if slb else '[混装]'
        else:
            tag = ''
        log(f"  合同={contract}, 出货货号={item_no}, 简货号={simple_key}, "
            f"出货={ship_qty}, 件数={ship_cases}, 出货日期={ship_date}, "
            f"{tag + ' ' if tag else ''}"
            f"候选行={len(candidates)}, 可用行={len(available)}")

        if candidates and not available:
            _collect_available(
                ws, candidates, col_qty, col_beizhu, col_contract,
                col_item, max_col, is_slb=slb, log_func=log)

        if not available:
            if candidates:
                # 有候选行但都是子行（合同列为空）→ 正常跳过，不计入错误
                stats['sub_skipped'] = stats.get('sub_skipped', 0) + 1
                log(f"    - 子行跳过: 合同={contract}, "
                    f"货号={item_no} (候选{len(candidates)}行均为子行)")
            else:
                stats['not_found'] += 1
                inventory_errors.append({
                    'contract': contract, 'item_no': item_no,
                    'simple_no': simple_key, 'ship_qty': ship_qty,
                    'available_qty': 0,
                    'reason': '接单表未找到该合同+货号',
                })
                log(f"    ✗ 未找到匹配: 合同={contract}, "
                    f"出货货号={item_no}, 简货号={simple_key}")
            continue

        # 组合货号（SLB/SLD/SLT/SK 或混装 7153/7154/25257）
        if is_group:
            matched = False

            if slb:
                # ---- SLB卡板：按配对匹配（SLB行+产品行为一对） ----
                # 先收集所有SLB配对信息
                slb_pairs = []
                for main_r, main_qty, needs_fill in available:
                    pr, pq, aux = _find_slb_pair(
                        ws, main_r, col_item, col_contract, col_qty,
                        ws.max_row)
                    if pr is not None and pq > 0:
                        slb_pairs.append(
                            (main_r, main_qty, needs_fill, pr, pq, aux))
                    else:
                        log(f"      行{main_r}: 无配对产品行，跳过")

                # 第1轮：找精确匹配
                for (main_r, main_qty, needs_fill,
                     product_row, product_qty, aux_rows) in slb_pairs:
                    if product_qty == ship_qty:
                        fill_tag = (" (需补填合同)"
                                    if needs_fill else "")
                        pair_subs = [product_row] + aux_rows
                        actions.append(
                            (main_r, ship_date, needs_fill, contract))
                        used_rows.add(main_r)
                        sub_rows_cache[main_r] = pair_subs
                        stats['processed'] += 1
                        log(f"    ✓ 行{main_r}: SLB配对精确 "
                            f"产品行{product_row} qty="
                            f"{product_qty}={ship_qty}{fill_tag}")
                        # 若选中行是needs_fill（合同列为空），
                        # 且前面存在有合同的未选中行 → 记录行组交换
                        if needs_fill:
                            preceding = [
                                (m2, mq2, nf2, pr2, pq2, ax2)
                                for (m2, mq2, nf2, pr2, pq2, ax2) in slb_pairs
                                if m2 < main_r and not nf2
                            ]
                            swap_done = False
                            if preceding:
                                # 取最近的未选中有合同行组
                                imm = max(preceding, key=lambda x: x[0])
                                g_a = sorted(
                                    [imm[0], imm[3]] + list(imm[5]))
                                g_b = sorted(
                                    [main_r, product_row] + list(aux_rows))
                                if min(g_b) == max(g_a) + 1:
                                    row_reorders.append(
                                        {'group_a': g_a, 'group_b': g_b})
                                    log(f"      ↺ 行组重排: "
                                        f"{g_b}→{g_a}前")
                                    swap_done = True
                            if not swap_done:
                                # 兜底：同合同无先行行时，找同一视觉块内（100行内）
                                # 跨合同同简货号的先行未用行
                                BLOCK_PROXIMITY = 100
                                cross_rows = [
                                    r for (c2, sk2), rows2
                                    in row_index.items()
                                    if sk2 == simple_key
                                    for r in rows2
                                    if main_r - BLOCK_PROXIMITY <= r < main_r
                                    and r not in used_rows
                                ]
                                if cross_rows:
                                    # 优先使用同合同行作为目标（避免跳到品类边界）
                                    same_contract_rows = [
                                        r for (c2, sk2), rows2
                                        in row_index.items()
                                        if sk2 == simple_key
                                        and c2 == contract
                                        for r in rows2
                                        if main_r - BLOCK_PROXIMITY <= r < main_r
                                        and r not in used_rows
                                    ]
                                    target_before = (
                                        min(same_contract_rows)
                                        if same_contract_rows
                                        else min(cross_rows)
                                    )
                                    g_b = sorted(
                                        [main_r, product_row]
                                        + list(aux_rows))
                                    row_reorders.append({
                                        'type': 'move_before',
                                        'rows_to_move': g_b,
                                        'before_row': target_before,
                                    })
                                    log(f"      ↺ 行组重排(跨合同): "
                                        f"{g_b}→行{target_before}前")
                        matched = True
                        break

                # 第1.5轮：look-ahead合并 —— 连续同合同+同货号SLB记录
                # 逐条累加，检查累计后是否能精确匹配某个产品行
                if not matched:
                    look_idx = rec_idx + 1
                    combined_qty = ship_qty
                    combined_cases = ship_cases
                    found_combined = False
                    while look_idx < len(records):
                        nc, ni, nq, nd, ncs, *_ = records[look_idx]
                        if (nc == contract
                                and _to_simple_key(ni) == simple_key
                                and look_idx not in merged_indices):
                            combined_qty += nq
                            combined_cases += ncs
                            # 检查合并后是否精确匹配某个SLB配对
                            for (m_r, m_qty, nf,
                                 p_r, p_q, ax) in slb_pairs:
                                if p_q == combined_qty:
                                    fill_tag = (" (需补填合同)"
                                                if nf else "")
                                    pair_subs = [p_r] + ax
                                    actions.append(
                                        (m_r, ship_date, nf, contract))
                                    used_rows.add(m_r)
                                    sub_rows_cache[m_r] = pair_subs
                                    stats['processed'] += 1
                                    n_merged = look_idx - rec_idx
                                    log(f"    ✓ 行{m_r}: SLB合并精确 "
                                        f"合并{n_merged+1}条记录 "
                                        f"产品行{p_r} "
                                        f"qty={p_q}={combined_qty} "
                                        f"件数={combined_cases}"
                                        f"{fill_tag}")
                                    # 若选中行是needs_fill，检查前方有合同行 → 行组重排
                                    if nf:
                                        preceding = [
                                            (m2, mq2, nf2, pr2, pq2, ax2)
                                            for (m2, mq2, nf2, pr2,
                                                 pq2, ax2) in slb_pairs
                                            if m2 < m_r and not nf2
                                        ]
                                        if preceding:
                                            imm = max(preceding,
                                                      key=lambda x: x[0])
                                            g_a = sorted(
                                                [imm[0], imm[3]]
                                                + list(imm[5]))
                                            g_b = sorted(
                                                [m_r, p_r] + list(ax))
                                            if min(g_b) == max(g_a) + 1:
                                                row_reorders.append(
                                                    {'group_a': g_a,
                                                     'group_b': g_b})
                                                log(f"      ↺ 行组重排: "
                                                    f"{g_b}→{g_a}前")
                                    # 标记已合并的后续记录索引
                                    for mi in range(
                                            rec_idx + 1, look_idx + 1):
                                        merged_indices.add(mi)
                                    matched = True
                                    found_combined = True
                                    break
                            if found_combined:
                                break
                            look_idx += 1
                        else:
                            break

                # 第2轮：没有精确→找部分匹配
                if not matched:
                    for (main_r, main_qty, needs_fill,
                         product_row, product_qty,
                         aux_rows) in slb_pairs:
                        if product_qty > ship_qty:
                            fill_tag = (" (需补填合同)"
                                        if needs_fill else "")
                            # ship_cases是出货资料的件数列（总箱数），
                            # SLB主行存的是卡板数(main_qty)。
                            # 只有当件数≤主行卡板数时才认为是卡板数，
                            # 否则用比例计算，防止件数=540写入卡板数=3的行。
                            if ship_cases > 0 and ship_cases <= main_qty:
                                shipped_main = ship_cases
                            else:
                                shipped_main = round(
                                    main_qty * ship_qty / product_qty)
                            remain_main = main_qty - shipped_main

                            shipped_qtys = {
                                main_r: shipped_main,
                                product_row: ship_qty,
                            }
                            remain_qtys = {
                                main_r: remain_main,
                                product_row: product_qty - ship_qty,
                            }
                            ratio = ship_qty / product_qty
                            for ar in aux_rows:
                                try:
                                    aq = int(float(
                                        ws.cell(row=ar, column=col_qty
                                                ).value or 0))
                                except (ValueError, TypeError):
                                    aq = 0
                                s = round(aq * ratio)
                                shipped_qtys[ar] = s
                                remain_qtys[ar] = aq - s

                            pair_subs = [product_row] + aux_rows
                            row_group = [main_r] + pair_subs
                            row_splits.append({
                                'start_row': main_r,
                                'row_group': row_group,
                                'shipped_qtys': shipped_qtys,
                                'remain_qtys': remain_qtys,
                                'qty_col': col_qty,
                                'date_str': ship_date,
                                'needs_fill': needs_fill,
                                'contract': contract,
                            })
                            actions.append(
                                (main_r, ship_date, needs_fill,
                                 contract))
                            used_rows.add(main_r)
                            sub_rows_cache[main_r] = pair_subs
                            stats['processed'] += 1
                            log(f"    ✓ 行{main_r}: SLB配对部分 "
                                f"产品行{product_row} "
                                f"{product_qty}>出货{ship_qty}"
                                f", 件数={shipped_main}"
                                f"{fill_tag}")
                            matched = True
                            break
                        else:
                            log(f"      行{main_r}: 产品qty="
                                f"{product_qty}<出货"
                                f"{ship_qty}，跳过")

            else:
                # ---- 混装（7153/7154/25257）：保持原逻辑 ----
                for main_r, main_qty, needs_fill in available:
                    sub_rows = _find_sub_rows(
                        ws, main_r, col_contract, ws.max_row)
                    if not sub_rows:
                        continue
                    n_subs = len(sub_rows)
                    sub_qtys = _get_sub_qtys(ws, sub_rows, col_qty)
                    sub_total = sum(sub_qtys.values())
                    fill_tag = " (需补填合同)" if needs_fill else ""

                    if sub_total == ship_qty:
                        actions.append(
                            (main_r, ship_date, needs_fill, contract))
                        used_rows.add(main_r)
                        stats['processed'] += 1
                        sub_rows_cache[main_r] = sub_rows
                        log(f"    ✓ 行{main_r}: 组合精确 "
                            f"子行总{sub_total}={ship_qty}{fill_tag}")
                        matched = True
                        break

                    elif sub_total > ship_qty:
                        each = ship_qty // n_subs
                        rem = ship_qty - each * n_subs
                        if ship_cases > 0 and ship_cases <= main_qty:
                            shipped_main = ship_cases
                        else:
                            shipped_main = round(
                                main_qty * ship_qty / sub_total)
                        remain_main = main_qty - shipped_main
                        shipped_qtys = {main_r: shipped_main}
                        remain_qtys = {main_r: remain_main}
                        for i, sr in enumerate(sub_rows):
                            s = each + (1 if i < rem else 0)
                            shipped_qtys[sr] = s
                            remain_qtys[sr] = sub_qtys[sr] - s
                        row_splits.append({
                            'start_row': main_r,
                            'row_group': [main_r] + sub_rows,
                            'shipped_qtys': shipped_qtys,
                            'remain_qtys': remain_qtys,
                            'qty_col': col_qty,
                            'date_str': ship_date,
                            'needs_fill': needs_fill,
                            'contract': contract,
                        })
                        actions.append(
                            (main_r, ship_date, needs_fill, contract))
                        used_rows.add(main_r)
                        stats['processed'] += 1
                        sub_rows_cache[main_r] = sub_rows
                        log(f"    ✓ 行{main_r}: 组合部分 "
                            f"子行总{sub_total}>出货{ship_qty}, "
                            f"件数={shipped_main}{fill_tag}")
                        matched = True
                        break

                    else:
                        log(f"      行{main_r}: 子行总"
                            f"{sub_total}<出货{ship_qty}，跳过")

            if not matched:
                avail_qty = max((q for _, q, _ in available), default=0)
                stats['not_found'] += 1
                inventory_errors.append({
                    'contract': contract, 'item_no': item_no,
                    'simple_no': simple_key, 'ship_qty': ship_qty,
                    'available_qty': avail_qty,
                    'reason': '组合货号: 子行数量不足或无子行',
                })
                log(f"    ✗ 组合货号匹配失败: 合同={contract}, 货号={item_no}")
            continue

        # 普通货号：按数量优先级匹配: 精确 → 最接近够扣(拆行) → look-ahead合并精确 → 不够报错
        row_idx, row_qty, needs_fill, match_type = _select_match(
            available, ship_qty)

        if match_type is None:
            # 所有单行可用量均不够扣 → 先尝试look-ahead合并（同合同同货号多记录之和精确匹配某行）
            # 注意：同合同同货号的记录可能不连续（中间夹杂其他货号），用continue跳过
            combined_qty = ship_qty
            found_combined = False
            to_merge = []
            for li in range(rec_idx + 1, len(records)):
                if li in merged_indices:
                    continue
                nc, ni, nq, nd, ncs, *_ = records[li]
                if nc != contract or ni != item_no:
                    continue
                combined_qty += nq
                to_merge.append(li)
                for r2, q2, nf2 in available:
                    if q2 == combined_qty and r2 not in used_rows:
                        fill_tag = " (需补填合同)" if nf2 else ""
                        actions.append((r2, ship_date, nf2, contract))
                        used_rows.add(r2)
                        if not nf2:
                            all_exact_subs2 = _find_sub_rows(
                                ws, r2, col_contract, ws.max_row,
                                col_item=col_item)
                            r2_simple = _extract_simple_no(
                                _normalize(ws.cell(row=r2, column=col_item).value))
                            exact_subs2 = [
                                sr for sr in all_exact_subs2
                                if _extract_simple_no(
                                    _normalize(ws.cell(row=sr, column=col_item).value)
                                ) == r2_simple
                                or (contract, _extract_simple_no(
                                    _normalize(ws.cell(row=sr, column=col_item).value)
                                )) not in row_index
                            ]
                            used_rows.update(exact_subs2)
                            sub_rows_cache[r2] = exact_subs2
                        else:
                            sub_rows_cache[r2] = []
                        stats['processed'] += 1
                        for mi in to_merge:
                            merged_indices.add(mi)
                        n_merged = len(to_merge)
                        log(f"    ✓ 行{r2}: 普通合并精确 合并{n_merged+1}条记录 "
                            f"qty={q2}={combined_qty}{fill_tag}")
                        found_combined = True
                        break
                if found_combined:
                    break
            if found_combined:
                continue

            # 主行不够扣 → 兜底检查子行（合同列为空的候选行）
            # SLB类货号的子行在base key下（如77711SLB→子行在77711下）
            sub_match = _match_sub_rows(
                ws, candidates, col_qty, col_beizhu, col_contract,
                ship_qty, used_rows)
            if sub_match:
                sub_r, sub_qty, sub_type = sub_match
                if sub_type == 'exact':
                    actions.append((sub_r, ship_date, True, contract))
                    used_rows.add(sub_r)
                    stats['processed'] += 1
                    log(f"    ✓ 子行{sub_r}: 精确匹配 qty={sub_qty}"
                        f" (补填合同)")
                elif sub_type == 'partial':
                    # 子行部分扣减 → 拆行（遇到已用行停止，保证split group连续）
                    all_sub_subs = _find_sub_rows(
                        ws, sub_r, col_contract, ws.max_row,
                        col_item=col_item)
                    sub_subs = []
                    for sr in all_sub_subs:
                        if sr in used_rows:
                            break
                        sub_subs.append(sr)
                    sub_qtys = _get_sub_qtys(ws, sub_subs, col_qty)
                    ratio = ship_qty / sub_qty
                    row_group = [sub_r] + sub_subs
                    shipped = {sub_r: ship_qty}
                    remain = {sub_r: sub_qty - ship_qty}
                    for sr, sq in sub_qtys.items():
                        s = round(sq * ratio)
                        shipped[sr] = s
                        remain[sr] = sq - s
                    row_splits.append({
                        'start_row': sub_r,
                        'row_group': row_group,
                        'shipped_qtys': shipped,
                        'remain_qtys': remain,
                        'qty_col': col_qty,
                        'date_str': ship_date,
                        'needs_fill': True,
                        'contract': contract,
                    })
                    sub_rows_cache[sub_r] = sub_subs
                    actions.append((sub_r, ship_date, True, contract))
                    used_rows.add(sub_r)
                    stats['processed'] += 1
                    log(f"    ✓ 子行{sub_r}: 部分扣减"
                        f" 出货{ship_qty}/{sub_qty} (补填合同)")
            else:
                # 子行单独也不够 → 尝试组合匹配（主行+子行总量）
                group_match = _match_group(
                    ws, available, col_qty, col_contract, ship_qty,
                    used_rows, ws.max_row)
                if group_match:
                    grp_rows, grp_total, grp_type = group_match
                    if grp_type == 'group_exact':
                        # 整组精确匹配 → 标蓝整组
                        main_r = grp_rows[0]
                        actions.append(
                            (main_r, ship_date, needs_fill, contract))
                        used_rows.add(main_r)
                        stats['processed'] += 1
                        log(f"    ✓ 组合匹配: 行{main_r}+"
                            f"{len(grp_rows)-1}子行, "
                            f"总量{grp_total}={ship_qty}")
                    else:
                        # group_partial: 组总量>出货量，暂不拆组
                        main_r = grp_rows[0]
                        actions.append(
                            (main_r, ship_date, needs_fill, contract))
                        used_rows.add(main_r)
                        stats['processed'] += 1
                        log(f"    ✓ 组合匹配(有余): 行{main_r}+"
                            f"{len(grp_rows)-1}子行, "
                            f"总量{grp_total}>出货{ship_qty}")
                else:
                    max_avail = max(available, key=lambda x: x[1])
                    stats['not_found'] += 1
                    inventory_errors.append({
                        'contract': contract, 'item_no': item_no,
                        'simple_no': simple_key, 'ship_qty': ship_qty,
                        'available_qty': max_avail[1],
                        'reason': f'数量不足: 出货{ship_qty}, '
                                  f'最大可用{max_avail[1]}',
                    })
                    log(f"    ✗ 数量不足: 出货{ship_qty}, "
                        f"最大可用行{max_avail[0]}={max_avail[1]}")
            continue

        if match_type == 'exact':
            # 精确匹配 → 直接标蓝
            # 若匹配的是B=None行(needs_fill)，先检查后续记录是否会匹配B-filled主行。
            # 如果是，跳过此次B=None独立匹配，让主行先被处理，子行由主行组带走。
            if needs_fill:
                b_filled_avail = [(r2, q2) for r2, q2, n2 in available
                                  if not n2]
                if b_filled_avail:
                    future_covers_parent = False
                    for li in range(rec_idx + 1, len(records)):
                        if li in merged_indices:
                            continue
                        nc, ni, nq = records[li][0], records[li][1], records[li][2]
                        if nc != contract or _to_simple_key(ni) != simple_key:
                            break
                        if any(q2 == nq for _, q2 in b_filled_avail):
                            future_covers_parent = True
                            break
                    if future_covers_parent:
                        stats['sub_skipped'] = stats.get('sub_skipped', 0) + 1
                        log(f"    - B=None行{row_idx}跳过: "
                            f"等待后续记录匹配主行")
                        continue
                    else:
                        # 无后续记录精确匹配B-filled父行 → 将父行一并封锁，
                        # 防止后续记录对其做错误的partial split
                        used_rows.update(r2 for r2, _ in b_filled_avail)
                        log(f"    - 封锁B-filled父行: "
                            f"{[r2 for r2, _ in b_filled_avail]}")
            actions.append((row_idx, ship_date, needs_fill, contract))
            used_rows.add(row_idx)
            # B=None续行本身没有物理子行，不收集exact_subs（避免把同合同其他续行误消耗）
            # 只有B已填的主行才收集子行
            if not needs_fill:
                all_exact_subs = _find_sub_rows(ws, row_idx, col_contract, ws.max_row,
                                                col_item=col_item)
                exact_subs = [
                    sr for sr in all_exact_subs
                    if _extract_simple_no(
                        _normalize(ws.cell(row=sr, column=col_item).value)
                    ) == simple_key
                    or (contract, _extract_simple_no(
                        _normalize(ws.cell(row=sr, column=col_item).value)
                    )) not in row_index
                ]
                used_rows.update(exact_subs)
                sub_rows_cache[row_idx] = exact_subs
            else:
                sub_rows_cache[row_idx] = []
            stats['processed'] += 1
            fill_tag = " (需补填合同)" if needs_fill else ""
            log(f"    ✓ 行{row_idx}: 精确匹配 qty={row_qty}{fill_tag}")

        elif match_type == 'partial':
            # 拆行前先检查look-ahead合并：当前+后续同合同同货号记录之和若精确等于行总量，
            # 则按合并精确匹配处理，不拆行（如3000+1200=4200，2808+600=3408）
            combined_qty = ship_qty
            to_merge_pre = []
            found_pre_combine = False
            for li in range(rec_idx + 1, len(records)):
                if li in merged_indices:
                    continue
                nc, ni, nq = records[li][0], records[li][1], records[li][2]
                if nc != contract or ni != item_no:
                    continue  # 跳过其他合同/货号，继续找同合同同货号
                combined_qty += nq
                to_merge_pre.append(li)
                if combined_qty == row_qty:
                    fill_tag = " (需补填合同)" if needs_fill else ""
                    actions.append((row_idx, ship_date, needs_fill, contract))
                    used_rows.add(row_idx)
                    if not needs_fill:
                        all_pre_subs = _find_sub_rows(ws, row_idx, col_contract,
                                                      ws.max_row, col_item=col_item)
                        pre_subs = [
                            sr for sr in all_pre_subs
                            if _extract_simple_no(
                                _normalize(ws.cell(row=sr, column=col_item).value)
                            ) == simple_key
                            or (contract, _extract_simple_no(
                                _normalize(ws.cell(row=sr, column=col_item).value)
                            )) not in row_index
                        ]
                        used_rows.update(pre_subs)
                        sub_rows_cache[row_idx] = pre_subs
                    else:
                        sub_rows_cache[row_idx] = []
                    stats['processed'] += 1
                    for mi in to_merge_pre:
                        merged_indices.add(mi)
                    log(f"    ✓ 行{row_idx}: look-ahead合并(拆→精确) "
                        f"合并{len(to_merge_pre)+1}条记录 "
                        f"qty={row_qty}{fill_tag}")
                    found_pre_combine = True
                    break
                elif combined_qty > row_qty:
                    break  # 已超出，放弃合并
            if found_pre_combine:
                continue

            # B=None行不拆行：B=None行是占位行，由块逻辑统一补填合同和备注，不改行量
            row_contract_val = ws.cell(row=row_idx, column=col_contract).value
            if not str(row_contract_val or '').strip():
                continue

            # 部分扣减 → 需要拆行
            # 遇到已用行时停止（保证split group连续，_xml_split_rows要求连续块）
            all_sub_rows = _find_sub_rows(
                ws, row_idx, col_contract, ws.max_row, col_item=col_item)
            sub_rows = []
            for sr in all_sub_rows:
                if sr in used_rows:
                    break  # 遇到已用行停止，不跳过（保持连续性）
                sub_rows.append(sr)
            sub_qtys = _get_sub_qtys(ws, sub_rows, col_qty)

            # 计算拆分比例
            ratio = ship_qty / row_qty
            shipped_main = ship_qty
            remain_main = row_qty - ship_qty
            shipped_subs = {}
            remain_subs = {}
            for sr, sq in sub_qtys.items():
                s = round(sq * ratio)
                shipped_subs[sr] = s
                remain_subs[sr] = sq - s

            # 查找后续同合同+同货号记录是否精确覆盖余量
            # 跳过不同合同/货号，遇到第一个同合同同货号记录即判断（不链式）
            remainder_covered_by = None
            for li in range(rec_idx + 1, len(records)):
                if li in merged_indices:
                    continue
                nc, ni, nq = records[li][0], records[li][1], records[li][2]
                if nc != contract or ni != item_no:
                    continue  # 跳过不同合同/货号
                if nq == remain_main:
                    remainder_covered_by = li
                break  # 找到第一个同合同同货号记录即停止

            row_group = [row_idx] + sub_rows
            split = {
                'start_row': row_idx,
                'row_group': row_group,
                'shipped_qtys': {row_idx: shipped_main, **shipped_subs},
                'remain_qtys': {row_idx: remain_main, **remain_subs},
                'qty_col': col_qty,
                'date_str': ship_date,
                'needs_fill': needs_fill,
                'contract': contract,
                'beizhu_col': col_beizhu,
                'contract_col': col_contract,
                # 余量标记：后续记录精确覆盖余量时，克隆行也打日期
                'remainder_mark': ({
                    'date_str': mark_date or records[remainder_covered_by][3],
                    'needs_fill': needs_fill,
                    'contract': contract,
                } if remainder_covered_by is not None else None),
            }
            row_splits.append(split)

            if remainder_covered_by is not None:
                merged_indices.add(remainder_covered_by)
                stats['processed'] += 1
                log(f"    ✓ 余量={remain_main} 由记录{remainder_covered_by}覆盖"
                    f"(货号={records[remainder_covered_by][1]},"
                    f"qty={records[remainder_covered_by][2]})")

            # 出货部分标蓝；缓存sub_rows避免blue_rows阶段重新计算错误的子行集
            sub_rows_cache[row_idx] = sub_rows
            actions.append((row_idx, ship_date, needs_fill, contract))
            used_rows.add(row_idx)
            stats['processed'] += 1
            log(f"    ✓ 行{row_idx}: 部分扣减 出货{ship_qty}/{row_qty}, "
                f"子行{len(sub_rows)}个(排除已用{len(all_sub_rows)-len(sub_rows)}个), "
                f"拆行处理")

    # B=None 行首次匹配后处理：同(合同,简货号)中只有行号最小的保留needs_fill=True
    # 只追踪needs_fill=True的B=None行；B-filled行不参与（不影响B=None行的日期写入）
    _b_none_nf = {}  # (contract, simple_key) → 最小行号（仅B=None行）
    for row_idx, date_str, needs_fill, contract in actions:
        if not needs_fill:
            continue  # 只统计B=None行
        full_item_val = _normalize(ws.cell(row=row_idx, column=col_item).value)
        if _is_set_item(full_item_val):
            continue  # SLB等集套货号不参与此逻辑
        sk = row_to_simple_key.get(row_idx, '')
        key = (contract, sk)
        if key not in _b_none_nf or row_idx < _b_none_nf[key]:
            _b_none_nf[key] = row_idx
    revised_actions = []
    for row_idx, date_str, needs_fill, contract in actions:
        if needs_fill:
            # B=None续行也需要写备注日期（和主行一样标记出货）
            pass
        revised_actions.append((row_idx, date_str, needs_fill, contract))
    actions = revised_actions

    # 收集需要蓝色填充的行（标记行 + 其下方真正的辅助子行）
    # 使用 _find_blue_sub_rows 而非缓存：缓存包含同货号独立行，不应上色
    # 注意：date_str=None 的非首行同样需要蓝色填充，不可跳过
    blue_rows = set()
    for row_idx, date_str, needs_fill, contract in actions:
        blue_rows.add(row_idx)
        subs = _find_blue_sub_rows(
            ws, row_idx, col_contract, col_item, ws.max_row)
        blue_rows.update(subs)

    # ── 普通货号行组重排序后处理 ──────────────────────────────────────
    # 已匹配行（B-filled 或 B=None 均包含）需排在同 item 未匹配的 B-filled 行之前
    matched_row_set = {r for r, _, _, _ in actions}  # 真正进入actions的行
    _reorder_checked = set()
    for row_idx, date_str, needs_fill, contract in actions:
        row_item_val = _normalize(ws.cell(row=row_idx, column=col_item).value)
        if _is_set_item(row_item_val):  # SLB等集套货号在匹配阶段处理
            continue
        simple_key_val = _to_simple_key(row_item_val)
        reorder_key = (contract, simple_key_val)
        if reorder_key in _reorder_checked:
            continue
        _reorder_checked.add(reorder_key)

        candidates_for_key = row_index.get(reorder_key, [])
        # 找同一 (合同, 简货号) 中未被匹配的 B-filled 行
        # 注意：used_rows包含"封锁"的B-filled父行，它们未匹配但在used_rows里
        # 需要用matched_row_set而非used_rows，否则封锁行被错误排除
        # 排除备注已有值的行：已处理行（上次运行留有N日期）不应视为"未匹配"触发重排
        unmatched_b_filled = sorted([
            r for r in candidates_for_key
            if r not in matched_row_set
            and ws.cell(row=r, column=col_contract).value is not None
            and str(ws.cell(row=r, column=col_contract).value).strip()
            and (not col_beizhu
                 or not ws.cell(row=r, column=col_beizhu).value
                 or not str(ws.cell(row=r, column=col_beizhu).value).strip())
        ])
        if not unmatched_b_filled:
            continue
        first_unmatched = unmatched_b_filled[0]

        # 找所有在 first_unmatched 之后、已匹配的行（B-filled 和 B=None 均包含，含其子行）
        # 注意：只处理同一视觉块内的行（距离≤50），防止跨区块错位
        # 拆行的行（row_splits的start_row）不参与重排：拆行后clone会插到其后，重排会导致clone位置错乱
        split_start_rows = {sp['start_row'] for sp in row_splits}
        matched_after = []
        for r2, ds2, nf2, c2 in actions:
            if c2 != contract:
                continue
            r2_item = _normalize(ws.cell(row=r2, column=col_item).value)
            if _is_set_item(r2_item):
                continue
            if _to_simple_key(r2_item) != simple_key_val:
                continue
            if r2 <= first_unmatched:
                continue
            if r2 in split_start_rows:
                continue  # 拆行不参与重排，避免clone插入位置混乱
            matched_after.append(r2)
            matched_after.extend(sub_rows_cache.get(r2, []))

        matched_after = sorted(set(r for r in matched_after if r > first_unmatched))
        if not matched_after:
            continue

        # 距离过远（>50行）说明不在同一视觉块，跳过避免错位
        if min(matched_after) - first_unmatched > 50:
            continue

        row_reorders.append({
            'type': 'move_before',
            'rows_to_move': matched_after,
            'before_row': first_unmatched,
        })
        log(f"    ↺ 普通货号行组重排: 行{matched_after}→行{first_unmatched}前")

    # B=None块级重排序 + 合同补填
    # 触发方式：遍历所有B-filled行作为块头，扫描其下方B=None块
    # 比"遍历actions"更全面，能处理块内无匹配B=None行但有匹配B-filled行的情况（如77711GQ4行7015）
    # 自动识别原则：块内行必须在(合同,简货号)≥2行的组中，单行组自动排除
    _bnone_block_done = set()   # 已处理的块头行号，防重复
    _bnone_block_fills = []     # [(row, contract_val, order_date), ...] 需补填合同的行

    for block_head in range(header_row + 1, ws.max_row + 1):
        bv = ws.cell(row=block_head, column=col_contract).value
        if bv is None or not str(bv).strip():
            continue  # 只以B-filled行作为块头
        if block_head in _bnone_block_done:
            continue
        _bnone_block_done.add(block_head)
        block_contract_val = str(bv).strip()
        # 块头货号：只有与块头货号完全相同的B=None行才参与重排/补填
        # 排除同合同下其他货号的子行（混装子件、PDQ、收藏指南等）
        block_head_item = _normalize(ws.cell(row=block_head, column=col_item).value or '')

        # 找块内所有"多行组"的B=None排期行
        # 自动排除：单行组（SLB主行/配对产品行/辅助行等）、B-filled行、不在row_index的行
        # 新增排除：货号与块头不同的子行（混装子件、PDQ等同合同其他货号）
        block_rows = []
        for r in range(block_head + 1, ws.max_row + 1):
            inner_bv = ws.cell(row=r, column=col_contract).value
            if inner_bv is not None and str(inner_bv).strip():
                break  # 遇到下一个B-filled行，块结束
            r_item = _normalize(ws.cell(row=r, column=col_item).value or '')
            if r_item != block_head_item:
                continue  # 货号与块头不同，跳过（混装子件、附件等）
            sk_r = row_to_simple_key.get(r)
            if sk_r and len(row_index.get((block_contract_val, sk_r), [])) >= 2:
                block_rows.append(r)

        if not block_rows:
            continue

        # block_rows中所有行的B列本来就是None（循环只收集B=None行）
        # unmatched_for_reorder：未被look-ahead匹配的行（用于重排序判断）
        # 注意：即使行被look-ahead合并匹配进matched_row_set，B列仍是None，仍需补填合同
        matched_in_block = [r for r in block_rows if r in matched_row_set]
        unmatched_for_reorder = [r for r in block_rows if r not in matched_row_set]

        # 关键门控：块内货号组必须有已匹配记录（防止无关B=None行被误填合同）
        # 只要块内任意行所属组有匹配行，即确认此合同本次有出货记录
        group_has_match = any(
            any(gr in matched_row_set
                for gr in row_index.get((block_contract_val, row_to_simple_key.get(br, '')), []))
            for br in block_rows
        )
        if not group_has_match:
            continue

        # 确定需补填合同的第一行：
        # 有未匹配行 → 取第一个未匹配行（同时做重排序）
        # 全部已匹配（如look-ahead combine场景） → 取第一个B=None行
        if unmatched_for_reorder:
            first_to_fill = min(unmatched_for_reorder)
            matched_after_first = sorted([r for r in matched_in_block if r > first_to_fill])
            if matched_after_first:
                # SLB主行要带上配对子行一起移动
                rows_with_subs = []
                for mr in matched_after_first:
                    rows_with_subs.append(mr)
                    for sr in sub_rows_cache.get(mr, []):
                        if sr not in rows_with_subs:
                            rows_with_subs.append(sr)
                rows_with_subs.sort()
                row_reorders.append({
                    'type': 'move_before',
                    'rows_to_move': rows_with_subs,
                    'before_row': first_to_fill,
                })
                log(f"    ↺ B=None块重排: 行{rows_with_subs}→行{first_to_fill}前 (块头行{block_head})")
        else:
            # 块内全部已匹配（如look-ahead combine），B列仍为None，仍需补填合同
            first_to_fill = min(block_rows)

        od = contract_order_dates.get(block_contract_val)
        _bnone_block_fills.append((first_to_fill, block_contract_val, od))
        log(f"    + 行{first_to_fill}: 补填合同={block_contract_val}, 接单日期={od}")

    # ---- 收集出货金额表数据（wb关闭前读取，保留ws有效） ----
    amount_records = []
    if amount_table_path and actions:
        from collections import defaultdict
        # 建立 (合同, 简货号) → [(qty, shipping_method), ...] 队列
        # 排除已被look-ahead合并的记录（它们不产生独立action）
        _recs_by_key = defaultdict(list)
        for _idx, _rec in enumerate(records):
            if _idx in merged_indices:
                continue
            _c, _i, _q, _d, _cs, _sm = _rec
            _sk = _to_simple_key(_i)
            _recs_by_key[(_c, _sk)].append((_q, _sm))
        _recs_pos = defaultdict(int)

        split_start_rows = {sp['start_row'] for sp in row_splits}

        for row_idx, date_str, needs_fill, contract in actions:
            full_item_val = _normalize(ws.cell(row=row_idx, column=col_item).value)
            sk = row_to_simple_key.get(row_idx, _to_simple_key(full_item_val))
            key = (contract, sk)
            rec_list = _recs_by_key.get(key, [])
            pos = _recs_pos[key]
            if pos < len(rec_list):
                _ship_qty, _sm = rec_list[pos]
                _recs_pos[key] += 1
            else:
                _ship_qty, _sm = None, ''

            # 拆行用出货记录qty，其余用接单表qty（look-ahead合并已合计）
            if row_idx in split_start_rows:
                qty_for_amt = _ship_qty
            else:
                qty_for_amt = ws.cell(row=row_idx, column=col_qty).value

            def _cv(col):
                return ws.cell(row=row_idx, column=col).value if col else None

            _order_date_val = ws.cell(row=row_idx, column=col_order_date).value
            # 若该行无接单日期，从合同→接单日期字典补填（子行也能拿到所属合同的日期）
            if _order_date_val is None:
                _order_date_val = contract_order_dates.get(contract)
            amount_records.append({
                'contract':        contract,
                'item_simple':     sk,
                'item_full':       full_item_val,
                'ship_qty':        qty_for_amt,
                'shipping_method': _sm,
                'order_date':      _order_date_val,
                'is_main':         _order_date_val is not None,
                'delivery':        _cv(col_delivery),
                'price':           _cv(col_price),
                'amount_usd':      _cv(col_amount_usd),
                'amount_hk':       _cv(col_amount_hk),
                'country':         _cv(col_country),
                'factory':         _cv(col_factory),
            })

            # SLB卡板行：把缓存子行（产品行+辅助行）也追加进来
            # 普通行的子行有各自独立的出货记录，已经在actions里，不需要从缓存补
            if _is_set_item(full_item_val):
                for sub_r in sub_rows_cache.get(row_idx, []):
                    sub_item = _normalize(
                        ws.cell(row=sub_r, column=col_item).value)
                    sub_sk = _to_simple_key(sub_item)
                    def _scv(col, _r=sub_r):
                        return ws.cell(row=_r, column=col).value if col else None
                    amount_records.append({
                        'contract':        contract,
                        'item_simple':     sub_sk,
                        'item_full':       sub_item,
                        'ship_qty':        _scv(col_qty),
                        'shipping_method': '',
                        'order_date':      None,
                        'is_main':         False,
                        'delivery':        _scv(col_delivery),
                        'price':           _scv(col_price),
                        'amount_usd':      _scv(col_amount_usd),
                        'amount_hk':       _scv(col_amount_hk),
                        'country':         _scv(col_country),
                        'factory':         _scv(col_factory),
                    })

    # 关闭只读工作簿
    wb.close()

    # ---- XML手术式写入 ----
    log("")
    log("=" * 50)
    log("步骤4: 写入出货日期到备注列")
    log("=" * 50)

    if output_path is None:
        dir_name = os.path.dirname(order_path)
        output_path = os.path.join(dir_name, "接单表_更新.xlsx")

    if not col_beizhu:
        log("  ✗ 未检测到备注列，无法写入出货日期")
        shutil.copy2(order_path, output_path)
    elif not actions:
        log("  没有需要标记的行")
        shutil.copy2(order_path, output_path)
    else:
        cell_edits = []
        for row_idx, date_str, needs_fill, contract in actions:
            if not date_str:
                log(f"  跳过行{row_idx}: 出货文件无日期")
                continue
            # 写备注日期（文本格式）
            cell_edits.append((row_idx, col_beizhu, date_str))
            stats['rows_marked'] += 1
            log(f"  行{row_idx}: 备注 ← '{date_str}'")
            # SLB行需补填合同号和接单日期
            if needs_fill:
                cell_edits.append((row_idx, col_contract, contract))
                od = contract_order_dates.get(contract)
                if od is not None:
                    cell_edits.append((row_idx, col_order_date, od))
                log(f"  行{row_idx}: 补填合同={contract}, 接单日期={od}")

        # B=None块级重排序：补填合同号和接单日期（行号用原始行号，_surgical_xlsx_write会自动按reorder_map平移）
        for fill_row, fill_contract_val, fill_od in _bnone_block_fills:
            cell_edits.append((fill_row, col_contract, fill_contract_val))
            if fill_od is not None:
                cell_edits.append((fill_row, col_order_date, fill_od))
            log(f"  行{fill_row}: B=None块补填合同={fill_contract_val}, 接单日期={fill_od}")

        try:
            _surgical_xlsx_write(order_path, output_path,
                                 sheet_title, cell_edits, blue_rows,
                                 row_splits=row_splits,
                                 row_reorders=row_reorders)
            log(f"  ✓ XML手术式写入完成，格式100%保留")
        except Exception as e:
            log(f"  XML写入失败: {e}，回退到openpyxl保存")
            # 兜底：用openpyxl写（可能有格式损失）
            _fallback_openpyxl_write(
                order_path, output_path, cell_edits, log)

    log(f"保存成功: {output_path}")

    # ---- 写入出货金额表（可选功能） ----
    if amount_table_path and amount_records:
        log("")
        log("=" * 50)
        log("步骤5: 写入出货金额表")
        log("=" * 50)
        _write_amount_table(amount_table_path, amount_records, log)

    # ---- 库存错误报告 ----
    report_path = None
    if inventory_errors:
        try:
            report_dir = os.path.dirname(order_path)
            report_path = _generate_error_report(inventory_errors, report_dir)
            stats['error_report'] = report_path
            log("")
            log(f"库存错误报告: {report_path}")
            log(f"  共 {len(inventory_errors)} 条异常，请人工检查")
        except Exception as e:
            log(f"生成错误报告失败: {e}")

    log("")
    log("=" * 50)
    log("处理结果")
    log("=" * 50)
    log(f"  成功匹配: {stats['processed']} 组")
    log(f"  子行跳过: {stats.get('sub_skipped', 0)} 组")
    log(f"  未找到匹配: {stats['not_found']} 组")
    log(f"  标记出货行: {stats['rows_marked']} 行")
    if row_splits:
        log(f"  拆行处理: {len(row_splits)} 组")
    if inventory_errors:
        log(f"  *** 异常: {len(inventory_errors)} 条 → 见错误报告")

    return output_path, stats


def _fallback_openpyxl_write(order_path, output_path, cell_edits, log):
    """兜底：用openpyxl写入（格式可能有损失）"""
    wb = load_workbook(order_path)
    ws = wb.active
    for row_num, col_num, value in cell_edits:
        ws.cell(row=row_num, column=col_num).value = value
    wb.save(output_path)
    wb.close()
    log("  (兜底) openpyxl保存完成")


def _write_amount_table(amount_table_path, amount_records, log_func=None):
    """
    生成全新出货记录 Excel 文件，存放于与出货金额表同目录。
    不修改原有金额表任何内容。
    结构与接单表一致：只有主行（接单表中有接单日期的行）才写发票日期/合同/出货方式，
    且出货方式只写每组合同的第一个主行。
    """
    def log(msg):
        logger.info(msg)
        if log_func:
            log_func(msg)

    if not amount_records:
        return

    # ── 预筛重复（合同+完整货号组合去重）──
    seen_keys = set()
    new_records = []
    for rec in amount_records:
        key = (str(rec.get('contract') or '').strip(),
               str(rec.get('item_full') or '').strip())
        if key in seen_keys:
            log(f"  跳过重复 合同={key[0]} 货号={key[1]}")
        else:
            seen_keys.add(key)
            new_records.append(rec)

    if not new_records:
        log("  出货记录: 无新数据")
        return

    # ── 按合同分组：每组第一行写发票日期/合同/出货方式 ──
    # 先汇总每个合同的接单日期（取第一个有值的）和出货方式
    contract_order_date = {}
    contract_shipping   = {}
    for rec in new_records:
        c  = str(rec.get('contract') or '').strip()
        od = rec.get('order_date')
        sm = rec.get('shipping_method') or ''
        if od and c not in contract_order_date:
            contract_order_date[c] = od
        if sm and c not in contract_shipping:
            contract_shipping[c] = sm

    # ── 创建新工作簿 ──
    from openpyxl import Workbook
    from openpyxl.styles import Font as _Font, PatternFill as _PFill, Alignment as _Align, Border as _Border, Side as _Side

    wb = Workbook()
    ws = wb.active
    ws.title = '出货记录'

    # 表头
    HEADERS = ['发票日期', '合同', '简货号', '货号', '走货期',
               '数量', '单价USD', '金额USD', '金额HK', '出货方式', '国家', '车间']
    COL_WIDTHS = [12, 16, 10, 14, 10, 8, 10, 12, 12, 12, 10, 12]

    hdr_fill = _PFill(fill_type='solid', fgColor='BDD7EE')   # 淡蓝色表头
    hdr_font = _Font(bold=True, name='宋体', size=10)
    thin = _Side(style='thin', color='BFBFBF')
    border = _Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font   = hdr_font
        cell.fill   = hdr_fill
        cell.border = border
        cell.alignment = _Align(horizontal='center', vertical='center')
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 18

    # 数据行填充色：主行淡绿，子行白色
    main_fill = _PFill(fill_type='solid', fgColor='E2EFDA')
    sub_fill  = _PFill(fill_type='solid', fgColor='FFFFFF')
    data_font = _Font(name='宋体', size=10)

    contract_written = set()   # 已写过第一行的合同
    new_count = 0
    for rec in new_records:
        contract  = str(rec.get('contract') or '').strip()
        is_first  = contract not in contract_written   # 本合同的第一行
        if is_first:
            contract_written.add(contract)

        # SLB/SLD主行：单价和金额为0或None时置空（卡板数行没有单价）
        price_val = rec.get('price')
        usd_val   = rec.get('amount_usd')
        hk_val    = rec.get('amount_hk')
        full_item = str(rec.get('item_full') or '')
        if _is_set_item(full_item):
            # 卡板主行：如果单价为0则清空价格和金额列
            if not price_val or (isinstance(price_val, (int, float)) and price_val == 0):
                price_val = None
                usd_val   = None
                hk_val    = None

        row_data = [
            contract_order_date.get(contract) if is_first else None,   # 发票日期
            contract                          if is_first else None,   # 合同
            rec.get('item_simple'),
            rec.get('item_full'),
            rec.get('delivery'),
            rec.get('ship_qty'),
            price_val,
            usd_val,
            hk_val,
            contract_shipping.get(contract)   if is_first else None,   # 出货方式
            rec.get('country'),
            rec.get('factory'),
        ]

        r = ws.max_row + 1
        fill = main_fill if is_first else sub_fill
        # 列1=发票日期, 列5=走货期 需要显示为日期格式
        DATE_COLS = {1, 5}
        DATE_FMT  = 'm"月"d"日"'
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font   = data_font
            cell.fill   = fill
            cell.border = border
            cell.alignment = _Align(horizontal='left', vertical='center')
            if col_idx in DATE_COLS and val is not None:
                cell.number_format = DATE_FMT

        new_count += 1

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 输出到与金额表同目录，文件名加 _新增数据 后缀
    base      = os.path.splitext(os.path.basename(amount_table_path))[0]
    out_path  = os.path.join(os.path.dirname(amount_table_path),
                             base + '_新增数据.xlsx')
    try:
        wb.save(out_path)
        log(f"  出货记录新文件: 写入{new_count}行 → {out_path}")
    except Exception as e:
        log(f"  出货记录新文件保存失败: {e}")
    finally:
        wb.close()


# ============================================================
# 命令行入口
# ============================================================

if __name__ == '__main__':
    import sys

    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s %(message)s',
        datefmt='%H:%M:%S',
    )

    if len(sys.argv) < 3:
        print("用法: python shipment_processor.py <接单表.xlsx> <出货文件夹>")
        sys.exit(1)

    order_file = sys.argv[1]
    ship_folder = sys.argv[2]

    if not os.path.isfile(order_file):
        print(f"错误: 接单表不存在: {order_file}")
        sys.exit(1)

    if not os.path.isdir(ship_folder):
        print(f"错误: 出货文件夹不存在: {ship_folder}")
        sys.exit(1)

    out, info = process(order_file, ship_folder)
    if out:
        print(f"\n完成 → {out}")
    else:
        print(f"\n处理失败: {info}")
        sys.exit(1)
